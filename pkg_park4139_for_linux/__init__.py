# -*- coding: utf-8 -*-  # python 3.x 하위버전 호환을 위한코드
__author__ = 'PARK4139 : Jung Hoon Park'

import asyncio
import base64
import inspect
import json
import os.path
import platform
import random
import re
import secrets
import shutil
import string
import subprocess
import sys
import threading
import time
import traceback
import urllib.parse as urllib_parser
from datetime import datetime, timedelta, UTC
from enum import Enum
from functools import partial
from pathlib import Path
from typing import Literal, Optional, TypeVar
from uuid import uuid4, UUID

# import jwt  # pip install pyJWT
import keyboard
import mutagen
import numpy
import pandas as pd
import pyglet
import send2trash  # pip install send2trash
import toml
from colorama import Fore
from fastapi import HTTPException
from fastapi.middleware.cors import CORSMiddleware
from googletrans import Translator
from gtts import gTTS  # Google TTS 적용
from jose import jwt
# from moviepy.video.io.VideoFileClip import VideoFileClip
# from moviepy.editor import *
from mutagen.mp3 import MP3
from pydantic import BaseModel, field_validator
# from screeninfo import get_monitors
from sqlalchemy import Column, Integer, String, text as sqlalchecdmy_text, VARCHAR, select, DateTime
from sqlalchemy import create_engine
from sqlalchemy.exc import ResourceClosedError
from sqlalchemy.orm import Session
from sqlalchemy.orm import sessionmaker, declarative_base

# windows 기반
# import pyautogui
# from BlurWindow.blurWindow import GlobalBlur
# import BlurWindow.blurWindow
# import win32gui
# import win32process
# from PySide6 import QtCore, QtWidgets
# from PySide6.QtCore import QEvent
# from PySide6.QtCore import Qt, QTimer, QThread, Signal, QCoreApplication
# from PySide6.QtGui import QCursor
# from PySide6.QtGui import QScreen, QIcon, QShortcut, QKeySequence, QFontDatabase, QFont, QGuiApplication
# from PySide6.QtWidgets import QWidget, QApplication, QGridLayout, QLabel, QPushButton, QVBoxLayout, QLineEdit, QDialog, QScrollArea, QTextBrowser

# from venv import create
# from app.routes import index, auth
# ! import 에 주석은 import 백업임 지우지말자. 오름차순 정리를 하자. 백업했으면 ctrl alt o를 누르자


# from venv import create

# logger = logging.getLogger('park4139_test_logger')
# hdlr = logging.FileHandler('park4139_logger.log')
# hdlr.setFormatter(logging.Formatter('%(asctime)s %(levelname)s %(message)s'))
# logger.addHandler(hdlr)
# logger.setLevel(logging.INFO)


T = TypeVar('T')  # 타입 힌팅 설정


class StateManagementUtil:
    members = []  # 리스트에 저장, 런타임 중에만 저장이 유지됨, 앱종료 시 데이터 삭제
    todos = []  # 리스트에 저장, 런타임 중에만 저장이 유지됨, 앱종료 시 데이터 삭제
    PROJECT_DIRECTORY = str(Path(__file__).parent.parent.absolute())  # __init__ 해당파일로 되어 있기 때문에. 위치가 복잡하게 되어 버렸다. 이를 __file__ 로 기준을 잡아서 경로를 수정하였다. 빌드를 하면서 알게되었는데 상대경로의 사용은 필연적인데, 상대경로로 경로를 설정할때 기준이 되는 절대경로 하나는 반드시 필요한 것 같다.
    GIT_HUB_ADDRESS = "https://github.com/PARK4139"
    DIRECTORYIES_GATHERED = rf"E:\[noe] [8TB] [local]/`/directories_gathered"
    EMPTY_DIRECTORYIES = rf"E:/[noe] [8TB] [local]/`/$directories_empty"
    USELESS_DIRECTORIES = rf"E:/[noe] [8TB] [local]/`/$directories_useless"
    USERPROFILE = os.environ.get('USERPROFILE')  # 환경변수를 통해 데이터 보안 강화
    SERVICE_DIRECTORY = os.path.dirname(PROJECT_DIRECTORY)
    PARK4139_ARCHIVE_TOML = rf'{PROJECT_DIRECTORY}/park4139_archive.toml'
    LOCAL_PKG_CACHE_FILE = rf'{PROJECT_DIRECTORY}/pkg_park4139_for_linux/__pycache__/__init__.cpython-312.pyc'
    ICON_PNG = rf"{PROJECT_DIRECTORY}/$cache_png/icon.PNG"
    SUCCESS_LOG = rf'{PROJECT_DIRECTORY}/$cache_log/success.log'
    MACRO_LOG = rf'{PROJECT_DIRECTORY}/$cache_log/macro.log'
    DIRSYNC_LOG = rf'{PROJECT_DIRECTORY}/$cache_log/dirsync.log'
    LOG_DIRECTORY = rf'{PROJECT_DIRECTORY}/$cache_log'
    YT_DLP_CMD = rf"{PROJECT_DIRECTORY}/pkg_yt_dlp/yt-dlp.cmd"
    JQ_WIN64_EXE = rf"{PROJECT_DIRECTORY}/pkg_jq/jq-win64.exe"
    FFMPEG_EXE = rf"{PROJECT_DIRECTORY}/$cache_tools/dev_tools_exe/LosslessCut-win-x64/resources/ffmpeg.exe"
    DB_TOML = rf"{PROJECT_DIRECTORY}/$cache_database/db.toml"
    MUSIC_FOR_WORK = rf"{PROJECT_DIRECTORY}/$cache_work_for_music/PotPlayer64.dpl"
    RDP_82106_BAT = rf"{PROJECT_DIRECTORY}/$cache_prison/rdp-82106.bat"
    USELESS_FILES = rf"{PROJECT_DIRECTORY}/$cache_database/useless_file_names.txt"
    POP_SOUND_WAV = rf"{PROJECT_DIRECTORY}/$cache_sound/pop_sound.wav"
    JSON_DIRECTORY = rf'{PROJECT_DIRECTORY}/$cache_json'
    DB_JSON = rf"{PROJECT_DIRECTORY}/$cache_json/db.json"
    BOOKS_JSON = rf"{PROJECT_DIRECTORY}/$cache_json/books.json"
    USERS_JSON = rf"{PROJECT_DIRECTORY}/$cache_json/users.json"
    NAV_ITEMS_JSON = rf"{PROJECT_DIRECTORY}/$cache_json/nav_items.json"
    PYCHARM64_EXE: str
    is_do_routine_processing = False
    # db_template = {
    # 'park4139_archive_log_line_cnt': 0,
    # }
    UNDERLINE_PROMISED = '_' * 59  # 제목작성 시 앞부분에 적용되는 기준인데 pep8 최대권장길이(79)를 기준으로 20 자 내외로 제목작성을 작성
    INDENTATION_PROMISED = ' ' * 5
    time_s = 0.0
    time_e = 0.0
    biggest_targets = []  # 300 MB 이상 빽업대상
    smallest_targets = []

    # speak() 쓰레드 상태관리용
    # is_speak_as_async_running = False
    pyglet_player = None
    playing_sounds = []
    previous_mp3_length_used_in_speak_as_async = 0

    routines_promised = [
        '물 한컵',
        '선풍기로 방환기 1분 이상',
        '세수',
        '로션',
        '아침식사',
        "프로폴리스 한알",
        '물가글(혹시 구내염 있으시면 가글양치)',
        '양치 WITHOUT TOOTH PASTE',
        '양치 WITH TOOTH PASTE',
        '양치 WITH GARGLE',
        '물가글',
        '즐거운일 1개',
        '스트레칭 밴드 V 유지 런지 30개',
        '계단 스쿼트 30 개',
        '푸쉬업 30 개',
        '점심식사',
        '저녁식사',
        # '음악틀기',
    ]
    COUNT_OF_MAKE_ME_GO_TO_SLEEP = []
    WATCH_KEYWORDS_LIST = ['AAPL', 'TQQQ', 'QQQ', 'ARM','AVGO','AMZN', 'NVDA','NDAQ','SOXL','S&P 500', '삼성전자', '삼성전자우', '펄어비스', '엔씨소프트']
    DIRECTORY_XLS_TO_MERGE = rf'{PROJECT_DIRECTORY}/pkg_xlsx/to_merge'
    DIRECTORY_XLS_MERGED = rf'{PROJECT_DIRECTORY}/pkg_xlsx/merged'
    FILE_MERGED_EXCEL_XLSX = rf'{PROJECT_DIRECTORY}/pkg_xlsx/merged/엑셀단일시트병합결과물.xlsx'
    DIRECTORY_PKG_CLOUD = rf'{PROJECT_DIRECTORY}/pkg_cloud'
    DIRECTORY_PKG_XLSX = f"{PROJECT_DIRECTORY}/pkg_xlsx"
    DIRECTORY_CODING_TEST_RESULT = rf'{PROJECT_DIRECTORY}/pkg_coding_test_submit'
    DIRECTORY_PKG_HTML = f"{PROJECT_DIRECTORY}/pkg_html"
    DIRECTORY_PKG_PNG = f"{PROJECT_DIRECTORY}/pkg_png"

    is_op_mode = True

    @classmethod
    def __init__(cls, self):
        try:
            if FileSystemUtil.is_os_windows():
                os.system('chcp 65001 >nul')
                BusinessLogicUtil.PYCHARM64_EXE = rf'{os.environ.get("PyCharm Community Edition").replace(";", "")}/pycharm64.exe'
            if not FileSystemUtil.is_os_windows():
                cls.DIRECTORY_XLS_TO_MERGE = cls.DIRECTORY_XLS_TO_MERGE.replace("\\", "/")
                cls.DIRECTORY_XLS_MERGED = cls.DIRECTORY_XLS_MERGED.replace("\\", "/")
                cls.FILE_MERGED_EXCEL_XLSX = cls.FILE_MERGED_EXCEL_XLSX.replace("\\", "/")
                cls.DIRECTORY_PKG_CLOUD = cls.DIRECTORY_PKG_CLOUD.replace("\\", "/")

        except AttributeError:
            pass
        except Exception:
            DebuggingUtil.trouble_shoot("%%%FOO%%%")


class SecurityUtil:
    @staticmethod
    def aes_encrypt(key, plaintext):
        from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
        from cryptography.hazmat.primitives import padding
        from cryptography.hazmat.backends import default_backend
        """AES 암호화: 대칭키암호화 ECB 모드, 보안취약"""
        # 키 생성
        cipher = Cipher(algorithms.AES(key), modes.ECB(), backend=default_backend())
        encryptor = cipher.encryptor()

        # 패딩 추가
        padder = padding.PKCS7(128).padder()
        padded_data = padder.update(plaintext) + padder.finalize()

        # 암호화 수행
        ciphertext = encryptor.update(padded_data) + encryptor.finalize()
        return ciphertext

    @staticmethod
    def aes_decrypt(key, ciphertext):
        from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
        from cryptography.hazmat.primitives import padding
        from cryptography.hazmat.backends import default_backend
        """AES 복호화"""
        # 키 생성
        cipher = Cipher(algorithms.AES(key), modes.ECB(), backend=default_backend())
        decryptor = cipher.decryptor()

        # 복호화 수행
        decrypted_data = decryptor.update(ciphertext) + decryptor.finalize()

        # 언패딩
        unpadder = padding.PKCS7(128).unpadder()
        plaintext = unpadder.update(decrypted_data) + unpadder.finalize()
        return plaintext

    @staticmethod
    def encode_as_lzw_algorizm(plaintext: str):
        r"""
        LZW 알고리즘을 사용하여 문자열을 압축하는 함수
            def compress_string(original_string):
                 return compressed_string

        LZW 알고리즘을 사용하여 압축된 문자열을 해제하는 함수
            def decompress_string(compressed_string):
                 return original_string

        문자열 압축
            current_directory_state = compress_string(current_directory_state)

        문자열 압축해제
            current_directory_state = decompress_string(current_directory_state)
        """
        import zlib
        return zlib.compress(plaintext.encode('utf-8'))

    @staticmethod
    def decode_as_lzw_algorizm(encrypted_text):
        import zlib
        return zlib.decompress(encrypted_text.decode('utf-8'))

    @staticmethod
    def convent_bytes_to_str(target: bytes):
        return target.decode('utf-8')

    @staticmethod
    def convent_str_to_bytes(target: str):
        return target.encode('utf-8')

    @staticmethod
    def get_random_alphabet():
        # DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        return random.choice(string.ascii_letters)

    @staticmethod
    def get_random_bytes():
        return secrets.token_bytes(16)  # 16바이트의 보안적으로 안전한 난수 생성

    @staticmethod
    def get_random_int():
        # DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        return secrets.randbelow(100)  # 0부터 99까지의 난수 생성

    @staticmethod
    def get_random_name():
        return secrets.choice([
            "김민지", "이준호", "박지영", "정성민", "홍길동", "김지현", "박민수", "이서연", "정현우", "한지원", "박정훈"
        ])

    @staticmethod
    def get_random_phone_number():
        return secrets.choice([
            "010-1234-5678", "02-9876-5432", "031-111-2222", "051-555-7777", "070-1234-5678",
            # "+1-123-456-7890", "+44-20-1234-5678", "+81-3-1234-5678", "+86-10-1234-5678", "+61-2-1234-5678"
        ])

    @staticmethod
    def get_random_e_mail():
        return secrets.choice([
            'john.doe@gmail.com', 'jane.smith@yahoo.com', 'david.wilson@hotmail.com', 'emily.johnson@outlook.com', 'michael.brown@aol.com', 'sarah.jones@icloud.com', 'robert.davis@protonmail.com', 'lisa.thomas@zoho.com', 'william.martin@yandex.com', 'jessica.anderson@mail.com', 'matthew.harris@live.com', 'laura.miller@gmx.com', 'james.jackson@fastmail.com', 'olivia.rodriguez@inbox.com',
            'benjamin.carter@ymail.com', 'mia.walker@rocketmail.com', 'ethan.white@tutanota.com', 'ava.hall@rediffmail.com', 'alexander.lee@mailinator.com', 'sophia.green@protonmail.ch', 'jacob.adams@yandex.ru', 'emma.baker@outlook.com', 'daniel.cook@zoho.eu', 'madison.lopez@google.com', 'logan.morgan@yahoo.co.uk', 'chloe.roberts@icloud.com', 'jayden.kelly@mail.com', 'grace.bennett@fastmail.fm',
            'samuel.rogers@protonmail.com', 'harper.edwards@outlook.com'
        ])

    @staticmethod
    def get_random_address():
        import random
        def generate_random_address_usa():
            street_number = random.randint(1, 9999)
            street_name = random.choice(['Main Street', 'Park Avenue', 'Oak Lane', 'Maple Avenue', 'Cedar Road'])
            city = random.choice(['New York', 'Los Angeles', 'Chicago', 'Houston', 'Phoenix'])
            state = random.choice(['California', 'Texas', 'Florida', 'New York', 'Illinois'])
            zip_code = random.randint(10000, 99999)
            return f"{street_number} {street_name}, {city}, {state} {zip_code}"

        def generate_random_address_kor():
            street = random.choice(['서울특별시', '부산광역시', '대구광역시', '인천광역시', '광주광역시', '대전광역시', '울산광역시', '세종특별자치시', '경기도', '강원도', '충청북도', '충청남도', '전라북도', '전라남도', '경상북도', '경상남도', '제주특별자치도'])
            city = random.choice(['서초구', '강남구', '송파구', '강동구', '관악구', '강서구', '영등포구', '구로구', '금천구', '양천구', '마포구', '서대문구', '은평구', '동작구', '광진구', '성동구', '중랑구', '동대문구', '성북구', '강북구', '도봉구', '노원구', '중구', '종로구'])
            dong = random.choice(['반포동', '삼성동', '청담동', '논현동', '압구정동', '서초동', '잠실동', '천호동', '신림동', '구로동', '영등포동', '신도림동', '여의도동', '목동', '신정동', '신촌동', '홍대입구동', '이태원동', '성수동', '왕십리동'])
            street_number = random.randint(1, 200)
            building_name = random.choice(['아파트', '빌라', '주택', '오피스텔'])
            return f"{street} {city} {dong} {street_number}-{random.randint(1, 20)}, {building_name}"

        return random.choice([generate_random_address_kor() for _ in range(100)])

    @staticmethod
    def get_random_id(length_limit: int):
        characters = string.ascii_letters + string.digits
        return ''.join(random.choice(characters) for _ in range(length_limit))

    @staticmethod
    def get_random_pw(length_limit: int):
        characters = string.ascii_letters + string.digits
        return ''.join(random.choice(characters) for _ in range(length_limit))

    @staticmethod
    def get_random_korean(length_limit: int):
        result = ''
        for _ in range(length_limit):
            result += random.choice(string.printable)
        return result

    @staticmethod
    def get_random_date():
        # 현재 날짜 가져오기
        current_date = datetime.now()
        # 시작 날짜 설정 (예: 1970년 1월 1일)
        start_date = datetime(1970, 1, 1)
        # 현재 날짜와 시작 날짜 사이의 일수 계산
        days_diff = (current_date - start_date).days
        # 랜덤하게 선택된 일수 더하기
        random_date = start_date + timedelta(days=random.randint(0, days_diff))
        # "yyyy-yy-yy" 형식으로 포맷팅
        formatted_date = random_date.strftime("%Y-%y-%y")
        return formatted_date

    @staticmethod
    def get_random_special_character(length_limit: int):
        import string
        result = ''
        for _ in range(length_limit):
            result += random.choice(string.printable)
        return

    @staticmethod
    def get_random_user_trial_input_case():
        options = [
            SecurityUtil.get_random_name(),
            SecurityUtil.get_random_phone_number(),
            SecurityUtil.get_random_id(30),
            SecurityUtil.get_random_special_character(30),
            string.punctuation,
            SecurityUtil.get_random_special_character(30),
            # SecurityUtil.get_random_hex(),
            # SecurityUtil.get_random_bytes(),
            SecurityUtil.get_random_date(),
            "None",
            None,
            ""
        ]
        return random.choice(options)

    @staticmethod
    def get_random_hex():
        return secrets.token_hex(16)  # 16바이트의 난수를 16진수 문자열로 생성

    @staticmethod
    def get_random_urlsafe():
        return secrets.token_urlsafe(16)  # 16바이트의 난수를 URL-safe 문자열로 생성

    @staticmethod
    def is_sql_injection_safe_simply(data: str):
        sql_pattern = r"(SELECT|INSERT|UPDATE|DELETE|CREATE|ALTER|DROP|TRUNCATE|GRANT|REVOKE)"
        match = re.search(sql_pattern, data, re.IGNORECASE)
        if match:
            return False
        return True

    @staticmethod
    def get_jwt(data: dict, secret_key: str, expires_delta: Optional[timedelta] = None):
        payload = data.copy()
        if expires_delta:
            # 기준시간 UTC 아닌 KST 적용
            # expire = datetime.now(UTC) + expires_delta
            utc_now = datetime.now(UTC)
            kst_now = utc_now + timedelta(hours=9)
            expire = kst_now + expires_delta
        else:
            # expire = datetime.now(UTC) + timedelta(minutes=15)  # jwt 토큰 만료 시간 설정 (15분)
            # 기준시간 UTC 아닌 KST 적용
            utc_now = datetime.now(UTC)
            kst_now = utc_now + timedelta(minutes=15)
            expire = kst_now + expires_delta

        payload.update({"exp": expire})
        for key, value in payload.items():
            DebuggingUtil.print_magenta(rf'''{key} : {value}''')

        jwt_encoded = jwt.encode(payload, secret_key, algorithm="HS256")
        return jwt_encoded

    @staticmethod
    def decode_jwt_token(token):
        """jwt 유효성 디버깅"""
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")

        # JWT 토큰을 점(.)으로 분리하여 헤더, 페이로드, 서명 부분을 구분합니다.
        header, payload, signature = token.split('.')

        # 헤더와 페이로드를 Base64 디코딩하여 바이트 문자열로 변환합니다.
        decoded_header = base64.urlsafe_b64decode(header + '=' * (4 - len(header) % 4))
        decoded_payload = base64.urlsafe_b64decode(payload + '=' * (4 - len(payload) % 4))

        # 디코딩된 헤더와 페이로드를 문자열로 변환하여 출력합니다.
        decoded_header_str = decoded_header.decode('utf-8')
        decoded_payload_str = decoded_payload.decode('utf-8')

        DebuggingUtil.print_magenta(f"Decoded Header:{decoded_header_str}")
        DebuggingUtil.print_magenta(f"Decoded Payload:{decoded_payload_str}")

    @staticmethod
    def decoded_jwt(jwt_, secret_key, algorithms):
        return jwt.decode(jwt_, secret_key, algorithms=algorithms)


class TimeUtil:
    @staticmethod
    def get_time_as_(pattern: str):
        #  %Y-%m-%d %H:%M:%S 의 형태도 모두 쓸 수 있고, 특정해둔 키워드도 쓸 수 있다
        # 어느날 보니 이 함수가 의도한대로 동작이 안됬다, 테스트를 해서 고쳤는데, for 문이 다른 값을 반환하는 것이 문제였다.
        # 테스트를 하지 않고 넘어 갔던 것이 문제였다. 꼭 함수를 만들었을 때는 함수에 대한 기능 테스트 하자
        # time 은 ms 지원않한다. 즉 %f 사용불가...
        # datetime 은 ms 지원!.  즉 %f 가능
        import time
        now = time
        localtime = now.localtime()
        from datetime import datetime
        time_styles = {
            'now': str(now.time()),
            'yyyy': str(localtime.tm_year),
            'MM': str(localtime.tm_mon),
            'dd': str(localtime.tm_mday),
            'HH': str(localtime.tm_hour),
            'mm': str(localtime.tm_min),
            'ss': str(localtime.tm_sec),
            'weekday': str(localtime.tm_wday),
            'elapsed_days_from_jan_01': str(localtime.tm_yday),  # 이거 테스트해보자 네이밍으로는 금년 1월 1일 부터 오늘까지 몇일이 지났는가를 출력하는 것 같아 보인다
        }
        for key, value in time_styles.items():
            if key == pattern:
                return time_styles[pattern]
        return str(datetime.now().strftime(pattern))


class ColoramaUtil(Enum):
    RED = "RED"
    GREEN = "GREEN"
    BLACK = "BLACK"
    YELLOW = "YELLOW"
    BLUE = "BLUE"
    MAGENTA = "MAGENTA"
    CYAN = "CYAN"
    WHITE = "WHITE"
    RESET = "RESET"
    LIGHTBLACK_EX = "LIGHTBLACK_EX"
    LIGHTRED_EX = "LIGHTRED_EX"
    LIGHTGREEN_EX = "LIGHTGREEN_EX"
    LIGHTYELLOW_EX = "LIGHTYELLOW_EX"
    LIGHTBLUE_EX = "LIGHTBLUE_EX"
    LIGHTMAGENTA_EX = "LIGHTMAGENTA_EX"
    LIGHTCYAN_EX = "LIGHTCYAN_EX"
    LIGHTWHITE_EX = "LIGHTWHITE_EX"


class TestUtil:
    is_first_test_lap = True
    test_results = []

    @staticmethod
    # 이해한 게 문제가 있는지 상속에 대한 실험은 꼭 진행해보도록 하자.
    # 하지만 부모 class 로 만든 인스턴스에 영향이 없도록(값의 공유가 되지 않도록) 사용하는 것이 기본적인 방법으로
    # 심도있게 예측해야할 상황은 field 가 공유되도록 해야 될때 이다.
    # 예시로 Account 번호를 DB에 넣는 객체는 singletone으로 유지.
    # Parent().name    parent.name   Child().name   child.name
    def pause():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        BusinessLogicUtil.print_today_time_info()
        os.system('pause >nul')

    @staticmethod
    def measure_seconds_performance_nth(function):
        """시간성능 측정 데코레이터 코드"""
        # DebuggingUtil.print_ment_light_yellow(f"{StateManagementUtil.UNDERLINE_PROMISED}TEST TRYING..."), FAIL
        # def wrapper(*args, **kwargs):
        def wrapper(**kwargs):  # **kwargs, keyword argument, dictionary 로 parameter 를 받음. named parameter / positional parameter 를 받을 사용가능?
            # def wrapper(*args):# *args, arguments, tuple 로 parameter 를 받음.
            # def wrapper():
            # n = 10  # 테스트 루프 반복 횟수 설정
            # n = 3
            n = 1
            if TestUtil.is_first_test_lap:
                DebuggingUtil.print_ment_light_yellow(f"{StateManagementUtil.UNDERLINE_PROMISED}TEST READY")
                ment = rf"총 {n}번의 시간성능측정 테스트를 시도합니다"
                DebuggingUtil.print_ment_light_yellow(ment)
                TextToSpeechUtil.speak_ment(ment=ment, sleep_after_play=1)
                TestUtil.is_first_test_lap = False
                DebuggingUtil.print_ment_light_yellow(f"{StateManagementUtil.UNDERLINE_PROMISED}TEST START")
            seconds_performance_test_results = TestUtil.test_results
            import time
            time_s = time.time()
            try:
                DebuggingUtil.print_ment_light_yellow(f"{StateManagementUtil.UNDERLINE_PROMISED}TEST LOOP {len(seconds_performance_test_results)} STARTED")
                # function(*args, **kwargs)
                function(**kwargs)
                # function(*args)
                # function()
                DebuggingUtil.print_ment_light_yellow(f"{StateManagementUtil.UNDERLINE_PROMISED}TEST LOOP {len(seconds_performance_test_results)} ENDED")
                # n = n +1
            except:
                DebuggingUtil.trouble_shoot("%%%FOO%%%")
                traceback.print_exc(file=sys.stdout)
                # TestUtil.pause()
            time_e = time.time()
            mesured_seconds = time_e - time_s
            seconds_performance_test_results.append(f"{round(mesured_seconds, 2)}sec")
            if len(seconds_performance_test_results) == n:
                ment = rf"총 {len(seconds_performance_test_results)}번의 시간성능측정 테스트가 성공적으로 이루어졌습니다"
                TextToSpeechUtil.speak_ment(ment=ment, sleep_after_play=0.55)
                DebuggingUtil.print_ment_light_yellow(rf'seconds_performance_test_results = {seconds_performance_test_results}')
                DebuggingUtil.print_ment_light_yellow(rf'len(seconds_performance_test_results) : {len(seconds_performance_test_results)}')
                DebuggingUtil.print_ment_light_yellow(ment)
                DebuggingUtil.print_ment_light_yellow(f"{StateManagementUtil.UNDERLINE_PROMISED}TEST END")
                TestUtil.pause()

            # api 테스트 시 여러번 빠르게 api 요청 시 거절 되는 것 같음. 이를 해결하기 위한 코드. 불필요 시 주석, api 제공할 때 어떤방법을 써서 3 번까지만 허용시켜둔 것 같다.
            BusinessLogicUtil.sleep(milliseconds=random.randint(4000, 5100), print_mode=True)

        return wrapper

    @staticmethod
    def measure_seconds_performance_once(function):
        """시간성능 측정 데코레이터 코드"""

        # def wrapper(*args, **kwargs):
        def wrapper(**kwargs):  # **kwargs, keyword argument, dictionary 로 parameter 를 받음. named parameter / positional parameter 를 받을 사용가능?
            # def wrapper(*args):# *args, arguments, tuple 로 parameter 를 받음.
            # def wrapper():
            # DebuggingUtil.print_ment_via_colorama(f"{StateManagementUtil.UNDERLINE_PROMISED}TEST START", colorama_color=ColoramaUtil.LIGHTWHITE_EX)
            if TestUtil.is_first_test_lap:
                TestUtil.is_first_test_lap = False
            seconds_performance_test_results = TestUtil.test_results
            import time
            time_s = time.time()
            # function(*args, **kwargs)
            function(**kwargs)
            # function(*args)
            # function()
            time_e = time.time()
            mesured_seconds = time_e - time_s
            DebuggingUtil.print_ment_via_colorama(rf'mesured_seconds = {round(mesured_seconds, 2)}', colorama_color=ColoramaUtil.BLUE)
            DebuggingUtil.print_ment_via_colorama(f"{StateManagementUtil.UNDERLINE_PROMISED}TEST END", colorama_color=ColoramaUtil.LIGHTWHITE_EX)

        return wrapper

    @staticmethod
    def measure_milliseconds_performance(function):
        """시간성능 측정 코드"""

        def wrapper():
            # def wrapper(*args):# *args, arguments, tuple 로 parameter 를 받음.
            # def wrapper(*args, **kwargs):
            DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
            test_cycle_max_limit = 5
            milliseconds_performance_test_result = []
            import time
            time_s = time.time()
            function()
            # function(*args)
            # function(**kwargs)
            # function(*args, **kwargs)
            time_e = time.time()
            mesured_seconds = time_e - time_s
            DebuggingUtil.commentize(rf"시간성능측정 결과")
            milliseconds_performance_test_result.append(round(mesured_seconds * 1000, 5))
            print(rf'milliseconds_performance_test_result : {milliseconds_performance_test_result}')
            print(rf'type(milliseconds_performance_test_result) : {type(milliseconds_performance_test_result)}')
            print(rf'len(milliseconds_performance_test_result) : {len(milliseconds_performance_test_result)}')
            if len(milliseconds_performance_test_result) == test_cycle_max_limit:
                TextToSpeechUtil.speak_ments("시간성능측정이 성공 되었습니다", sleep_after_play=0.65)
                TestUtil.pause()

        return wrapper


# class UiUtil:
#     class CustomQdialog(QDialog):
#         """순환참조 를 회피하기 위해서 객체를 복제했다."""
#
#         def update_btn_text_clicked(self, text_of_clicked_button):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             self.btn_text_clicked = text_of_clicked_button
#
#         def update_btn_text_clicked_and_close(self, text_of_clicked_button):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             FileSystemUtil.play_wav_file(file_abspath=StateManagementUtil.POP_SOUND_WAV)
#             self.btn_text_clicked = text_of_clicked_button
#             self.close()
#
#         def set_shortcut(self, key_plus_key: str, function):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             self.shortcut = QShortcut(QKeySequence(key_plus_key), self)
#             self.shortcut.activated.connect(function)
#
#         # 스택 마우스 제작 참고용
#         # def eventFilter(self, obj, event):
#
#         #     # 드래그 가능한 레이블의 이벤트 구현
#         #     if isinstance(obj, DraggableLabel) and event.type() == QEvent.MouseButtonPress:
#         #         obj.mousePressEvent(event)
#         #         return True
#         #     elif isinstance(obj, DraggableLabel) and event.type() == QEvent.MouseMove:
#         #         obj.mouseMoveEvent(event)
#         #         return True
#         #
#         #     return super().eventFilter(obj, event)
#
#         def __init__(self, ment: str, btns=None, parent=None, is_input_box=False, input_box_text_default="", title="", auto_click_negative_btn_after_seconds: int = None, auto_click_positive_btn_after_seconds: int = None):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             super().__init__(parent)
#             FileSystemUtil.play_wav_file(file_abspath=StateManagementUtil.POP_SOUND_WAV)
#             self.input_text_default = input_box_text_default
#             self.context = ment
#             self.title = title
#             if self.title == "":
#                 self.setWindowTitle(".")
#             else:
#                 self.setWindowTitle(self.title)
#             self.is_input_text_box = is_input_box
#             self.setStyleSheet("background-color: rgba(0, 0, 0, 0); color: white;")  # 메시지박스창 스타일시트 적용 설정
#             ICON_PNG = StateManagementUtil.ICON_PNG
#             self.setWindowIcon(QIcon(ICON_PNG))  # 메시지박스창 아이콘 설정
#             self.setAttribute(Qt.WA_TranslucentBackground)  # 메시지박스창 블러 설정
#             # self.setWindowFlags(Qt.WindowType.FramelessWindowHint)  # 메시지박스창 최상단 프레임레스 설정
#             BlurWindow.blurWindow.GlobalBlur(self.winId(), hexColor=False, Acrylic=False, Dark=True, QWidget=self)
#             # self.setWindowFlags(Qt.WindowTitleHint | Qt.WindowCloseButtonHint)  # 메시지박스창 최대화 최소화 버튼숨기기
#             # self.setWindowFlag(Qt.WindowCloseButtonHint, False)  # 메시지박스창 닫기 버튼 disable
#             # self.setWindowFlag(Qt.WindowStaysOnTopHint)  # 모든 창들 중 가장 앞에 메시지박스창 위치하도록 설정
#             # self.setStyleSheet("background-color: rgba(0, 0, 0, 0)")  # 메시지박스창 스타일시트 적용 설정
#
#             self.display_width = FileSystemUtil.get_display_info()['width'],
#             self.display_height = FileSystemUtil.get_display_info()['height'],
#             # self.pop_up_window_width_default = int(int(self.display_width[0]) * 0.4)
#             # self.pop_up_window_height_default = int(int(self.display_height[0]) * 0.4)
#             # self.resize(self.pop_up_window_width_default, self.pop_up_window_height_default)
#             if len(ment.split("\n")) < 20:
#                 # self.resize(500, 250)
#                 self.resize(int(self.display_width[0] * 0.3), int(self.display_height[0] * 0.2))
#             else:
#                 # self.resize(int(500 * 2.5), 250 * 2)
#                 # self.resize(int(self.display_width[0] * 0.8), int(self.display_height[0] * 0.6))
#                 self.resize(int(self.display_width[0] * 0.4), int(self.display_height[0] * 0.6))
#                 # self.resize(int(self.display_width[0] * 0.9), int(self.display_height[0] * 0.6))
#
#             # 창을 화면의 상단가운데로 이동
#             screen = QGuiApplication.primaryScreen()
#             screen_geometry = screen.availableGeometry()
#             center_point = screen_geometry.center()
#             self.move(center_point.x() - self.width() / 2, screen_geometry.top() + (center_point.y() * 0.1))
#
#             # pyside6 표준버튼 설정
#             # self.setStandardButtons(QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel)
#             # self.setButtonText(QMessageBox.StandardButton.Ok, "확인", )
#             # self.setButton.setStyleSheet("background-color: rgba(0, 0, 0, 0);  color: white;")
#             # self.setButtonText(QMessageBox.StandardButton.Cancel, "취소")
#             self.btn_text_clicked = None
#             self.layout_horizontal = None
#             self.btn_positive = None
#             self.btn_negative = None
#             self.btn_third = None
#
#             # 스크롤지역 설정
#             self.scroll_area = None
#
#             # 버튼 설정
#             btn_to_copy = QPushButton(ment)
#             btn_to_copy.setStyleSheet("background-color: rgba(0, 0, 0, 0);  color: white;")
#             btn_to_copy.setFocusPolicy(Qt.NoFocus)
#             btn_to_copy.clicked.connect(self.copy_label_text_to_clipboard)
#             # Park4139.debug_as_cli(context=len(contents.strip()))
#             if 30 < len(ment.strip()):
#                 # btn_to_copy.setStyleSheet("text-align: left; font-size: 8px")
#                 btn_to_copy.setStyleSheet("text-align: left; font-size: 16px")
#             else:
#                 btn_to_copy.setStyleSheet("text-align: center; font-size: 20px")
#             btn_to_copy.setFont(Pyside6Util.get_font_for_pyside6(font_path=FontsUtil.GMARKETSANSTTFLIGHT_TTF))
#             self.set_shortcut(function=self.copy_label_text_to_clipboard, key_plus_key="alt+C")  # 단축키 설정
#
#             self.scroll_area = QScrollArea()
#             # self.scroll_area.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
#             self.scroll_area.setWidget(btn_to_copy)
#             self.scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
#             self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
#             self.scroll_area.setAlignment(Qt.AlignVCenter | Qt.AlignHCenter)
#             self.scroll_area.setStyleSheet(" border: none;")
#             # if 100 < len(contents):
#             # self.scroll_area.setStyleSheet("text-align: center; border: none;")
#             # else:
#             # self.scroll_area.setStyleSheet("text-align: left; border: none; ")
#             self.scroll_area.setFocusPolicy(Qt.NoFocus)  # focus 가 필요 없는 부분에는 이렇게 설정을 해두어야 원하는 곳으로 처음 창이 열렸을 때 focus 되도록 유도할 수 있었다
#             # self.scroll_area.setMaximumSize(3000, 1000)
#
#             # 입력 텍스트 박스 설정
#             if self.is_input_text_box == True:
#                 self.input_box = QLineEdit()
#                 self.input_box.setText(self.input_text_default)
#                 self.input_box.setFocusPolicy(Qt.StrongFocus)
#                 self.input_box.setFocus()  # 창이 나타났을 때 focus 가 다른데 말고 입력 텍스트 박스에 있도록
#
#             if btns != None:
#                 self.btns = btns
#             else:
#                 self.btns = [""]
#             self.btns_etc = [None] * (len(self.btns) - 3)
#             try:
#                 # 버튼 설정 / 단축키 설정
#                 if self.btns[0]:
#                     self.btn_positive = QPushButton(f'{self.btns[0]} (F1)')
#                     self.btn_positive.setStyleSheet("background-color: rgba(0, 0, 0, 0);  color: white;")
#                     self.btn_positive.clicked.connect(partial(self.update_btn_text_clicked_and_close, self.btns[0]))
#                     font_path = FontsUtil.GMARKETSANSTTFLIGHT_TTF
#                     # font_path = FONTS.RUBIKDOODLESHADOW_REGULAR_TTF
#                     self.btn_positive.setFont(Pyside6Util.get_font_for_pyside6(font_path=font_path))
#                     # self.set_shortcut(function=partial(self.update_btn_text_clicked_and_close, self.btns[0]), key_plus_key="alt+y")  # 단축키 설정
#                     self.set_shortcut(function=partial(self.update_btn_text_clicked_and_close, self.btns[0]), key_plus_key="F1")  # 단축키 설정
#                 if self.btns[1]:
#                     self.btn_negative = QPushButton(f'{self.btns[1]} (F2)')
#                     self.btn_negative.setStyleSheet("background-color: rgba(0, 0, 0, 0);  color: white;")
#                     self.btn_negative.clicked.connect(partial(self.update_btn_text_clicked_and_close, self.btns[1]))
#                     font_path = FontsUtil.GMARKETSANSTTFLIGHT_TTF
#                     self.btn_negative.setFont(Pyside6Util.get_font_for_pyside6(font_path=font_path))
#                     self.set_shortcut(function=partial(self.update_btn_text_clicked_and_close, self.btns[1]), key_plus_key="F2")  # 단축키 설정
#                 if self.btns[2]:
#                     self.btn_third = QPushButton(f'{self.btns[2]} (F3)')
#                     self.btn_third.setStyleSheet("background-color: rgba(0, 0, 0, 0);  color: white;")
#                     self.btn_third.clicked.connect(partial(self.update_btn_text_clicked_and_close, self.btns[2]))
#                     font_path = FontsUtil.GMARKETSANSTTFLIGHT_TTF
#                     self.btn_third.setFont(Pyside6Util.get_font_for_pyside6(font_path=font_path))
#                     self.set_shortcut(function=partial(self.update_btn_text_clicked_and_close, self.btns[2]), key_plus_key="F3")  # 단축키 설정
#                 try:
#
#                     for n, item in enumerate(iterable=self.btns[3:]):
#                         # 여기서 Unexpected type(s): (int, QPushButton | QPushButton) Possible type(s): (SupportsIndex, None) (slice, Iterable[None]) 이 메세지가
#                         # self.btns_etc[n]  에서 나타났는데  결국 원인을 찾지 못하였고, n 에 대한 노란밑줄을 무시처리했다.
#                         self.btns_etc[n] = QPushButton(f'{self.btns[3:][n]}')  # noqa
#                         self.btns_etc[n].setStyleSheet("background-color: rgba(0, 0, 0, 0);  color: white;")  # noqa
#                         self.btns_etc[n].clicked.connect(partial(self.update_btn_text_clicked_and_close, self.btns[3:][n]))  # noqa
#                         self.btns_etc[n].setFont(Pyside6Util.get_font_for_pyside6(font_path=FontsUtil.GMARKETSANSTTFLIGHT_TTF))  # noqa
#
#                 except:
#                     DebuggingUtil.trouble_shoot("%%%FOO%%%")
#                 print(rf'self.btns_etc : {self.btns_etc}')
#             except IndexError:
#                 # btns=[]의 index 를 초과하거나 부족한 경우
#                 pass
#             except TypeError:
#                 # btns=None 인 경우
#                 pass
#
#             # 타이머에 의한 자동 버튼 클릭 설정
#             if auto_click_negative_btn_after_seconds != None:
#                 self.seconds_remaining_until_auto_click = auto_click_negative_btn_after_seconds
#                 self.is_closing_timer = True
#             else:
#                 self.is_closing_timer = False
#             if auto_click_positive_btn_after_seconds != None:
#                 self.seconds_remaining_until_auto_click = auto_click_positive_btn_after_seconds
#                 self.is_starting_timer = True
#             else:
#                 self.is_starting_timer = False
#             # print(rf'self.is_closing_timer : {self.is_closing_timer}')
#             # print(rf'auto_closing_seconds : {auto_click_negative_btn_after_seconds}')
#             # print(rf'self.is_starting_timer : {self.is_starting_timer}')
#             # print(rf'auto_starting_seconds : {auto_click_positive_btn_after_seconds}')
#             if self.is_closing_timer == True and self.is_starting_timer == True:
#                 DebuggingUtil.print_magenta("closing_timer 와 starting_timer 는 동시에 설정 할 수 없습니다")
#                 sys.exit()
#
#             if self.is_closing_timer == True or self.is_starting_timer == True:
#                 self.auto_choice_timer = QTimer()
#                 if auto_click_negative_btn_after_seconds != None:
#                     self.auto_choice_timer.timeout.connect(self.countdown_and_click_negative_btn)  # noqa
#                 elif auto_click_positive_btn_after_seconds != None:
#                     self.auto_choice_timer.timeout.connect(self.countdown_and_click_positive_btn)  # noqa
#                 self.auto_choice_timer.start(1000)
#
#             # 레이아웃 설정
#             self.layout_horizontal = QGridLayout()
#
#             try:
#                 if self.btns[0]:
#                     self.layout_horizontal.addWidget(self.btn_positive, 0, 0)
#                 if self.btns[1]:
#                     self.layout_horizontal.addWidget(self.btn_negative, 0, 1)
#                 if self.btns[2]:
#                     self.layout_horizontal.addWidget(self.btn_third, 0, 2)
#                 try:
#                     for n, item in enumerate(iterable=self.btns_etc):
#                         self.layout_horizontal.addWidget(self.btns_etc[n], 0, n + 3)
#                 except IndexError:
#                     pass
#                 except:
#                     DebuggingUtil.trouble_shoot("%%%FOO%%%")
#
#
#             except IndexError:
#                 # btns=[]의 index 를 초과하거나 부족한 경우
#                 pass
#             except TypeError:
#                 # btns=None 인 경우
#                 pass
#             layout_vertical = QVBoxLayout()
#             # if len(contents.split("\n")) < 10:
#             #     layout_vertical.addWidget(self.label)
#             # else:
#             layout_vertical.addWidget(self.scroll_area)
#             if is_input_box:
#                 layout_vertical.addWidget(self.input_box)
#             layout_vertical.addLayout(self.layout_horizontal)
#             self.setLayout(layout_vertical)
#
#             self.bring_this_window()
#
#         # deprecating test
#         # def centerOnScreen(self):
#         #     # 현재 화면의 가운데 좌표를 계산
#         #     screen_geometry = QScreen().geometry()
#         #     x = (screen_geometry.width() - self.width()) // 2
#         #     y = (screen_geometry.height() - self.height()) // 2
#         #     self.move(x, y)
#
#         def copy_label_text_to_clipboard(self):
#             clipboard.copy(self.context)
#             FileSystemUtil.play_wav_file(StateManagementUtil.POP_SOUND_WAV)
#
#         def countdown_and_click_positive_btn(self):
#             mins, secs_remaining = MathUtil.get_minuites_and_remaining_secs(seconds=self.seconds_remaining_until_auto_click)
#             if secs_remaining != 0:
#                 secs_remaining_with_unit = f"{secs_remaining}초"
#             else:
#                 secs_remaining_with_unit = ""
#             if mins != 0:
#                 mins_with_unit = f"{mins}분"
#             else:
#                 mins_with_unit = ""
#             self.btn_positive.setText(f"{self.btns[0]} ({mins_with_unit} {secs_remaining_with_unit} 뒤 자동클릭)")
#
#             if self.seconds_remaining_until_auto_click == 0:
#                 if self.btns[0]:
#                     self.update_btn_text_clicked_and_close(self.btns[0])
#             self.seconds_remaining_until_auto_click = self.seconds_remaining_until_auto_click - 1
#
#         def countdown_and_click_negative_btn(self):
#             mins, secs_remaining = MathUtil.get_minuites_and_remaining_secs(seconds=self.seconds_remaining_until_auto_click)
#             if secs_remaining != 0:
#                 secs_remaining_with_unit = f"{secs_remaining}초"
#             else:
#                 secs_remaining_with_unit = ""
#             if mins != 0:
#                 mins_with_unit = f"{mins}분"
#             else:
#                 mins_with_unit = ""
#             self.btn_negative.setText(f"{self.btns[1]} ({mins_with_unit} {secs_remaining_with_unit} 뒤 자동클릭)")
#
#             if self.seconds_remaining_until_auto_click == 0:
#                 if self.btns[1]:
#                     self.update_btn_text_clicked_and_close(self.btns[1])
#             self.seconds_remaining_until_auto_click = self.seconds_remaining_until_auto_click - 1
#
#         def bring_this_window(self):
#             # self.activateWindow() 와 self.show() 의 위치는 서로 바뀌면 의도된대로 동작을 하지 않는다
#             self.show()
#             self.activateWindow()
#             # import win32gui
#             # active_window = win32gui.GetForegroundWindow()
#             # win32gui.SetForegroundWindow(active_window)
#
#     class CustomDialog():
#         def __init__(self, q_application: QApplication, q_wiget: QWidget, is_app_instance_mode=False, is_exec_mode=True):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             """
#             이 함수는 특별한 사용요구사항이 있습니다
#             pyside6 앱 내에서 해당 함수를 호출할때는 is_app_instance_mode 를 파라미터에 넣지 않고 쓰는 것을 default 로 디자인했습니다.
#             pyside6 앱 밖에서 해당 함수를 호출할때는 is_app_instance_mode 를 True 로 설정하고 쓰십시오.
#
#             그 사용요구사항이 생기게 된 이유는 다음과 같습니다
#             pyside6 는 app을 singletone으로 instance를 구현합니다. 즉, instance는 반드시 pyside6 app 내에서 하나여야 합니다.
#             pyside6의 QApplication()이 앱 내/외에서도 호출이 될 수 있도록 디자인했습니다.
#             앱 내에서 호출 시에는 is_app_instance_mode 파라미터를 따로 설정하지 않아도 되도록 디자인되어 있습니다.
#             앱 외에서 호출 시에는 is_app_instance_mode 파라미터를 True 로 설정해야 동작하도록 디자인되어 있습니다.
#             앱 외에서 호출 시에는 반드시 CustomDialog() 인스턴스로 close()를 호출해 pyside6 app instance 를 종료 해야 합니다.
#             """
#             # 테스트 해보니, QApplication가 먼저 생성된 뒤에 QDialog 는 instance 가 생성되어야 하는 것같다.
#             # 그래서 QDialog instance 를 CustomDialog 생성자의 파라미터로 못받는 것 같다.
#             # 다른 Qwiget 을 받을 생각 이었는데...
#             # 갑자기 든 생각인데. QApplication 도 같이 넘겨주면 되지 않을까?
#             # 된다! 내 생각이 맞은 것 같다. QApplicaion() QDialog() 인스턴스의 생성순서만 바꿨는데 동작하지 않는다. 아무튼 QDialog 를 instance 인자로 받을 수 있다.
#             FileSystemUtil.play_wav_file(file_abspath=StateManagementUtil.POP_SOUND_WAV)
#             self.is_app_instance_mode = is_app_instance_mode
#             if is_app_instance_mode == True:
#                 self.app_instance = q_application
#             # if is_exec_mode == False:
#             #     global dialog
#             dialog = q_wiget
#             if is_exec_mode == True:
#                 dialog.exec()  # noqa
#             self.btn_text_clicked = dialog.btn_text_clicked  # noqa
#
#         def close(self):
#             if self.is_app_instance_mode == True:
#                 if isinstance(self.app_instance, QApplication):
#                     self.app_instance.exec()
#             if self.is_app_instance_mode == True:
#                 # self.app_instance.quit()  # QApplication 인스턴스 제거시도 : fail
#                 self.app_instance.shutdown()  # QApplication 인스턴스 파괴시도 : success  # 성공요인은 app.shutdown()이 호출이 되면서 메모리를 해제까지 수행해주기 때문
#                 # del self.app_instance  # QApplication 인스턴스 파괴시도 : fail
#                 # self.app_instance.deleteLater()  # QApplication 인스턴스 파괴시도 : fail
#                 # self.app_instance = None  # QApplication 인스턴스 파괴시도 : fail
#                 # sys.exit()
#
#     class RpaProgramMainWindow(QWidget):
#         # class RpaProgramMainWindow(QDialog):
#         # class RpaProgramMainWindow(QMainWindow):
#         # def __init__(self, shared_obj): # shared_obj 는 창간 통신용 공유객체 이다. pyside6 app 의 상태관리
#         def __init__(self, q_application):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             FileSystemUtil.play_wav_file(file_abspath=StateManagementUtil.POP_SOUND_WAV)
#             super().__init__()
#             self.app = q_application
#
#             #  창간 통신 설정
#             # self.shared_obj = shared_obj  # 창간 통신용 객체
#             # self.shared_obj.rpa_program_main_window = self  # 메인 창(self)을 다른 창에서도 공유 할 수 있도록 설정 # 안됨.
#             # shared_obj.prompt_window = self.prompt_window
#
#             # deprecated test started at 2023 12 23 01 28
#             # self.prompt_window = None
#             # self.sub_window = None
#             # self.question = None
#
#             #  앱 전역 변수 설정
#             self.text = "text"
#             self.pw = "`"
#             self.id = "`"
#             # self.is_window_maximized = False
#             self.display_width = FileSystemUtil.get_display_info()['width'],
#             self.display_height = FileSystemUtil.get_display_info()['height'],
#             # self.display_width_default = int(int(self.display_width[0]) * 0.106)
#             self.display_width_default = int(int(self.display_width[0]) * 0.045)
#             self.display_height_default = int(int(self.display_height[0]) * 0.2)
#
#             #  메인창 설정
#             self.setWindowTitle('.')
#             icon_png = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_png\icon.PNG"
#             self.setWindowIcon(QIcon(icon_png))  # 메인창 아이콘 설정
#             # self.setAttribute(Qt.WA_TranslucentBackground) # 메인창 블러 설정
#             # self.setWindowFlags(Qt.WindowType.FramelessWindowHint) # 메인창 최상단 프레임레스 설정
#             GlobalBlur(self.winId(), hexColor=False, Acrylic=False, Dark=True, QWidget=self)
#             self.setWindowFlags(Qt.WindowTitleHint | Qt.WindowCloseButtonHint)  # 최대화 최소화 버튼 숨기기
#             self.setWindowFlags(Qt.WindowStaysOnTopHint)  # 모든 창 앞에 위치하도록 설정
#             self.setStyleSheet("background-color: rgba(0, 0, 0, 0)")
#             # self.setStyleSheet(pyqt6css.qss)
#             # self.setAttribute(Qt.WA_TranslucentBackground)
#             # blur(self.winId())
#             # self.scale = 1/1
#             # self.scale = 1/2
#             # self.scale = 1/3
#             # self.scale = 1 / 4
#             self.scale = 1 / 10
#             # self.setGeometry(0, 0, int(self.display_width_default * self.scale), int(self.display_height_default * self.scale))
#             # self.setGeometry(0, 0,self.display_width_default, self.display_height_default)
#             # self.resize(self.display_width_default, self.display_height_default)
#
#             # self.setGeometry(0, 0, int(self.display_width_default * self.scale), int(self.display_height[0]))
#             self.windows_size_mode = 1  # 창크기 모드 설정  #0 ~ 3
#             QCoreApplication.setAttribute(Qt.AA_EnableHighDpiScaling)  # 고해상도 스케일링을 활성화합니다.
#             self.screens = QGuiApplication.screens()  # 사용 가능한 모든 화면을 가져옵니다.
#
#             # inputbox 설정
#             # self.inputbox = QLineEdit(self)
#             # self.inputbox.setStyleSheet("color: rgba(255,255,255, 0.9);")
#             # self.inputbox.setText("0,0")
#             # self.inputbox.setFixedWidth(120)
#             # # self.inputbox.setStyleSheet("text-shadow: 1px 1px 7px rgba(1, 1, 1, 1);") #텍스트에 그림자 넣고 싶었는데 안된다.
#             # self.inputbox.textChanged.connect(self.inputbox_changed)
#             # self.inputbox.editingFinished.connect(self.inputbox_edit_finished)
#             # self.inputbox.returnPressed.connect(self.inputbox_return_pressed)
#
#             # 이벤트 가 많으면 프로그램이 늦어지는 것 같아보였다
#
#             # monitor_mouse_position 이벤트 설정
#             # self.listener = pynput.mouse.Listener(on_move=self.monitor_mouse_position) # 아주 빠르게 마우스 움직임 감지
#             # self.listener.start()
#
#             # monitor_mouse_position_per_second 이벤트 설정 (마우스 멈춤 감지 이벤트/ 5초간 마우스 중지 시 메인화면 자동 숨김 이벤트)
#             self.mouse_positions = []
#             self.previous_position = None
#             self.current_position = None
#             self.timer = QTimer()
#             self.timer.timeout.connect(self.monitor_mouse_position_per_second)  # noqa
#             self.timer.start(1000)
#
#             # console_blurred 내 단축키 설정
#             self.available_shortcut_list = {
#                 # 'Ctrl+F4' 는 설정하지 마는 것이 좋겠다.
#                 # 'KEYMAP': '?',
#                 'HIDE': 'Esc',  # GHOST MODE?
#                 'EXIT': 'Q',
#                 'TEST': '1',
#                 'BACK UP TARGET': 'S',
#                 'ASK AI QUESTION': 'A',
#                 'SHOOT SCREENSHOT FULL': 'F',
#                 'SHOOT SCREENSHOT CUSTOM': 'C',
#                 'SHOOT SCREENSHOT FOR RPA': '4',
#                 'ANI': '5',
#                 'rdp-82106': 'R+1',
#                 # 'DOWNLOAD YOUTUBE(wav)': 'S',
#                 # 'DOWNLOAD YOUTUBE(webm) ONLY SOUND': 'Alt+S',
#                 'WEATHER': 'W',
#                 'PROJECT DIRECTORY': 'P',
#                 # 'NO PASTE MEMO': 'M',
#                 'ENG TO KOR': 'K',
#                 'KOR TO ENG': 'E',
#                 'RECORD MACRO': 'M',
#                 'UP AND DOWN GAME': 'F1',
#                 "CLASSIFY SPECIAL FILES": "F2",
#                 "GATHER EMPTY DIRECTORY": "F3",
#                 "GATHER SPECIAL FILES": "F4",
#                 "GATHER USELESS FILES": "F5",
#                 "MERGE DIRECTORIES": "F6",
#                 "CONVERT MKV TO WAV": "F7",
#                 'DOWNLOAD YOUTUBE(webm)': 'F8',
#                 'DOWNLOAD YOUTUBE(webm)_': 'F9',
#                 'DOWNLOAD VIDEO FROM WEB1': 'Alt+F1',
#                 'DOWNLOAD VIDEO FROM WEB2': 'Alt+F2',
#                 'ROTATE WINDOW MODE': 'Alt+W',
#                 'SYSTEM REBOOT': 'Alt+9',
#                 'SYSTEM SHUTDOWN': 'Alt+]',
#                 'SYSTEM POWER SAVING MODE': 'Alt+[',
#                 # 'LOGIN': 'Alt+F7',
#                 'EMPTY RECYCLE BIN': 'Alt+E',
#                 # 'RUN CMD.EXE AS ADMIN': 'Alt+C',
#                 'NAVER MAP': 'Alt+N',
#
#                 # 버튼없는 SHORCUT
#                 # 'WEB CRAWL HREF': "`+F1", # fail
#                 'WEB CRAWL HREF': "Ctrl+F1",
#                 'WEB CRAWL YOUTUBE VIDEO TITLE AND URL': "Ctrl+F2",
#                 'WEB CRAWL YOUTUBE VIDEO PLAYLIST': "Ctrl+F3",
#                 'EXPLORER': 'O',
#                 'SYNC SERVICES': '`',
#                 # mkr
#             }
#             # self.set_shortcut('RUN CMD.EXE AS ADMIN', self.run_cmd_exe)
#             # self.set_shortcut('DOWNLOAD YOUTUBE(wav)', self.download_youtube_as_wav)
#             # self.set_shortcut('DOWNLOAD YOUTUBE(webm) ONLY SOUND', self.download_youtube_as_webm_only_sound)
#             self.set_shortcut('DOWNLOAD YOUTUBE(webm)_', self.should_i_download_youtube_as_webm_alt)
#             # self.set_shortcut('HIDE', self.hide_windows_of_this_app)
#             # self.set_shortcut('LOGIN', self.login)
#             # self.set_shortcut('NO PASTE MEMO', self.run_no_paste_memo)
#             self.set_shortcut('PROJECT DIRECTORY', self.open_project_directory)
#             self.set_shortcut('rdp-82106', self.should_i_connect_to_rdp1)
#             self.set_shortcut('TEST', self.should_i_start_test)
#             self.set_shortcut('ASK AI QUESTION', self.ask_something_to_ai)
#             self.set_shortcut('BACK UP TARGET', self.should_i_back_up_target)
#             self.set_shortcut('SHOOT SCREENSHOT FOR RPA', self.shoot_screenshot_for_rpa)
#             self.set_shortcut('DOWNLOAD YOUTUBE(webm)', self.should_i_download_youtube_as_webm)
#             self.set_shortcut('EMPTY RECYCLE BIN', self.should_i_empty_trash_can)
#             self.set_shortcut('ENG TO KOR', self.should_i_translate_eng_to_kor)
#             self.set_shortcut('KOR TO ENG', self.should_i_translate_kor_to_eng)
#             self.set_shortcut('HIDE', self.toogle_rpa_window)
#             self.set_shortcut('EXIT', self.should_i_exit_this_program)
#             self.set_shortcut('SYSTEM POWER SAVING MODE', self.should_i_enter_to_power_saving_mode)
#             self.set_shortcut('SYSTEM REBOOT', self.should_i_reboot_this_computer)
#             self.set_shortcut('SYSTEM SHUTDOWN', self.should_i_shutdown_this_computer)
#             self.set_shortcut('ROTATE WINDOW MODE', self.rotate_window_size_mode)
#             self.set_shortcut('SHOOT SCREENSHOT CUSTOM', self.shoot_screenshot_custom)
#             self.set_shortcut('SHOOT SCREENSHOT FULL', self.shoot_screenshot_full)
#             self.set_shortcut('NAVER MAP', self.should_i_find_direction_via_naver_map)
#             self.set_shortcut('WEATHER', self.show_weather_from_web)
#             self.set_shortcut('ANI', self.should_i_show_animation_information_from_web)
#             self.set_shortcut('DOWNLOAD VIDEO FROM WEB1', self.download_video_from_web1)
#             self.set_shortcut('DOWNLOAD VIDEO FROM WEB2', self.download_video_from_web2)
#             self.set_shortcut('RECORD MACRO', self.should_i_record_macro)
#             self.set_shortcut('UP AND DOWN GAME', self.run_up_and_down_game)
#             self.set_shortcut("CLASSIFY SPECIAL FILES", self.should_i_classify_special_files)
#             self.set_shortcut("GATHER EMPTY DIRECTORY", self.should_i_gather_empty_directory)
#             # self.set_shortcut("GATHER SPECIAL FILES", self.should_i_gather_special_files)
#             self.set_shortcut("GATHER USELESS FILES", self.should_i_gather_useless_files)
#             self.set_shortcut("MERGE DIRECTORIES", self.should_i_merge_directories)
#             self.set_shortcut("CONVERT MKV TO WAV", self.should_i_convert_mkv_to_wav)
#
#             # 버튼없는 SHORCUT
#             self.set_shortcut("WEB CRAWL HREF", self.should_i_crawl_a_tag_href)
#             self.set_shortcut("WEB CRAWL YOUTUBE VIDEO TITLE AND URL", self.should_i_crawl_youtube_video_title_and_url)
#             self.set_shortcut("WEB CRAWL YOUTUBE VIDEO PLAYLIST", self.should_i_crawl_youtube_playlist)
#             self.set_shortcut("EXPLORER", self.should_i_explorer)
#             self.set_shortcut("SYNC SERVICES", self.should_i_sync)
#
#             # 약속된 버튼명인 버튼 설정
#             self.btn_to_show_weather_from_web = self.get_btn(self.get_btn_name_promised('WEATHER'), self.show_weather_from_web)
#             #  self.btn_to_run_cmd_exe= self.get_btn(self.get_btn_name_promised('RUN CMD.EXE AS ADMIN'), self.run_cmd_exe)
#             #  self.btn_to_download_youtube_as_wav= self.get_btn(self.get_btn_name_promised('DOWNLOAD YOUTUBE(wav)'), self.download_youtube_as_wav)
#             #  self.btn_to_download_youtube_as_webm_only_sound= self.get_btn(self.get_btn_name_promised('DOWNLOAD YOUTUBE(webm) ONLY SOUND'), self.download_youtube_as_webm_only_sound)
#             self.btn_to_should_i_download_youtube_as_webm_alt = self.get_btn(self.get_btn_name_promised('DOWNLOAD YOUTUBE(webm)_'), self.should_i_download_youtube_as_webm_alt)
#             #  self.btn_to_hide_windows_of_this_app= self.get_btn(self.get_btn_name_promised('HIDE'), self.hide_windows_of_this_app)
#             #  self.btn_to_login= self.get_btn(self.get_btn_name_promised('LOGIN'), self.login)
#             #  self.btn_to_run_no_paste_memo= self.get_btn(self.get_btn_name_promised('NO PASTE MEMO'), self.run_no_paste_memo)
#             self.btn_to_open_project_directory = self.get_btn(self.get_btn_name_promised('PROJECT DIRECTORY'), self.open_project_directory)
#             self.btn_to_connect_to_rdp1 = self.get_btn(self.get_btn_name_promised('rdp-82106'), self.should_i_connect_to_rdp1)
#             self.btn_to_test = self.get_btn(self.get_btn_name_promised('TEST'), self.should_i_start_test)
#             #  self.btn_to_test2= self.get_btn(self.get_btn_name_promised('TEST 2'), self.test2)
#             self.btn_to_ask_something_to_ai = self.get_btn(self.get_btn_name_promised('ASK AI QUESTION'), self.ask_something_to_ai)
#             self.btn_to_back_up_target = self.get_btn(self.get_btn_name_promised('BACK UP TARGET'), self.should_i_back_up_target)
#             self.btn_to_shoot_screenshot_for_rpa = self.get_btn(self.get_btn_name_promised('SHOOT SCREENSHOT FOR RPA'), self.shoot_screenshot_for_rpa)
#             self.btn_to_should_i_download_youtube_as_webm = self.get_btn(self.get_btn_name_promised('DOWNLOAD YOUTUBE(webm)'), self.should_i_download_youtube_as_webm)
#             self.btn_to_should_i_empty_trash_can = self.get_btn(self.get_btn_name_promised('EMPTY RECYCLE BIN'), self.should_i_empty_trash_can)
#             self.btn_to_should_i_translate_eng_to_kor = self.get_btn(self.get_btn_name_promised('ENG TO KOR'), self.should_i_translate_eng_to_kor)
#             self.btn_to_should_i_translate_kor_to_eng = self.get_btn(self.get_btn_name_promised('KOR TO ENG'), self.should_i_translate_kor_to_eng)
#             self.btn_to_toogle_rpa_window = self.get_btn(self.get_btn_name_promised('HIDE'), self.toogle_rpa_window)
#             self.btn_to_should_i_exit_this_program = self.get_btn(self.get_btn_name_promised('EXIT'), self.should_i_exit_this_program)
#             self.btn_to_should_i_enter_to_power_saving_mode = self.get_btn(self.get_btn_name_promised('SYSTEM POWER SAVING MODE'), self.should_i_enter_to_power_saving_mode)
#             self.btn_to_should_i_reboot_this_computer = self.get_btn(self.get_btn_name_promised('SYSTEM REBOOT'), self.should_i_reboot_this_computer)
#             self.btn_to_should_i_shutdown_this_computer = self.get_btn(self.get_btn_name_promised('SYSTEM SHUTDOWN'), self.should_i_shutdown_this_computer)
#             self.btn_to_rotate_window_size_mode = self.get_btn(self.get_btn_name_promised('ROTATE WINDOW MODE'), self.rotate_window_size_mode)
#             self.btn_to_shoot_screenshot_custom = self.get_btn(self.get_btn_name_promised('SHOOT SCREENSHOT CUSTOM'), self.shoot_screenshot_custom)
#             self.btn_to_shoot_screenshot_full = self.get_btn(self.get_btn_name_promised('SHOOT SCREENSHOT FULL'), self.shoot_screenshot_full)
#             self.btn_to_should_i_find_direction_via_naver_map = self.get_btn(self.get_btn_name_promised('NAVER MAP'), self.should_i_find_direction_via_naver_map)
#             self.btn_to_should_i_show_animation_information_from_web = self.get_btn(self.get_btn_name_promised('ANI'), self.should_i_show_animation_information_from_web)
#             self.btn_to_download_video_from_web1 = self.get_btn(self.get_btn_name_promised('DOWNLOAD VIDEO FROM WEB1'), self.download_video_from_web1)
#             self.btn_to_download_video_from_web2 = self.get_btn(self.get_btn_name_promised('DOWNLOAD VIDEO FROM WEB2'), self.download_video_from_web2)
#             self.btn_to_record_macro = self.get_btn(self.get_btn_name_promised('RECORD MACRO'), self.should_i_record_macro)
#             self.btn_to_run_up_and_down_game = self.get_btn(self.get_btn_name_promised('UP AND DOWN GAME'), BusinessLogicUtil.run_up_and_down_game)
#             self.btn_to_classify_special_files = self.get_btn(self.get_btn_name_promised("CLASSIFY SPECIAL FILES"), BusinessLogicUtil.should_i_classify_special_files)
#             self.btn_to_gather_empty_directory = self.get_btn(self.get_btn_name_promised("GATHER EMPTY DIRECTORY"), BusinessLogicUtil.should_i_gather_empty_directory)
#             # self.btn_to_gather_special_files = self.get_btn(self.get_btn_name_promised("GATHER SPECIAL FILES"), BusinessLogicUtil.should_i_gather_special_files)
#             self.btn_to_gather_useless_files = self.get_btn(self.get_btn_name_promised("GATHER USELESS FILES"), BusinessLogicUtil.should_i_gather_useless_files)
#             self.btn_to_merge_directories = self.get_btn(self.get_btn_name_promised("MERGE DIRECTORIES"), BusinessLogicUtil.should_i_merge_directories)
#             self.btn_to_convert_mkv_to_wav = self.get_btn(self.get_btn_name_promised("CONVERT MKV TO WAV"), BusinessLogicUtil.should_i_convert_mkv_to_wav)
#
#             # 약속된 단축키명 버튼 설정
#             self.btn_to_show_weather_from_web_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('WEATHER'), function=self.show_weather_from_web)
#             #  self.btn_to_run_cmd_exe_only_shortcut_name= self.get_btn(btn_text_align = "right", btn_name  = self.get_shortcut_name_promised()('RUN CMD.EXE AS ADMIN'), function= self.run_cmd_exe)
#             #  self.btn_to_download_youtube_as_wav_only_shortcut_name= self.get_btn(btn_text_align = "right", btn_name  = self.get_shortcut_name_promised()('DOWNLOAD YOUTUBE(wav)'), function= self.download_youtube_as_wav)
#             #  self.btn_to_download_youtube_as_webm_only_sound_only_shortcut_name= self.get_btn(btn_text_align = "right", btn_name  = self.get_shortcut_name_promised()('DOWNLOAD YOUTUBE(webm) ONLY SOUND'), function= self.download_youtube_as_webm_only_sound)
#             self.btn_to_should_i_download_youtube_as_webm_alt_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('DOWNLOAD YOUTUBE(webm)_'), function=self.should_i_download_youtube_as_webm_alt)
#             #  self.btn_to_hide_windows_of_this_app_only_shortcut_name= self.get_btn(btn_text_align = "right", btn_name  = self.get_shortcut_name_promised()('HIDE'), function= self.hide_windows_of_this_app)
#             #  self.btn_to_login_only_shortcut_name= self.get_btn(btn_text_align = "right", btn_name  = self.get_shortcut_name_promised()('LOGIN'), function= self.login)
#             #  self.btn_to_run_no_paste_memo_only_shortcut_name= self.get_btn(btn_text_align = "right", btn_name  = self.get_shortcut_name_promised()('NO PASTE MEMO'), function= self.run_no_paste_memo)
#             self.btn_to_open_project_directory_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('PROJECT DIRECTORY'), function=self.open_project_directory)
#             self.btn_to_connect_to_rdp1_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('rdp-82106'), function=self.should_i_connect_to_rdp1)
#             self.btn_to_test_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('TEST'), function=self.should_i_start_test)
#             #  self.btn_to_test2_only_shortcut_name= self.get_btn(btn_text_align = "right", btn_name  = self.get_shortcut_name_promised()('TEST 2'), function= self.test2)
#             self.btn_to_ask_something_to_ai_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('ASK AI QUESTION'), function=self.ask_something_to_ai)
#             self.btn_to_back_up_target_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('BACK UP TARGET'), function=self.should_i_back_up_target)
#             self.btn_to_shoot_screenshot_for_rpa_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('SHOOT SCREENSHOT FOR RPA'), function=self.shoot_screenshot_for_rpa)
#             self.btn_to_should_i_download_youtube_as_webm_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('DOWNLOAD YOUTUBE(webm)'), function=self.should_i_download_youtube_as_webm)
#             self.btn_to_should_i_empty_trash_can_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('EMPTY RECYCLE BIN'), function=self.should_i_empty_trash_can)
#             self.btn_to_should_i_translate_eng_to_kor_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('ENG TO KOR'), function=self.should_i_translate_eng_to_kor)
#             self.btn_to_should_i_translate_kor_to_eng_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('KOR TO ENG'), function=self.should_i_translate_kor_to_eng)
#             self.btn_to_toogle_rpa_window_only_shorcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('HIDE'), function=self.toogle_rpa_window)
#             self.btn_to_should_i_exit_this_program_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('EXIT'), function=self.should_i_exit_this_program)
#             self.btn_to_should_i_enter_to_power_saving_mode_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('SYSTEM POWER SAVING MODE'), function=self.should_i_enter_to_power_saving_mode)
#             self.btn_to_should_i_reboot_this_computer_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('SYSTEM REBOOT'), function=self.should_i_reboot_this_computer)
#             self.btn_to_should_i_shutdown_this_computer_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('SYSTEM SHUTDOWN'), function=self.should_i_shutdown_this_computer)
#             self.btn_to_rotate_window_size_mode_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('ROTATE WINDOW MODE'), function=self.rotate_window_size_mode)
#             self.btn_to_shoot_screenshot_custom_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('SHOOT SCREENSHOT CUSTOM'), function=self.shoot_screenshot_custom)
#             self.btn_to_shoot_screenshot_full_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('SHOOT SCREENSHOT FULL'), function=self.shoot_screenshot_full)
#             self.btn_to_should_i_find_direction_via_naver_map_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('NAVER MAP'), function=self.should_i_find_direction_via_naver_map)
#             self.btn_to_should_i_show_animation_information_from_web_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('ANI'), function=self.should_i_show_animation_information_from_web)
#             self.btn_to_download_video_from_web1_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('DOWNLOAD VIDEO FROM WEB1'), function=self.download_video_from_web1)
#             self.btn_to_download_video_from_web2_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('DOWNLOAD VIDEO FROM WEB2'), function=self.download_video_from_web2)
#             self.btn_to_record_macro_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('RECORD MACRO'), function=self.should_i_record_macro)
#             self.btn_to_run_up_and_down_game_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised('UP AND DOWN GAME'), function=BusinessLogicUtil.run_up_and_down_game)
#             self.btn_to_classify_special_files_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised("CLASSIFY SPECIAL FILES"), function=BusinessLogicUtil.should_i_classify_special_files)
#             self.btn_to_gather_empty_directory_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised("GATHER EMPTY DIRECTORY"), function=BusinessLogicUtil.should_i_gather_empty_directory)
#             # self.btn_to_gather_special_files_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised("GATHER SPECIAL FILES"), function=BusinessLogicUtil.should_i_gather_special_files)
#             self.btn_to_gather_useless_files_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised("GATHER USELESS FILES"), function=BusinessLogicUtil.should_i_gather_useless_files)
#             self.btn_to_merge_directories_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised("MERGE DIRECTORIES"), function=BusinessLogicUtil.should_i_merge_directories)
#             self.btn_to_convert_mkv_to_wav_only_shortcut_name = self.get_btn(btn_text_align="right", btn_name=self.get_shortcut_name_promised("CONVERT MKV TO WAV"), function=BusinessLogicUtil.should_i_convert_mkv_to_wav)
#
#             # view = QWebEngineView()
#             # view.load(QUrl("https://qt-project.org/"))
#             # view.resize(1024, 750)
#             # view.show()
#
#             # pyside6 에서는 제거되었다고 함...
#             # url = "https://google.com"
#             # browser = QWebEngineView()
#             # browser.load(QUrl(url))
#
#             # 레이블로 설정 mkmkmkmk
#             html = "<html><body><h1 style='color: white;'>Hello, World!</h1></body></html>"
#             text_browser = QTextBrowser()
#             text_browser.setHtml(html)
#             text_browser.resize(800, 600)
#
#             html = "<html><body><h1 style='color: white;'>Hello, World!</h1></body></html>"
#             text_browser2 = QTextBrowser()
#             text_browser2.setHtml(html)
#             text_browser2.resize(800, 600)
#
#             html = "<html><body><h1 style='color: white;'>Hello, World!</h1></body></html>"
#             text_browser3 = QTextBrowser()
#             text_browser3.setHtml(html)
#             text_browser3.resize(800, 600)
#
#             html = "<html><body><h1 style='color: white;'>Hello, World!</h1></body></html>"
#             text_browser4 = QTextBrowser()
#             text_browser4.setHtml(html)
#             # text_browser4.setStyleSheet("background-color: rgba(0, 0, 0, 0); color: white;")  # 메시지박스창 스타일시트 적용 설정
#             text_browser4.resize(800, 600)
#
#             btns = [
#                 [text_browser, text_browser2],
#                 [text_browser3, text_browser4],
#                 [self.btn_to_should_i_exit_this_program, self.btn_to_should_i_exit_this_program_only_shortcut_name],
#                 [self.btn_to_toogle_rpa_window, self.btn_to_toogle_rpa_window_only_shorcut_name],
#                 [self.btn_to_show_weather_from_web, self.btn_to_show_weather_from_web_only_shortcut_name],
#                 [self.btn_to_should_i_download_youtube_as_webm_alt, self.btn_to_should_i_download_youtube_as_webm_alt_only_shortcut_name],
#                 [self.btn_to_should_i_download_youtube_as_webm, self.btn_to_should_i_download_youtube_as_webm_only_shortcut_name],
#                 [self.btn_to_download_video_from_web1, self.btn_to_download_video_from_web1_only_shortcut_name],
#                 [self.btn_to_download_video_from_web2, self.btn_to_download_video_from_web2_only_shortcut_name],
#                 [self.btn_to_should_i_translate_kor_to_eng, self.btn_to_should_i_translate_kor_to_eng_only_shortcut_name],
#                 [self.btn_to_should_i_translate_eng_to_kor, self.btn_to_should_i_translate_eng_to_kor_only_shortcut_name],
#                 [self.btn_to_ask_something_to_ai, self.btn_to_ask_something_to_ai_only_shortcut_name],
#                 [self.btn_to_open_project_directory, self.btn_to_open_project_directory_only_shortcut_name],
#                 [self.btn_to_connect_to_rdp1, self.btn_to_connect_to_rdp1_only_shortcut_name],
#                 [self.btn_to_back_up_target, self.btn_to_back_up_target_only_shortcut_name],
#                 [self.btn_to_rotate_window_size_mode, self.btn_to_rotate_window_size_mode_only_shortcut_name],
#                 # [  self.btn_to_run_cmd_exe,      self.btn_to_run_cmd_exe_only_shortcut_name],
#                 # [  self.btn_to_download_youtube_as_wav,      self.btn_to_download_youtube_as_wav_only_shortcut_name],
#                 # [  self.btn_to_download_youtube_as_webm_only_sound,      self.btn_to_download_youtube_as_webm_only_sound_only_shortcut_name],
#                 # [  self.btn_to_hide_windows_of_this_app,      self.btn_to_hide_windows_of_this_app_only_shortcut_name],
#                 # [  self.btn_to_login,      self.btn_to_login_only_shortcut_name],
#                 # [  self.btn_to_run_no_paste_memo,      self.btn_to_run_no_paste_memo_only_shortcut_name],
#                 [self.btn_to_test, self.btn_to_test_only_shortcut_name],
#                 [self.btn_to_should_i_empty_trash_can, self.btn_to_should_i_empty_trash_can_only_shortcut_name],
#
#                 [self.btn_to_should_i_enter_to_power_saving_mode, self.btn_to_should_i_enter_to_power_saving_mode_only_shortcut_name],
#                 [self.btn_to_should_i_reboot_this_computer, self.btn_to_should_i_reboot_this_computer_only_shortcut_name],
#                 [self.btn_to_should_i_shutdown_this_computer, self.btn_to_should_i_shutdown_this_computer_only_shortcut_name],
#                 [self.btn_to_shoot_screenshot_for_rpa, self.btn_to_shoot_screenshot_for_rpa_only_shortcut_name],
#                 [self.btn_to_shoot_screenshot_custom, self.btn_to_shoot_screenshot_custom_only_shortcut_name],
#                 [self.btn_to_shoot_screenshot_full, self.btn_to_shoot_screenshot_full_only_shortcut_name],
#                 [self.btn_to_should_i_find_direction_via_naver_map, self.btn_to_should_i_find_direction_via_naver_map_only_shortcut_name],
#                 [self.btn_to_should_i_show_animation_information_from_web, self.btn_to_should_i_show_animation_information_from_web_only_shortcut_name],
#                 [self.btn_to_record_macro, self.btn_to_record_macro_only_shortcut_name],
#                 [self.btn_to_run_up_and_down_game, self.btn_to_run_up_and_down_game_only_shortcut_name],
#                 [self.btn_to_classify_special_files, self.btn_to_classify_special_files_only_shortcut_name],
#                 [self.btn_to_gather_empty_directory, self.btn_to_gather_empty_directory_only_shortcut_name],
#                 # [self.btn_to_gather_special_files, self.btn_to_gather_special_files_only_shortcut_name],
#                 [self.btn_to_gather_useless_files, self.btn_to_gather_useless_files_only_shortcut_name],
#                 [self.btn_to_merge_directories, self.btn_to_merge_directories_only_shortcut_name],
#                 [self.btn_to_convert_mkv_to_wav, self.btn_to_convert_mkv_to_wav_only_shortcut_name],
#             ]
#
#             # GRID SETTING
#             grid = QtWidgets.QGridLayout(self)
#
#             # GRID COORDINATION REFERENCE (ROW, COLUMN)
#             #        0,0  0,1  0,2
#             #        1,0  1,1  1,2
#             #        2,0  2,1  2,2
#
#             # spaver
#
#             # GRID_COLUMN 1 폭 조절용 policy
#             # size_policy = QSizePolicy()
#             # grid.setHorizontalPolicy(QSizePolicy.Minimum)  # 그리드의 열의 폭을 최소로 설정합니다.
#             grid.setVerticalSpacing(9)
#             grid.setHorizontalSpacing(5)
#             grid.setColumnMinimumWidth(1, 125)
#             grid.setColumnMinimumWidth(2, 45)
#
#             btns_grid = btns
#             line_no = 0
#             for btn in btns_grid:
#                 grid.addWidget(btn[0], line_no, 0)  # GRID_COLUMN 0 설정
#                 grid.addWidget(btn[1], line_no, 2)  # GRID_COLUMN 1 설정
#                 line_no = line_no + 1
#
#             # 화면실행
#             def do_once():
#                 self.rotate_window_size_mode()
#                 self.timer2.stop()
#
#             self.timer2 = QTimer()
#             self.timer2.timeout.connect(do_once)  # noqa
#             # self.timer2.start(500)
#             self.timer2.start(50)
#
#             # TEST
#             # self.inputbox = QPlainTextEdit(self)
#             # self.inputbox = QTextEdit(self)
#             # self.ta1 = QTableWidget(self)
#             # self.ta1.resize(500, 500)
#             # self.ta1.setColumnCount(3)
#             # self.ta1.setStyleSheet("color: rgba(255,255,255, 0.9);")
#             # self.ta1.setStyleSheet("background-color: rgba(255,255,255, 0.9);")
#             # table_column = ["첫번째 열", "두번째 열", "Third 열"]
#             # self.ta1.setHorizontalHeaderLabels(table_column)
#
#             # # 행 2개 추가
#             # self.ta1.setRowCount(2)
#
#             # # 추가된 행에 데이터 채워넣음
#             # self.ta1.setItem(0, 0, QTableWidgetItem("(0,0)"))
#             # self.ta1.setItem(0, 1, QTableWidgetItem("(0,1)"))
#             # self.ta1.setItem(1, 0, QTableWidgetItem("(1,0)"))
#             # self.ta1.setItem(1, 1, QTableWidgetItem("(1,1)"))
#
#             # 마지막에 행 1개추가
#             # self.ta1.insertRow(2)
#             # self.ta1.setItem(2, 0, QTableWidgetItem("New Data"))
#
#             # 셀의 텍스트 변경
#             # self.ta1.item(1, 1).setText("데이터 변경")
#
#             # 셀에 있는 텍스트 출력
#             # print(self.ta1.item(0, 1).text())
#
#             # 테이블 데이터 전부 삭제
#             # self.ta1.clear()
#
#             # 테이블 행전부 삭제
#             # self.ta1.setRowCount(0)
#
#             self.scroll_area = QScrollArea()
#             self.scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
#             self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
#             self.scroll_area.setAlignment(Qt.AlignVCenter | Qt.AlignHCenter)
#             self.scroll_area.setStyleSheet(f" border: none; width: {self.display_width_default}px; height: {self.display_height_default}px")
#             self.scroll_area.setLayout(grid)
#
#             # 레이아웃 설정
#             layout = QVBoxLayout(self)
#             layout.addWidget(self.scroll_area)
#
#             self.setMouseTracking(True)  # pyside6 창 밖에서도 마우스 추적 가능 설정 # 마우스 움직임 이벤트 감지 허용 설정
#             # self.toogle_rpa_window()
#             # Park4139.press("alt", "w")
#
#             self.move_this_window_to_front()
#
#             self.btn_text_clicked = "foo"
#
#             # self.event_loop = QEventLoop()
#             # self.event_loop.exec()
#
#             # self.exec()
#
#         def monitor_mouse_position_per_second(self):
#             """마우스 움직임 감지 함수"""
#             x, y = pyautogui.position()  # 현재 마우스 위치 ( 이게 내가 원하던 수치 )
#             # print(x, y)
#             self.current_position = QCursor.pos()  # 현재 마우스 커서 위치 ( 이건 내가 원하는 수치는 아닌데... 뭘 의미하는 거지? )
#             # print(self.current_position)
#             if self.previous_position is not None and self.previous_position == self.current_position:
#                 # print("마우스가 멈췄습니다")
#                 # print(f"self.mouse_positions : {self.mouse_positions}") # 마우스 위치 리스트
#                 if 10 <= len(self.mouse_positions):
#                     # 10번 연속 mouse 중지 감지
#                     # self.mouse_positions 에 등록된 모든 self.current_position 가 동일하면 10번 연속으로 움직이지 않은 것으로 판단
#                     if len(self.mouse_positions) == 10:
#                         # 동일한 10개 원소를 갖는 리스트 내에서 요소의 중복을 제거하면 중복이 제거된 리스트의 요소의 수는 1개가 나올 것을 기대
#                         mouse_positions_removed_duplicatd_elements = list(set(self.mouse_positions))  # orderless way
#                         if len(mouse_positions_removed_duplicatd_elements) == 1:
#                             # print("10번 연속 중지 감지")
#                             self.hide()
#                             pass
#
#                 if 5 <= len(self.mouse_positions):
#                     self.mouse_positions = []  # 감지값들이 5개 이상이면 감지값목록 초기화
#
#                 elif len(self.mouse_positions) < 5:
#                     self.mouse_positions.append(self.current_position)
#                 pass
#             else:
#                 # print("마우스가 움직였습니다")
#
#                 # 우측하단 꼭지점 부근 네비게이션
#                 if 3440 - 25 <= x and 1440 - 25 <= y:
#                     BusinessLogicUtil.press("win", "d", interval=0.15)
#
#                 # 우측 네비게이션
#                 # if 3440 - 50 <= x <= 3440 and 300 <= y <= 1440:
#                 #     try:
#                 #         Park4139.explorer(Park4139.PYCHARM64_EXE)
#                 #     except:
#                 #         DebuggingUtil.trouble_shoot("%%%FOO%%%")
#                 #         pass
#
#                 # 좌측 네비게이션
#                 elif 0 <= x <= 15 and 0 <= y <= 1440:
#                     self.toogle_rpa_window()
#                     pass
#
#
#                 else:
#                     pass
#             self.previous_position = self.current_position
#
#         def monitor_mouse_position(self, x, y):
#             # 상단 네비게이션
#             if 0 <= x <= 3440 and 0 <= y <= 25:
#                 pass
#             else:
#                 pass
#
#         @staticmethod
#         # def rpa_program_method_decorator(function):
#         def rpa_program_method_decorator(function: Callable[[T], None]):
#             def wrapper(self):
#                 FileSystemUtil.play_wav_file(file_abspath=StateManagementUtil.POP_SOUND_WAV)
#                 self.hide()  # 비동기 전까지는 사용자가 다른 명령을 하지 못하도록 이 코드를 사용
#                 function(self)
#                 self.move_this_window_to_front()
#
#             return wrapper
#
#         def eventFilter(self, obj, event):
#             # if event.type() == QEvent.MouseMove:
#             #     x = event.globalX()
#             #     y = event.globalY()
#             #     print(f"마우스 이동 - X: {x}, Y: {y}")
#             # return super().eventFilter(obj, event)
#             if event.type() == QEvent.MouseButtonPress and not self.rect().contains(event.pos()):
#                 print("pyside6 창 외부 클릭 되었습니다")
#             return super().eventFilter(obj, event)
#
#         def mousePressEvent(self, e):
#             if e.button() == Qt.LeftButton:  # 왼쪽 버튼 클릭 시 동작
#                 print("왼쪽 버튼 클릭")
#             #     print(f"마우스 좌표: ({x}, {y})")
#             #     print(f"마우스 좌표:\n : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.BackButton:
#                 print(f"mouse event monitor:\nBackButton : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton1:
#                 print(f"mouse event monitor:\nExtraButton1 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton10:
#                 print(f"mouse event monitor:\nExtraButton10 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton11:
#                 print(f"mouse event monitor:\nExtraButton11 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton12:
#                 print(f"mouse event monitor:\nExtraButton12 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton13:
#                 print(f"mouse event monitor:\nExtraButton13 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton14:
#                 print(f"mouse event monitor:\nExtraButton14 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton15:
#                 print(f"mouse event monitor:\nExtraButton15 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton16:
#                 print(f"mouse event monitor:\nExtraButton16 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton17:
#                 print(f"mouse event monitor:\nExtraButton17 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton18:
#                 print(f"mouse event monitor:\nExtraButton18 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton19:
#                 print(f"mouse event monitor:\nExtraButton19 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton2:
#                 print(f"mouse event monitor:\nExtraButton2 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton20:
#                 print(f"mouse event monitor:\nExtraButton20 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton21:
#                 print(f"mouse event monitor:\nExtraButton21 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton22:
#                 print(f"mouse event monitor:\nExtraButton22 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton23:
#                 print(f"mouse event monitor:\nExtraButton23 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton24:
#                 print(f"mouse event monitor:\nExtraButton24 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton3:
#                 print(f"mouse event monitor:\nExtraButton3 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton4:
#                 print(f"mouse event monitor:\nExtraButton4 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton5:
#                 print(f"mouse event monitor:\nExtraButton5 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton6:
#                 print(f"mouse event monitor:\nExtraButton6 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton7:
#                 print(f"mouse event monitor:\nExtraButton7 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton8:
#                 print(f"mouse event monitor:\nExtraButton8 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ExtraButton9:
#                 print(f"mouse event monitor:\nExtraButton9 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.ForwardButton:
#                 print(f"mouse event monitor:\nForwardButton : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.LeftButton:
#                 print(f"mouse event monitor:\nLeftButton : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.MiddleButton:
#                 print(f"mouse event monitor:\nMiddleButton : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.NoButton:
#                 print(f"mouse event monitor:\nNoButton : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.RightButton:
#                 print(f"mouse event monitor:\nRightButton : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.TaskButton:
#                 print(f"mouse event monitor:\nTaskButton : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.XButton1:
#                 print(f"mouse event monitor:\nXButton1 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#             elif e.button() == Qt.XButton2:
#                 print(f"mouse event monitor:\nXButton2 : {str(e.pos().x())} ,{str(e.pos().y())} ")
#
#         # def mouseReleaseEvent(self, e):
#         #     self.label10.setText(f"event monitor:\nmouseReleaseEvent : {str(e.pos().x())} ,{str(e.pos().y())} ")
#         #
#         # def mouseDoubleClickEvent(self, e):
#         #     self.label10.setText(f"event monitor:\nmouseDoubleClickEvent : {str(e.pos().x())} ,{str(e.pos().y())} ")
#
#         def keyPressEvent(self, e):
#             self.mouse_positions = []  # 키보드가 눌리면 사용자가 사용중인 것으로 간주하고 마우스 위치 값 목록 초기화
#
#         # def keyPressEvent(self, e):
#         #     # these keys refered from https://doc.qt.io/qtforpython-6/PySide6/QtCore/Qt.html
#         #     # 테스트 결과 한/영, 한자 인식안됨.
#         #     if e.key() == Qt.Key_Return:
#         #         self.label11.setText(f"keyboard event monitor:\nKey_Return : Key_Return ")
#         #         self.showMinimized()
#         #         park4139.press("space")
#         #         self.showMaximized()
#         #     elif e.key() == Qt.Key_0:
#         #         self.label11.setText(f"keyboard event monitor:\nKey_0 : Key_0 ")
#         # deprecated test
#         # def inputbox_changed(self):
#         #     DebuggingUtil.commentize("inputbox 텍스트 change event 감지 되었습니다")
#         #     print(self.inputbox.ment())
#         #
#         # def inputbox_edit_finished(self):
#         #     DebuggingUtil.commentize("inputbox edit finish event 감지 되었습니다")
#         #
#         # def inputbox_return_pressed(self):
#         #     DebuggingUtil.commentize("inputbox return pressed event 감지 되었습니다")
#
#         def get_btn_name_with_shortcut_name(self, button_name_without_shortcut):
#             # 버튼명과 shourtcut 명을 을 적당한 간격으로 띄워서 string 으로 반환하는 코드, 폰트 가 고정폭폰트 가 아니면 무용지물인 함수
#             numbers = []
#             for key, value in self.available_shortcut_list.items():
#                 numbers.append(len(value) + len(key))
#             max_len_value = max(numbers)
#             button_name_with_short_cut = ""
#             for key, value in self.available_shortcut_list.items():
#                 if key == button_name_without_shortcut:
#                     space_between = " " * (max_len_value - len(key) - len(value) + 1)
#                     # space_between = "\t"
#                     # button_name_with_short_cut = button_name_with_short_cut + f"{key}{space_between}{value}".strip()
#                     # button_name_with_short_cut = button_name_with_short_cut + f"{key}{space_between}( {value} )".strip()
#                     button_name_with_short_cut = button_name_with_short_cut + f"{value}{space_between}( {key} )".strip()
#             print(button_name_with_short_cut)
#             return button_name_with_short_cut
#
#         def get_btn_name_promised(self, button_name_without_shortcut):
#             button_name_with_short_cut = ""
#             for key, value in self.available_shortcut_list.items():
#                 if key == button_name_without_shortcut:
#                     button_name_with_short_cut = button_name_with_short_cut + f"{key}".strip()
#             return button_name_with_short_cut
#
#         def get_shortcut_name_promised(self, button_name_without_shortcut):
#             button_name_with_short_cut = ""
#             for key, value in self.available_shortcut_list.items():
#                 if key == button_name_without_shortcut:
#                     button_name_with_short_cut = button_name_with_short_cut + f"{value}".strip()
#             return button_name_with_short_cut
#
#         # def show_available_shortcut_list(self):
#         #     # global max
#         #     # global 을 설정하면, 이 변수는 함수의 실행이 끝난 다음에도 없어지지 않는다.
#         #     # 이 값을 나중에 함수 끝나고도 또 쓸려면 이렇게 쓰면 되겠다. @staticmethod 의 경우에는 변수 간의 값에 간섭이 되지 않도록 굳이 쓰지 않는 것이 좋겠다.
#         #     # global 많이 쓰면 이는 변수가 전역화 되니까 메모리의 성능이 저하되는 것이 아닐까?
#         #     # 그렇다면 함수 내에서만 전역적으로 변수를 쓰는 경우에, global 을 쓰지 않는 것이 성능을 위해서는 좋은 선택이겠다. 굳이 함수가 끝난 뒤에 밖에서 써야한다면 global 을 써야 겠지만, 나는 무척이나 이게 헷갈릴 것 같다
#         #     # 그동안의 경험으로는 코드 맥락 상, global 선언을 하지 않아도 전역변수 처럼 작동 되는 것 같아 보인다.... 아니다 이게 global max 를 선언하지 않았다고 가정하면 max 는 show_available_shortcut_list() 가 종료되면 max 는 사라진다. 그런데 global max를 선언하면 max 는 유지된다!
#         #     # 혹시 객체의 인스턴스 같은 것을 global 을 통해서 변수에 저장하고 쓰면 싱글톤 처럼 쓸 수 있는 것일까? 메모리 효율은 많이 나빠질까?
#         #
#         #     numbers = []
#         #     for key_shortcut, value in self.available_shortcut_list.items():
#         #         numbers.append(len(value))
#         #     max_no: int
#         #     max_no = max(numbers)
#         #     for key_shortcut, value in self.available_shortcut_list.items():
#         #         print(f"{{0: <{max_no}}} : {key_shortcut}".format(value))
#
#         def rotate_window_size_mode(self):
#             if self.windows_size_mode == 0:
#                 self.resize(self.display_width_default, self.display_height_default)
#                 self.move_window_to_center()  # 불필요 하면 주석하는 게 나쁘지 않겠다
#                 self.windows_size_mode = self.windows_size_mode + 1
#             elif self.windows_size_mode == 1:
#                 self.setGeometry(0, 0, int(self.display_width_default), int(self.display_height[0]))
#                 # self.setWindowFlags(Qt.WindowStaysOnTopHint)  # 모든 창 앞에 위치하도록 설정
#                 self.windows_size_mode = self.windows_size_mode + 1
#             elif self.windows_size_mode == 2:
#                 self.showMaximized()
#                 self.windows_size_mode = self.windows_size_mode + 1
#             elif self.windows_size_mode == 3:
#                 self.setGeometry(1600 - int(self.display_width_default), 0, int(self.display_width_default), int(self.display_height[0]))
#                 self.windows_size_mode = 0
#
#         def set_shortcut(self, shortcut_key_promised, function):
#             # self.shortcut = QShortcut(QKeySequence(self.available_shortcut_list[btn_name_promised]), self)  # ctrl+1 2개 키들의 조합 설정
#             self.shortcut = QShortcut(self.available_shortcut_list[shortcut_key_promised], self)  # ctrl+n+d 3개 키들의 조합 설정 시도
#             self.shortcut.activated.connect(function)
#             pass
#
#         def get_btn(self, btn_name, function, btn_text_align="left"):
#             button = QPushButton(btn_name, self)  # alt f4 로 가이드 해도 되겠다. 이건 그냥 설정 되어 있는 부분.
#             button.clicked.connect(function)
#
#             # 2023년 12월 14일 (목) 16:28:15
#             # 결론, fixed width font로 시도해볼 수 있다 자릿 수를 맞출 수 있다.
#             # non-fixed width font 이슈 JAVA 에서도 구현했을 때 마딱드렸던 내용인데,
#             # 분명히 문장 전체 길이를 단어 사이의 공백의 수를 결정짓는 함수를 테스트 했음에도 자릿수가 맞지 않았는데
#             # 이는 고정 폭이 아님이기 때문이었다 따라서 고정 폭 폰트로 출력되는 콘솔에서는 정상, 비고정 폭 폰트로 출력되는 콘솔에서는 비정상,
#             # 이 경우에는 콘솔이 아니라 pyside6 로 만든 UI 에서 나타났다.
#             # 새벽에 이 문제를 만나서 잠깐 넋나갔는데 아침에 다시보니 그때 경험이 떠올라서 실험해보니 잘 해결되었다. 덕분에 pyside6에서 위젯에 폰트 적용하는 법도 터득
#
#             # pyside6 버튼 내장폰트 설정
#             # pyside6 built in fixed width font
#             # font = QFont("Monospace")
#             # font = QFont("Ubuntu Mono")
#             # font = QFont("Inconsolata")
#             # font = QFont("Monaco")
#             # font = QFont("Courier")
#             # font = QFont("Courier 10 Pitch")
#             # font = QFont("Courier Prime")
#             # font = QFont("Droid Sans Mono")
#             # font = QFont("Fira Mono")
#             # font = QFont("Hack")
#             # font = QFont("Menlo")
#             # font = QFont("Monofur")
#             # font = QFont("Noto Mono")
#             # font = QFont("PT Mono")
#             # font = QFont("Roboto Mono")
#             # font = QFont("Source Code Pro")
#             # font = QFont("Victor Mono")
#             # font = QFont("Courier New")
#             # font = QFont("Liberation Mono")
#             # font = QFont("DejaVu Sans Mono")
#             # font = QFont("Consolas")  # 그나마 가장 마음에 드는 폰트
#
#             # pyside6 버튼 외부폰트 설정
#             button.setFont(Pyside6Util.get_font_for_pyside6(font_path=FontsUtil.GMARKETSANSTTFLIGHT_TTF))
#             if btn_text_align == "right":
#                 # button.setStyleSheet("QPushButton { text-align: right; color: rgba(255,255,255, 0.9); height: 20px; font-size: 8px}")
#                 # button.setStyleSheet("QPushButton { text-align: right; color: rgba(255,255,255, 0.9);               font-size: 8px}")
#                 button.setStyleSheet("QPushButton { text-align: right; color: rgba(255,255,255, 0.9);               font-size: 13px}")
#                 # button.setStyleSheet("QPushButton { text-align: right; color: rgba(255,255,255, 0.9); height: 20px; font-size: 13px}")
#                 button.setFixedWidth(65)
#                 # button.setMinimumWidth(button.sizeHint().width())
#             else:
#                 # button.setStyleSheet("QPushButton { text-align: left; color: rgba(255,255,255, 0.9); height: 20px; font-size: 8px}")
#                 # button.setStyleSheet("QPushButton { text-align: left; color: rgba(255,255,255, 0.9);                 font-size: 8px}")
#                 button.setStyleSheet("QPushButton { text-align: left; color: rgba(255,255,255, 0.9);                 font-size: 13px}")
#                 # button.setStyleSheet("QPushButton { text-align: left; color: rgba(255,255,255, 0.9); height: 20px; font-size: 13px}")
#                 button.setFixedWidth(225)
#                 # button.setMinimumWidth(button.sizeHint().width())
#             return button
#
#         def move_window_to_center(self):
#             center = QScreen.availableGeometry(self.app.primaryScreen()).center()
#             geo = self.frameGeometry()
#             geo.moveCenter(center)
#             self.move(geo.topLeft())
#             # if self.screens:
#             #     primary_screen = self.screens[0]  # 첫 번째 화면을 기본 화면으로 설정합니다.
#             #     center = primary_screen.availableGeometry().center()  # 기본 화면의 중앙 좌표를 가져옵니다.
#             #     # 화면을 가운데로 이동시키는 코드를 작성하세요.
#             #     # 예시로 윈도우를 생성하고 중앙 좌표를 이용하여 위치를 설정합니다.
#             #     self.setGeometry(100, 100, 500, 300)  # 윈도우의 초기 위치와 크기를 설정합니다.
#             #     self.move(center - self.rect().center())  # 윈도우를 화면 중앙으로 이동시킵니다.
#
#         @rpa_program_method_decorator
#         def show_weather_from_web(self):
#             BusinessLogicUtil.get_comprehensive_weather_information_from_web()
#
#         @rpa_program_method_decorator
#         def run_no_paste_memo(self):
#             BusinessLogicUtil.speak_that_service_is_in_preparing()
#
#         @rpa_program_method_decorator
#         def should_i_reboot_this_computer(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_reboot_this_computer()
#
#         @QtCore.Slot()
#         def login(self):
#             BusinessLogicUtil.speak_that_service_is_in_preparing()
#
#         @rpa_program_method_decorator
#         def should_i_show_animation_information_from_web(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_show_animation_information_from_web()
#
#         @rpa_program_method_decorator
#         def should_i_exit_this_program(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_exit_this_program()
#
#         @rpa_program_method_decorator
#         def should_i_find_direction_via_naver_map(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_find_direction_via_naver_map()
#
#         @rpa_program_method_decorator
#         def download_youtube_as_wav(self):
#             BusinessLogicUtil.speak_that_service_is_in_preparing()
#
#         @rpa_program_method_decorator
#         def should_i_download_youtube_as_webm(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_download_youtube_as_webm()
#
#         @rpa_program_method_decorator
#         def should_i_download_youtube_as_webm_alt(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_download_youtube_as_webm_alt()
#
#         @rpa_program_method_decorator
#         def download_youtube_as_webm_only_sound(self):
#             BusinessLogicUtil.speak_that_service_is_in_preparing()
#
#         @rpa_program_method_decorator
#         def should_i_shutdown_this_computer(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_shutdown_this_computer()
#
#         @rpa_program_method_decorator
#         def shoot_screenshot_custom(self):
#             while True:
#                 BusinessLogicUtil.shoot_custom_screenshot()
#                 break
#
#         @rpa_program_method_decorator
#         def shoot_screenshot_full(self):
#             while True:
#                 BusinessLogicUtil.shoot_full_screenshot()
#                 break
#
#         @rpa_program_method_decorator
#         def shoot_screenshot_for_rpa(self):
#             while True:
#                 BusinessLogicUtil.shoot_img_for_rpa()
#                 break
#
#         @rpa_program_method_decorator
#         def should_i_back_up_target(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_back_up_target()
#
#         @rpa_program_method_decorator
#         def should_i_start_test(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_start_test_core()
#
#         @rpa_program_method_decorator
#         def open_project_directory(self):
#             project_directory = str(StateManagementUtil.PROJECT_DIRECTORY)
#             FileSystemUtil.explorer(project_directory)
#
#         @rpa_program_method_decorator
#         def should_i_enter_to_power_saving_mode(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_enter_to_power_saving_mode()
#
#         @rpa_program_method_decorator
#         def should_i_translate_eng_to_kor(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_translate_eng_to_kor()
#
#         @rpa_program_method_decorator
#         def should_i_translate_kor_to_eng(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_translate_kor_to_eng()
#
#         @rpa_program_method_decorator
#         def should_i_empty_trash_can(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_do(ment="쓰레기통을 비울까요?", function=BusinessLogicUtil.empty_recycle_bin, auto_click_negative_btn_after_seconds=10)
#
#         @rpa_program_method_decorator
#         def run_cmd_exe_as_admin(self):
#             BusinessLogicUtil.run_cmd_exe_as_admin()
#
#         @rpa_program_method_decorator
#         def ask_something_to_ai(self):
#             BusinessLogicUtil.ask_something_to_ai()
#
#         @rpa_program_method_decorator
#         def should_i_connect_to_rdp1(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_connect_to_rdp1()
#
#         @rpa_program_method_decorator
#         def should_i_record_macro(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_record_macro()
#
#         @rpa_program_method_decorator
#         def run_up_and_down_game(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.run_up_and_down_game()
#
#         @rpa_program_method_decorator
#         def should_i_classify_special_files(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_classify_special_files()
#
#         @rpa_program_method_decorator
#         def should_i_gather_empty_directory(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_gather_empty_directory()
#
#         @rpa_program_method_decorator
#         def should_i_gather_useless_files(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_gather_useless_files()
#
#         @rpa_program_method_decorator
#         def should_i_merge_directories(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_merge_directories()
#
#         @rpa_program_method_decorator
#         def should_i_convert_mkv_to_wav(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_convert_mkv_to_wav()
#
#         @rpa_program_method_decorator
#         def should_i_crawl_a_tag_href(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_crawl_a_tag_href()
#
#         @rpa_program_method_decorator
#         def should_i_crawl_youtube_video_title_and_url(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_crawl_youtube_video_title_and_url()
#
#         @rpa_program_method_decorator
#         def should_i_crawl_youtube_playlist(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_crawl_youtube_playlist()
#
#         @rpa_program_method_decorator
#         def should_i_explorer(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             BusinessLogicUtil.should_i_explorer()
#
#         @rpa_program_method_decorator
#         def should_i_sync(self):
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             # BusinessLogicUtil.should_i_sync()
#             BusinessLogicUtil.should_i_sync_V_0_0_2()
#
#         @rpa_program_method_decorator
#         def download_video_from_web1(self):
#             BusinessLogicUtil.download_video_from_web1()
#
#         @rpa_program_method_decorator
#         def download_video_from_web2(self):
#             BusinessLogicUtil.download_video_from_web2()
#
#         def toogle_rpa_window(self):
#             # Park4139.debug_as_cli(f"def {inspect.currentframe().f_code.co_name}() is called...")
#             if self.isHidden() and not self.isVisible():
#                 # console_blurred 프로그램 창 활성화
#                 self.move_this_window_to_front()
#             else:
#                 self.hide()
#
#         def move_this_window_to_front(self):
#             # self.activateWindow() 와 self.show() 의 위치는 서로 바뀌면 의도된대로 동작을 하지 않는다
#             self.show()
#             self.activateWindow()
#             # import win32gui
#             # active_window = win32gui.GetForegroundWindow()
#             # win32gui.SetForegroundWindow(active_window)
#
#     class MacroWindow(QDialog):
#         # class MacroWindow(QWidget):
#         def __init__(self):
#             super().__init__()
#             FileSystemUtil.play_wav_file(file_abspath=StateManagementUtil.POP_SOUND_WAV)
#
#             #  매크로 창 전역 변수 설정
#             self.previus_pressed_keys = []
#             self.display_width = FileSystemUtil.get_display_info()['width'],
#             self.display_height = FileSystemUtil.get_display_info()['height'],
#             # self.display_width_default = int(int(self.display_width[0]) * 0.106)
#             self.display_width_default = int(int(self.display_width[0]) * 0.06)
#             self.display_height_default = int(int(self.display_height[0]) * 0.2)
#
#             # 마우스 위치, 클릭 정보, 시간 정보를 저장할 리스트
#             # self.positions = []
#
#             # 녹화 시작 시간
#             self.time_recording_start = time.time()
#             # self.time_recording_start_rel = 0.0
#             self.elapsed_full_recording_time = 0.0
#             self.previous_time = time.time()
#             self.time_recording_end = 0.0
#             #  메인창 설정
#             self.setWindowTitle('.')
#             self.setWindowIcon(QIcon(StateManagementUtil.ICON_PNG))  # 메인창 아이콘 설정
#             # self.setAttribute(Qt.WA_TranslucentBackground) # 메인창 블러 설정
#             # self.setWindowFlags(Qt.WindowType.FramelessWindowHint) # 메인창 최상단 프레임레스 설정
#             GlobalBlur(self.winId(), hexColor=False, Acrylic=False, Dark=True, QWidget=self)
#             self.setWindowFlags(Qt.WindowTitleHint | Qt.WindowCloseButtonHint)  # 최대화 최소화 버튼 숨기기
#             self.setStyleSheet("background-color: rgba(0, 0, 0, 0)")
#             self.scale = 1 / 10
#             self.windows_size_mode = 0  # 창크기 모드 설정  #0 ~ 3
#
#             QCoreApplication.setAttribute(Qt.AA_EnableHighDpiScaling)  # 고해상도 스케일링을 활성화합니다.
#             self.screens = QGuiApplication.screens()  # 사용 가능한 모든 화면을 가져옵니다.
#             self.rotate_window_size_mode()
#
#             # label 설정
#             self.label = QLabel(self)
#             self.label.setStyleSheet("color: rgba(255,255,255, 0.9);")
#             self.label.setText(f"녹화 중...\n녹화경과시간: None\n")
#
#             # 단축키 설정
#             self.available_shortcut_list = {
#                 'MACRO EXCUTE': 'Ctrl+E',
#                 'MACRO EXIT': 'Ctrl+Q',
#             }
#             # self.set_shortcut('MACRO RECORD',self.record_macro)
#             self.set_shortcut('MACRO EXCUTE', self.excute_macro)
#             self.set_shortcut('MACRO EXIT', self.exit_macro)
#
#             # 버튼 설정
#             btn_to_excute_macro = self.get_btn(self.get_btn_name_with_shortcut_name('MACRO EXCUTE'), self.excute_macro)
#             btn_to_exit_macro = self.get_btn(self.get_btn_name_with_shortcut_name('MACRO EXIT'), self.exit_macro)
#
#             # GRID SETTING
#             grid = QtWidgets.QGridLayout(self)
#             btns = [
#                 self.label,
#                 btn_to_excute_macro,
#                 btn_to_exit_macro
#             ]
#             cnt = 0
#             for i in btns:
#                 grid.addWidget(i, cnt, 0)
#                 cnt = cnt + 1
#
#             # 레이아웃 설정
#             layout = QGridLayout(self)
#             layout.addLayout(grid, 0, 0)
#
#             # 단일 단축키가 같이 눌리는 문제 ( ctrl + v 를 누르면   ctrl + v 에 바인딩된 함수만 호출되길 기대하는데, ctrl 에 바인딩된 함수도 호출되는 문제 )
#             # 이벤트를 나누어 만들어 하나의 이벤트가 호출되면 다른 하나의 이벤트는 호출되지 않도록 설정 시도
#             self.is_processing_event = False
#
#             # Condition 객체 생성
#             # self.condition = threading.Condition()
#
#             # 이벤트 우선순위 설정
#             # self.is_event_shortcut_3_processing = False # highest priority
#             # self.is_event_shortcut_1_processing = False
#
#             # 마우스 이동 이벤트 핸들러 설정
#             listener = pynput.mouse.Listener(on_move=self.on_mouse_move)
#             listener.start()
#
#             # 마우스 버튼 클릭 이벤트 핸들러 설정
#             listener2 = pynput.mouse.Listener(on_click=self.on_mouse_btn_clicked)
#             listener2.start()
#
#             # 키보드 이벤트 핸들러 설정 ( 3개 조합 단축키 , 2개 조합 단축키 , 단일 단축키 모두가능)
#             self.keyboard_main_listener = pynput.keyboard.Listener(on_press=self.on_keys_down, on_release=self.on_keys_up)
#             self.keyboard_main_listener.start()
#             # 충돌이 문제가 아니고 두 이벤트가 중복이 되면 안되고 이벤트에 우선순위를 더 두어서 두 이벤트 호출 시 우선순위가 높은 이벤트만 실행되도록
#
#             # self.is_processing_event = False
#
#             # 이벤트 핸들러 스레드 생성
#             # event1_thread = threading.Thread(target=self.listener3)
#             # event2_thread = threading.Thread(target=self.listner_2_combination_shorcuts)
#
#             # 이벤트 핸들러 스레드 시작
#             # event1_thread.start()
#             # event2_thread.start()
#
#             # 키보드 이벤트 핸들러 설정 ( 단일 단축키 )
#             # self.listener3 = pynput.keyboard.Listener(on_press=self.on_keboard_press, suppress=True)
#             # self.listener3.start()
#
#             # 모니터링 이벤트 설정
#             # self.mouse_positions = []
#             # self.previous_position = None
#             # self.current_position = None
#             # self.timer = QTimer()
#             # self.timer.timeout.connect(self.on_left_mouse_btn_clicked)
#             # self.timer.start(900)
#
#             # 녹화 경과시간 업데이트
#             self.timer2 = QTimer()
#             self.timer2.timeout.connect(self.update_label)  # noqa # 해당 검사에 대해서는 예외 목록에 "connect"를 넣는 것을 권장합니다. jetbrain 오류로 다년간 지속된 것으로 생각 중이다,
#             self.timer2.start(1000)
#
#             TextToSpeechUtil.speak_ments(ment="매크로녹화를 시작합니다", sleep_after_play=0.65)
#
#             # 매크로녹화시작 로깅
#             log_title = "매크로녹화시작"
#             contents = f"{StateManagementUtil.UNDERLINE_PROMISED}{TimeUtil.get_time_as_('%Y-%m-%d_%H:%M:%S')}{log_title}"
#             DebuggingUtil.print_magenta(contents)
#             self.save_macro_log(contents=contents)
#
#             self.bring_this_window()
#
#             # event_loop = QEventLoop()
#             # event_loop.exec()
#
#         def on_mouse_move(self, x, y):  # 아주 빠르게 감지
#             # 이방식으로 매크로 중지를 할까?
#             # print(f"마우스 이동 - X: {x}, Y: {y}")
#             # print("마우스가 움직였습니다")
#             # 상단 네비게이션
#             if 0 <= x <= 3440 and 0 <= y <= 25:
#                 self.rotate_window_size_mode()
#             else:
#                 pass
#
#         def eventFilter(self, obj, event):
#             if event.type() == QEvent.MouseMove:
#                 x = event.globalX()
#                 y = event.globalY()
#                 print(f"pyside6 창 외부 마우스 이동 감지 시도 - X: {x}, Y: {y}")
#             return super().eventFilter(obj, event)
#
#             # if event.type() == QEvent.MouseButtonPress and not self.rect().contains(event.pos()):
#             #     print("pyside6 창 외부 클릭 되었습니다")
#             # return super().eventFilter(obj, event)
#
#         def get_btn_name_with_shortcut_name(self, button_name_without_shortcut):
#             numbers = []
#             for key, value in self.available_shortcut_list.items():
#                 numbers.append(len(value) + len(key))
#             max_len_value = max(numbers)
#             # print(max_len_value)
#
#             button_name_with_short_cut = ""
#             for key, value in self.available_shortcut_list.items():
#                 # print(key == button_name_without_shortcut)
#                 # print(key)
#                 # print(button_name_without_shortcut)
#
#                 if key == button_name_without_shortcut:
#                     space_between = " " * (max_len_value - len(key) - len(value) + 1)
#                     # space_between = " "
#                     # space_between = str(max_len_value - len(key) - len(value) + 1)
#                     # button_name_with_short_cut = button_name_with_short_cut + f"{key}{space_between}{value}\n"
#                     button_name_with_short_cut = button_name_with_short_cut + f"{key}{space_between}{value}".strip()
#             print(button_name_with_short_cut)
#             return button_name_with_short_cut
#
#         def rotate_window_size_mode(self):
#             if self.windows_size_mode == 0:
#                 self.bring_this_window()
#                 self.setGeometry(0, 0, int(self.display_width_default / 2), int(self.display_height[0] / 5))
#                 self.move_window_to_center()
#                 self.windows_size_mode = self.windows_size_mode + 1
#             elif self.windows_size_mode == 1:
#                 self.showMinimized()
#                 self.windows_size_mode = 0
#
#         def set_shortcut(self, btn_name_promised, function):
#             self.shortcut = QShortcut(self.available_shortcut_list[btn_name_promised], self)  # ctrl+n+d 3개 키들의 조합 설정 시도
#             self.shortcut.activated.connect(function)
#             pass
#
#         def get_btn(self, btn_name, function):
#             button = QPushButton(btn_name, self)
#             button.clicked.connect(function)
#
#             # 폰트 설정
#             button.setFont(Pyside6Util.get_font_for_pyside6(font_path=FontsUtil.RUBIKDOODLESHADOW_REGULAR_TTF))  # 입체감있는 귀여운 영어 폰트
#             button.setStyleSheet("QPushButton { text-align: left; color: rgba(255,255,255, 0.9); height: 20px ; font-size: 10px}")
#             # button.setLayoutDirection(QtCore.Qt.)
#             return button
#
#         def move_window_to_center(self):
#             # center = QScreen.availableGeometry(app.primaryScreen()).center()
#             # geo = self.frameGeometry()
#             # geo.moveCenter(center)
#             # self.move(geo.topLeft())
#             if self.screens:
#                 primary_screen = self.screens[0]  # 첫 번째 화면을 기본 화면으로 설정합니다.
#                 center = primary_screen.availableGeometry().center()  # 기본 화면의 중앙 좌표를 가져옵니다.
#                 # 화면을 가운데로 이동시키는 코드를 작성하세요.
#                 # 예시로 윈도우를 생성하고 중앙 좌표를 이용하여 위치를 설정합니다.
#                 self.setGeometry(100, 100, 500, 300)  # 윈도우의 초기 위치와 크기를 설정합니다.
#                 self.move(center - self.rect().center())  # 윈도우를 화면 중앙으로 이동시킵니다.
#
#                 # self.show()
#
#         def excute_macro(self):
#             pass
#
#         def exit_macro(self):
#             # 매크로녹화종료 로깅
#             log_title = "매크로녹화종료"
#             self.time_recording_end = self.elapsed_full_recording_time
#             contents = f"{StateManagementUtil.UNDERLINE_PROMISED}{TimeUtil.get_time_as_('%Y-%m-%d_%H:%M:%S')}{log_title}"
#             DebuggingUtil.print_magenta(contents)
#             self.save_macro_log(contents=contents)
#
#             # 매크로 로그 확인
#             TextToSpeechUtil.speak_ments(ment="저장된 매크로 로그를 확인합니다", sleep_after_play=0.65)
#             FileSystemUtil.explorer(StateManagementUtil.MACRO_LOG)
#
#         # @Slot() # 최신 버전의 PySide6에서는 @Slot() 데코레이터를 사용하지 않고도 Slot 메서드를 정의할 수 있습니다. 이렇게 되면 자동으로 Slot으로 인식됩니다. 즉 최신버전 pyside6 에서는 쓸 필요 없다.
#         def update_label(self):
#             try:
#                 self.elapsed_full_recording_time = int(time.time() - self.time_recording_start)
#                 self.label.setText(f"녹화 중...\n녹화경과시간: {self.elapsed_full_recording_time} secs \n")
#             except:
#                 traceback.print_exc(file=sys.stdout)
#
#         def on_mouse_btn_clicked(self, x, y, button, pressed):
#
#             # 현재 시간과 녹화 시작 시간의 차이 계산
#             current_time = time.time()
#             elapsed_time = int((current_time - self.previous_time) * 1000)
#             # print(f"current_time : {current_time}")
#             # print(f"self.previous_time : {self.previous_time}")
#
#             # x, y = pyautogui.position() # parameter 에서 오는 값과 동일하므로 대안으로 남겨둠.
#
#             if button == pynput.mouse.Button.left and pressed:
#                 info = f"Park4139.sleep({elapsed_time})   %%%FOO%%%    Park4139.click_mouse_left_btn(abs_x={x},abs_y={y}) "
#                 DebuggingUtil.print_magenta(info)
#                 self.save_macro_log(contents=f"{TimeUtil.get_time_as_('%Y-%m-%d_%H:%M:%S')} {info}")
#             elif button == pynput.mouse.Button.right and pressed:
#                 info = f"Park4139.sleep({elapsed_time})   %%%FOO%%%    Park4139.click_mouse_right_btn(abs_x={x},abs_y={y}) "
#                 DebuggingUtil.print_magenta(info)
#                 self.save_macro_log(contents=f"{TimeUtil.get_time_as_('%Y-%m-%d_%H:%M:%S')} {info}")
#
#             # 현재시간을 이전시간에 저장
#             self.previous_time = current_time
#
#         def save_macro_log(self, contents: str):
#             macro_log = StateManagementUtil.MACRO_LOG
#             FileSystemUtil.make_leaf_file(StateManagementUtil.MACRO_LOG)
#             with open(macro_log, "a", encoding="utf-8") as f:
#                 f.write(f"{contents}\n")
#
#         def on_keboard_press(self, key):
#             # global is_processing_event
#             # if self.is_processing_event != False:
#             print(f'키보드 입력: {key}')
#
#             # isinstance()
#             # all(list) # (iterable) 객체의 모든 요소가 참(True)인지 확인
#             # all(dict) # (iterable) 객체의 모든 요소가 참(True)인지 확인
#             # all(tuple) # (iterable) 객체의 모든 요소가 참(True)인지 확인
#
#         def on_keys_down(self, key):
#             # 현재 시간과 녹화 시작 시간의 차이 계산
#             current_time = time.time()
#             elapsed_time = int((current_time - self.previous_time) * 1000)
#
#             # for hotkey in self.HOTKEYS:
#             #     hotkey.release(self.keyboard_listener1.canonical(key))
#             #     print(f"key : {key}")
#
#             # 키이름 여러 형식으로 출력
#             # try:
#             #     print(key)
#             #     print(key.value)
#             #     print(key.value.vk)
#             # except:
#             #     print(key)
#             #     pass
#
#             # 키이름 전처리
#             key: str = str(key)
#             key: str = key.lower()
#             key: str = key.replace("\'", "")
#             key: str = key.replace("\"", "")
#             key: str = key.replace("key.", "")
#             key: str = key.replace("<25>", "한자 or ctrl_r")
#             key: str = key.replace("<21>", "한영 or alt_r")
#             key: str = key.replace("12", "텐키 5")  # 텐키
#             key: str = key.replace("cmd", "win")
#             key: str = key.replace("page", "pg")
#             key: str = key.replace("down", "dn")
#             if key != "num_lock":
#                 key: str = key.replace("_l", "")
#             key: str = key.replace("_r", "")
#
#             key: str = key.replace(" ", "")
#             key: str = key.replace("<", "")
#             key: str = key.replace(">", "")
#             # key: str = key.replace("_", "")
#
#             # 전처리 후 출력
#             # print(str(key))
#
#             # if key == "ctrl+alt":
#             log_title = "키보드인풋"
#             info = f"Park4139.sleep({elapsed_time})\nPark4139.keyDown('{key}') "
#             DebuggingUtil.print_magenta(info)
#             self.save_macro_log(contents=f"{TimeUtil.get_time_as_('%Y-%m-%d_%H:%M:%S')}{log_title} {info}")
#
#             # 현재시간을 이전시간에 저장
#             self.previous_time = current_time
#
#             # if self.store = []
#             # global is_processing_event
#             # if self.is_processing_event == False:
#             # self.is_processing_event = True
#             # self.is_processing_event = False
#
#             # pynput.keyboard.Listener()는 다시 self.listener3.start() 할 수 없다.
#             # 새로 만들어야 한다
#             # if self.listener3.is_alive():
#             #     self.listener3.stop()
#             # else:
#
#             # self.listener3 = pynput.keyboard.Listener(on_press=self.on_keboard_press)
#             # self.listener3.start()
#
#             # self.condition.notify()  # event1에게 동작 신호 보내기
#
#         def on_single_key_pressed(self, key):
#
#             # 현재 시간과 녹화 시작 시간의 차이 계산
#             current_time = time.time()
#             elapsed_time = int((current_time - self.previous_time) * 1000)
#
#             # print(str(key))
#             if pynput.keyboard.GlobalHotKeys.name == "<ctrl>":
#                 log_title = "키보드단일키인풋"
#                 info = f"  Park4139.sleep({elapsed_time})   %%%FOO%%%    Park4139.press({str(key)}) "
#                 DebuggingUtil.print_magenta(info)
#                 self.save_macro_log(contents=f"{TimeUtil.get_time_as_('%Y-%m-%d_%H:%M:%S')}{log_title} {info}")
#
#             # 현재시간을 이전시간에 저장
#             self.previous_time = current_time
#
#             # try:
#             #     # alphanumeric key
#             #     print('{0}'.format(key.char))
#             # except AttributeError:
#             #     # special key
#             #     print('{0}'.format(key.char))
#
#         def on_keys_up(self, key):
#             # 현재 시간과 녹화 시작 시간의 차이 계산
#             current_time = time.time()
#             elapsed_time = int((current_time - self.previous_time) * 1000)
#
#             # for hotkey in self.HOTKEYS:
#             #     if hotkey
#             #     hotkey.press(self.keyboard_listener1.canonical(key))
#             #     print(f"key : {key}")
#
#             # 키이름 전처리
#             key: str = str(key)
#             key: str = key.lower()
#             key: str = key.replace("\'", "")
#             key: str = key.replace("\"", "")
#             key: str = key.replace("key.", "")
#             key: str = key.replace("<25>", "한자 or ctrl_r")
#             key: str = key.replace("<21>", "한영 or alt_r")
#             key: str = key.replace("12", "텐키 5")  # 텐키
#             key: str = key.replace("cmd", "win")
#             key: str = key.replace("page", "pg")
#             key: str = key.replace("down", "dn")
#             if key != "num_lock":
#                 key: str = key.replace("_l", "")
#             key: str = key.replace("_r", "")
#
#             key: str = key.replace(" ", "")
#             key: str = key.replace("<", "")
#             key: str = key.replace(">", "")
#             # key: str = key.replace("_", "")
#
#             # 전처리 후 출력
#             # print(str(key))
#
#             log_title = "키보드릴리즈"
#             info = f"Park4139.sleep({elapsed_time})\nPark4139.keyUp('{key}') "
#             DebuggingUtil.print_magenta(info)
#             self.save_macro_log(contents=f"{TimeUtil.get_time_as_('%Y-%m-%d_%H:%M:%S')}{log_title} {info}")
#
#             # 현재시간을 이전시간에 저장
#             self.previous_time = current_time
#             pass
#
#         def bring_this_window(self):
#             self.show()
#             self.activateWindow()
#             # import win32gui
#             # active_window = win32gui.GetForegroundWindow()
#             # win32gui.SetForegroundWindow(active_window)
#
#     class CustomQthread(QThread):
#         """pyside6 app 전용 객체"""
#         finished = Signal()  # run()의 성공 신호를 보내기 위해 필요, run() 은 상속된 QThread 객체의 run()을 통해서 오버라이딩 되는 것 보인다.
#
#         def __init__(self, q_application):  # 이 생성자는 q_application 를 받기 위해서 내가 작성 q_application 가 필요없는 void method QThread 이용 시 그냥 없애면 되는 생성자.
#             super().__init__()
#             self.q_application = q_application
#
#         def run(self):
#             # QThread 동작 테스트 코드
#             print("QThread 초기화 시작")
#             for i in range(10):
#                 print(i)
#                 self.msleep(30)
#
#             def tmp2(q_application: QApplication):
#                 global dialog2
#                 dialog2 = UiUtil.CustomDialog(q_application=q_application, q_wiget=UiUtil.CustomQdialog(ment="테스트", btns=["실행", "실행하지 않기"]), is_app_instance_mode=False)
#                 if dialog2.btn_text_clicked == "실행":
#                     TextToSpeechUtil.speak_ments("정상적으로 동작합니다", sleep_after_play=0.65)
#
#             def tmp(string: str):
#                 TextToSpeechUtil.speak_ments(string, sleep_after_play=0.65)
#                 # global tmpc
#                 # dialog = CustomDialog(q_application=q_application, q_wiget=RpaProgramMainWindow(), is_app_instance_mode=True, is_exec_mode=True)
#                 # global dialog
#                 # dialog = UiUtil.CustomQdialog(context="테스트", buttons=["확인"], is_input_text_box=True, input_text_default="???")
#                 # dialog.show()
#                 # dialog = CustomDialog(q_application=q_application, q_wiget=UiUtil.CustomQdialog(context="테스트", buttons=["실행", "실행하지 않기"], starting_timer=True, closing_timer=False), is_app_instance_mode=True)
#
#             # 프로그램 외 단축키 이벤트 설정
#             shortcut_keys_up_promised = {
#                 # "<ctrl>+<cmd>": self.close, #success
#                 # "<ctrl>+T": partial(tmp, "test"), #success
#                 # "<ctrl>+Q": partial(tmp2, self.q_application), #fail
#                 # "<ctrl>+h": partial(q_wiget.toogle_rpa_window, "<ctrl>+h"), #fail
#             }
#             keyboard_main_listener = pynput.keyboard.GlobalHotKeys(hotkeys=shortcut_keys_up_promised)
#             keyboard_main_listener.start()
#
#             self.finished.emit()  # 작업이 성공되었음을 신호로 알림 # flutter 상태관리와 비슷
#             print("QThread 실행 증...")
#
#     # class SharedObject(QObject):
#     #     """pyside6 app 전용 공유객체 (pyside6 app 상태관리용)"""
#     #     dataChanged = Signal(str)  # 데이터 변경을 알리는 시그널
#     #
#     #     def __init__(self):
#     #         super().__init__()
#     #         self._data = ""
#     #         self.answer = ""
#     #         self.question = ""
#     #         self.rpa_program_main_window = None
#     #
#     #     #  /////////////////////////////////////////////// SharedObject.data
#     #     @property
#     #     def data(self):
#     #         return self._data
#     #
#     #     @data.setter
#     #     def data(self, value):
#     #         self._data = value
#     #         self.dataChanged.emit(self._data)  # 데이터 변경 시 시그널 발생
#     #
#     #     #  /////////////////////////////////////////////// 공유객체 SharedObject.answer
#     #     @property
#     #     def data(self):
#     #         return self.answer
#     #
#     #     @data.setter
#     #     def data(self, value):
#     #         self.answer = value
#     #         self.dataChanged.emit(self.answer)
#     #
#     #     #  SharedObject.question
#     #     @property
#     #     def data(self):
#     #         return self.question
#     #
#     #     @data.setter
#     #     def data(self, value):
#     #         self.question = value
#     #         self.dataChanged.emit(self.question)
#     #
#     #     #  SharedObject.rpa_program_main_window ....이래도 되나 모르겠네... 성능 이슈 있을 수도...
#     #     @property
#     #     def data(self):
#     #         return self.rpa_program_main_window
#     #
#     #     @data.setter
#     #     def data(self, value):
#     #         self.rpa_program_main_window = value
#     #         self.dataChanged.emit(self.rpa_program_main_window)
#
#     # @staticmethod
#     # def pop_up_as_complete(title: str, ment: str, auto_click_positive_btn_after_seconds: int, input_text_default=""):
#     #     """결과팝업, 자동클릭설정가능"""
#     #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#     #     FileSystemUtil.play_wav_file(file_abspath=StateManagementUtil.POP_SOUND_WAV)
#     #
#     #     app_foo = None
#     #
#     #     app = QApplication.instance()
#     #     if app is None:
#     #         app_foo = QApplication()
#     #     if input_text_default == "":
#     #         is_input_text_box = False
#     #     else:
#     #         is_input_text_box = True
#     #
#     #     dialog_ = UiUtil.CustomQdialog(title=title, ment=ment, btns=["확인"], is_input_box=is_input_text_box, input_box_text_default=input_text_default, auto_click_positive_btn_after_seconds=auto_click_positive_btn_after_seconds)
#     #     dialog_.exec()
#     #     btn_text_clicked = dialog_.btn_text_clicked
#     #
#     #     if btn_text_clicked == "":
#     #         DebuggingUtil.print_magenta()(f'버튼 눌렸습니다 입니다 {btn_text_clicked}')
#     #     if app == True:
#     #         if isinstance(app_foo, QApplication):
#     #             app_foo.exec()
#     #     if app == True:
#     #         app_foo.shutdown()
#     # 백업
#
#     @staticmethod
#     def pop_up_as_complete(title_: str, ment: str, auto_click_positive_btn_after_seconds: int, input_text_default=""):
#         """결과팝업, 자동클릭설정가능"""
#         if FileSystemUtil.is_os_windows():
#             DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
#             FileSystemUtil.play_wav_file(file_abspath=StateManagementUtil.POP_SOUND_WAV)
#
#             app_foo = None
#
#             # QApplication 인스턴스 확인
#             app = QApplication.instance()
#             if app is None:
#                 app_foo = QApplication()
#             if input_text_default == "":
#                 is_input_text_box = False
#             else:
#                 is_input_text_box = True
#
#             # Shadows name 'dialog' from outer scope"는 변수명이 외부 범위에서 선언된 'dialog'와 동일한 이름으로 선언되어서 발생한 오류입니다.
#             dialog_ = UiUtil.CustomQdialog(title=title_, ment=ment, btns=["확인"], is_input_box=is_input_text_box, input_box_text_default=input_text_default, auto_click_positive_btn_after_seconds=auto_click_positive_btn_after_seconds)
#             dialog_.exec()
#             btn_text_clicked = dialog_.btn_text_clicked
#
#             if btn_text_clicked == "":
#                 DebuggingUtil.print_magenta(f'버튼 눌렸습니다 입니다 {btn_text_clicked}')
#             if app == True:  # .....app 은 bool 이 아닌데. 동작 되고있는데..
#                 DebuggingUtil.print_magenta("여기는 좀 확인을 해야하는데. 호출 안되면 좋겠는데1")
#                 if isinstance(app_foo, QApplication):
#                     DebuggingUtil.print_magenta("여기는 좀 확인을 해야하는데. 호출 안되면 좋겠는데2")
#                     app_foo.exec()
#             if app == True:
#                 # app_foo.quit()# QApplication 인스턴스 제거시도 : fail
#                 # app_foo.deleteLater()# QApplication 인스턴스 파괴시도 : fail
#                 # del app_foo # QApplication 인스턴스 파괴시도 : fail
#                 # app_foo = None # QApplication 인스턴스 파괴시도 : fail
#                 DebuggingUtil.print_magenta("여기는 좀 확인을 해야하는데. 호출 안되면 좋겠는데3")
#                 app_foo.shutdown()  # QApplication 인스턴스 파괴시도 : success  # 성공요인은 app.shutdown()이 호출이 되면서 메모리를 해제까지 수행해주기 때문
#                 # sys.exit()
#         else:
#             DebuggingUtil.print_ment_success(f"{ment}")
#
#
# class Pyside6Util:
#     @staticmethod
#     def get_font_for_pyside6(font_path):
#         font_id = QFontDatabase.addApplicationFont(font_path)
#         font_name = QFontDatabase.applicationFontFamilies(font_id)[0]
#         font = QFont(font_name)
#         font.setFixedPitch(True)
#         return font


class DebuggingUtil:
    # @afterpause # gui 가 뜨면 pause() 수행되는 것과 동일한 효과가 나타난다.
    @staticmethod
    def trouble_shoot(trouble_id: str):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        ment = f"TROUBLE SHOOT REPORT:\n{traceback.format_exc()}\ntrouble_id : {trouble_id}"
        print(ment)
        BusinessLogicUtil.debug(ment, input_text_default=trouble_id, auto_click_positive_btn_after_seconds=600)
        return trouble_id

    @staticmethod
    def commentize(ment):
        # print(f'{StateManagementUtil.LINE_LENGTH_PROMISED} {ment}')
        # DebuggingUtil.print_ment_via_colorama(f'{StateManagementUtil.LINE_LENGTH_PROMISED} {ment}', colorama_color=ColoramaColorUtil.LIGHTGREEN_EX)
        DebuggingUtil.print_ment_light_black(f'{StateManagementUtil.UNDERLINE_PROMISED} {ment}')
        # self.speak(title) # 생각보다 너무 말이 많아 주석처리
        return f'{StateManagementUtil.UNDERLINE_PROMISED} {ment}'

    # @staticmethod
    # 시작로깅(json 형태로 넣을 수 있도록 코드 업데이트 할것)
    # def log_s(log_title="시작로깅"):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     starting_directory = os.getcwd()
    #     FileSystemUtil.get_cmd_output('chcp 65001 >nul')  # 한글 엔코딩 설정
    #     BusinessLogicUtil.time_s = time.time()  # 서버라이프사이클 계산용 변수 설정
    #     server_time = TimeUtil.get_time_as_(f'%Y-%m-%d %H:%M:%S')
    #     cmd = rf'echo "server_time  : {server_time} ,  starting_directory  : {starting_directory},  __file__  : {__file__},  log_title : {log_title} " >> "{BusinessLogicUtil.LOG_DIRECTORY}\success.log"'
    #     FileSystemUtil.get_cmd_output(cmd)

    @staticmethod
    def log_mid(log_title="중간로깅"):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        starting_directory = os.getcwd()
        server_time = TimeUtil.get_time_as_(f'%Y-%m-%d %H:%M:%S')
        cmd = rf'echo "server_time  : {server_time} ,  starting_directory  : {starting_directory},  __file__  : {__file__},  log_title : {log_title} " >> "{StateManagementUtil.LOG_DIRECTORY}\success.log"'
        FileSystemUtil.get_cmd_output(cmd)

    # @staticmethod
    # def log_e(log_title="종료로깅"):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     starting_directory = os.getcwd()
    #     BusinessLogicUtil.time_e = time.time()  # 서버라이프사이클 계산용 변수 설정
    #     server_time = TimeUtil.get_time_as_(f'%Y-%m-%d %H:%M:%S')
    #     server_life_cycle = BusinessLogicUtil.time_e - BusinessLogicUtil.time_s
    #     lines = subprocess.check_output(
    #         rf'echo "server_time  : {server_time} ,  starting_directory  : {starting_directory},  __file__  : {__file__},  log_title : {log_title},  server_life_cycle : {server_life_cycle}  " >> "{BusinessLogicUtil.LOG_DIRECTORY}\success.log"',
    #         shell=True).decode('utf-8').split(
    #         '\n')
    # os.system('cls')
    @staticmethod
    # def print_ment_via_colorama(ment, color:Union[ColoramaColorUtil.GREEN, ColoramaColorUtil.RED]):
    def print_ment_via_colorama(ment, colorama_color: Enum):
        # colorama 초기화
        # init(autoreset=True)  # 색상전이를 막을 수 있음, 최적화를 한다면 연산되어 콘솔에 출력되는 것을 모두 중단처리. StateManagementUtil.is_op_mode == True 에서 동작하도록 하는 것도 방법이다.
        # 혹시나 싶었는데 console_blurred 에서 팝업 기능과 충돌이 되는 것으로 보인다.

        # colorama_color 값에 해당하는 Fore 값으로 매핑, 리펙토리 후
        colorama_to_fore = {
            ColoramaUtil.BLACK: Fore.BLACK,
            ColoramaUtil.RED: Fore.RED,
            ColoramaUtil.GREEN: Fore.GREEN,
            ColoramaUtil.YELLOW: Fore.YELLOW,
            ColoramaUtil.BLUE: Fore.BLUE,
            ColoramaUtil.MAGENTA: Fore.MAGENTA,
            ColoramaUtil.CYAN: Fore.CYAN,
            ColoramaUtil.WHITE: Fore.WHITE,
            ColoramaUtil.RESET: Fore.RESET,
            ColoramaUtil.LIGHTBLACK_EX: Fore.LIGHTBLACK_EX,
            ColoramaUtil.LIGHTRED_EX: Fore.LIGHTRED_EX,
            ColoramaUtil.LIGHTGREEN_EX: Fore.LIGHTGREEN_EX,
            ColoramaUtil.LIGHTYELLOW_EX: Fore.LIGHTYELLOW_EX,
            ColoramaUtil.LIGHTBLUE_EX: Fore.LIGHTBLUE_EX,
            ColoramaUtil.LIGHTMAGENTA_EX: Fore.LIGHTMAGENTA_EX,
            ColoramaUtil.LIGHTCYAN_EX: Fore.LIGHTCYAN_EX,
            ColoramaUtil.LIGHTWHITE_EX: Fore.LIGHTWHITE_EX
        }
        try:
            colorama_color = colorama_to_fore.get(colorama_color, Fore.RESET)
        finally:
            print(f"{colorama_color}{ment}")
        # colorama_color = colorama_to_fore.get(ColoramaUtil.LIGHTBLACK_EX, Fore.RESET)
        # print(f"{colorama_color}")

    @staticmethod
    def print_ment_fail(ment):
        DebuggingUtil.print_ment_via_colorama(ment, colorama_color=ColoramaUtil.RED)

    @staticmethod
    def print_ment_success(ment):
        DebuggingUtil.print_ment_via_colorama(ment, colorama_color=ColoramaUtil.LIGHTGREEN_EX)

    @staticmethod
    def print_magenta(ment):
        DebuggingUtil.print_ment_via_colorama(ment, colorama_color=ColoramaUtil.LIGHTMAGENTA_EX)

    @staticmethod
    def print_ment_light_white(ment):
        DebuggingUtil.print_ment_via_colorama(ment, colorama_color=ColoramaUtil.LIGHTWHITE_EX)

    @staticmethod
    def print_ment_light_yellow(ment):
        DebuggingUtil.print_ment_via_colorama(ment, colorama_color=ColoramaUtil.LIGHTYELLOW_EX)

    @staticmethod
    def print_ment_yellow(ment):
        DebuggingUtil.print_ment_via_colorama(ment, colorama_color=ColoramaUtil.YELLOW)

    @staticmethod
    def print_ment_blue(ment):
        DebuggingUtil.print_ment_via_colorama(ment, colorama_color=ColoramaUtil.BLUE)

    @staticmethod
    def print_ment_light_blue(ment):
        DebuggingUtil.print_ment_via_colorama(ment, colorama_color=ColoramaUtil.LIGHTBLUE_EX)

    @staticmethod
    def print_ment_red(ment):
        DebuggingUtil.print_ment_via_colorama(ment, colorama_color=ColoramaUtil.RED)

    @staticmethod
    def print_ment_green(ment):
        DebuggingUtil.print_ment_via_colorama(ment, colorama_color=ColoramaUtil.LIGHTGREEN_EX)

    @staticmethod
    def print_ment_light_black(ment):
        DebuggingUtil.print_ment_via_colorama(ment, colorama_color=ColoramaUtil.LIGHTBLACK_EX)

    @staticmethod
    def print_ment_cyan(ment):
        DebuggingUtil.print_ment_via_colorama(ment, colorama_color=ColoramaUtil.CYAN)


class CustomErrorUtil(Exception):
    def __init__(self, message):
        self.message = message
        super().__init__(self.message)


class MentsUtil:
    NOT_PREPARED_YET = "아직 준비되지 않은 서비스입니다"
    NO = "아니"
    YES = "응"
    CHECKED = "확인"
    OK_I_WILL_DO_IT_NOW = "지금할게"
    DONE = "했어"
    I_WANT_TO_TO_DO_NEXT_TIME = "다음에 할게"


class DataStructureUtil:
    @staticmethod
    def get_list_added_elements_alternatively(list_for_odd, list_for_even):  # from [1, 2, 3] + [ x, y, z] to [1,x,2,y,3,z]
        """
        두 리스트를 더할때 사용하는 함수
        list, ndarray 모두 사용이 가능한 함수이다.

        result = get_list_added_elements_alternatively(list1 = [1, 2, 3], list_to_added_as_even_order = ['x', 'y', 'z'])
        """
        result = []
        for i in range(len(list_for_odd)):
            result.append(list_for_odd[i])
            if i < len(list_for_even):
                result.append(list_for_even[i])
        return result

    @staticmethod
    def print_list_each_two_elements(list: []):  # print(rf'list[index] list[index+1] : {list[index]} {list[index+1]}') without out of index
        for index, item in enumerate(list):  # enumerate 로 리스트의 원소를 2개씩 출력
            if index + 1 < len(list):
                print(rf'list[index] list[index+1] : {list[index]} {list[index + 1]}')

    @staticmethod
    def get_list_each_two_elements_joined(list: []):  # from ["a", "b", "c", "d"] to ["ab", "cd"]
        return [f"{list[i]}{list[i + 1]}" for i in range(0, len(list), 2)]

    @staticmethod
    def get_nested_list_grouped_by_each_two_elements_in_list(list: []):  # from ["a", "b", "c", "d"] to [["a","b"], ["c","d"]]
        result = []
        for i in range(0, len(list), 2):
            if i + 1 < len(list):
                result.append([list[i], list[i + 1]])
        return result

    @staticmethod
    def get_column_of_2_dimension_list(list_2_dimension: [], column_no):  # return 은 list 아니고 ndarray 일 수 있다
        import numpy as np
        arr = np.array(list_2_dimension)
        second_column = arr[:, column_no]
        print(second_column)

    @staticmethod
    def get_nested_list_converted_from_ndarray(ndarray: numpy.ndarray):  # ndarray 에서 list 로 변환 # ndarray 에서 list 로 변환 # ndarray to nested []  from [[1 2]] to [[1, 2]] # from [ [str str] [str str] ]  to  [ [str, str], [str, str] ]
        return ndarray.tolist()

    @staticmethod
    def get_nested_list_sorted_by_column_index(nested_list: [[str]], column_index: int, decending_order=False):  # tree depth(sample[1]) 에 대한 내림차순 정렬 # list 2차원 배열의 특정 열의 내림차순 정렬 # from [[str, str]] to [[str, str]]
        """
         중첩리스트를 첫 번째 행을 기준으로 내림차순으로 정렬
         from
         [
             [ 1 a z ]
             [ 7 a b ]
             [ 4 a c ]
         ]
         to
         [
             [ 7 a b ]
             [ 4 a c ]
             [ 1 a z ]
         ]
         튜플도 된다
         from
         [
             ( 1 a z )
             ( 7 a b )
             ( 4 a c )
         ]
         to
         [
             ( 7 a b )
             ( 4 a c )
             ( 1 a z )
         ]
        """

        def bubble_sort_nested_list(nested_list, column_index):
            """버블 소팅 테스트"""
            if type(nested_list[0][column_index]) == float:
                column_index = int(column_index)
                n = len(nested_list)
                for i in range(n):
                    for j in range(0, n - i - 1):
                        if decending_order == True:
                            if int(str(nested_list[j][column_index]).replace(".", "")) < int(str(nested_list[j + 1][column_index]).replace(".", "")):
                                nested_list[j], nested_list[j + 1] = nested_list[j + 1], nested_list[j]
                        elif decending_order == False:
                            if int(str(nested_list[j][column_index]).replace(".", "")) > int(str(nested_list[j + 1][column_index]).replace(".", "")):
                                nested_list[j], nested_list[j + 1] = nested_list[j + 1], nested_list[j]
                return nested_list
            elif type(nested_list[0][column_index]) == int or nested_list[0][column_index].isdigit():  # 에고 힘들다. 머리아팠다
                column_index = int(column_index)
                n = len(nested_list)
                for i in range(n):
                    for j in range(0, n - i - 1):
                        if decending_order == True:
                            if int(nested_list[j][column_index]) < int(nested_list[j + 1][column_index]):
                                # print(rf'{nested_list[j][column_index]}>{nested_list[j + 1][column_index]}')
                                # print(rf'nested_list[j][column_index] > nested_list[j+1][column_index] : {int(nested_list[j][column_index]) > int(nested_list[j+1][column_index])}')
                                nested_list[j], nested_list[j + 1] = nested_list[j + 1], nested_list[j]
                        elif decending_order == False:
                            if int(nested_list[j][column_index]) > int(nested_list[j + 1][column_index]):
                                # print(rf'{nested_list[j][column_index]}>{nested_list[j + 1][column_index]}')
                                # print(rf'nested_list[j][column_index] > nested_list[j+1][column_index] : {int(nested_list[j][column_index]) > int(nested_list[j+1][column_index])}')
                                nested_list[j], nested_list[j + 1] = nested_list[j + 1], nested_list[j]
                return nested_list
            elif type(nested_list[0][column_index]) == str:
                column_index = int(column_index)
                n = len(nested_list)
                for i in range(n):
                    for j in range(0, n - i - 1):
                        if decending_order == True:
                            if nested_list[j][column_index] < nested_list[j + 1][column_index]:
                                # print(rf'{nested_list[j][column_index]}>{nested_list[j + 1][column_index]}')
                                # print(rf'nested_list[j][column_index] > nested_list[j+1][column_index] : {int(nested_list[j][column_index]) > int(nested_list[j+1][column_index])}')
                                nested_list[j], nested_list[j + 1] = nested_list[j + 1], nested_list[j]
                        if decending_order == False:
                            if nested_list[j][column_index] > nested_list[j + 1][column_index]:
                                # print(rf'{nested_list[j][column_index]}>{nested_list[j + 1][column_index]}')
                                # print(rf'nested_list[j][column_index] > nested_list[j+1][column_index] : {int(nested_list[j][column_index]) > int(nested_list[j+1][column_index])}')
                                nested_list[j], nested_list[j + 1] = nested_list[j + 1], nested_list[j]
                return nested_list

        return bubble_sort_nested_list(nested_list, column_index)

    @staticmethod
    def get_nested_list_removed_row_that_have_nth_element_dulplicated_by_column_index_for_ndarray(ndarray: [[]], column_index: int):  # 중복된 행 제거하는게 아니고 행의 2번째 요소가 중복되는 것을 제거
        seen = set()
        unique_rows = []
        for row in ndarray:
            if row[column_index] not in seen:
                unique_rows.append(row)
                seen.add(row[column_index])
        return unique_rows

    @staticmethod
    def get_nested_list_removed_row_that_have_nth_element_dulplicated_by_column_index(nested_list: [[]], column_index: int):  # 중복된 행 제거하는게 아니고 행의 2번째 요소가 중복되는 것을 제거
        seen = set()  # 중복된 행의 두 번째 요소를 추적하기 위한 집합(set) 생성
        unique_rows = []
        for row in nested_list:
            if row[column_index] not in seen:
                unique_rows.append(row)  # 중복된 행의 두 번째 요소가 중복되지 않는 행들만 추출하여 unique_rows 리스트에 추가
                seen.add(row[column_index])
        return unique_rows  # 중복된 행의 두 번째 요소가 중복되지 않는 행들만 남게 됩니다.

    @staticmethod
    def get_list_seperated_by_each_elements_in_nested_list(nested_list):
        return [item for sublist in nested_list for item in sublist]

    @staticmethod
    def is_two_lists_equal(list1, list2):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        # 두 리스트의 요소들이 동일한지만 확인 # 하나라도 발견되면 탐색 중지 # 두 리스트의 동일 인덱스에서 진행하다 달라도 다른거다
        if len(list1) != len(list2):
            print(f"두 리스트의 줄 수가 다릅니다 {len(list1)}, {len(list2)}")
            return False
        for i in range(len(list1)):
            if list1[i] != list2[i]:
                print(f"두 리스트의 같은 인덱스 중 다른 리스트 발견={list1[i]}, {list2[i]}")
                return False
        return True

    @staticmethod
    def get_elements_that_list1_only_have(list1, list2):  # 두 개의 리스트를 비교하여 특정 하나의 리스트 만 가진 요소만 모아서 리스트로 출력
        unique_elements_of_list1 = []
        for element in list1:
            if element not in list2:
                unique_elements_of_list1.append(element)
        return unique_elements_of_list1

    @staticmethod
    def get_different_elements(list1, list2):  # 두 개의 리스트를 비교하여 서로 다른 요소만 모아서 리스트로 출력
        different_elements = []
        # 두 리스트 중 더 긴 리스트의 길이를 구합니다.
        max_length = max(len(list1), len(list2))
        # 범위는 더 긴 리스트의 길이로 설정합니다.
        for i in range(max_length):
            # 리스트의 인덱스가 범위 내에 있는지 확인 후 비교합니다.
            if i < len(list1) and i < len(list2):
                if list1[i] != list2[i]:
                    different_elements.append(list1[i])
            else:
                # 한 리스트는 범위를 벗어났으므로 해당 요소를 추가합니다.
                if i >= len(list1):
                    different_elements.append(list2[i])
                else:
                    different_elements.append(list1[i])
        # 위의 코드는 5 초 경과

        different_elements = []
        # set1 = set(list1)  # list to set
        # set2 = set(list2)
        # different_elements = list(set1.symmetric_difference(set2))
        # 위의 코드는 12 초 경과
        return different_elements

    @staticmethod
    def get_common_elements(list1, list2):  # 두 개의 리스트를 비교하여 서로 동일한 요소만 새로운 리스트로 출력 # 중복값 색출
        common_elements = []
        for element in list1:
            if element in list2:
                common_elements.append(element)
        return common_elements

    @staticmethod
    def add_suffix_to_string_list(string_list, suffix):
        result = []
        for string in string_list:
            result.append(string + suffix)
        return result

    @staticmethod
    def add_prefix_to_string_list(string_list, prefix):
        result = []
        for string in string_list:
            result.append(prefix + string)
        return result


class PerformanceOptimizingUtil:
    @TestUtil.measure_seconds_performance_nth
    @staticmethod
    def generate_mp3_file_for_time_performance():  # time performance : 9028 초 /60 /60  =  2.5 시간
        """
        deprecated
        시간에 대한 mp3 파일 미리 만들어 놓음으로 생성시간 저감 기대 > speaks() 메소드 최적화
        """
        for HH in range(24, 0, -1):
            for mm in range(0, 60):
                DebuggingUtil.commentize(f'{int(HH)}시')
                DebuggingUtil.commentize(f'{int(mm)}분 입니다')

    @staticmethod
    def gen_dictionary_for_monitor_target_edited_and_back_up(directory_abspath):
        r"""
        deprecated
        monitor_target_edited_and_back_up() 성능개량용 코드
        lzw 알고리즘 딕셔너리 생성용 성능개량용 코드 > 이건 추후에 생각해보기로
        딕셔너리 업데이트 필요한 경우에 사용해서 콘솔에 출력된 딕셔너리 초기화용 코드를
        딕셔너리를 코드베이스에 하드코딩하여 사용

        생성된 딕셔너리 코드 예시
        dictionary_based_on_tri_structure = [
            "{PROJECT_DIRECTORY}\$cache_fonts\Montserrat\static",
            "{PROJECT_DIRECTORY}\$cache_fonts\Noto_Sans_KR\static",
            "{PROJECT_DIRECTORY}\$cache_fonts\GmarketSans",
            "{PROJECT_DIRECTORY}\$cache_fonts\Montserrat",
            "{PROJECT_DIRECTORY}\$cache_fonts\Noto_Sans_KR",
            "{PROJECT_DIRECTORY}\$cache_fonts\Poppins",
            "{PROJECT_DIRECTORY}\$cache_database",
            "{PROJECT_DIRECTORY}",
        ]
        """
        current_directory_state = FileSystemUtil.get_directory_files_files_for_monitoring_mtime_without_files_to_exclude(directory_abspath=directory_abspath)  # str to dict
        current_directory_state = [f"{key}" for key, value in current_directory_state.items()]  # from dict to ["key\n"]   # 이건 함수는 테스트해보고, # noqa 처리 해야할 듯
        print(rf'type(current_directory_state) : {type(current_directory_state)}')
        print(rf'len(current_directory_state) : {len(current_directory_state)}')

        current_target_files = current_directory_state
        # current_target_files = current_directory_state[0:100]  # 샘플 100개만 테스트
        dirnames = [os.path.dirname(sample) for sample in current_target_files]  # 파일 dirname 목록
        tree_levels = [FileSystemUtil.get_tree_depth_level(sample) for sample in current_target_files]  # 파일시스템 트리깊이레벨 목록
        dirnames_and_trees = DataStructureUtil.get_list_added_elements_alternatively(dirnames, tree_levels)  # from [][] to []
        dirnames_and_trees = DataStructureUtil.get_nested_list_grouped_by_each_two_elements_in_list(dirnames_and_trees)
        # [print(sample) for sample in dirnames_and_trees]
        # print(rf'dirnames_and_trees : {dirnames_and_trees}')
        # print(rf'type(dirnames_and_trees) : {type(dirnames_and_trees)}')
        # print(rf'len(dirnames_and_trees) : {len(dirnames_and_trees)}')
        dirnames_and_trees = DataStructureUtil.get_nested_list_sorted_by_column_index(nested_list=dirnames_and_trees, column_index=1, decending_order=True)  # tree depth를 의미하는 column_index=1 에 대한 내림차순 정렬
        # [print(sample) for sample in dirnames_and_trees]
        # print(rf'dirnames_and_trees : {dirnames_and_trees}')
        # print(rf'type(dirnames_and_trees) : {type(dirnames_and_trees)}')
        # print(rf'len(dirnames_and_trees) : {len(dirnames_and_trees)}')
        dirnames_and_trees = DataStructureUtil.get_nested_list_removed_row_that_have_nth_element_dulplicated_by_column_index(nested_list=dirnames_and_trees, column_index=0)  # from [[]] to [[]] # from [[1 2]] to [[1, 2]] # from [ [str str] [str str] ]  to  [ [str, str], [str, str]]
        # [print(sample) for sample in dirnames_and_trees]
        # print(rf'dirnames_and_trees : {dirnames_and_trees}')
        # print(rf'type(dirnames_and_trees) : {type(dirnames_and_trees)}')
        # print(rf'len(dirnames_and_trees) : {len(dirnames_and_trees)}')
        dirnames = [x[0] for x in dirnames_and_trees]  # dirnames_and_trees 의 첫번째 컬럼인 dirname 만 추출
        DebuggingUtil.commentize("딕셔너리 초기화용 코드 시작")
        print("dictionary_based_on_tri_structure = {")
        [print(rf'    r"{dirname}":"{index}빠릿{index}" ,') for index, dirname in enumerate(dirnames)]
        print("}")
        # print(rf'dirnames : {dirnames}')
        print(rf'type(dirnames) : {type(dirnames)}')
        print(rf'len(dirnames) : {len(dirnames)}')
        DebuggingUtil.commentize("딕셔너리 초기화용 코드 종료")
        TestUtil.pause()

    dictionary_for_monitoring_performance = {
        r"{PROJECT_DIRECTORY}\$cache_fonts\Montserrat\static": "이거지1",
        r"{PROJECT_DIRECTORY}\$cache_fonts\Noto_Sans_KR\static": "이거지2",
        r"{PROJECT_DIRECTORY}\$cache_fonts\GmarketSans": "이거지3",
        r"{PROJECT_DIRECTORY}\$cache_fonts\Montserrat": "이거지4",
        r"{PROJECT_DIRECTORY}\$cache_fonts\Noto_Sans_KR": "이거지5",
        r"{PROJECT_DIRECTORY}\$cache_fonts\Poppins": "이거지6",
        r"{PROJECT_DIRECTORY}\$cache_database": "이거지7",
        r"{PROJECT_DIRECTORY}": "이거지8",
    }

    @staticmethod
    def replace_words_based_on_tri_node(text, dictionary):
        """
        # 사용 예시
        dictionary = ["apple", "banana", "orange"]
        text = "I like apple and banana."
        replaced_text = replace_words(text, dictionary)
        print(replaced_text)
        """

        class TrieNode:
            def __init__(self):
                self.children = {}  # 자식 노드를 저장하는 딕셔너리
                self.is_end_of_word = False  # 단어의 끝인지 나타내는 플래그

        class Trie:
            def __init__(self):
                self.root = TrieNode()  # 루트 노드 생성

            def insert(self, word):
                current_node = self.root
                for char in word:
                    if char not in current_node.children:
                        current_node.children[char] = TrieNode()
                    current_node = current_node.children[char]
                current_node.is_end_of_word = True

            def search(self, word):
                current_node = self.root
                for char in word:
                    if char not in current_node.children:
                        return False
                    current_node = current_node.children[char]
                return current_node.is_end_of_word

        trie = Trie()
        for word in dictionary:
            trie.insert(word)

        words = text.split()
        replaced_words = []
        for word in words:
            if trie.search(word):
                replaced_words.append("REPLACED")
            else:
                replaced_words.append(word)

        return " ".join(replaced_words)


class FileSystemUtil:
    @staticmethod
    def get_tree_depth_level(file_abspath: str):
        return len(file_abspath.split("\\")) - 1

    @staticmethod
    def get_target_as_pnx(target_abspath):
        """디렉토리/파일 대상 테스트 성공"""
        return target_abspath

    @staticmethod
    def get_target_as_pn(target_abspath):
        """디렉토리/파일 대상 테스트 성공"""
        return rf"{os.path.dirname(target_abspath)}/{os.path.splitext(os.path.basename(target_abspath))[0]}"

    @staticmethod
    def get_target_as_nx(target_abspath):
        """디렉토리/파일 대상 테스트 성공"""
        return rf"{os.path.splitext(os.path.basename(target_abspath))[0]}{os.path.splitext(os.path.basename(target_abspath))[1]}"

    @staticmethod
    def get_target_as_x(target_abspath):
        """디렉토리/파일 대상 테스트 성공"""
        return rf"{os.path.splitext(os.path.basename(target_abspath))[1]}"

    @staticmethod
    def get_target_as_n(target_abspath):
        """디렉토리/파일 대상 테스트 성공"""
        return rf"{os.path.splitext(os.path.basename(target_abspath))[0]}"

    @staticmethod
    def convert_mkv_to_wav(file_mkv):
        FileSystemUtil.get_cmd_output(rf'ffmpeg -i "{file_mkv}" -ab 160k -ac 2 -ar 44100 -vn {FileSystemUtil.get_target_as_pn(file_mkv)}.wav')

    @staticmethod
    def convert_mp4_to_webm(target_abspath):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        '''테스트 필요'''
        DebuggingUtil.print_magenta(f'from : {target_abspath}')
        file_edited = f'{os.path.splitext(os.path.basename(target_abspath))[0]}.webm'
        DebuggingUtil.print_magenta(f'to   : {file_edited}')

        path_started = os.getcwd()
        os.system("chcp 65001 >nul")
        os.system('mkdir storage >nul')
        os.chdir('storage')
        os.system(f'"{StateManagementUtil.FFMPEG_EXE}" -i "{target_abspath}" -f webm -c:v libvpx -b:v 1M -acodec libvorbis "{file_edited}" -hide_banner')
        os.chdir(path_started)

    @staticmethod
    def convert_wav_to_flac(target_abspath):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        '''테스트 필요'''

        # 한글 인코딩 setting
        os.system("chcp 65001 >nul")

        DebuggingUtil.print_magenta(f'from : {target_abspath}')
        file_edited = f'{os.path.splitext(os.path.basename(target_abspath))[0]}.flac'
        DebuggingUtil.print_magenta(f'to   : {file_edited}')

        # ffmpeg location setting
        ffmpeg_exe = rf"{StateManagementUtil.USERPROFILE}\Desktop\`workspace\tool\LosslessCut-win-x64\resources\ffmpeg.exe"
        destination = 'storage'
        try:
            os.makedirs(destination)
        except Exception as e:
            pass
        os.chdir(destination)
        DebuggingUtil.print_magenta(f'"{ffmpeg_exe}" -i "{target_abspath}" -c:a flac "{file_edited}"        를 수행합니다.')
        subprocess.check_output(f'"{ffmpeg_exe}" -i "{target_abspath}" -c:a flac "{file_edited}"', shell=True)

    # @staticmethod
    # def convert_mp4_to_wav(target_abspath):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     '''테스트 필요'''
    #     DebuggingUtil.print_magenta(f'from : {target_abspath}')
    #     file_edited = f'{os.path.splitext(os.path.basename(target_abspath))[0]}.wav'
    #     DebuggingUtil.print_magenta(f'to   : {file_edited}')
    #
    #     path_started = os.getcwd()
    #
    #     os.system('mkdir storage')
    #     os.chdir('storage')
    #     if os.path.splitext(os.path.basename(target_abspath))[1] == '.mp4':
    #         videoclip = VideoFileClip(target_abspath)
    #         audioclip = videoclip.audio
    #
    #         # audioclip.write_audiofile(file_edited, fps= 8000 )
    #         audioclip.write_audiofile(file_edited, fps=44100)
    #         audioclip.close()
    #         videoclip.close()
    #
    #     os.chdir(path_started)

    @staticmethod
    def convert_mp4_to_flac(target_abspath):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        '''테스트 필요'''
        # 한글 인코딩 setting
        os.system("chcp 65001 >nul")
        DebuggingUtil.print_magenta(f'from : {target_abspath}')
        file_edited = f'{os.path.splitext(os.path.basename(target_abspath))[0]}.flac'
        DebuggingUtil.print_magenta(f'to   : {file_edited}')

        path_started = os.getcwd()

        # ffmpeg location setting
        ffmpeg_exe = rf"{StateManagementUtil.USERPROFILE}\Desktop\`workspace\tool\LosslessCut-win-x64\resources\ffmpeg.exe"
        destination = 'storage'
        try:
            os.makedirs(destination)
        except Exception as e:
            pass
        os.chdir(destination)
        DebuggingUtil.print_magenta(f'"{ffmpeg_exe}" -i "{target_abspath}" -c:a flac "{file_edited}"        를 수행합니다.')
        subprocess.check_output(f'"{ffmpeg_exe}" -i "{target_abspath}" -c:a flac "{file_edited}"', shell=True)

        os.chdir(path_started)

    @staticmethod
    def convert_mp3_to_flac(target_abspath):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        '''테스트 필요'''
        os.system("chcp 65001 >nul")

        DebuggingUtil.print_magenta(f'from : {target_abspath}')
        file_edited = f'{os.path.splitext(os.path.basename(target_abspath))[0]}.flac'
        DebuggingUtil.print_magenta(f'to   : {file_edited}')

        path_started = os.getcwd()

        # ffmpeg location setting
        ffmpeg_exe = rf"{StateManagementUtil.USERPROFILE}\Desktop\`workspace\tool\LosslessCut-win-x64\resources\ffmpeg.exe"
        destination = 'storage'
        try:
            os.makedirs(destination)
        except Exception as e:
            pass
        os.chdir(destination)
        DebuggingUtil.print_magenta(f'"{ffmpeg_exe}" -i "{target_abspath}" -c:a flac "{file_edited}"        를 수행합니다.')
        subprocess.check_output(f'"{ffmpeg_exe}" -i "{target_abspath}" -c:a flac "{file_edited}"', shell=True)

        os.chdir(path_started)

    @staticmethod
    def convert_xls_to_xlsx(target_abspath):
        """
        2024-02-12 15:45 작성 함수 템플릿 샘플
        """
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")

        import pandas as pd
        new_file_xlsx = f'{FileSystemUtil.get_target_as_pn(target_abspath)}.xlsx'

        try:
            if ".xls" == FileSystemUtil.get_target_as_x(target_abspath):
                if not os.path.exists(target_abspath):
                    DebuggingUtil.print_ment_fail(f"{function_name}(), fail, {FileSystemUtil.get_target_as_x(target_abspath)} 는 처리할 수 없는 확장자입니다.")
                    return
        except CustomErrorUtil as e:
            DebuggingUtil.print_ment_fail(f"{function_name}(), fail, {FileSystemUtil.get_target_as_x(target_abspath)} 는 처리할 수 없는 확장자입니다.")
            return

        from zipfile import BadZipFile
        try:
            df = pd.read_excel(target_abspath, engine='openpyxl')  # openpyxl 에만 적용이 되는 함수.. 에러 소지 있음.
            # if not os.path.exists(new_file_xlsx):
            df.to_excel(new_file_xlsx, index=False)
        except BadZipFile as e:
            # 유효데이터만 파싱하여 데이터프레임 으로 변환 후 Excel 파일로 저장
            # read_html() 를 이용하면 손상된 파일을 열 수 ...
            DebuggingUtil.print_ment_yellow(f"확장자가 {FileSystemUtil.get_target_as_x(target_abspath)} 손상된 파일 같습니다. 복구를 시도합니다")
            df = pd.read_html(target_abspath, encoding='utf-8')
            # DebuggingUtil.print_magenta(df)
            # DebuggingUtil.print_magenta(df[0]) # 이번 데이터 구조의 특성상, 불필요 데이터
            # DebuggingUtil.print_magenta(df[1]) # 이번 데이터 구조의 특성상, 불필요 데이터
            # DebuggingUtil.print_magenta(df[2]) # 이번 데이터 구조의 특성상, 유효 데이터 , 데이터프레임 모든컬럼
            # DebuggingUtil.print_magenta(df[2].get(0)) # 데이터프레임 첫번쨰컬럼
            # DebuggingUtil.print_magenta(df[2].get(1)) # 데이터프레임 두번쨰컬럼
            # DebuggingUtil.print_magenta(df[2].get(2)) #
            # DebuggingUtil.print_magenta(df[2].get(6)) #
            # 이번 데이터에서 필요한 데이터, # 0 ~ 6 컬럼까지 유효
            df = df[2]  # 이번 데이터 구조의 특성상, 유효 데이터
            # df_selected = df[[0, 1, 2, 3, 4, 5, 6]]
            # df_selected = df[df.columns[0:7]]
            # df_selected = df[df.columns[7:]]
            # df_selected = df[df.columns[:5]]
            # df_selected = df[df.columns[1:5]]
            df = df[df.columns[:]]  # 모든 컬럼
            # print(rf'df : {df}')
            df.to_excel(new_file_xlsx, index=False)

            # # txt 로 변환
            # file_recovery = f"{FileSystemUtil.get_target_as_pn(target_abspath)}.txt"
            # print(rf'''file_recovery : {file_recovery}''')
            # shutil.copy(target_abspath,file_recovery)

            # # html 로 변환하여, HTML 테이블을 데이터프레임으로 저장
            # file_recovery = f"{FileSystemUtil.get_target_as_pn(target_abspath)}.html"
            # print(rf'''file_recovery : {file_recovery}''')
            # shutil.copy(target_abspath,file_recovery)
            # with open(file_recovery, 'r', encoding='utf-8') as file:
            #     html_content = file.read()
            # soup = BeautifulSoup(html_content, "lxml")
            # # results = soup.find_all(href=re.compile("magnet"), id='link1') # <a class="sister" href="http://example.com/magnet" id="link1">Elsie</a>
            # # results = soup.find_all("table", class_="datatable")  # <table class="datatable">foo!</div>
            # # results = soup.find_all("body")
            # # results = soup.find_all("html")
            # tables = soup.find_all("table")
            # table_selected = tables[2]  # 3번째 테이블 선택
            # df = pd.read_html(str(table_selected))[0]
            # DebuggingUtil.print_magenta(df)

        except Exception as e:
            DebuggingUtil.print_ment_fail(f"{function_name}(), fail, \n {traceback.format_exc()}")
        DebuggingUtil.print_ment_success(f"{function_name}(), success")

    @staticmethod
    def convert_as_zip_with_timestamp(target_abspath):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        starting_directory = ""
        try:
            starting_directory = os.getcwd()

            target_dirname = os.path.dirname(target_abspath)
            target_dirname_dirname = os.path.dirname(target_dirname)
            target_basename = os.path.basename(target_abspath).split(".")[0]
            target_zip = rf'$zip_{target_basename}.zip'
            target_yyyy_mm_dd_HH_MM_SS_zip = rf'{target_basename} - {TimeUtil.get_time_as_("%Y %m %d %H %M %S")}.zip'
            # DebuggingUtil.commentize(rf'# target_dirname_dirname 로 이동')
            os.chdir(target_dirname_dirname)
            # DebuggingUtil.commentize(rf'부모디렉토리로 빽업')
            cmd = f'bandizip.exe c "{target_zip}" "{target_abspath}"'
            FileSystemUtil.get_cmd_output(cmd)
            # DebuggingUtil.commentize(rf'이름변경')
            cmd = f'ren "{target_zip}" "$deleted_{target_yyyy_mm_dd_HH_MM_SS_zip}"'
            FileSystemUtil.get_cmd_output(cmd)
            # DebuggingUtil.commentize(rf'부모디렉토리에서 빽업될 디렉토리로 이동')
            cmd = f'move "$deleted_{target_yyyy_mm_dd_HH_MM_SS_zip}" "{target_dirname}"'
            FileSystemUtil.get_cmd_output(cmd)
            # DebuggingUtil.commentize(rf'빽업될 디렉토리로 이동')
            os.chdir(target_dirname)
            # DebuggingUtil.commentize("os.getcwd()")
            # Park4139.debug_as_cli(os.getcwd())
            # DebuggingUtil.commentize("원본파일삭제")
            os.remove(target_abspath)

        except:
            DebuggingUtil.trouble_shoot("202312030000c")
        finally:
            DebuggingUtil.commentize(rf'프로젝트 디렉토리로 이동')
            os.chdir(starting_directory)

    # @staticmethod
    # def convert_img_to_img_blurred(img_abspath):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     img_converted = Image.open(img_abspath).filter(ImageFilter.GaussianBlur(10))  # 가우시안 블러 적용 # 숫자크면 많이흐려짐
    #     img_converted.show()
    #
    # @staticmethod
    # def convert_img_to_img_grey(img_abspath):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     img_converted = Image.open(img_abspath).convert("L")
    #     img_converted.show()
    #
    # @staticmethod
    # def convert_img_to_img_resized(img_abspath, width_px, height_px):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     img_converted = Image.open(img_abspath).resize((width_px, height_px))
    #     img_converted.show()
    #
    # @staticmethod
    # def convert_img_to_img_cropped(img_abspath, abs_x: int, abs_y: int, width_px: int, height_px: int):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     img_converted = Image.open(img_abspath).crop((abs_x, abs_y, width_px, height_px))
    #     img_converted.show()
    #
    # @staticmethod
    # def convert_img_to_img_rotated(img_abspath, degree: int):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     img_converted = Image.open(img_abspath).rotate(degree)
    #     img_converted.show()
    #     img_converted.save(f"{os.path.dirname(img_abspath)}   {os.path.splitext(img_abspath)[0]}_$flipped_h{os.path.splitext(img_abspath)[1]}")
    #
    # @staticmethod
    # def convert_img_to_img_flipped_horizontally(img_abspath):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     img_converted = Image.open(img_abspath).transpose(Image.FLIP_LEFT_RIGHT)
    #     img_converted.show()
    #     img_converted.save(f"{os.path.dirname(img_abspath)}   {os.path.splitext(img_abspath)[0]}_$flipped_h{os.path.splitext(img_abspath)[1]}")
    #
    # @staticmethod
    # def convert_img_to_img_flipped_vertical(img_abspath):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     img_converted = Image.open(img_abspath).transpose(Image.FLIP_TOP_BOTTOM)
    #     img_converted.show()
    #     img_converted.save(f"{os.path.dirname(img_abspath)}   {os.path.splitext(img_abspath)[0]}_$flipped_v{os.path.splitext(img_abspath)[1]}")

    @staticmethod
    def convert_img_to_img_watermarked(img_abspath):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        from PIL import Image, ImageDraw, ImageFont

        # step2.워터마크 삽입할 이미지 불러오기
        img = Image.open('cat.jpg')
        width, height = img.size

        # step3.그림판에 이미지를 그대로 붙여넣는 느낌의 Draw() 함수
        draw = ImageDraw.Draw(img)

        # step4.삽입할 워터마크 문자
        text = "봵 워터마크"

        # step5.삽입할 문자의 폰트 설정
        font = ImageFont.truetype('/Users/sangwoo/Downloads/나눔 글꼴/나눔손글씨_펜/NanumPen.ttf', 30)

        # step6.삽입할 문자의 높이, 너비 정보 가져오기
        width_txt, height_txt = draw.textsize(text, font)  # noqa

        # step7.워터마크 위치 설정
        margin = 10
        x = width - width_txt - margin
        y = height - height_txt - margin

        # step8.텍스트 적용하기
        draw.text((x, y), text, fill='white', font=font)

        # step9.이미지 출력
        img.show()

        # step10.현재작업 경로에 완성 이미지 저장
        img.save("cat_watermakr.jpg")  # 절대경로 되는지 확인해보자.

    @staticmethod
    def kill_thread(thread_name):
        # 종료할 스레드 이름

        # 현재 실행 중인 모든 스레드 가져오기
        current_threads = threading.enumerate()

        # 종료할 스레드 찾기
        target_thread = None
        for thread in current_threads:
            if thread.name == thread_name:
                target_thread = thread
                break

        # 스레드 종료
        if target_thread:
            target_thread.join()
            print(f"{thread_name} 스레드가 종료되었습니다.")
        else:
            print(f"{thread_name} 스레드를 찾을 수 없습니다.")

    # @staticmethod
    # def is_file_changed(file_abspath):
    #     # 조건문을 반복 작성해서 기능을 분리할 수가 있었다.
    #     # 여기서 알게된 사실은 조건문의 구조를 반복해서 특정 기능들의 로직들을 분리할 수 있었다.
    #     key = DbTomlUtil.get_db_toml_key(file_abspath)
    #
    #     # 기존에 측정된 파일의 줄 수 : 없으면 새로 측정된 파일의 줄 수로 대신함.
    #     line_cnt_from_db = DbTomlUtil.select_db_toml(key=key)
    #     if line_cnt_from_db == None:
    #         DbTomlUtil.insert_db_toml(key=key, value=BusinessLogicUtil.get_line_cnt_of_file(file_abspath))  # 50000 줄의 str 다루는 것보다 50000 개의 list 다루는 것이 속도성능에 대하여 효율적이다.
    #         line_cnt_from_db = DbTomlUtil.select_db_toml(key=key)
    #     DebuggingUtil.print_magenta(f"line_cnt_from_db : {line_cnt_from_db}")
    #
    #     # 새로 측정된 파일의 줄 수
    #     line_cnt_measured = BusinessLogicUtil.get_line_cnt_of_file(file_abspath)
    #     DebuggingUtil.print_magenta(f"line_cnt_measured : {line_cnt_measured}")
    #
    #     # 로직분리 새로운 시도: 기능에 따라 조건문을 여러개 만들어 보았다.
    #     # commentize() 관련된 로직 분리
    #     if FileSystemUtil.is_file_edited(file_abspath) is None:
    #         DebuggingUtil.commentize("데이터베이스 타겟에 대한 key가 없어 key를 생성합니다")
    #     elif FileSystemUtil.is_file_edited(file_abspath):
    #         ment = f'모니터링 중 편집을 감지하였습니다.\n{os.path.basename(file_abspath)}\n 타겟빽업을 시도합니다.\n 타겟에 대한 key를 toml 데이터베이스에 업데이트합니다.'
    #         UiUtil.pop_up_as_complete(title_="모니터링감지보고", ment=ment, auto_click_positive_btn_after_seconds=3)
    #
    #     # db crud 관련된 로직 분리
    #     if FileSystemUtil.is_file_edited(file_abspath):
    #         DbTomlUtil.update_db_toml(key=key, value=BusinessLogicUtil.get_line_cnt_of_file(file_abspath))
    #     elif FileSystemUtil.is_file_edited(file_abspath) is None:
    #         DbTomlUtil.insert_db_toml(key=key, value=BusinessLogicUtil.get_line_cnt_of_file(file_abspath))
    #     if FileSystemUtil.is_file_edited(file_abspath):
    #         return True

    @staticmethod
    def sync_directory_remote(target_abspath):
        """이게 뭐냐면 외부망에 있는 디렉토리를 동기화 할 수 있다. 리눅스 rsync 에 의존하는 기술"""
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        TextToSpeechUtil.speak_ment(MentsUtil.NOT_PREPARED_YET)
        pass

    # @staticmethod
    # def sync_directory_local(target_abspath):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     try:
    #         import os
    #         import shutil
    #         from dirsync import sync
    #
    #         target_abspath = target_abspath
    #         future_abspath = rf"{target_abspath}_sync"
    #         DebuggingUtil.print_magenta(rf'future_abspath : {future_abspath}')
    #
    #         # 기존 작업 폴더가 없는 경우
    #         if not os.path.exists(future_abspath):
    #             shutil.copytree(target_abspath, future_abspath)
    #         else:
    #             FileSystemUtil.remove_target_parmanently(StateManagementUtil.DIRSYNC_LOG)
    #             # 로깅 설정 및 Park4139.DIRSYNC_LOG 생성
    #             import logging
    #             logging.basicConfig(filename=StateManagementUtil.DIRSYNC_LOG, level=logging.DEBUG, filemode='w', encoding='utf-8')
    #             dirsync_logger = logging.getLogger('dirsync')
    #             # result_sync = sync(sourcedir=target_abspath, targetdir=future_abspath, action='sync', verbose=True, logger=dirsync_logger) #success
    #             result_sync = sync(sourcedir=target_abspath, targetdir=future_abspath, action="sync", options=["--purge", "--verbose", "--force"], logger=dirsync_logger)
    #             # sync(sourcedir=future_abspath, targetdir=target_abspath , action='sync',verbose=True , logger = dirsync_logger)  # 양방향 으로 로컬동기화폴더를 만드려면 sync() 코드를 추가하여 sync() 함수가 총 2개가 targetdir 간에 sourcedir 서로 자리바뀌어 있도록 작성
    #             if result_sync:
    #                 # Park4139.DIRSYNC_LOG 내용 가져오기
    #                 if os.path.exists(StateManagementUtil.DIRSYNC_LOG):
    #                     lines = FileSystemUtil.get_lines_of_file(StateManagementUtil.DIRSYNC_LOG)[-4:-1]
    #                     lines = [sample.strip() for sample in lines]
    #                     for sample in lines:
    #                         # print(Park4139.translate_eng_to_kor_via_googletrans(sample))
    #                         DebuggingUtil.print_ment_light_white(rf'sample : {sample}')
    #                     DebuggingUtil.print_ment_light_white(rf'len(lines) : {len(lines)}')
    #                     FileSystemUtil.remove_target_parmanently(StateManagementUtil.DIRSYNC_LOG)
    #                     lines = [x for x in lines if x.strip("\n")]  # 리스트 요소 "" 제거,  from ["", A] to [A]       [""] to []
    #                     # TextToSpeechUtil.speak_ments(f"타겟의 동기화가 성공 되었습니다", sleep_after_play=0.65, thread_join_mode=True)
    #                     DebuggingUtil.print_ment_success("타겟동기화 성공")
    #                     UiUtil.pop_up_as_complete(title_="작업성공보고", ment=f"타겟의 동기화가 성공 되었습니다\n{future_abspath}", auto_click_positive_btn_after_seconds=3)
    #     except:
    #         DebuggingUtil.print_ment_fail("타겟동기화 실패")
    #         DebuggingUtil.trouble_shoot("%%%FOO%%%")
    #
    #     # from dirsync import sync
    #     # sources = [r'C:\Users\seon\Desktop\오리지널',
    #     #            r'C:\Users\seon\Desktop\오리지널2',
    #     #            r'C:\Users\seon\Desktop\오리지널3']
    #     # targets = [r'C:\Users\seon\Desktop\테스트',
    #     #            r'C:\Users\seon\Desktop\테스트2',
    #     #            r'C:\Users\seon\Desktop\테스트3']
    #     # total = dict(zip(sources, targets))
    #     # for source, target in total.items():
    #     #     sync(sourcedir=source, targetdir=target, action='sync', verbose=True, purge=True, create=True,  delete=True, update=True)  # 이것이 sync() default 파라미터들이다.
    #     #     sync(sourcedir=source, targetdir=target, action='sync', verbose=True, purge=True, create=True,  delete=True, update=True)  # purge=True 이면 targetdir 에 이물질 같은 파일이 있으면 삭제를 합니다, delete=False 이면 어떻게 되는거지? # verbose = True 이면 상세설명출력
    #
    #     # 윈도우 디렉토리 경로를 WSL 경로로 변환
    #     # try:
    #     #     server_time = Park4139.get_time_as_('%Y_%m_%d_%H_%M_%S_%f')
    #     #     target_abspath = rf"/mnt/c/{target_abspath}" \ 에서 / 로 바꿔야한다
    #     #     # future_abspath = rf"/mnt/c/{target_abspath}_{server_time}"
    #     #     future_abspath = rf"/mnt/c/{target_abspath}_sync"
    #     #     command = f"rsync -avz {target_abspath} {future_abspath}"
    #     #     print(command)
    #     #     subprocess.call(command, shell=True)
    #     # except:
    #     #     pass

    @staticmethod
    def get_lines_of_file(file_abspath):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        try:
            if os.path.exists(file_abspath):
                # with open(file_abspath, 'r', encoding="utf-8") as f:
                # with open(file_abspath, 'r', errors='ignore') as f:
                with open(file_abspath, 'r', encoding="utf-8", errors='ignore') as f:
                    lines = f.readlines()  # from file.ext to ["한줄","한줄","한줄"]
                    return lines
        except:
            DebuggingUtil.trouble_shoot("%%%FOO%%%")

    @staticmethod
    def remove_target_parmanently(target_abspath):
        """
        정말 로직적으로 삭제해도 상관 없는 경우가 아니면 이 함수는 쓰지 말것
        이 메소드를 사용하는 것은 데이터소실과 관계된 위험한 일이다.
        윈도우의 경우, move_target_to_trash_bin() 로 내부를 만일을 위해서 주석처리 후, 변경하였음.
        """
        if FileSystemUtil.is_os_windows():
            DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
            try:
                if BusinessLogicUtil.validate_and_return(value=target_abspath) is not False:
                    os.chdir(os.path.dirname(target_abspath))
                    if os.path.exists(target_abspath):
                        if os.path.isdir(target_abspath):
                            # shutil.rmtree(target_abspath)
                            # if os.path.isdir(target_abspath):
                            #     FileSystemUtil.get_cmd_output(rf'echo y | rmdir /s "{target_abspath}"')
                            FileSystemUtil.move_target_to_trash_bin(target_abspath)
                        elif os.path.isfile(target_abspath):
                            # os.remove(target_abspath)
                            # if os.path.isfile(target_abspath):
                            #     FileSystemUtil.get_cmd_output(rf'echo y | del /f "{target_abspath}"')
                            FileSystemUtil.move_target_to_trash_bin(target_abspath)
            except:
                DebuggingUtil.print_ment_fail(f"에러발생:\n{traceback.format_exc()}\ntrouble_id : %%%FOO%%%")
                DebuggingUtil.print_ment_fail("파일 삭제 중 에러가 발생했습니다")

            finally:
                os.chdir(StateManagementUtil.PROJECT_DIRECTORY)
        else:
            DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
            try:
                if BusinessLogicUtil.validate_and_return(value=target_abspath) is not False:
                    os.chdir(os.path.dirname(target_abspath))
                    if os.path.exists(target_abspath):
                        if os.path.isdir(target_abspath):
                            shutil.rmtree(target_abspath)
                            if os.path.isdir(target_abspath):
                                # FileSystemUtil.get_cmd_output(rf'리눅스로 코드 수정대기 | rmdir /s "{target_abspath}"')
                                pass
                        elif os.path.isfile(target_abspath):
                            os.remove(target_abspath)
                            if os.path.isfile(target_abspath):
                                # FileSystemUtil.get_cmd_output(rf'리눅스로 코드 수정대기 echo y | del /f "{target_abspath}"')
                                pass

            except:
                DebuggingUtil.print_ment_fail(f"에러발생:\n{traceback.format_exc()}\ntrouble_id : %%%FOO%%%")
                DebuggingUtil.print_ment_fail("파일 삭제 중 에러가 발생했습니다")

            finally:
                os.chdir(StateManagementUtil.PROJECT_DIRECTORY)

    @staticmethod
    def is_directory_changed(directory_abspath):
        # 현재 디렉토리 상태
        current_directory_state = None
        previous_directory_state = None
        try:
            current_directory_state = FileSystemUtil.get_directory_files_files_for_monitoring_mtime_without_files_to_exclude(directory_abspath=directory_abspath)  # from str to [["abspath", "mtime"]]
            current_directory_state = sorted(current_directory_state, key=lambda item: item[1], reverse=True)  # from [["abspath", "mtime"]] to [["abspath", "mtime"]] (딕셔너리를 value(mtime)를 기준으로 내림차순으로 정렬), 날짜를 제일 현재와 가까운 날짜를 선택하는 것은 날짜의 숫자가 큰 숫자에 가깝다는 이야기이다. 그러므로  큰 수부터 작은 수의 순서로 가는 내림차순으로 정렬을 해주었다(reverse=True).
            current_directory_state = DataStructureUtil.get_list_seperated_by_each_elements_in_nested_list(current_directory_state)  # from [["a","b"], ["c","d"]] to ["a", "b", "c", "d"]
            current_directory_state = DataStructureUtil.get_list_each_two_elements_joined(current_directory_state)  # from ["a", "b", "c", "d"] to ["ab", "cd"]
            current_directory_state = BusinessLogicUtil.get_list_replaced_from_list_that_have_special_characters(target=current_directory_state, replacement="")  # from [] to [] (특수문자 제거 시킨) # 이부분을 암호화 시키면 어떨까 싶은데 #딕셔너리에 있는부분을 replace
            [print(sample) for sample in current_directory_state[0:10]]
            print(rf'type(current_directory_state) : {type(current_directory_state)}')
            print(rf'len(current_directory_state) : {len(current_directory_state)}')
        # 이전 디렉토리 상태
        except:
            pass
        try:
            previous_directory_state = DbTomlUtil.select_db_toml(key=DbTomlUtil.get_db_toml_key(directory_abspath))
            if previous_directory_state == None:
                DbTomlUtil.insert_db_toml(key=DbTomlUtil.get_db_toml_key(directory_abspath), value=current_directory_state)  # 50000 줄의 str 다루는 것보다 50000 개의 list 다루는 것이 속도성능에 대하여 효율적이다. # DB에 [] 로 저장
                previous_directory_state = DbTomlUtil.select_db_toml(key=DbTomlUtil.get_db_toml_key(directory_abspath))
            [print(sample) for sample in previous_directory_state[0:10]]
            print(rf'type(previous_directory_state) : {type(previous_directory_state)}')
            print(rf'len(previous_directory_state) : {len(previous_directory_state)}')
        except:
            pass

        # 변화 감지 ( 러프, 빠름)
        # 두 디렉토리 상태 합쳐서 중복제거 해서 디렉토리 처음 상태의 절반이 나오면 중복이 없다는 생각.
        # 위의 생각은 맞는데 실제로 내가 숫자를 잘 못 파악했다.
        # set() 함수는 중복을 제거한 나머지를 리스트에 저장하는 함수이다.
        # 따라서 a != b 인 경우,  a == b 인 경우, 중복이 없는것
        # str을 줄단위로 쪼개 [str]로 리스트를 합한다.
        # try:
        #     merged_list = current_directory_state + previous_directory_state  # from [1, 2, 3] + [ x, y, z] to [1, x, 2, y, 3, z]
        #     [print(sample) for sample in merged_list[0:10]]
        #     print(rf'type(merged_list) : {type(merged_list)}')
        #     print(rf'len(merged_list) : {len(merged_list)}')
        #     a = len(merged_list)/2
        #     b = len(list(set(merged_list)))
        #     print(rf'a : {a}  b : {b} ')
        #     # duplicates = list(set([x for x in merged_list if merged_list.count(x) == 1])) # 중복되는 것만 추출 # 너무 느림 # 추가된파일, 삭제된 파일
        #     # print(rf'a : {a}  b : {b}   duplicates : {duplicates}')
        # except:
        #     pass
        # try:
        #     merged_list = Park4139List.get_list_added_elements_alternatively(current_directory_state, previous_directory_state)  # from [1, 2, 3] + [ x, y, z] to [1, x, 2, y, 3, z]
        #     [print(sample) for sample in merged_list[0:10]]
        #     print(rf'type(merged_list) : {type(merged_list)}')
        #     print(rf'len(merged_list) : {len(merged_list)}')
        #     a = len(merged_list) / 2
        #     b = len(list(set(merged_list)))
        #     print(rf'a : {a}  b : {b} ')
        #     print(rf'a : {a}  b : {b}   duplicates : {duplicates}')
        # finally:
        #     pass

        # 변화 감지 ( 디테일, 느림)
        # try:
        #     added_files = Park4139.get_added_files(previous_directory_state, current_directory_state)
        #     if added_files:
        #         print(f"추가된 파일 개수: {len(added_files)} 추가된 파일: {added_files}")
        #     deleted_files = Park4139.get_deleted_files(previous_directory_state, current_directory_state)
        #     if deleted_files:
        #         print(f"삭제된 파일 개수: {len(deleted_files)} 삭제된 파일: {deleted_files}")
        #     modified_files = Park4139.get_modified_files(previous_directory_state, current_directory_state)
        #     if modified_files:
        #         print(f"변경된 파일 개수: {len(modified_files)} 변경된 파일: {modified_files}")
        # except:
        #     pass

        # 변화 감지 ( 러프, 덜 느림)
        try:
            if DataStructureUtil.is_two_lists_equal(list1=current_directory_state, list2=previous_directory_state):
                print("디렉토리가 변경되지 않았습니다")
                return False
            else:
                print("디렉토리 변경이 감지되었습니다")
                DbTomlUtil.update_db_toml(key=DbTomlUtil.get_db_toml_key(directory_abspath), value=current_directory_state)
                return True
        except:
            pass

    @staticmethod
    def get_directory_files_files_for_monitoring_mtime_without_files_to_exclude(directory_abspath):
        files_to_exclude = [
            StateManagementUtil.DB_TOML,
            StateManagementUtil.SUCCESS_LOG,
            StateManagementUtil.LOCAL_PKG_CACHE_FILE,
            StateManagementUtil.DIRSYNC_LOG,
        ]
        files_of_directory = []
        for root, dirs, files in os.walk(directory_abspath):
            for file in files:
                if not file.endswith(".mp3"):  # 모든 mp3 파일을 배제합니다
                    file_path = os.path.join(root, file)
                    if file_path not in files_to_exclude:
                        # files_of_directory[rf"{file_path}"] = os.path.getmtime(file_path)
                        files_of_directory.append([file_path, os.path.getmtime(file_path)])
        return files_of_directory

    @staticmethod
    def get_added_files(previous_state, current_state):
        return DataStructureUtil.get_elements_that_list1_only_have(list1=current_state, list2=previous_state)

    @staticmethod
    def get_deleted_files(previous_state, current_state):
        return DataStructureUtil.get_elements_that_list1_only_have(list1=previous_state, list2=current_state)

    @staticmethod
    def get_modified_files(previous_state, current_state):
        return DataStructureUtil.get_different_elements(list1=current_state, list2=previous_state)

    @staticmethod
    def make_leaf_file(file_abspath):
        if not os.path.exists(file_abspath):
            try:
                os.makedirs(os.path.dirname(file_abspath))
            except FileExistsError:
                pass
            FileSystemUtil.get_cmd_output(rf'chcp 65001 >nul')
            FileSystemUtil.get_cmd_output(rf'echo. > "{file_abspath}"')

    @staticmethod
    def make_leaf_directory(leaf_directory_abspath):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        # DebuggingUtil.print_ment_via_colorama(ment=rf"leaf_directory_abspath : {leaf_directory_abspath}", colorama_color=ColoramaColorUtil.LIGHTWHITE_EX)
        if not os.path.exists(leaf_directory_abspath):
            os.makedirs(leaf_directory_abspath)

    @staticmethod
    def is_empty_directory(directory_abspath):
        try:
            if len(os.listdir(directory_abspath)) == 0:
                return True
            else:
                return False
        except FileNotFoundError:
            DebuggingUtil.print_ment_via_colorama(ment="FileNotFoundError", colorama_color=ColoramaUtil.RED)
            pass

    @staticmethod
    def rename_target(current_target_abspath, future_target_abspath):
        # directory 명을 변경하는 경우에 재귀적으로 바뀌어야 한는 것으로 생각되어 os.renames 를 테스트 후 적용하였다.
        # os.rename 와 os.renames 에는 큰 차이가 있는데
        # os.rename 사용 중에 old_path 가 directory  인 경우는 재귀적으로 변경이 안된다고 wrtn 은 말해주었다.
        os.renames(current_target_abspath, future_target_abspath)

    # @staticmethod
    # def merge_two_directories_without_overwrite(directory_a, directory_b):
    #     """디렉토리 머지용 기능"""
    #     # 빈디렉토리 머지 용도로도 많이 쓸것 같다
    #     while True:
    #         if directory_a == directory_b:
    #             TextToSpeechUtil.speak_ments("동일한 디렉토리는 머지하실 수 없습니다", sleep_after_play=0.65)
    #             break
    #         dst_p = rf"{os.path.dirname(directory_a)}"
    #         dst_n = rf"[{os.path.basename(directory_a)}] [{os.path.basename(directory_b)}]"
    #         dst_n = dst_n.replace("[[", "[")
    #         dst_n = dst_n.replace("]]", "]")
    #         # dst_n = dst_n.split("] [")
    #         # dst_n = re.sub(pattern=r'\$\d{19}', repl='', string=dst_n, count=1) # count 파라미터를 count=1 로 설정하여, 패턴이 여러번 나타나도 첫번째 카운트만 제거
    #         dst_n = re.sub(pattern=r'\$\d{20}', repl='', string=dst_n)
    #         dst_n = rf"{dst_n} ${TimeUtil.get_time_as_('%Y%m%d%H%M%S%f')}"
    #         dst = rf"{dst_p}/{dst_n}"
    #         FileSystemUtil.make_leaf_directory(dst)
    #         print(rf'dst : {dst}')
    #         if os.path.isdir(directory_b):
    #             for target in os.listdir(directory_a):  # os.listdir 은 without walking 으로 동작한다
    #                 print(rf'target : {directory_a}/{target}')
    #                 FileSystemUtil.move_target_without_overwrite(target_abspath=rf'{directory_a}/{target}', dst=dst)
    #                 # Park4139.move_target_to_trash_bin(directory_a)
    #             for target in os.listdir(directory_b):  # os.listdir 은 without walking 으로 동작한다
    #                 print(rf'target : {directory_b}/{target}')
    #                 FileSystemUtil.move_target_without_overwrite(target_abspath=rf'{directory_b}/{target}', dst=dst)
    #         else:
    #             TextToSpeechUtil.speak_ments("디렉토리만 머지하실 수 있습니다", sleep_after_play=0.65)
    #             break
    #         break
    # nagasarete airantou

    @staticmethod
    def merge_directories(directoryies: str):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")

        is_breaking = False
        # cmd /c dir /b /s /a:d | clip
        directoryies = directoryies.strip()
        directoryies = directoryies.strip("\t")
        directoryies_: [str] = directoryies.split("\n")
        directoryies_ = [x for x in directoryies_ if x.strip()]  # 리스트 요소 "" 제거,  from ["", A] to [A]    from [""] to []
        directoryies_ = [x.strip() for x in directoryies_]  # 리스트 각 요소 strip(),  ["   A", "B   "] from ["A", "B"]
        directoryies_ = [x.strip("\"") for x in directoryies_]  # 리스트 각 요소 strip("\""),  [""A""] from ["A"]
        directoryies_ = [x.strip("\'") for x in directoryies_]  # 리스트 각 요소 strip("\'),  ["'A'""] from ["A"]

        [print(rf'directory : {directory}') for directory in directoryies_]
        print(rf'type(directoryies_) : {type(directoryies_)}')
        print(rf'len(directoryies_) : {len(directoryies_)}')

        if 0 == len(directoryies_):
            TextToSpeechUtil.speak_ments("타겟경로가 아무것도 입력되지 않았습니다", sleep_after_play=0.65)
            return
        elif 1 == len(directoryies_):
            TextToSpeechUtil.speak_ments("하나의 타겟경로로는 머지를 시도할수 없습니다, 여러개의 타겟경로들을 입력해주세요", sleep_after_play=0.65)
            return
        elif 1 < len(directoryies_):
            for index, directory in enumerate(directoryies_):
                connected_drives = []
                for drive_letter in string.ascii_uppercase:
                    drive_path = drive_letter + ":\\"
                    if os.path.exists(drive_path):
                        connected_drives.append(drive_path)
                        if directory == drive_path:
                            TextToSpeechUtil.speak_ments("입력된 타겟경로는 너무 광범위하여, 진행할 수 없도록 설정되어 있습니다", sleep_after_play=0.65)
                            break

            # FileSystemUtil.make_leaf_directory(StateManagementUtil.EMPTY_DIRECTORYIES)

            # indices_to_remove = []  # 제거할 인덱스를 기록할 리스트
            # for index, directory in enumerate(directoryies_):
            #     if os.path.isdir(directory):
            #         if FileSystemUtil.is_empty_directory(directory):
            #             FileSystemUtil.move_target_without_overwrite(target_abspath=directory, dst=StateManagementUtil.EMPTY_DIRECTORYIES)
            #             indices_to_remove.append(index)
            # for index in indices_to_remove:
            #     directoryies_.pop(index)

            # for index, directory in enumerate(directoryies_):
            #     if directory == "":
            #         TextToSpeechUtil.speak_ments("하나 이상의 타겟경로가 공백으로 입력되었습니다", sleep_after_play=0.65)
            #         return
            #     if not os.path.exists(directory):
            #         TextToSpeechUtil.speak_ments("하나 이상의 타겟경로가 존재하지 않습니다", sleep_after_play=0.65)
            #         return

            DebuggingUtil.commentize("빈 트리 리프디렉토리별로 해체한 뒤 제거")
            files_to_move = []
            for index, directory in enumerate(directoryies_):
                for root, directories, files in os.walk(directory, topdown=True):
                    for file in files:
                        files_to_move.append(rf"{root}/{file}")
            [DebuggingUtil.print_ment_light_white(rf'file_to_move : {file_to_move}') for file_to_move in files_to_move]
            DebuggingUtil.print_ment_light_white(rf'type(files_to_move) : {type(files_to_move)}')
            DebuggingUtil.print_ment_light_white(rf'len(files_to_move) : {len(files_to_move)}')
            dst = rf"{os.path.dirname(directoryies_[0])}/{os.path.basename(directoryies_[0]).replace("_$merged", "")}_$merged"
            DebuggingUtil.print_ment_light_yellow(rf'dst : {dst}')
            FileSystemUtil.make_leaf_directory(dst)

            for file_to_move in files_to_move:
                FileSystemUtil.move_target_without_overwrite(target_abspath=file_to_move, dst=dst)

            # 이 함수는 캐싱문제를 해결하지 못함.
            def count_files_in_folder(folder_path):
                file_count = 0
                for _, _, files in os.walk(folder_path):
                    file_count += len(files)
                return file_count

            # 이 함수는 캐싱문제를 해결하지 못함.
            # def count_files_in_folder(folder_path):
            #     file_count = 0
            #     with os.scandir(folder_path) as entries:
            #         for entry in entries:
            #             if entry.is_file():
            #                 file_count += 1
            #     return file_count

            # 분명히 파일이 들어있는데 개수가 0개로 나오네  캐싱때문에 그럴 수 있어?
            # 운영체제 성능향상을 위한 폴더정보 캐싱으로 갱신이 안될 수 있음.
            # 폴더를 다른 위치로 이동한 후 다시 이동합니다. 이렇게 하면 파일 시스템이 폴더의 변경을 감지하고 캐시를 갱신할 수 있습니다.
            # current_path = os.getcwd()
            # os.chdir(os.path.dirname(current_path)) # 부모 디렉토리로 이동
            # os.chdir(current_path)

            # 분명히 가 틀렸다. 운영체제의 캐싱이 문제가 아닌, 로직을 잘못 만든 것이었다.
            # directory 변수를 엉뚱하게 초기화하였기 때문이다.
            # directory 를 directory_ 로 별도로 호출해서 해소하였다.

            # empty directory 리프단위로 분해하여 이동
            while True:
                directorys_to_move = []
                for index, directory in enumerate(directoryies_):
                    for root, directories, files in os.walk(directory, topdown=False):
                        for directory_ in directories:
                            file_count = count_files_in_folder(rf"{root}/{directory_}")
                            if file_count == 0:
                                directorys_to_move.append(rf"{root}/{directory_}")

                [DebuggingUtil.print_ment_light_white(rf'directory_to_move : {directory_to_move}') for directory_to_move in directorys_to_move]
                DebuggingUtil.print_ment_light_white(rf'type(directorys_to_move) : {type(directorys_to_move)}')
                DebuggingUtil.print_ment_light_white(rf'len(directorys_to_move) : {len(directorys_to_move)}')

                dst = StateManagementUtil.EMPTY_DIRECTORYIES
                DebuggingUtil.print_ment_light_yellow(rf'dst : {dst}')
                FileSystemUtil.make_leaf_directory(dst)

                if len(directorys_to_move) == 0:
                    break

                for directory_to_move in directorys_to_move:
                    FileSystemUtil.move_target_without_overwrite(target_abspath=directory_to_move, dst=dst)

            for directory in directoryies_:
                BusinessLogicUtil.gather_empty_directory(rf"{directory}")

            DebuggingUtil.print_ment_success(rf'디렉토리 머지를 완료했습니다')

    # @staticmethod
    # THIS IS BAD FUNCTION... I SHOULD CHECK TYPE, WHEN I MADE THIS
    # def get_display_info():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     # 디스플레이 정보 가져오기  # pyautogui.size?() 로 대체할것.
    #     global height, width
    #     from screeninfo import get_monitors
    #     for infos in get_monitors():
    #         for info in str(infos).split(','):
    #             if 'width=' in info.strip():
    #                 width = info.split('=')[1]
    #             elif 'height=' in info.strip():
    #                 height = info.split('=')[1]
    #     display_setting = {
    #         'height': int(height),  # ㅋㅋㅋㅋㅋㅋㅋㅋㅋㅋㅋ  , 실수로 써놨는데 런타임에러 안뜸 ㅋㅋㅋㅋㅋㅋㅋㅋㅋㅋ 코드수정 귀찮...ㅋ 언젠가 하자.
    #         'width': int(width)
    #     }
    #     return display_setting

    @staticmethod
    def reboot_this_computer():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        FileSystemUtil.get_cmd_output(rf'shutdown -r ')

    @staticmethod
    def shutdown_this_computer():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        FileSystemUtil.get_cmd_output(rf'%windir%\System32\Shutdown.exe -s ')
        # Park4139.get_cmd_output('Shutdown.exe -s ')

    @staticmethod
    def enter_power_saving_mode():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        cmd = rf'%windir%\System32\rundll32.exe powrprof.dll SetSuspendState'
        FileSystemUtil.get_cmd_output(cmd)

    @staticmethod
    def play_wav_file(file_abspath):
        # 종료가 되버린다
        # try:
        #     import playsound
        #     playsound.playsound(file_abspath)
        # except:
        #     pass

        # 너무 느리다.
        # try:
        #     import wave
        #     import pyaudio
        #     # WAV 파일 열기
        #     with wave.open(file_abspath, 'rb') as wav:
        #         audio = pyaudio.PyAudio()
        #         # 스트림 열기
        #         stream = audio.open(format=audio.get_format_from_width(wav.getsampwidth()),
        #                             channels=wav.getnchannels(),
        #                             rate=wav.getframerate(),
        #                             output=True)
        #         # 데이터 읽기
        #         data = wav.readframes(1024)
        #
        #         # 재생
        #         while data:
        #             stream.write(data)
        #             data = wav.readframes(1024)
        #
        #         # 스트림 닫기
        #         stream.stop_stream()
        #         stream.close()
        #
        #         # PyAudio 객체 종료
        #         audio.terminate()
        #
        #     # WAV 파일 닫기
        #     # wav.close()
        # except:
        #     pass

        # 소리 시끄러워서 주석처리, 소리 필요 시 주석해제
        # source = pyglet.media.load(file_abspath)
        # source.play()
        pass

    @staticmethod
    def get_cmd_output(cmd):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        if not FileSystemUtil.is_os_windows():
            cmd = cmd.replace("\\", "/")
        else:
            cmd = cmd.replace("/", "\\")
        DebuggingUtil.print_magenta(rf'{cmd}')
        # os.Popen 으로 print 가능하도록 할 수 있다는 것 같았는데 다른 방식으로 일단 되니까. 안되면 시도.
        # cmd = ['dir', '/b']
        # fd_popen = subprocess.Popen(cmd, stdout=subprocess.PIPE).stdout # 명령어 실행 후 반환되는 결과를 파일에 저장합니다.
        # fd_popen = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE).stdout # shell=True 옵션, cmd를 string 으로 설정
        # lines = fd_popen.read().strip().split('\n')# lines에 저장합니다.
        # fd_popen.close()# 파일을 닫습니다.
        lines = None
        try:
            if FileSystemUtil.is_os_windows():
                lines = subprocess.check_output('chcp 65001 >nul', shell=True).decode('utf-8').split('\n')
            lines = subprocess.check_output(cmd, shell=True).decode('utf-8').split('\n')
            return lines
        except UnicodeDecodeError:
            print("UnicodeDecodeError 가 발생했습니다. euc-kr 로 설정하여 명령어 실행을 재시도 합니다")
            try:
                lines = subprocess.check_output(cmd, shell=True).decode('euc-kr').split('\n')
            except Exception:
                print("????????  os.system(cmd)로 재시도 합니다")
                # os.system(cmd)
            return lines
        except subprocess.CalledProcessError:
            [print(sample) for sample in lines]
            print(rf'type(lines) : {type(lines)}')
            print(rf'len(lines) : {len(lines)}')
            DebuggingUtil.print_ment_fail(f"에러발생:\n{traceback.format_exc()}\ntrouble_id : %%%FOO%%%")
            # dialog = UiUtil.CustomQdialog(ment=f"에러코드[E%%%FOO%%%]\n\nsubprocess.CalledProcessError 가 발생했습니다", btns=["확인"], is_input_box=True)
            # dialog.exec()
            # os.system(cmd)
            return lines
        except Exception:
            print("예상되지 않은 예외가 발생했습니다.")
            DebuggingUtil.trouble_shoot("%%%FOO%%%")
            return lines

    @staticmethod
    def explorer(target_abspath: str):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        target_abspath = str(target_abspath)
        target_abspath = target_abspath.replace('\"', '')
        target_abspath = rf'explorer "{target_abspath}"'
        DebuggingUtil.print_magenta(rf'{target_abspath}')
        try:
            subprocess.check_output(target_abspath, shell=True).decode('utf-8').split('\n')
        except UnicodeDecodeError:
            subprocess.check_output(target_abspath, shell=True).decode('euc-kr').split('\n')
        except subprocess.CalledProcessError:
            DebuggingUtil.print_ment_fail(f"에러발생:\n{traceback.format_exc()}\ntrouble_id : %%%FOO%%%")
            DebuggingUtil.print_ment_fail("파일 삭제 중 에러가 발생했습니다")
            pass
        except UnboundLocalError and Exception as e:
            DebuggingUtil.trouble_shoot("202312030015a")

    @staticmethod
    # 프로그램 PID 출력
    def get_current_program_pid():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        pro = subprocess.check_output(
            rf'powershell (Get-WmiObject Win32_Process -Filter ProcessId=$PID).ParentProcessId', shell=True).decode(
            'utf-8')  # 실험해보니 subprocess.check_output(cmd,shell=True).decode('utf-8') 코드는 프로세스가 알아서 죽는 것 같다. 모르겠는데 " " 가 있어야 동작함
        lines = pro.split('\n')
        pids = []
        for line in lines:
            if "" != line.strip():
                pid = line
                pids.append(pid)
                DebuggingUtil.print_magenta(f'pid: {pid}')
        return pids

    @staticmethod
    def get_target_bite(start_path='.'):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        total_size = 0
        for dirpath, dirnames, filenames in os.walk(start_path):
            for f in filenames:
                fp = os.path.join(dirpath, f)
                # skip if it is symbolic link
                if not os.path.islink(fp):
                    total_size += os.path.getsize(fp)
        return total_size

    @staticmethod
    def get_target_megabite(target_path):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        return FileSystemUtil.get_target_bite(target_path.strip()) / 1024 ** 2

    @staticmethod
    def get_target_gigabite(target_path):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        return FileSystemUtil.get_target_bite(target_path.strip()) / 1024 ** 3

    @staticmethod
    def move_target_to_trash_bin(target_abspath):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")

        # 타겟 휴지통으로 이동
        send2trash.send2trash(target_abspath)

        # 휴지통 열기
        # Park4139.get_cmd_output('explorer.exe shell:RecycleBinFolder')

    @staticmethod
    def xcopy_with_overwrite(target_abspath, future_target_abspath):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        try:
            result = FileSystemUtil.get_cmd_output(rf'echo a | xcopy "{target_abspath}" "{future_target_abspath}" /e /h /k /y')
            if result == subprocess.CalledProcessError:
                if os.path.isfile(target_abspath):
                    FileSystemUtil.get_cmd_output(rf'echo f | xcopy "{target_abspath}" "{future_target_abspath}" /e /h /k /y')
                else:
                    FileSystemUtil.get_cmd_output(rf'echo d | xcopy "{target_abspath}" "{future_target_abspath}" /e /h /k /y')
        except Exception:
            DebuggingUtil.trouble_shoot('%%%FOO%%%')

    @staticmethod
    def xcopy_without_overwrite(target_abspath, future_target_abspath):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        try:
            if os.path.exists(future_target_abspath):
                future_target_abspath = rf"{os.path.dirname(future_target_abspath)}/{os.path.basename(target_abspath)[0]}_{TimeUtil.get_time_as_('%Y_%m_%d_%H_%M_%S_%f')}{os.path.basename(target_abspath)[1]}"
            FileSystemUtil.get_cmd_output(rf'echo a | xcopy "{target_abspath}" "{future_target_abspath}" /e /h /k')
            if os.path.isfile(target_abspath):
                FileSystemUtil.get_cmd_output(rf'echo f | xcopy "{target_abspath}" "{future_target_abspath}" /e /h /k')
            else:
                FileSystemUtil.get_cmd_output(rf'echo d | xcopy "{target_abspath}" "{future_target_abspath}" /e /h /k')
        except Exception:
            print(rf"subprocess.CalledProcessError 가 발생하여 재시도를 수행합니다 {inspect.currentframe().f_code.co_name}")
            DebuggingUtil.trouble_shoot('%%%FOO%%%')

    @staticmethod
    def is_file_edited(target_abspath: str):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        try:
            key = DbTomlUtil.get_db_toml_key(target_abspath)
            db = DbTomlUtil.read_db_toml()
            line_cnt_measured = BusinessLogicUtil.get_line_cnt_of_file(target_abspath)
            if line_cnt_measured != db[key]:
                DebuggingUtil.commentize("파일편집 감지되었습니다")
            else:
                DebuggingUtil.commentize("파일편집 감지되지 않았습니다")
                pass
            if line_cnt_measured != db[key]:
                return True
            else:
                return False
        except KeyError:
            DbTomlUtil.insert_db_toml(key=DbTomlUtil.get_db_toml_key(target_abspath), value=BusinessLogicUtil.get_line_cnt_of_file(target_abspath))
            DebuggingUtil.commentize("파일편집 확인 중 key 에러가 발생하였습니다")
        except Exception:  # except Exception:으로 작성하면 KeyError 를 뺀 나머지 에러처리 설정 , except: 를 작성하면 모든 에러처리 설정
            DebuggingUtil.commentize("데이터베이스 확인 중 예상되지 않은 에러가 감지되었습니다")
            DebuggingUtil.trouble_shoot("202312030000b")

    @staticmethod
    def move_target_without_overwrite(target_abspath, dst):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        try:
            # 시간 포멧팅 ms 까지 설정
            # get_time_as_() 로 %f 를 pettern 인자에 str 으로서 덧붙여 전달하면, 에러가 났는데
            # time 모듈이 %f 를 지원하지 않아서 나던 것.
            # %f를 지원하는 datetime 으로 교체했다
            target_dirname = os.path.dirname(target_abspath)
            spaceless_time_pattern = rf"{TimeUtil.get_time_as_('%Y%m%d%H%M%S%f')}"
            target_n = FileSystemUtil.get_target_as_n(target_abspath)
            target_x = FileSystemUtil.get_target_as_x(target_abspath)
            # future_abspath = rf"{dst}/{target_n}{target_x}"
            # future_abspath = copy.deepcopy(future_abspath)

            target_n = re.sub(pattern=r'\$\d{22}', repl='', string=target_n)
            target_abspath_with_timestamp = rf'{target_dirname}/{target_n}${spaceless_time_pattern}{random.randint(10, 99)}{target_x}'  # 시각적으로 _ 가 늘어나니까 좋을 수 도 있음.
            if dst != os.path.dirname(target_abspath_with_timestamp):
                DebuggingUtil.print_ment_via_colorama(ment=rf"target_abspath_with_timestamp : {target_abspath_with_timestamp}", colorama_color=ColoramaUtil.LIGHTWHITE_EX)
                DebuggingUtil.print_ment_via_colorama(ment=rf"dst                           : {dst}", colorama_color=ColoramaUtil.LIGHTWHITE_EX)
                if os.path.isfile(target_abspath):
                    try:
                        os.rename(target_abspath, target_abspath_with_timestamp)
                        shutil.move(src=target_abspath_with_timestamp, dst=dst)
                        DebuggingUtil.print_ment_via_colorama(ment=rf"파일이동성공", colorama_color=ColoramaUtil.LIGHTCYAN_EX)
                    except Exception:
                        # traceback.print_exc(file=sys.stdout)
                        DebuggingUtil.print_ment_via_colorama(ment=rf"파일이동실패", colorama_color=ColoramaUtil.RED)
                elif os.path.isdir(target_abspath):
                    try:
                        os.rename(target_abspath, target_abspath_with_timestamp)
                        shutil.move(src=target_abspath_with_timestamp, dst=dst)
                        DebuggingUtil.print_ment_via_colorama(ment=rf"디렉토리이동성공", colorama_color=ColoramaUtil.LIGHTCYAN_EX)
                    except Exception:
                        # traceback.print_exc(file=sys.stdout)
                        DebuggingUtil.print_ment_via_colorama(ment=rf"디렉토리이동실패", colorama_color=ColoramaUtil.RED)
        except:
            DebuggingUtil.trouble_shoot("202312030021")

    @staticmethod
    def move_without_overwrite_via_robocopy(target_abspath, dst):  # 명령어 자체가 안되는데 /mir 은 되는데 /move 안된다
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        try:
            DebuggingUtil.commentize(f'타겟이동 시도')
            # Park4139.get_cmd_output(rf'robocopy "{target_abspath}" "{dst}" /MOVE')
            if os.path.exists(rf'{dst}/{os.path.dirname(target_abspath)}'):
                FileSystemUtil.move_target_to_trash_bin(target_abspath)

        except Exception:
            DebuggingUtil.trouble_shoot("202312030021")

    @staticmethod
    def is_empty_tree(directory_path):
        '''디렉토리의 내부를 순회하며 돌면서 모든 하위 디렉토리에 파일이 하나도 없는지를 판단합니다'''
        try:
            # 디렉토리 내부의 파일과 하위 디렉토리 목록을 가져옵니다.
            contents = os.listdir(directory_path)

            # 파일이 존재하는 경우,
            for content in contents:
                [print(rf'content : {content}') for content in contents]
                content_path = os.path.join(directory_path, content)
                if os.path.isfile(content_path):
                    # DebuggingUtil.print_ment_light_white("빈 트리 디렉토리 아닙니다")
                    return False
            # DebuggingUtil.print_ment_light_white("빈 트리 디렉토리 있니다.")
            return True
        except FileNotFoundError:
            DebuggingUtil.print_ment_fail('FileNotFoundError')
        except OSError:
            DebuggingUtil.print_ment_fail('OSError')
        except UnboundLocalError:
            DebuggingUtil.print_ment_fail('UnboundLocalError')

    @staticmethod
    def is_leaf_directory(directory_path):
        # leaf 디렉토리란, 하위 디렉토리를 가지지 않는 가장 마지막 단계의 디렉토리를 말합니다

        # 디렉토리 내부의 파일과 하위 디렉토리 목록을 가져옵니다.
        try:
            contents = os.listdir(directory_path)

            # 디렉토리 내에 파일이 존재하면 leaf 디렉토리가 아닙니다.
            if len(contents) > 0:
                return False

            # 디렉토리 내에 하위 디렉토리가 존재하는지 확인합니다.
            for content in contents:
                content_path = os.path.join(directory_path, content)
                if os.path.isdir(content_path):
                    return False

            # 디렉토리가 leaf 디렉토리입니다.
            return True
        except FileNotFoundError:
            DebuggingUtil.print_ment_via_colorama('FileNotFoundError', colorama_color=ColoramaUtil.RED)
        except:
            DebuggingUtil.print_ment_via_colorama(f'{inspect.currentframe().f_code.co_name}() 기타에러발생', colorama_color=ColoramaUtil.RED)

    @staticmethod
    def add_text_to_file(file_abspath, text):
        with open(file_abspath, 'a') as file:
            file.write(text)

    @staticmethod
    def get_cnts_of_line_of_file(file_abspath):
        with open(file_abspath, 'r') as file:
            return sum(1 for line in file)

    @staticmethod
    def is_letters_cnt_zero(file_abspath):
        try:
            with open(file_abspath, 'r') as file:
                contents = file.read().strip()
                # print(rf'len(contents) : {len(contents)}')
                if len(contents) == 0:
                    return True
        except FileNotFoundError:
            print("파일을 찾을 수 없습니다.")
            return False
        except UnicodeDecodeError:
            with open(file_abspath, 'r', encoding='utf-8') as file:
                contents = file.read().strip()
                # print(rf'len(contents) : {len(contents)}')
                if len(contents) == 0:
                    return True
            return False
        except:
            DebuggingUtil.trouble_shoot("%%%FOO%%%")
            print("오류가 발생했습니다.")
            return False

    @staticmethod
    def is_os_windows():
        # DebuggingUtil.print_ment_red(rf'''INFO:{StateManagementUtil.INDENTATION_PROMISED}윈도우 기반 리눅스 개발 모드를 실행시도''')
        # return False

        if platform.system() == 'Windows':
            return True
        else:
            return False

    @staticmethod
    def get_os():
        if platform.system() == 'Windows':
            return 'Windows'
        elif platform.system() == 'Linux':
            return 'Linux'
        else:
            return 'Unknown'

    @staticmethod
    def truncate_tree(directory_abspath):
        if os.path.exists(directory_abspath):
            shutil.rmtree(directory_abspath)
        if not os.path.exists(directory_abspath):
            FileSystemUtil.make_leaf_directory(directory_abspath)


class FontsUtil:
    MONTSERRAT_THIN_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Montserrat\static\Montserrat-Thin.ttf"
    NOTOSANSKR_VARIABLEFONT_WGHT_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Noto_Sans_KR\NotoSansKR-VariableFont_wght.ttf"
    NOTOSANSKR_BLACK_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Noto_Sans_KR\static\NotoSansKR-Black.ttf"
    NOTOSANSKR_BOLD_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Noto_Sans_KR\static\NotoSansKR-Bold.ttf"
    NOTOSANSKR_EXTRABOLD_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Noto_Sans_KR\static\NotoSansKR-ExtraBold.ttf"
    NOTOSANSKR_EXTRALIGHT_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Noto_Sans_KR\static\NotoSansKR-ExtraLight.ttf"
    NOTOSANSKR_LIGHT_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Noto_Sans_KR\static\NotoSansKR-Light.ttf"
    NOTOSANSKR_MEDIUM_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Noto_Sans_KR\static\NotoSansKR-Medium.ttf"
    NOTOSANSKR_REGULAR_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Noto_Sans_KR\static\NotoSansKR-Regular.ttf"
    NOTOSANSKR_SEMIBOLD_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Noto_Sans_KR\static\NotoSansKR-SemiBold.ttf"
    NOTOSANSKR_THIN_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Noto_Sans_KR\static\NotoSansKR-Thin.ttf"
    GMARKETSANSTTFBOLD_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\GmarketSans\GmarketSansTTFBold.ttf"
    GMARKETSANSTTFLIGHT_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\GmarketSans\GmarketSansTTFLight.ttf"
    GMARKETSANSTTFMEDIUM_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\GmarketSans\GmarketSansTTFMedium.ttf"
    ITALIC_VARIABLEFONT_WGHT_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Montserrat\Montserrat-Italic-VariableFont_wght.ttf"
    MONTSERRAT_VARIABLEFONT_WGHT_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Montserrat\Montserrat-VariableFont_wght.ttf"
    MONTSERRAT_BLACK_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Montserrat\static\Montserrat-Black.ttf"
    MONTSERRAT_BLACKITALIC_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Montserrat\static\Montserrat-BlackItalic.ttf"
    MONTSERRAT_BOLD_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Montserrat\static\Montserrat-Bold.ttf"
    MONTSERRAT_BOLDITALIC_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Montserrat\static\Montserrat-BoldItalic.ttf"
    MONTSERRAT_EXTRABOLD_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Montserrat\static\Montserrat-ExtraBold.ttf"
    MONTSERRAT_EXTRABOLDITALIC_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Montserrat\static\Montserrat-ExtraBoldItalic.ttf"
    MONTSERRAT_EXTRALIGHT_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Montserrat\static\Montserrat-ExtraLight.ttf"
    MONTSERRAT_EXTRALIGHTITALIC_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Montserrat\static\Montserrat-ExtraLightItalic.ttf"
    MONTSERRAT_ITALIC_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Montserrat\static\Montserrat-Italic.ttf"
    MONTSERRAT_LIGHT_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Montserrat\static\Montserrat-Light.ttf"
    MONTSERRAT_LIGHTITALIC_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Montserrat\static\Montserrat-LightItalic.ttf"
    MONTSERRAT_MEDIUM_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Montserrat\static\Montserrat-Medium.ttf"
    MONTSERRAT_MEDIUMITALIC_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Montserrat\static\Montserrat-MediumItalic.ttf"
    MONTSERRAT_REGULAR_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Montserrat\static\Montserrat-Regular.ttf"
    MONTSERRAT_SEMIBOLD_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Montserrat\static\Montserrat-SemiBold.ttf"
    MONTSERRAT_SEMIBOLDITALIC_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Montserrat\static\Montserrat-SemiBoldItalic.ttf"
    MONTSERRAT_THINITALIC_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Montserrat\static\Montserrat-ThinItalic.ttf"
    POPPINS_BLACK_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Poppins\Poppins-Black.ttf"
    POPPINS_BLACKITALIC_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Poppins\Poppins-BlackItalic.ttf"
    POPPINS_BOLD_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Poppins\Poppins-Bold.ttf"
    POPPINS_BOLDITALIC_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Poppins\Poppins-BoldItalic.ttf"
    POPPINS_EXTRABOLD_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Poppins\Poppins-ExtraBold.ttf"
    POPPINS_EXTRABOLDITALIC_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Poppins\Poppins-ExtraBoldItalic.ttf"
    POPPINS_EXTRALIGHT_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Poppins\Poppins-ExtraLight.ttf"
    POPPINS_EXTRALIGHTITALIC_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Poppins\Poppins-ExtraLightItalic.ttf"
    POPPINS_ITALIC_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Poppins\Poppins-Italic.ttf"
    POPPINS_LIGHT_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Poppins\Poppins-Light.ttf"
    POPPINS_LIGHTITALIC_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Poppins\Poppins-LightItalic.ttf"
    POPPINS_MEDIUM_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Poppins\Poppins-Medium.ttf"
    POPPINS_MEDIUMITALIC_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Poppins\Poppins-MediumItalic.ttf"
    POPPINS_REGULAR_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Poppins\Poppins-Regular.ttf"
    POPPINS_SEMIBOLD_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Poppins\Poppins-SemiBold.ttf"
    POPPINS_SEMIBOLDITALIC_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Poppins\Poppins-SemiBoldItalic.ttf"
    POPPINS_THIN_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Poppins\Poppins-Thin.ttf"
    POPPINS_THINITALIC_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Poppins\Poppins-ThinItalic.ttf"
    RUBIKDOODLESHADOW_REGULAR_TTF = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_fonts\Rubik_Doodle_Shadow\RubikDoodleShadow-Regular.ttf"  # 너무 귀여운 입체감 있는 영어폰트


class TextToSpeechUtil:
    @staticmethod
    def is_containing_kor(text):
        pattern = "[ㄱ-ㅎ가-힣]"
        if re.search(pattern, text):
            return True
        else:
            return False

    @staticmethod
    def is_containing_special_characters_with_thread(text: str):
        # 비동기 처리 설정 ( advanced  )
        nature_numbers = [n for n in range(1, 101)]  # from 1 to 100
        work_quantity = len(text)
        n = 4  # thread_cnt # interval_cnt
        d = work_quantity // n  # interval_size
        r = work_quantity % n
        start_1 = 0
        end_1 = d - 1
        starts = [start_1 + (n - 1) * d for n in nature_numbers[:n]]  # 등차수열 공식
        ends = [end_1 + (n - 1) * d for n in nature_numbers[:n]]
        remained_start = ends[-1] + 1
        remained_end = work_quantity

        # print(rf'nature_numbers : {nature_numbers}')  # 원소의 길이의 합이 11넘어가면 1에서 3까지만 표기 ... 의로 표시 그리고 마지막에서 3번째에서 마지막에서 0번째까지 표기 cut_list_proper_for_pretty()
        # print(rf'work_quantity : {work_quantity}')
        # print(rf'n : {n}')
        # print(rf'd : {d}')
        # print(rf'r : {r}')
        # print(rf'start_1 : {start_1}')
        # print(rf'end_1 : {end_1}')
        # print(rf'starts : {starts}')
        # print(rf'ends : {ends}')
        # print(rf'remained_start : {remained_start}')
        # print(rf'remained_end : {remained_end}')

        # 비동기 이벤트 함수 설정 ( advanced  )
        async def is_containing_special_characters(start_index: int, end_index: int, text: str):
            pattern = "[~!@#$%^&*()_+|<>?:{}]"  # , 는 제외인가?
            if re.search(pattern, text[start_index:end_index]):
                # print(f"쓰레드 {start_index}에서 {end_index}까지 작업 성공 True return")
                result_list.append(True)
                # return True
            else:
                result_list.append(False)
                # print(f"쓰레드 {start_index}에서 {end_index}까지 작업 성공 False return")
                # return False

        # 비동기 이벤트 루프 설정
        def run_async_event_loop(start_index: int, end_index: int, text: str):
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop.run_until_complete(is_containing_special_characters(start_index, end_index, text))

        # 스레드 객체를 저장할 리스트 생성
        threads = []

        # 각 스레드의 결과를 저장할 리스트
        # result_list = [None] * work_quantity
        result_list = []

        # 주작업 처리용 쓰레드
        for n in range(0, n):
            start_index = starts[n]
            end_index = ends[n]
            thread = threading.Thread(target=run_async_event_loop, args=(start_index, end_index, text))
            thread.start()
            threads.append(thread)

        # 남은 작업 처리용 쓰레드
        if remained_end <= work_quantity:
            start_index = remained_start
            end_index = remained_end
            thread = threading.Thread(target=run_async_event_loop, args=(start_index, end_index, text))
            thread.start()
            threads.append(thread)
        else:
            start_index = remained_start
            end_index = start_index  # end_index 를 start_index 로 하면 될 것 같은데 테스트필요하다.
            thread = threading.Thread(target=run_async_event_loop, args=(start_index, end_index, text))
            thread.start()
            threads.append(thread)

        # 모든 스레드의 작업이 종료될 때까지 대기
        for thread in threads:
            thread.join()

        # 먼저 종료된 스레드가 있는지 확인하고, 나머지 스레드 중지
        # for thread in threads:
        #     if not thread.is_alive():
        #         for other_thread in threads:
        #             if other_thread != thread:
        #                 other_thread.cancel()
        #         break

        # 바뀐 부분만 결과만 출력, 전체는 abspaths_and_mtimes 에 반영됨
        # print(rf'result_list : {result_list}')
        # print(rf'type(result_list) : {type(result_list)}')
        # print(rf'len(result_list) : {len(result_list)}')

        if all(result_list):
            # print("쓰레드 작업결과 result_list의 모든 요소가 True이므로 True를 반환합니다")
            return True
        else:
            # print("쓰레드 작업결과 result_list에 False인 요소가 있어 False를 반환합니다")
            return False  # 불필요시 주석

    @staticmethod
    def is_containing_jpn(text):
        # pattern = r"^[\p{Script=Hiragana}\p{Script=Katakana}\p{Script=Han}ー〜・]+$"
        # pattern = r"^\P{Script=Hiragana}\P{Script=Katakana}\P{Script=Han}ー〜・]+$"
        # pattern = r"^[\p{Script=Hiragana}\p{Script=Katakana}\p{Script=Han}ー〜・]+$"
        # pattern = r"^[^\p{Script=Hiragana}^\p{Script=Katakana}^\p{Script=Han}ー〜・]+$"
        # pattern = r"^[^\p{Script=Hiragana}^\p{Script=Katakana}^\p{Script=Han}ー〜・]+$"
        pattern = r"^[^\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FFFー〜・]+$"
        if re.search(pattern, text, flags=re.U):
            return True
        else:
            return False

    @staticmethod
    def is_containing_eng(text):
        pattern = "[a-zA-Z]"
        if re.search(pattern, text):
            return True
        else:
            return False

    @staticmethod
    def is_containing_number(text):
        pattern = "[0-9]"
        if re.search(pattern, text):
            return True
        else:
            return False

    @staticmethod
    def is_only_no(text):
        pattern = "^[0-9]+$"
        if re.search(pattern, text):
            return True
        else:
            return False

    @staticmethod
    def is_only_speacial_characters(text):
        pattern = "^[~!@#$%^&*()_+|<>?:{}]+$"
        if re.search(pattern, text):
            return True
        else:
            return False

    @staticmethod
    def is_only_eng_and_kor_and_no_and_speacial_characters(text):
        pattern = "^[ㄱ-ㅎ가-힣0-9a-zA-Z~!@#$%^&*()_+|<>?:{}]+$"
        if re.search(pattern, text):
            return True
        else:
            return False

    @staticmethod
    def is_only_eng_and_no_and_speacial_characters(text):
        pattern = "^[0-9a-zA-Z~!@#$%^&*()_+|<>?:{}]+$"
        if re.search(pattern, text):
            return True
        else:
            return False

    @staticmethod
    def is_only_eng_and_speacial_characters(text):
        pattern = "^[a-zA-Z~!@#$%^&*()_+|<>?:{}]+$"
        if re.search(pattern, text):
            return True
        else:
            return False

    @staticmethod
    def is_only_eng_and_no(text):
        pattern = "^[0-9a-zA-Z]+$"
        if re.search(pattern, text):
            return True
        else:
            return False

    @staticmethod
    def is_only_eng(text):
        pattern = "^[a-zA-Z]+$"
        if re.search(pattern, text):
            return True
        else:
            return False

    @staticmethod
    def is_eng_or_kor_ja(text: str):
        """
        한글처리 :
            한영숫특 :
            한숫특 :
            한숫
            한특
            숫특
            특
            숫
        영어처리 :
            영숫특
            영특
            영숫
            영
        일어처리 :
        빠진거있나? 일단 여기까지

        문자구성 판별기가 필요하다
            return "eng, kor, jap, special_characters ", ... 이런식 > what_does_this_consist_of() 를 만들었다.
        """
        if TextToSpeechUtil.is_only_speacial_characters(text=text):
            return "ko"
        elif TextToSpeechUtil.is_only_no(text=text):
            return "ko"
        elif TextToSpeechUtil.is_containing_kor(text=text):
            return "ko"
        if TextToSpeechUtil.is_only_eng_and_no_and_speacial_characters(text=text):
            return "en"
        elif TextToSpeechUtil.is_only_eng_and_speacial_characters(text=text):
            return "en"
        elif TextToSpeechUtil.is_only_eng_and_no(text=text):
            return "en"
        elif TextToSpeechUtil.is_only_eng(text=text):
            return "en"
        else:
            return "ko"

    @staticmethod
    def what_does_this_consist_of(text: str):
        result = []
        if TextToSpeechUtil.is_containing_kor(text=text):
            result.append("kor")
        if TextToSpeechUtil.is_containing_eng(text=text):
            result.append("eng")
        if TextToSpeechUtil.is_containing_number(text=text):
            result.append("number")
        if TextToSpeechUtil.is_containing_special_characters_with_thread(text=text):
            result.append("special_characters")
        if TextToSpeechUtil.is_containing_jpn(text=text):
            result.append("jpn")
        print(rf'text : {text}')
        print(rf'result : {result}')
        return result

    @staticmethod
    def get_length_of_mp3(target_abspath: str):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        try:
            audio = MP3(target_abspath)
            return audio.info.length
        except mutagen.MutagenError:
            # DebuggingUtil.trouble_shoot("%%%FOO%%%") # gtts 모듈 불능? mutagen 모듈 불능? license 찾아보자 으로 어쩔수 없다.
            return
        except Exception:
            DebuggingUtil.trouble_shoot("%%%FOO%%%")

    @staticmethod
    def run_loop_for_speak_as_async_deprecated(ment):
        async def speak_as_async(ment):
            DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
            # pyglet_player 가 play() 하고 있으면 before_mp3_length_used_in_speak_as_async 만큼 대기
            DebuggingUtil.print_magenta(rf'before_mp3_length_used_in_speak_as_async 만큼 재생 대기')

            # Park4139.pyglet_player 을 del 을 한 경우:
            # Park4139.pyglet_player 에 대한 참조가 불가능해진다. 다른 참조법을 적용해보자. 시도해보니 Park4139.pyglet_player 는 더이상 쓸 수가 없다.
            # 재활용해야하는 Park4139.pyglet_player는 더이상 쓸 수 없어졌다.
            if StateManagementUtil.pyglet_player == None:
                StateManagementUtil.pyglet_player = pyglet.media.Player()
            # if hasattr(Park4139, 'pyglet_player') and Park4139.pyglet_player is None:
            #     # Park4139.pyglet_player가 None인 경우
            #     Park4139.pyglet_player = pyglet.media.Player()
            #     pass
            if StateManagementUtil.pyglet_player != None:
                if StateManagementUtil.pyglet_player.playing == True:
                    length_of_mp3 = BusinessLogicUtil.previous_mp3_length_used_in_speak_as_async
                    time.sleep(length_of_mp3 * 0.65)  # 이렇면 소리가 겹치지 않는 장점이 있으나, 실제로 일을 미리하고 나중에 보고하는 단점이 있다
                    # asyncio.sleep(length_of_mp3 * 0.65) # 이렇면 실제로 일을 하자마자 보고하는 장점이 있으나, 소리가 겹치는 않는 단점이 있다
                    # time.sleep(length_of_mp3 * 0.75)
                    # time.sleep(length_of_mp3 * 0.85)
                    # time.sleep(length_of_mp3 * 0.95)
                    # time.sleep(length_of_mp3 * 1.05)
                    # time.sleep(length_of_mp3 * 1.00)
            try:
                while True:
                    if type(ment) == str:
                        # while Park4139.is_speak_as_async_running:
                        #     Park4139.debug_as_cli(f"def {inspect.currentframe().f_code.co_name}() 에 대한 다른 쓰레드를 기다리는 중입니다")
                        #     pass
                        # Park4139.is_speak_as_async_running = True # 쓰레드상태 사용 중으로 업데이트

                        cache_mp3 = rf'{StateManagementUtil.PROJECT_DIRECTORY}\$cache_mp3'
                        FileSystemUtil.make_leaf_directory(leaf_directory_abspath=cache_mp3)

                        # DebuggingUtil.print_magenta("ment 전처리, 윈도우 경로명에 들어가면 안되는 문자들 공백으로 대체")
                        ment = BusinessLogicUtil.get_str_replaced_special_characters(target=ment, replacement=" ")
                        ment = ment.replace("\n", " ")

                        # DebuggingUtil.print_magenta(rf'파일 없으면 생성')
                        ment__mp3 = rf'{cache_mp3}/{ment}_.mp3'
                        ment_mp3 = rf'{cache_mp3}/{ment}.mp3'
                        if not os.path.exists(ment_mp3):
                            if not os.path.exists(ment__mp3):
                                if "special_characters" in TextToSpeechUtil.what_does_this_consist_of(text=ment):
                                    gtts = gTTS(text=ment, lang='ko')
                                    gtts.save(ment__mp3)
                                elif "kor" in TextToSpeechUtil.what_does_this_consist_of(text=ment):
                                    gtts = gTTS(text=ment, lang='ko')
                                    gtts.save(ment__mp3)
                                elif "eng" in TextToSpeechUtil.what_does_this_consist_of(text=ment):
                                    gtts = gTTS(text=ment, lang='en')
                                    gtts.save(ment__mp3)
                                elif "jpn" in TextToSpeechUtil.what_does_this_consist_of(text=ment):
                                    gtts = gTTS(text=ment, lang='ja')
                                    gtts.save(ment__mp3)
                                else:
                                    gtts = gTTS(text=ment, lang='ko')
                                    gtts.save(ment__mp3)

                        try:
                            # 앞부분 소리가 들리지 않는 현상 해결 아이디어.
                            # 앞부분의 단어가 들리지 않은채로 음악파일의 중간부터가 소리가 들린다. 전체 소리가 다 들려야 하는데.
                            # 음악파일의 앞부분에 빈소리를 추가해주려고 한다. ffmpeg를 이용하여 silent.mp3(1초간 소리가 없는 mp3 파일)을 소리를 재생해야할 mp3 파일의 앞부분에 합쳐서 재생시도.
                            silent_mp3 = rf"{cache_mp3}\silent.mp3"
                            if not os.path.exists(silent_mp3):
                                DebuggingUtil.print_magenta("사일런트 mp3 파일이 없습니다")
                                break
                            if not os.path.exists(ment_mp3):
                                cmd = rf'echo y | "ffmpeg" -i "concat:{os.path.abspath(silent_mp3)}|{os.path.abspath(ment__mp3)}" -acodec copy -metadata "title=Some Song" "{os.path.abspath(ment_mp3)}" -map_metadata 0:-1  >nul 2>&1'
                                FileSystemUtil.get_cmd_output(cmd)
                        except Exception:
                            DebuggingUtil.trouble_shoot("%%%FOO%%%")

                        try:
                            # 자꾸 프로세스 세션을 빼앗기 때문에 불편한 코드
                            # os.system(rf'start /b cmd /c call "{ment_mp3}" >nul 2>&1')

                            # 프로세스 세션을 빼앗지 않고 mp3재생하는 다른 방법
                            # playsound.playsound(os.path.abspath(ment_mp3))

                            # Park4139.debug_as_cli(rf'프로세스 세션을 빼앗지 않고 mp3재생')
                            ment_mp3 = os.path.abspath(ment_mp3)
                            DebuggingUtil.print_magenta(rf'{ment_mp3}')
                            try:
                                # 재생
                                # pyglet_player = Park4139.pyglet_player

                                pyglet_player = pyglet.media.Player()
                                source = pyglet.media.load(ment_mp3)
                                # Park4139.multiprocessed_source_play = source
                                pyglet_player.queue(source)

                                # 재생추적

                                # 재생 중인 사운드 리소스를 추적하는 이벤트 핸들러 함수
                                def on_player_eos():
                                    player = pyglet.media.Player()
                                    playing_sounds.remove(player)

                                pyglet_player.push_handlers(on_eos=on_player_eos)
                                # pyglet_player.play()  # 이거 재생 안되는데, 이게 공식문서에 나온 방식인데, media 객채 생성 순서가 문제였나 보다.이젠 된다
                                pyglet_player.play()
                                source.play()  # 웃긴다 이거 이렇게는 재생이 된다.
                                playing_sounds = StateManagementUtil.playing_sounds
                                playing_sounds.append(pyglet_player)
                                # 사운드 재생을 시작할 때마다 이벤트 핸들러를 등록하여 playing_sounds 리스트에 추가

                                # FAIL
                                # import  multiprocessing
                                # Park4139.multiprocessed_source_play = multiprocessing.Process(target= source.play , args=None)
                                # Park4139.multiprocessed_source_play.start()

                                if StateManagementUtil.pyglet_player.playing == False:
                                    StateManagementUtil.pyglet_player = None
                                    BusinessLogicUtil.previous_mp3_length_used_in_speak_as_async = 0

                                # Park4139.is_speak_as_async_running = False # 쓰레드상태 사용종료로 업데이트
                            except FileNotFoundError:
                                DebuggingUtil.print_magenta(f"{ment_mp3} 재생할 파일이 없습니다")
                            except:
                                DebuggingUtil.trouble_shoot("%%%FOO%%%")
                                return

                            DebuggingUtil.commentize("before_mp3_length_used_in_speak_as_async 업데이트")
                            length_of_mp3 = round(float(TextToSpeechUtil.get_length_of_mp3(ment_mp3)), 1)
                            BusinessLogicUtil.previous_mp3_length_used_in_speak_as_async = length_of_mp3

                        except Exception:
                            DebuggingUtil.trouble_shoot("%%%FOO%%%")

                        # Park4139.debug_as_cli(rf'불필요 파일 삭제')
                        os.system(f'echo y | del /f "{ment__mp3}" >nul 2>&1')

                        # Park4139.debug_as_cli(rf'중간로깅')
                        DebuggingUtil.log_mid(log_title=f"TTS 재생시도")
                        return
                    return
            except Exception:
                DebuggingUtil.trouble_shoot("%%%FOO%%%")

        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.run_until_complete(speak_as_async(ment))

    @staticmethod
    def speak_ments(ment, sleep_after_play=1.00, thread_join_mode=False):
        if FileSystemUtil.is_os_windows():
            # if thread_for_speak_as_async is not None:
            #     # 이전에 생성된 쓰레드가 있다면 종료
            #     thread_for_speak_as_async.join()

            # 이미 실행 중인 speak()관련 쓰레드가 있다면 종료시킴
            # if threading.Event() != None:
            #     global exit_event
            #     exit_event = threading.Event()
            # else:
            #     exit_event.set()
            #     exit_event.clear()  # 종료 이벤트 객체 초기화

            # pyglet을 실행하는 스레드 종료
            # Park4139.kill_thread(thread_name="thread_for_speak")
            # Park4139.kill_thread(thread_name="thread_for_run_loop_for_speak_as_async")

            # 재생 중인 사운드 리소스를 중지하는 함수
            def stop_all_sounds():
                playing_sounds = StateManagementUtil.playing_sounds
                for player in playing_sounds:
                    player.pause()  # 또는 player.stop()

            stop_all_sounds()
            if StateManagementUtil.pyglet_player != None:
                # if Park4139.pyglet_player.playing == True:
                # pyglet.app.exit()
                # Park4139.pyglet_player.delete()
                # Park4139.pyglet_player = None
                # Park4139.pyglet_player.pause()
                # pyglet.app.platform_event_loop.pause()
                # Park4139.pyglet_player.pause()
                # Park4139.pyglet_player.delete()

                # Park4139.pyglet_player 을 del 을 한 경우. 의도한대로 pyglet 의 play() 동작은 중지가 되는데.
                # Park4139.pyglet_player 에 대한 참조가 불가능해진다. 다른 참조법을 적용해보자. 시도해보니 Park4139.pyglet_player 는 더이상 쓸 수가 없다.
                # 재활용해야하는 Park4139.pyglet_player는 더이상 쓸 수 없어졌다.
                # del Park4139.pyglet_player
                # 혹시 몰라서 tmp 변수에 참조시켜서 del 시도 해보았는데, Park4139.pyglet_player에 대해 참조는 가능해 보인다.
                # tmp = Park4139.pyglet_player
                # del tmp
                # del 쓰는 것을 포기하자

                # TBD
                # sound = pyglet.resource.media('sound.wav')
                # sound.stop()
                pass
            else:
                import pyglet
                StateManagementUtil.pyglet_player = pyglet.media.Player()
                if StateManagementUtil.pyglet_player.playing == True:
                    StateManagementUtil.pyglet_player.pause()
                    # Park4139.pyglet_player.delete()
                    # del Park4139.pyglet_player
                    # tmp = Park4139.pyglet_player
        else:
            DebuggingUtil.print_magenta(ment=f"{inspect.currentframe().f_code.co_name}() 는 리눅스 환경에서 테스트가 필요합니다.")

        # 비동기 이벤트 함수 설정 (simple for void async processing)
        async def process_thread_speak(ment):
            # while not exit_event.is_set(): # 쓰레드 이벤트 제어
            DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
            ment = str(ment)
            ment = ment.strip()
            if ment == "":
                return None
            if TextToSpeechUtil.is_containing_special_characters_with_thread(text=ment):
                ment = TextToSpeechUtil.remove_special_characters(text=ment)
            while True:
                ments = [x.strip() for x in ment.replace(".", ",").split(",")]  # from "abc,abc.abc,abc." to ["abc","abc","abc","abc"] # , or . 를 넣으면 나누어 읽도록 업데이트
                ments = [x for x in ments if x.strip()]  # 리스트 요소 "" 제거,  from ["", A] to [A]
                # [print(sample) for sample in ments ]
                # print(rf'ments : {ments}')
                # print(rf'type(ments) : {type(ments)}')
                # print(rf'len(ments) : {len(ments)}')
                for ment in ments:
                    # Park4139Park4139.Tts.speak(ment)
                    # thread = threading.Thread(target=partial(Park4139Park4139.Tts.run_loop_for_speak_as_async, ment), name ="thread_for_run_loop_for_speak_as_async")
                    # thread.start()
                    # if thread.is_alive():
                    #     thread.join()
                    TextToSpeechUtil.speak_ment_without_async_and_return_last_word_mp3_length(ment, sleep_after_play=sleep_after_play)
                    pass
                return None

        # 비동기 이벤트 루프 설정 (simple for void async processing)
        def process_thread_loop(ment):
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop.run_until_complete(process_thread_speak(ment))

        # 비동기 이벤트 루프 실행할 쓰레드 설정 (simple for void async processing)
        thread = threading.Thread(target=partial(process_thread_loop, ment), name="thread_for_speak")  # Q: 왜 thread.name 의 case 는 다른거야 ?  wrtn: 네, 스레드의 이름은 일반적으로 대소문자를 구분하지 않습니다.
        thread.start()
        if thread_join_mode == True:
            thread.join()

    @staticmethod
    def speak_ment(ment, sleep_after_play=1.00):  # 많이 쓸 수록 프로그램이 느려진다
        # 무음 모드 적용
        return

        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        ment = str(ment)
        ment = ment.strip()
        if BusinessLogicUtil.is_containing_special_characters(text=ment):
            ment = TextToSpeechUtil.remove_special_characters(text=ment)
        if ment == "":
            return None
        try:
            while True:
                ments = []
                if "," in ment:  # , 를 넣으면 나누어 읽도록 업데이트
                    ments = ment.split(",")
                    for ment in ments:
                        TextToSpeechUtil.speak_ments(ment, sleep_after_play=0.65)
                    break
                if type(ment) == str:
                    cache_mp3 = rf'{StateManagementUtil.PROJECT_DIRECTORY}\$cache_mp3'
                    FileSystemUtil.make_leaf_directory(leaf_directory_abspath=cache_mp3)

                    # DebuggingUtil.print_magenta("ment 전처리, 윈도우 경로명에 들어가면 안되는 문자들 공백으로 대체")
                    ment = BusinessLogicUtil.get_str_replaced_special_characters(target=ment, replacement=" ")
                    ment = ment.replace("\n", " ")

                    # DebuggingUtil.print_magenta(rf'파일 없으면 생성')
                    ment__mp3 = rf'{cache_mp3}/{ment}_.mp3'
                    ment_mp3 = rf'{cache_mp3}/{ment}.mp3'
                    if not os.path.exists(ment_mp3):
                        if not os.path.exists(ment__mp3):
                            if "special_characters" in TextToSpeechUtil.what_does_this_consist_of(text=ment):
                                gtts = gTTS(text=ment, lang='ko')
                                gtts.save(ment__mp3)
                            elif "kor" in TextToSpeechUtil.what_does_this_consist_of(text=ment):
                                gtts = gTTS(text=ment, lang='ko')
                                gtts.save(ment__mp3)
                            elif "eng" in TextToSpeechUtil.what_does_this_consist_of(text=ment):
                                gtts = gTTS(text=ment, lang='en')
                                gtts.save(ment__mp3)
                            elif "jpn" in TextToSpeechUtil.what_does_this_consist_of(text=ment):
                                gtts = gTTS(text=ment, lang='ja')
                                gtts.save(ment__mp3)
                            else:
                                gtts = gTTS(text=ment, lang='ko')
                                gtts.save(ment__mp3)
                    try:
                        silent_mp3 = rf"{cache_mp3}\silent.mp3"
                        if not os.path.exists(silent_mp3):
                            DebuggingUtil.print_magenta("사일런트 mp3 파일이 없습니다")
                            break
                        if not os.path.exists(ment_mp3):
                            cmd = rf'echo y | "ffmpeg" -i "concat:{os.path.abspath(silent_mp3)}|{os.path.abspath(ment__mp3)}" -acodec copy -metadata "title=Some Song" "{os.path.abspath(ment_mp3)}" -map_metadata 0:-1  >nul 2>&1'
                            FileSystemUtil.get_cmd_output(cmd)
                    except Exception:
                        DebuggingUtil.trouble_shoot("%%%FOO%%%")
                    try:
                        ment_mp3 = os.path.abspath(ment_mp3)
                        DebuggingUtil.print_magenta(rf'{ment_mp3}')
                        try:
                            source = pyglet.media.load(ment_mp3)
                            source.play()

                            length_of_mp3 = TextToSpeechUtil.get_length_of_mp3(ment_mp3)
                            time.sleep(length_of_mp3 * sleep_after_play)
                            return length_of_mp3
                        except FileNotFoundError:
                            DebuggingUtil.print_magenta(f"{ment_mp3} 재생할 파일이 없습니다")
                        except:
                            DebuggingUtil.trouble_shoot("%%%FOO%%%")
                            break
                        DebuggingUtil.commentize("before_mp3_length_used_in_speak_as_async 업데이트")
                        length_of_mp3 = round(float(TextToSpeechUtil.get_length_of_mp3(ment_mp3)), 1)
                        BusinessLogicUtil.previous_mp3_length_used_in_speak_as_async = length_of_mp3
                    except Exception:
                        DebuggingUtil.trouble_shoot("%%%FOO%%%")

                    os.system(f'echo y | del /f "{ment__mp3}" >nul 2>&1')
                    DebuggingUtil.log_mid(log_title=f"TTS 재생시도")
                break
        except Exception:
            DebuggingUtil.trouble_shoot("%%%FOO%%%")

    @staticmethod
    def speak_ment_without_async_and_return_last_word_mp3_length(ment, sleep_after_play=1.00):
        '''
        이 함수를 많이 쓸 수록 프로그램이 느려진다, 왜냐하면 말하는 속도 < 처리 속도
        '''

        # 임시 음소거를 위한 설정 처리, 필요시 ment = "`" 삭제
        DebuggingUtil.print_ment_light_yellow("음소거 모드를 실행중입니다")
        ment = "`"

        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")  # noqa , pycharm 에서 개별검사 억제설정 via noqa
        ment = str(ment)
        ment = ment.strip()
        if TextToSpeechUtil.is_containing_special_characters_with_thread(text=ment):
            ment = TextToSpeechUtil.remove_special_characters(text=ment)
        if ment == "":
            return None
        try:
            while True:
                ments = []
                if "," in ment:  # , 를 넣으면 나누어 읽도록 업데이트
                    ments = ment.split(",")
                    for ment in ments:
                        TextToSpeechUtil.speak_ments(ment, sleep_after_play=0.65)
                    break
                if type(ment) == str:
                    # while Park4139.is_speak_as_async_running:
                    #     Park4139.debug_as_cli(f"def {inspect.currentframe().f_code.co_name}() 에 대한 다른 쓰레드를 기다리는 중입니다")
                    #     pass
                    # Park4139.is_speak_as_async_running = True # 쓰레드상태 사용 중으로 업데이트

                    cache_mp3 = rf'{StateManagementUtil.PROJECT_DIRECTORY}\$cache_mp3'
                    FileSystemUtil.make_leaf_directory(leaf_directory_abspath=cache_mp3)

                    # DebuggingUtil.print_magenta("ment 전처리, 윈도우 경로명에 들어가면 안되는 문자들 공백으로 대체")
                    ment = BusinessLogicUtil.get_str_replaced_special_characters(target=ment, replacement=" ")
                    ment = ment.replace("\n", " ")

                    # DebuggingUtil.print_magenta(rf'파일 없으면 생성')
                    ment__mp3 = rf'{cache_mp3}/{ment}_.mp3'
                    ment_mp3 = rf'{cache_mp3}/{ment}.mp3'
                    if not os.path.exists(ment_mp3):
                        if not os.path.exists(ment__mp3):
                            if "special_characters" in TextToSpeechUtil.what_does_this_consist_of(text=ment):
                                gtts = gTTS(text=ment, lang='ko')
                                gtts.save(ment__mp3)
                            elif "kor" in TextToSpeechUtil.what_does_this_consist_of(text=ment):
                                gtts = gTTS(text=ment, lang='ko')
                                gtts.save(ment__mp3)
                            elif "eng" in TextToSpeechUtil.what_does_this_consist_of(text=ment):
                                gtts = gTTS(text=ment, lang='en')
                                gtts.save(ment__mp3)
                            elif "jpn" in TextToSpeechUtil.what_does_this_consist_of(text=ment):
                                gtts = gTTS(text=ment, lang='ja')
                                gtts.save(ment__mp3)
                            else:
                                gtts = gTTS(text=ment, lang='ko')
                                gtts.save(ment__mp3)

                    try:
                        # 앞부분 소리가 들리지 않는 현상 해결 아이디어.
                        # 앞부분의 단어가 들리지 않은채로 음악파일의 중간부터가 소리가 들린다. 전체 소리가 다 들려야 하는데.
                        # 음악파일의 앞부분에 빈소리를 추가해주려고 한다. ffmpeg를 이용하여 silent.mp3(1초간 소리가 없는 mp3 파일)을 소리를 재생해야할 mp3 파일의 앞부분에 합쳐서 재생시도.
                        silent_mp3 = rf"{cache_mp3}\silent.mp3"
                        if not os.path.exists(silent_mp3):
                            DebuggingUtil.print_magenta("사일런트 mp3 파일이 없습니다")
                            break
                        if not os.path.exists(ment_mp3):
                            cmd = rf'echo y | "ffmpeg" -i "concat:{os.path.abspath(silent_mp3)}|{os.path.abspath(ment__mp3)}" -acodec copy -metadata "title=Some Song" "{os.path.abspath(ment_mp3)}" -map_metadata 0:-1  >nul 2>&1'
                            FileSystemUtil.get_cmd_output(cmd)
                    except Exception:
                        DebuggingUtil.trouble_shoot("%%%FOO%%%")

                    try:
                        # 자꾸 프로세스 세션을 빼앗기 때문에 불편한 코드
                        # os.system(rf'start /b cmd /c call "{ment_mp3}" >nul 2>&1')

                        # 프로세스 세션을 빼앗지 않고 mp3재생하는 다른 방법
                        # playsound.playsound(os.path.abspath(ment_mp3))

                        # Park4139.debug_as_cli(rf'프로세스 세션을 빼앗지 않고 mp3재생')
                        ment_mp3 = os.path.abspath(ment_mp3)
                        DebuggingUtil.print_magenta(rf'{ment_mp3}')
                        try:
                            # 재생
                            source = pyglet.media.load(ment_mp3)
                            # pyglet_player = Park4139.pyglet_player
                            # pyglet_player.queue(source)
                            source.play()  # 웃긴다 이거 이렇게는 재생이 된다.

                            length_of_mp3 = TextToSpeechUtil.get_length_of_mp3(ment_mp3)
                            # time.sleep(length_of_mp3 * 0.65)
                            # time.sleep(length_of_mp3 * 0.75)
                            # time.sleep(length_of_mp3 * 0.85)
                            # time.sleep(length_of_mp3 * 0.95)
                            # time.sleep(length_of_mp3 * 1.05)
                            # time.sleep(length_of_mp3 * 1.00)
                            time.sleep(length_of_mp3 * sleep_after_play)

                            return length_of_mp3
                            # Park4139.is_speak_as_async_running = False # 쓰레드상태 사용종료로 업데이트
                        except FileNotFoundError:
                            DebuggingUtil.print_magenta(f"{ment_mp3} 재생할 파일이 없습니다")
                        except:
                            DebuggingUtil.trouble_shoot("%%%FOO%%%")
                            break

                        DebuggingUtil.commentize("before_mp3_length_used_in_speak_as_async 업데이트")
                        length_of_mp3 = round(float(TextToSpeechUtil.get_length_of_mp3(ment_mp3)), 1)
                        BusinessLogicUtil.previous_mp3_length_used_in_speak_as_async = length_of_mp3

                    except Exception:
                        DebuggingUtil.trouble_shoot("%%%FOO%%%")

                    # Park4139.debug_as_cli(rf'불필요 파일 삭제')
                    os.system(f'echo y | del /f "{ment__mp3}" >nul 2>&1')

                    # Park4139.debug_as_cli(rf'중간로깅')
                    DebuggingUtil.log_mid(log_title=f"TTS 재생시도")
                break
        except Exception:
            DebuggingUtil.trouble_shoot("%%%FOO%%%")

    @staticmethod
    def remove_special_characters(text):
        # pattern = r"[^\w\s]" #  \s(whitespace characters = [" ", "\t", "\n", ".", "\n", "_" ])      # space = " " #공백
        # pattern = r"[^\w.,]" # \w(알파벳, 숫자, 밑줄 문자)와 "." 그리고 ","를 제외한 모든문자를 선택, \t, \n, _ 도 삭제되는 설정
        pattern = r"[~!@#$%^&*()_+|<>?:{}]"
        return re.sub(pattern, "", text)

    @staticmethod
    def speak_server_hh_mm():  # '몇 시야' or usr_input_txt == '몇시야':
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        server_hour = TimeUtil.get_time_as_('%H')
        server_minute = TimeUtil.get_time_as_('%M')
        TextToSpeechUtil.speak_ments(f'{server_hour}시 {server_minute}분 입니다', sleep_after_play=0.65)

    @staticmethod
    def speak_server_ss():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        server_seconds = TimeUtil.get_time_as_('%S')
        TextToSpeechUtil.speak_ments(f'{server_seconds}초', sleep_after_play=0.65)

    @staticmethod
    def speak_today_time_info():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        yyyy = TimeUtil.get_time_as_('%Y')
        MM = TimeUtil.get_time_as_('%m')
        dd = TimeUtil.get_time_as_('%d')
        HH = TimeUtil.get_time_as_('%H')
        mm = TimeUtil.get_time_as_('%M')
        week_name = BusinessLogicUtil.return_korean_week_name()
        TextToSpeechUtil.speak_ment_without_async_and_return_last_word_mp3_length(ment=f'현재 시각 {int(yyyy)}년 {int(MM)}월 {int(dd)}일 {week_name}요일 {int(HH)}시 {int(mm)}분', sleep_after_play=0.75)


class MathUtil:
    pi = 3.141592  # 원주율은 여기까지만 외웠기 때문 ㅋ

    @staticmethod
    def get_minuites_from_(seconds):  # 큰단위로 단위변환 시 숫자는 작아지니 숫자가 작아지도록 약속된 숫자를 곱하면 된다. # seconds -> min  인 경우 큰단위로 변환되고 있고 약속된 숫자는 60 이다. 작은 숫자를 만들기 위해서 1/60 을 곱한다.
        return round(seconds / 60, 2)

    @staticmethod
    def get_minuites_and_remaining_secs(seconds):  # 큰단위로 단위변환 시 숫자는 작아지니 숫자가 작아지도록 약속된 숫자를 곱하면 된다.
        mins = seconds // 60
        secs_remaining = seconds % 60
        return mins, secs_remaining

    @staticmethod
    def get_value_splited_integer_part_and_decimal_part_from_(value):  # 큰단위로 단위변환 시 숫자는 작아지니 숫자가 작아지도록 약속된 숫자를 곱하면 된다.
        import math
        decimal_part, integer_part = math.modf(value)
        decimal_part = abs(decimal_part)  # 음수 부호 제거
        return [integer_part, decimal_part]


class SeleniumUtil:
    @staticmethod
    def get_driver_for_selenium():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        import selenium.webdriver as webdriver
        # DebuggingUtil.commentize("웹 드라이버 설정")
        options = SeleniumUtil.get_webdriver_options_customed()

        # # driver = webdriver.Chrome(options=options,executable_path=rf"{os.getcwd()}\pkg_chrome_driver\chrome-win64\chrome.exe")
        # # executable_url 를 지워야겠다, because you have installed the latest version of Selenium if you have selenium above the 4.6.0
        # # you don't need to add executable_url and in the latest version of Selenium you don't need to download webdriver.
        # # stackoverflow 이렇다고 한다. chromedriver의 추가 설치가 불필요해졌다

        driver = webdriver.Chrome(options=options)
        driver.get('about:blank')
        driver.execute_script("Object.defineProperty(navigator, 'plugins', {get: function() {return[1, 2, 3, 4, 5]}})")  # hide plugin 0EA
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        driver.execute_script("Object.defineProperty(navigator, 'languages', {get: function() {return ['ko-KR', 'ko']}})")  # hide own lanuages
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        driver.execute_script("const getParameter = WebGLRenderingContext.getParameter;WebGLRenderingContext.prototype.getParameter = function(parameter) {if (parameter === 37445) {return 'NVIDIA Corporation'} if (parameter === 37446) {return 'NVIDIA GeForce GTX 980 Ti OpenGL Engine';}return getParameter(parameter);};")  # hide own gpu # WebGL렌더러를 Nvidia회사와 GTX980Ti엔진인 ‘척’ 하는 방법입니다.
        return driver

    @staticmethod
    def get_webdriver_options_customed():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        import selenium.webdriver as webdriver
        options = webdriver.ChromeOptions()

        # 백그라운드 실행 설정
        # headless 옵션은 브라우저를 화면에 표시하지 않고 백그라운드에서 실행하는 모드 설정
        options.add_argument('headless')

        # options.add_argument('window-size=1920x1080')
        options.add_argument('window-size=3440x1440')  # size 하드코딩 되어 있는 부분을 pyautogui 에서 size() 로 대체하는 것이 좋겠다.
        options.add_argument("lang=ko_KR")  # 한국어!
        options.add_argument("disable-gpu")  # 문제 있을 수 있음, 아래는 그 대안코드
        # options.add_argument("--disable-gpu")
        # hide headless, UserAgent headless 탐지 방지
        # UserAgent 값 얻는 방법, 일반 크롬으로 https://www.whatismybrowser.com/detect/what-is-my-user-agent/ 에 접속해서 얻어온다
        user_agent_value = "Mozilla/5.0(Windows NT 10.0; Win64; x64) AppleWebKit/537.36(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        options.add_argument(f"user-agent={user_agent_value}")
        # options.add_argument("proxy-server=localhost:8080")
        return options


class FastapiUtil:
    # def import_pkg_park4139_for_linux():  # 명령어 자체가 안되는데 /mir 은 되는데 /move 안된다
    # try:
    # 파이썬 프로젝트 외부 패키지 import 시키는 방법 : 일반적으로는 불가능하다.
    # services_directory_abspath = os.path.dirname(os.path.dirname(__file__))
    # external_pkg_abspath = rf'{services_directory_abspath}\archive_py\pkg_park4139_for_linux'
    # print(rf'external_pkg_abspath : {external_pkg_abspath}')
    # os.environ['pkg_park4139_for_linux_ABSPATH'] = rf'{external_pkg_abspath}'
    # sys.path.append(external_pkg_abspath)
    # print(rf'sys.path : {sys.path}')
    # from ..archive_py.pkg_park4139_for_linux import FileSystemUtil, StateManagementUtil, TestUtil
    # except:
    #     pass
    # try:
    # 대안 global pkg 로 복사하고 쓰고 실행후에는 프로젝트 내에서 패키지를 삭제한다 > 자동화할것
    # 의존성을 추가하기 위해서, pkg_park4139_for_linux의 .venv 도 따라 복사한다.
    # CURRENT_PROJECT_DIRECTORY= os.path.dirname(__file__)
    # SERVICES_DIRECTORY = os.path.dirname(CURRENT_PROJECT_DIRECTORY)
    # EXTERNAL_PKG_ABSPATH = rf'{SERVICES_DIRECTORY}\archive_py\pkg_park4139_for_linux'
    # # VIRTUAL_PKG_ABSPATH = rf'{CURRENT_PROJECT_DIRECTORY}\.venv\Lib\site-packages'
    # INTERNAL_PKG_ABSPATH = rf'{CURRENT_PROJECT_DIRECTORY}\pkg_park4139_for_linux'
    # print(rf'VIRTUAL_PKG_ABSPATH : {INTERNAL_PKG_ABSPATH}')
    # os.system(rf'chcp 65001')
    # os.system(rf'robocopy "{EXTERNAL_PKG_ABSPATH}" "{INTERNAL_PKG_ABSPATH}" /MIR')
    # # os.system("pause")
    # print("파이썬 프로젝트 외부 패키지 import 시도에 성공")
    # except Exception:
    #     print("파이썬 프로젝트 외부 패키지 import 시도에 실패")
    #     sys.exit()
    # try:
    # 그냥 pkg 내에 집어 넣었다. 가상환경도 동일하게 개발한다.
    # except:
    #     pass
    # pass

    @staticmethod
    def convert_file_from_cp_949_to_utf_8(file_abspath):
        # import codecs
        # file_abspath_converted = rf"{FileSystemUtil.get_target_as_pn(file_abspath)}_utf_8{FileSystemUtil.get_target_as_x(file_abspath)}"
        # print(file_abspath_converted)
        # # 기존 파일을 'utf-8'로 읽어오고, 새로운 파일에 'utf-8'로 저장
        # with codecs.open(file_abspath, 'r', encoding='cp949') as file_previous:
        #     contents = file_previous.read()
        #     with codecs.open(file_abspath_converted, 'w', encoding='utf-8') as new_file:
        #         new_file.write(contents)
        pass

    @staticmethod
    def init_cors_policy(app):
        # dev
        # op 에서는 이 함수는 사용하면 않기로 결정, nginx 에서 CORS 설정을 할 것 이므로
        # CORS 두개 설정하면 multi header 로 인한 Control-Allow-Origin' header contains multiple values '*, https://www.pjh4139.store', but only one is allowed. 에러가 발생할 것이다.
        app.add_middleware(
            # success, fastapi CORS allow_origins 동적 할당은 대안 못찾았고, # 초기 origins 값을 와일드카드로 설정
            CORSMiddleware,  # 노란줄 원인 뭘까? #_noqa 를 적용해야 할지 고민 중
            allow_credentials=True,  # cookie 포함 여부를 설정. 기본은 False
            allow_origins=['*'],
            allow_methods=["*"],  # 허용할 method를 설정할 수 있으며, 기본값은 'GET'이다. OPTIONS request ?
            allow_headers=["*"],  # 허용할 http header 목록을 설정할 수 있으며 Content-Type, Accept, Accept-Language, Content-Language은 항상 허용된다.

            # try, 보안강화
            # fastapi CORS allow_origins 정적 할당 # 톳시하나 틀리면 안됨.
            # allow_origins=[,
            #     # "http://localhost",  # fail
            #     # "http://localhost/",  # fail
            #     # "http://localhost:3000",# success
            #     "http://localhost:3000/service-dev-diary",
            #     # "http://127.0.0.1:3000",#
            #     "https://e-magazine-jung-hoon-parks-projects.vercel.app",
            #     # "https://e-magazine-jung-hoon-parks-projects.vercel.app/",
            #     "https://e-magazine-jung-hoon-parks-projects.vercel.app/````````````````````",
            # ],
        )

    @staticmethod
    def init_Logging_via_middleware(app):
        # 미들웨어를 통한 로깅
        # @app.middleware("http")
        # async def log_requests(request: Request, call_next):
        #     logging.debug(f"요청: {request.method} {request.url}")
        #     response = await call_next(request)
        #     logging.debug(f"응답: {response.status_code}")
        #     return response
        pass

    @staticmethod
    def init_domain_address_allowed(app):
        # 접속허용도메인 주소 제한, 이 서버의 api 를 특정 도메인에서만 호출 가능하도록 설정
        # @app.middleware("http")
        # async def domain_middleware(request: Request, call_next):
        #     allowed_domains = ["example.com", "subdomain.example.com"]
        #     client_host = request.client.host
        #     client_domain = client_host.split(":")[0]  # 도메인 추출
        #     if client_domain not in allowed_domains:
        #         raise HTTPException(status_code=403, detail="Forbidden")
        #     response = await call_next(request)
        #     return response
        pass

    @staticmethod
    def init_ip_address_allowed(app):
        # 접속허용IP 주소 제한
        # @app.middleware("http")
        # async def ip_middleware(request: Request, call_next):
        #     allowed_ips = [
        #         "127.0.0.1",
        #         "10.0.0.1",
        #     ]
        #     client_ip = request.client.host
        #     if client_ip not in allowed_ips:
        #         raise HTTPException(status_code=403, detail="Forbidden")
        #     response = await call_next(request)
        #     return response
        pass

    @staticmethod
    async def preprocess_after_request(request):
        # DebuggingUtil.commentize(f"{str(request.url)} 로 라우팅 시도 중...")
        pass

    @staticmethod
    async def preprocess_before_response_return(request, response):
        # DebuggingUtil.commentize(f"{str(request.url)} 로 라우팅 되었습니다")
        pass

    @staticmethod
    def init_and_update_json_file(json_file, objects=None):
        if objects is None:
            objects = []
        try:
            FileSystemUtil.make_leaf_file(file_abspath=json_file)
            if os.path.exists(json_file):
                if FileSystemUtil.is_letters_cnt_zero(file_abspath=json_file) == True:
                    FileSystemUtil.add_text_to_file(file_abspath=json_file, text="[]\n")  # 이러한 형태로 객체를 받을 수 있도록 작성해 두어야 받을 수 있음.
                else:
                    if not os.path.isfile(json_file):
                        with open(json_file, "w", encoding='utf-8') as f:
                            # json.dump(objects, f, ensure_ascii=False)  # ensure_ascii=False 는 encoding 을 그대로 유지하는 것 같다. ascii 로 변환하는게 안전할 지도 모르겠다.
                            json.dump(objects, f)  # ensure_ascii=False 는 encoding 을 그대로 유지하는 것 같다. ascii 로 변환하는게 안전할 지도 모르겠다.
                    else:
                        with open(json_file, "r", encoding='utf-8') as f:
                            # DebuggingUtil.print_ment_via_colorama(f"{BOOKS_FILE} 업로드 되었습니다", colorama_color=ColoramaColorUtil.LIGHTWHITE_EX)
                            objects = json.load(f)
                        return objects
        except IOError as e:
            print("파일 작업 중 오류가 발생했습니다:", str(e))

    @staticmethod
    def test_client_post_request():  # swagger 로 해도 되지만, test 자동화 용도
        # def save_data_via_api_server():
        import requests
        import json
        url = "https://park4139.store/api/volatile/items"
        url = "https://park4139.store/api/db-json/items"
        url = "https://park4139.store/api/db-maria/items"

        # 저장할 데이터
        data = {
            "name": "John Doe",
            "email": "johndoe@example.com",
            "password": "password",
        }
        # 데이터를 JSON 문자열로 변환합니다.
        json_data = json.dumps(data)

        # POST 요청을 생성합니다.
        headers = {"Content-Type": "application/json"}
        response = requests.post(url, headers=headers, data=json_data)

        # 응답을 확인합니다.
        if response.status_code == 201:
            print("데이터가 성공적으로 저장되었습니다.")
        else:
            print("데이터 저장에 실패했습니다.")

    # Pydantic은 Python의 데이터 유효성 검사 및 직렬화를 위한 라이브러리,
    # BaseModel은 Pydantic에 built in 되어 있다,
    # 데이터 모델을 정의하고 유효성을 검사하며 직렬화/역직렬화를 수행하는 기능을 제공합니다
    # auto increment 하고 싶어 pydantic model 로 구현
    class Board(BaseModel):
        id: int
        title: str
        content: str

        class Config:
            populate_by_name = True

    @staticmethod
    def is_board_validated(board: Board) -> bool:
        # Board(게시물)의 유효성을 검사하는 커스텀 비지니스 로직을 구현, create/update 의 경우에 사용
        if not board.title:
            return False
        if not board.content:
            return False
        return True

    class Book(BaseModel):  # BaseModel 를 상속받은 Book 은 일반적인 객체가 아니다. type 을 출력해봐도 pydantic 의 하위 객체를 상속한 것으로 보인다
        id: Optional[str] = uuid4().hex  # Optional 을 설정하면 nullable 되는 거야? # 여기서 할당을 시키면 put() 동작하며 id가 바뀌어버린다. put()에서 업데이트되도록 따로 설정했다.
        name: str
        genre: Literal["러브코메디", "러브픽션", "액션"]  # string literal validation 설정, 이 중 하나만 들어갈 수 있음
        # price: float
        price: int

    class TodoItem(BaseModel):
        id: str
        title: str
        completed: bool

    class User(BaseModel):  # 여기에 validation 해두면 docs에서 post request 시 default 값 지정해 둘 수 있음.
        # id: str = uuid4().hex + TimeUtil.get_time_as_('%Y%m%d%H%M%S%f') + SecurityUtil.get_random_alphabet()  # 진짜id 는 이렇게 default 로 들어가게하고, 사용자 id 는 e-mail
        id: str
        pw: str
        name: str
        date_join: str = TimeUtil.get_time_as_('%Y-%m-%d %H:%M %S%f')
        date_logout_last: str
        address_house: str
        address_e_mail: str
        number_phone: str

        @classmethod
        @field_validator('id')
        def validate_id(cls, id):
            if 53 < len(id) or 53 < len(id):
                return {"fail", "id 는 53 자리 이상 이거나 이하 일 수 없습니다"}
            return id

        @classmethod
        @field_validator('pw')
        def validate_pw(cls, pw):
            # 다시 작성
            return pw

        @classmethod
        @field_validator('name')
        def validate_name(cls, name):
            if 30 < len(name):
                return {"fail", "name 은 30 자리 이상 일 수 없습니다"}
            return name

        @classmethod  # class 간 종속 관계가 있을 때 하위 class 에 붙여 줘야하나?, cls, 파라미터와 함께? , instance를 생성하지 않고 호출 가능해?
        @field_validator('date_join')
        def validate_date_join(cls, date_join):
            # datetime.strptime(date_join, '%Y-%m-%d %H:%M %S%f')
            # 다시 작성
            return date_join

        @classmethod
        @field_validator('address_e_mail')
        def validate_address_e_mail(cls, address_e_mail):
            # 다시 작성
            return address_e_mail

    class LoginForm(BaseModel):
        id: str
        pw: str

    class JoinForm(BaseModel):  # 회원가입 유효성검사 설정
        id: str  # 필수항목
        pw: str
        pw2: str
        name: str
        pw: str
        phone_no: str
        address: str
        e_mail: str
        age: str
        birthday: str
        date_joined: Optional[str]  # nullable
        date_canceled: Optional[str]
        fax_no: Optional[str]
        business_registration_no: Optional[str]
        company_name: Optional[str]
        department: Optional[str]
        position: Optional[str]
        company_address: Optional[str]

    class jinja_data(BaseModel):  # 모든 데이터를 수집하고, Optional 로 할당.
        id: Optional[str]  # nullable
        pw: Optional[str]
        prefix_promised: Optional[str]
        random_bytes: Optional[str]
        random_str: Optional[str]
        pw2: Optional[str]
        name: Optional[str]
        pw: Optional[str]
        phone_no: Optional[str]
        address: Optional[str]
        e_mail: Optional[str]
        age: Optional[str]
        date_joined: Optional[str]
        date_canceled: Optional[str]
        fax_no: Optional[str]
        business_registration_no: Optional[str]
        company_name: Optional[str]
        department: Optional[str]
        position: Optional[str]
        company_address: Optional[str]
        birthday: Optional[str]
        items: Optional[list]


class UvicornUtil:
    class Settings:
        # class 를 사용하면 tuple 로 오며, str(tuple) 이렇게 사용할 수 없고, tuple[0] 으로 가져와야 하네.
        # js 의 destructon 문법처럼 py의 unpacking 을 사용하는 방법이 있으나 변수 새로 생성해야함
        # 위의 생각은 틀렸다 잘못 코드를 작성한 것이다 , 가 문제 였다. 이 오류는 IDE 에서 알려주지 않는다.
        protocol_type = "http"  # success
        # protocol_type =  "https"
        # host =  "0.0.0.0"
        host = "127.0.0.1"  # success , localhost
        port = 8080  # success
        url = f"{protocol_type}://{host}:{port}"


class DbTomlUtil:
    @staticmethod
    def create_db_toml():
        try:
            db_abspath = StateManagementUtil.DB_TOML
            # db_template =Park4139.db_template
            DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
            if not os.path.exists(os.path.dirname(db_abspath)):
                os.makedirs(os.path.dirname(db_abspath))  # 이거 파일도 만들어지나? 테스트 해보니 안만들어짐 디렉토리만 만들어짐
            # DebuggingUtil.commentize("db 파일 존재 검사 시도")
            if not os.path.isfile(db_abspath):
                # with open(db_abspath, "w") as f2:
                #     toml.dump(db_template, f2)
                FileSystemUtil.make_leaf_file(db_abspath)
                DebuggingUtil.print_magenta(f"DB를 만들었습니다 {db_abspath}")
            else:
                DebuggingUtil.print_magenta(f"DB가 이미존재합니다 {db_abspath}")
        except:
            DebuggingUtil.trouble_shoot("%%%FOO%%%")

    @staticmethod
    def read_db_toml():
        try:
            DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
            print("DB 의 모든 자료를 가져왔습니다")
            return toml.load(StateManagementUtil.DB_TOML)
        except:
            DebuggingUtil.trouble_shoot("%%%FOO%%%")

    @staticmethod
    def get_db_toml_key(unique_id):
        return BusinessLogicUtil.get_str_replaced_special_characters(target=unique_id, replacement="_")

    @staticmethod
    def select_db_toml(key):
        try:
            key = str(key)
            DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
            db_abspath = StateManagementUtil.DB_TOML
            try:
                # return toml.load(DB_TOML)[key]
                with open(db_abspath, 'r', encoding='utf-8') as f:
                    db = toml.load(f)
                return db[key]

            except KeyError:
                print(f"DB 에 해당키가 존재하지 않습니다 {key}")
                return None
        except:
            DebuggingUtil.trouble_shoot("%%%FOO%%%")

    @staticmethod
    def update_db_toml(key, value):
        try:
            db_abspath = StateManagementUtil.DB_TOML
            key: str = str(key)
            DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
            # 기존의 DB 의 데이터
            with open(db_abspath, 'r', encoding='utf-8') as f:
                db = toml.load(f)
                # DebuggingUtil.commentize("DB 업데이트 전 ")
                # Park4139.debug_as_cli(context=str(db))

            # 데이터 업데이트
            if key in db:
                db[key] = value
            else:
                # db[key] = value
                print(f"DB에 key가 없어서 업데이트 할 수 없습니다 key:{key} value:{value}")

            # TOML 파일에 데이터 쓰기
            with open(db_abspath, 'w', encoding='utf-8') as f:
                toml.dump(db, f)
                print(f"DB에 데이터가 업데이트되었습니다")

            # DB 변경 확인
            # with open(db_abspath, 'r') as f:
            #     db = toml.load(f)
            #     DebuggingUtil.commentize("DB 변경 확인")
            #     Park4139.debug_as_cli(context=str(db))
        except:
            DebuggingUtil.trouble_shoot("%%%FOO%%%")

    @staticmethod
    def insert_db_toml(key, value):
        try:
            key = str(key)
            db_abspath = StateManagementUtil.DB_TOML
            DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
            db_abspath = StateManagementUtil.DB_TOML
            # 기존의 DB 의 데이터
            with open(db_abspath, 'r', encoding='utf-8') as f:
                db = toml.load(f)
                # DebuggingUtil.commentize("DB 업데이트 전 ")
                # Park4139.debug_as_cli(context=str(db))

            # 데이터 삽입
            if key in db:
                # db[key] = value
                print(f"이미 값이 존재하여 삽입할 수 없습니다 key:{key} value:{value}")
            else:
                db[key] = value

            # TOML 파일에 데이터 쓰기
            with open(db_abspath, 'w', encoding='utf-8') as f:
                toml.dump(db, f)
                print(f"DB에 데이터가 삽입되었습니다")

            # DB 변경 확인
            # with open(db_abspath, 'r') as f:
            #     db = toml.load(f)
            #     DebuggingUtil.commentize("DB 변경 확인")
            #     Park4139.debug_as_cli(context=str(db))
        except:
            DebuggingUtil.trouble_shoot("%%%FOO%%%")

    # @staticmethod
    # def back_up_db_toml():
    #     try:
    #         DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #         BusinessLogicUtil.back_up_target(StateManagementUtil.DB_TOML)
    #     except:
    #         DebuggingUtil.trouble_shoot("%%%FOO%%%")

    @staticmethod
    def delete_db_toml():
        try:
            DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
            FileSystemUtil.convert_as_zip_with_timestamp(StateManagementUtil.DB_TOML)
        except:
            DebuggingUtil.trouble_shoot("%%%FOO%%%")

    @staticmethod
    def is_accesable_local_database():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        if not os.path.exists(StateManagementUtil.DB_TOML):
            DbTomlUtil.create_db_toml()
            return False
        else:
            return True


class MySqlUtil:
    """
    mysql 에 의존하는 유틸리티 객체
    """

    class Settings:
        protocol_type = "http"
        # db_driver = "mysql+mysqlconnector" # sqlalchemy mysql driver 설정 #  mysql-connector-python 라이브러리 에 의존
        db_driver = "mysql+pymysql"  # sqlalchemy 의 mysql driver 는 여러 개 built in 되어 있다. #  pymysql 라이브러리 에 의존
        # host = "doc"  # fail
        # host = "docker_image_maker-db-1"  # fail, 도커컴포즈 로 생성되면 컨테이너명이 docker_image_maker-db-1 로 자동생성된다고 했는데 잘안됬음.
        # host = "db-1"  # fail,
        # host = "www"  # fail # https://yes-admit.tistory.com/80
        # host = "127.0.0.1"  # success, 로컬에서 실행 시 사용
        # host = "localhost"  # success, 로컬에서 실행 시 사용
        # host = "host.docker.internal" # success, 도커컨테이너 형태로 app 과 db를 설정한 경우. app 에서 db url의 host를  localhost로 호출을 하면 호출 을 할 수 없다, 도커컨테이너 간 네트워크 설정(win/linux/mac) linux 는 도커컨테이너를 실행할 때 --add-host host.docker.internal:host-gateway 옵션을 의존, 불편하다.
        host = "mysql_container" # success,  도커네트워크 를 도커컨테이너 간에 설정하면, 도커DNS 에 의해 도커컨테이너 명으로 동적 맵핑이 이뤄진다, fastapi 도커컨테이너 에서는 해당 DB의 도커컨테이너 명으로 찾아갈 것이다. 로컬에서 dveaber 로는 도커컨테이너 명으로 연결 안되었음, localhost 로 는 가능했음, 도커빌드 시에 사용.
        port = 3306
        id = "root"
        pw = "admin123!"
        database_name = 'test_db'
        # database_url = rf'{protocol_type}://{host}:{port}' # database 에 바로 연결하지 않음, use DB명령어를 써야할 수있음
        uri = rf"{db_driver}://{id}:{pw}@{host}:{port}/{database_name}?charset=utf8"

    # Base 가 여기에서 설정되어야 동작
    engine = create_engine(Settings.uri)
    # engine = create_engine(Settings.uri, poolclass=QueuePool, pool_size=100)
    # engine = create_engine(Settings.uri, pool_timeout=60)
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    Base = declarative_base()

    @staticmethod
    def get_session():  # ?  generator 로 되어있는데..?
        db = MySqlUtil.SessionLocal()
        try:
            yield db
        finally:
            db.close()

    @staticmethod
    def get_session_local():
        db = MySqlUtil.SessionLocal()
        return db


    @staticmethod
    def execute(native_query: str):  # without sqlarchemy
        try:
            # engine = create_engine(Settings.url)
            connection = MySqlUtil.engine.connect()
            DebuggingUtil.print_ment_light_white(native_query)
            result = connection.execute(sqlalchecdmy_text(native_query))
            # connection.execute(text(native_query))  # 결과를 반환하지 않는 쿼리 실행
            # from sqlalchemy import text
            # result = connection.execute(text(native_query))
            # return result
            for word_reserved in ['truncate']:
                if word_reserved not in native_query.lower():
                    df = pd.DataFrame(result.fetchall(), columns=result.keys())
                    # Result를 DataFrame으로 변환
                    return df
                else:
                    DebuggingUtil.print_magenta(rf'''word_reserved : {word_reserved}''')
        except Exception:
            traceback.print_exc(file=sys.stdout)
        finally:
            connection.close()




class MemberUtil:
    """
    mysql / sqlalchemy / fastapi 의존하는 유틸리티 객체
    """

    class Member(MySqlUtil.Base):  # orm 설정에는 id 있음
        __tablename__ = "members"
        __table_args__ = {'extend_existing': True}
        # __table_args__ = {'extend_existing': True, 'mysql_collate': 'utf8_general_ci'} # encoding 안되면 비슷하게 방법을 알아보자  mysql 에 적용이 가능한 코드로 보인다.
        Member_id = Column(Integer, primary_key=True, autoincrement=True)
        id = Column(VARCHAR(length=30))
        name = Column(VARCHAR(length=30))
        e_mail = Column(VARCHAR(length=30))
        phone_no = Column(VARCHAR(length=13))
        address = Column(VARCHAR(length=255))
        birthday = Column(VARCHAR(length=50))
        pw = Column(VARCHAR(length=100))
        date_joined = Column(VARCHAR(length=50))
        date_canceled = Column(VARCHAR(length=50))

    class MemberBase(BaseModel):  # pydantic validator 설정에는 Member_id 없음
        name: str
        pw: str
        phone_no: str
        address: str
        e_mail: str
        age: str
        id: str
        date_joined: str
        date_canceled: str

        @staticmethod
        @field_validator('id')
        def validate_id(value):
            MemberUtil.validate_id(value)

        @staticmethod
        @field_validator('name')
        def validate_name(value):
            MemberUtil.validate_name(value)

        @staticmethod
        @field_validator('e_mail')
        def validate_e_mail(value):
            MemberUtil.validate_e_mail(value)

        @staticmethod
        @field_validator('phone_no')
        def validate_phone_no(value):
            MemberUtil.validate_phone_no(value)

        @staticmethod
        @field_validator('address')
        def validate_address(value):
            MemberUtil.validate_address(value)

        @staticmethod
        @field_validator('birthday')
        def validate_birthday(value):
            MemberUtil.validate_birthday(value)

        @staticmethod
        @field_validator('pw')
        def validate_pw(value):
            MemberUtil.validate_pw(value)

        @staticmethod
        @field_validator('date_joined')
        def validate_date_joined(value):
            MemberUtil.validate_date_joined(value)

        @staticmethod
        @field_validator('date_canceled')
        def validate_date_canceled(value):
            MemberUtil.validate_date_canceled(value)

        @staticmethod  # class 간 종속 관계가 있을 때 하위 class 에 붙여 줘야하나?, cls, 파라미터와 함께? , instance를 생성하지 않고 호출 가능해?
        @field_validator('date_join')
        def validate_date_join(value):
            MemberUtil.validate_date_join(value)
            # datetime.strptime(date_join, '%Y-%m-%d %H:%M %S%f')
            if len(value) != 18:
                raise HTTPException(status_code=400, detail="유효한 날짜가 아닙니다.")
            return value

        @staticmethod
        @field_validator('pw')
        def validate_pw(value):
            MemberUtil.validate_pw(value)
            if len(value) != MemberUtil.Member.__table__.c.pw.type.length:
                raise HTTPException(status_code=400, detail="유효한 이메일 주소가 아닙니다.")
            return value

    @staticmethod
    def get_member_validated(member):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")

        # Member 클래스의 필드 개수 확인
        field_count = len(MemberUtil.Member.__table__.c)
        print(rf'''field_count : {field_count}''')

        targets_validated = [
            {"field_en": 'name', "field_ko": "이름", "field_validation_func": MemberUtil.validate_name, "field_length_limit": MemberUtil.Member.__table__.c.name.type.length},
            {"field_en": 'id', "field_ko": "아이디", "field_validation_func": MemberUtil.validate_id, "field_length_limit": MemberUtil.Member.__table__.c.id.type.length},
            {"field_en": 'phone_no', "field_ko": "전화번호", "field_validation_func": MemberUtil.validate_phone_no, "field_length_limit": MemberUtil.Member.__table__.c.phone_no.type.length},
            {"field_en": 'e_mail', "field_ko": "이메일", "field_validation_func": MemberUtil.validate_e_mail, "field_length_limit": MemberUtil.Member.__table__.c.e_mail.type.length},
            {"field_en": 'pw', "field_ko": "비밀번호", "field_validation_func": MemberUtil.validate_pw, "field_length_limit": MemberUtil.Member.__table__.c.pw.type.length},
            {"field_en": 'address', "field_ko": "주소", "field_validation_func": MemberUtil.validate_address, "field_length_limit": MemberUtil.Member.__table__.c.address.type.length},
            {"field_en": 'birthday', "field_ko": "생년월일", "field_validation_func": MemberUtil.validate_birthday, "field_length_limit": MemberUtil.Member.__table__.c.birthday.type.length},
            {"field_en": 'date_joined', "field_ko": "가입일", "field_validation_func": MemberUtil.validate_date_joined, "field_length_limit": MemberUtil.Member.__table__.c.date_joined.type.length},
            {"field_en": 'date_canceled', "field_ko": "탈퇴일", "field_validation_func": MemberUtil.validate_date_canceled, "field_length_limit": MemberUtil.Member.__table__.c.date_canceled.type.length},
        ]
        for target in targets_validated:
            field_en = target["field_en"]
            field_ko = target["field_ko"]
            field_validation_func = target["field_validation_func"]
            field_length_limit = target["field_length_limit"]
            try:
                DebuggingUtil.print_magenta(rf'''member[field_en] : {member[field_en]}''')
                if len(member[field_en]) > target['field_length_limit']:
                    raise HTTPException(status_code=400, detail=f"{field_ko}({field_en})의 길이제한은 {field_length_limit}자 이하여야 합니다.")
                else:
                    field_validation_func(member[field_en])  # success, 호출할 수 없는 함수의 내부에 구현된 부분이 필요한것 이므로 내부에 구현된 것을 다른 클래스에 구현해서 참조하도록 로직 분리,
            except KeyError:
                DebuggingUtil.print_magenta(rf'''member[field_en] 에서 KeyError 발생했습니다, field_en={field_en}''')
                pass
        return member

    @staticmethod
    def validate_member(member):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        MemberUtil.get_member_validated(member)

    class MemberCreate(MemberBase):
        pass

    class MemberExtendedMemberBase(MemberBase):
        name: str
        pw: str
        phone_no: str
        address: str
        e_mail: str
        birthday: str
        id: str
        date_joined: str
        date_canceled: str

        class Config:
            # orm_mode = True
            from_attributes = True

    @staticmethod
    def get_members(db: Session):
        return db.query(MemberUtil.Member).all()

    @staticmethod
    def get_member(db: Session, id: int):
        # return db.query(MemberUtil.Member).filter(MemberUtil.Member.id == id).first() # success , 그러나 타입힌팅 에러가...
        MySqlUtil.execute(f'''SELECT * FROM members where id= {id} ORDER BY date_joined LIMIT 2;''')  # LIMIT 2 로 쿼리 성능 향상 기대, 2인 이유는 id가 2개면
        # 네이티브 쿼리를 한번 더 작성한 이유는 쿼리 디버깅
        return select(MemberUtil.Member).where(MemberUtil.Member.id.in_([id]))  # try

    @staticmethod
    def insert_member(db: Session, member):
        member_ = MemberUtil.Member(**member)
        db.add(member_)
        db.flush()  # flush() 메서드 없이 바로 commit() 메서드를 호출하면, 롤백할 수 있는 포인트가 만들어지지 않습니다. (# 나중에 롤백을 수행할 수 있는 포인트가 만들어짐)
        db.commit()
        db.refresh(member_)  # 데이터베이스에 업데이트된 최신내용을 세션에 가져오는 것.
        return member_

    @staticmethod
    def update_member(db: Session, member, updated_member):
        for key, value in updated_member.model_dump().members():
            setattr(member, key, value)
        db.commit()
        db.refresh(member)
        return member

    @staticmethod
    def delete_member(db: Session, member):
        db.delete(member)
        db.commit()

    @staticmethod
    def is_member_joined_by_id(id, request):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")

        result = MySqlUtil.get_session_local().query(MemberUtil.Member).filter(MemberUtil.Member.id == id).limit(2)
        for member in result:
            print(f"member.name: {member.name}, member.id: {member.id}")
        member_count = result.count()
        print(rf'''member_count : {member_count}''')
        if member_count == 1:
            for member_joined in result:
                print(f'member_joined.name: {member_joined.name}  member_joined.id: {member_joined.id}')
                request.session['name'] = member_joined.name
            return True
        else:
            return False

    @staticmethod
    def is_member_joined(id, pw, request):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")

        # native query
        # sql injection 에 취약,
        # 위험해서, 로그인 로직에 그냥은 못 쓴다.
        # rows = MySqlUtil.execute(f'''SELECT count(*) FROM members where id="{id}" and pw="{pw}" ORDER BY date_joined LIMIT 2;''')
        # id_count = rows.fetchone()[0]
        # print(rf'id_count : {id_count}')
        # if id_count == 1:
        #     return True
        # else:
        #     return False

        # orm
        # sql injection 에 강화됨.
        result = MySqlUtil.get_session_local().query(MemberUtil.Member).filter(MemberUtil.Member.id == id, MemberUtil.Member.pw == pw).limit(2)
        for member in result:
            print(f"member.name: {member.name}, member.id: {member.id}, member.pw: {member.pw}")
        member_count = result.count()
        print(rf'''member_count : {member_count}''')
        if member_count == 1:
            return True
        else:
            return False

    @staticmethod
    def get_member_name_joined(id, pw, request):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        result = MySqlUtil.get_session_local().query(MemberUtil.Member).filter(MemberUtil.Member.id == id, MemberUtil.Member.pw == pw).limit(2)
        member_count = result.count()
        print(rf'''member_count : {member_count}''')
        if member_count == 1:
            for member in result:
                return member.name

    @staticmethod
    def validate_id(value):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        BusinessLogicUtil.raise_exception_after_special_charcater_check(value, inspect.currentframe().f_code.co_name)
        return True

    @staticmethod
    def validate_name(value):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        BusinessLogicUtil.raise_exception_after_special_charcater_check(value, inspect.currentframe().f_code.co_name)
        return True

    @staticmethod
    def validate_e_mail(value):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        BusinessLogicUtil.raise_exception_after_special_charcater_check(value, inspect.currentframe().f_code.co_name, ignore_list=["@"])
        MemberUtil.validate_address_e_mail(value)
        pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
        if not re.match(pattern, value):
            # if not address_e_mail.endswith('@kakao.com'):
            #     raise HTTPException(status_code=400, detail="유효한 카카오 이메일이 아닙니다.")
            # return address_e_mail
            raise HTTPException(status_code=400, detail=f"유효한 이메일 주소가 아닙니다. {value}")
        return value

    @staticmethod
    def validate_phone_no(value):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        # r'^\d{3}-\d{3,4}-\d{4}$'
        # r'^\d{2}-\d{3,4}-\d{4}$' 둘다
        if not re.match(r'^\d{2,3}-\d{3,4}-\d{4}$', value):
            raise HTTPException(status_code=400, detail=f"유효한 전화번호가 아닙니다. {value}")
        return value

    @staticmethod
    def validate_address(value):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")

        return True

    @staticmethod
    def validate_birthday(value):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")

        return True

    @staticmethod
    def validate_pw(value):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")

        return True

    @staticmethod
    def validate_date_joined(value):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")

        return True

    @staticmethod
    def validate_date_canceled(value):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")

        return True

    @staticmethod
    def validate_date_join(value):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")

        return True

    @staticmethod
    def validate_address_e_mail(value):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")

        return True


class CommutationManagementUtil:
    """
    mysql / sqlalchemy / fastapi 의존하는 유틸리티 객체
    """

    class CommutationManagement(MySqlUtil.Base):
        __tablename__ = "commutation_management"
        __table_args__ = {'extend_existing': True}
        commutation_management_id = Column(Integer, primary_key=True, autoincrement=True)
        id = Column(VARCHAR(length=30))
        name = Column(VARCHAR(length=30))
        phone_no = Column(VARCHAR(length=13))
        time_to_go_to_office = Column(VARCHAR(length=100))
        time_to_leave_office = Column(VARCHAR(length=100))

    class CommutationManagementBase(BaseModel):
        id: str
        name: str
        phone_no: str
        time_to_go_to_office: str
        time_to_leave_office: str

        # @staticmethod
        # @field_validator('id')
        # def validate_id(value):
        #     CommutationManagementUtil.validate_id(value)

    @staticmethod
    def get_commutation_management_validated(commutation_management):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")

        # CommutationManagement 클래스의 필드 개수 확인
        field_count = len(CommutationManagementUtil.CommutationManagement.__table__.c)
        print(rf'''field_count : {field_count}''')

        targets_validated = [
            {"field_en": 'id', "field_ko": "아이디", "field_validation_func": CommutationManagementUtil.validate_id, "field_length_limit": CommutationManagementUtil.CommutationManagement.__table__.c.id.type.length},
            {"field_en": 'name', "field_ko": "이름", "field_validation_func": CommutationManagementUtil.validate_name, "field_length_limit": CommutationManagementUtil.CommutationManagement.__table__.c.name.type.length},
            {"field_en": 'phone_no', "field_ko": "전화번호", "field_validation_func": CommutationManagementUtil.validate_phone_no, "field_length_limit": CommutationManagementUtil.CommutationManagement.__table__.c.phone_no.type.length},
            {"field_en": 'time_to_go_to_office', "field_ko": "출근시간", "field_validation_func": CommutationManagementUtil.validate_time_to_go_to_office, "field_length_limit": CommutationManagementUtil.CommutationManagement.__table__.c.time_to_go_to_office.type.length},
            {"field_en": 'time_to_go_to_office', "field_ko": "퇴근시간", "field_validation_func": CommutationManagementUtil.validate_time_to_leave_office, "field_length_limit": CommutationManagementUtil.CommutationManagement.__table__.c.time_to_leave_office.type.length},
        ]
        for target in targets_validated:
            field_en = target["field_en"]
            field_ko = target["field_ko"]
            field_validation_func = target["field_validation_func"]
            field_length_limit = target["field_length_limit"]
            if len(commutation_management[field_en]) > target['field_length_limit']:
                raise HTTPException(status_code=400, detail=f"{field_ko}({field_en})의 길이제한은 {field_length_limit}자 이하여야 합니다.")
            else:
                field_validation_func(commutation_management[field_en])  # success, 호출할 수 없는 함수의 내부에 구현된 부분이 필요한것 이므로 내부에 구현된 것을 다른 클래스에 구현해서 참조하도록 로직 분리,
        return commutation_management

    @staticmethod
    def validate_commutation_management(commutation_management):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        CommutationManagementUtil.get_commutation_management_validated(commutation_management)

    class CommutationManagementCreate(CommutationManagementBase):
        pass

    class CommutationManagementExtendedCommutationManagementBase(CommutationManagementBase):
        id: str
        name: str
        phone_no: str
        time_to_go_to_office: str
        time_to_leave_office: str

        class Config:
            # orm_mode = True
            from_attributes = True

    @staticmethod
    def get_commutation_managements(db: Session):
        return db.query(CommutationManagementUtil.CommutationManagement).all()

    @staticmethod
    def get_commutation_management(db: Session, id: int):
        # return db.query(CommutationManagementUtil.CommutationManagement).filter(CommutationManagementUtil.CommutationManagement.id == id).first() # success , 그러나 타입힌팅 에러가...
        return select(CommutationManagementUtil.CommutationManagement).where(CommutationManagementUtil.CommutationManagement.id.in_([id]))  # try

    @staticmethod
    def insert_commutation_management(db: Session, commutation_management):
        commutation_management_ = CommutationManagementUtil.CommutationManagement(**commutation_management)
        db.add(commutation_management_)
        db.flush()
        db.commit()
        db.refresh(commutation_management_)
        return commutation_management_

    @staticmethod
    def update_commutation_management(db: Session, commutation_management, updated_commutation_management):
        for key, value in updated_commutation_management.model_dump().commutation_managements():
            setattr(commutation_management, key, value)
        db.commit()
        db.refresh(commutation_management)
        return commutation_management

    @staticmethod
    def validate_id(value):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        BusinessLogicUtil.raise_exception_after_special_charcater_check(value, inspect.currentframe().f_code.co_name)
        return True

    @staticmethod
    def validate_name(value):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        BusinessLogicUtil.raise_exception_after_special_charcater_check(value, inspect.currentframe().f_code.co_name)
        return True

    @staticmethod
    def validate_phone_no(value):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        # r'^\d{3}-\d{3,4}-\d{4}$'
        # r'^\d{2}-\d{3,4}-\d{4}$' 둘다
        if not re.match(r'^\d{2,3}-\d{3,4}-\d{4}$', value):
            raise HTTPException(status_code=400, detail=f"유효한 전화번호가 아닙니다. {value}")
        return value

    @staticmethod
    def validate_time_to_go_to_office(value):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        return True

    @staticmethod
    def validate_time_to_leave_office(value):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        return True


class ItemsUtil:
    class Item(MySqlUtil.Base):  # orm 설정에는 id 있음
        __tablename__ = "items"
        # __table_args__ = {'extend_existing': True}  # 이옵션을 쓰면 여기에 작성된 item 대로 테이블이 다르면 새로확장이되는 거란다.
        id = Column(Integer, primary_key=True, autoincrement=True)
        name = Column(String(30))
        # price = Column(Integer)

    class ItemBase(BaseModel):  # pydantic validator 설정에는 id 없음
        name: str

    class ItemCreate(ItemBase):
        pass

    class ItemExtendedItemBase(ItemBase):
        id: int

        class Config:
            # orm_mode = True
            from_attributes = True

    @staticmethod
    def get_items(db: Session):
        return db.query(ItemsUtil.Item).all()

    @staticmethod
    def get_item(db: Session, id: int):
        # return db.query(ItemsUtil.Item).filter(column("id") == id).first()
        # return db.query(ItemsUtil.Item).filter(ItemsUtil.Item.id == id).first() , success , 그러나 타입힌팅 에러가...
        # return select(ItemsUtil.Item).where(ItemsUtil.Item.id.in_(["id1","id2"]))
        return select(ItemsUtil.Item).where(ItemsUtil.Item.id.in_([id]))

    @staticmethod
    def create_item(db: Session, item):
        db_item = ItemsUtil.Item(**item.model_dump())
        db.add(db_item)
        db.commit()
        db.refresh(db_item)
        return db_item

    @staticmethod
    def update_item(db: Session, item, updated_item):
        for key, value in updated_item.model_dump().items():
            setattr(item, key, value)
        db.commit()
        db.refresh(item)
        return item

    @staticmethod
    def delete_item(db: Session, item):
        db.delete(item)
        db.commit()


class FaqBoardUtil:
    """
    mysql / sqlalchemy / fastapi 의존하는 유틸리티 객체
    """

    class FaqBoard(MySqlUtil.Base):  # orm 설정에는 id 있음
        __tablename__ = "faq_boards"
        __table_args__ = {'extend_existing': True}
        # __table_args__ = {'extend_existing': True, 'mysql_collate': 'utf8_general_ci'} # encoding 안되면 비슷하게 방법을 알아보자  mysql 에 적용이 가능한 코드로 보인다.
        FaqBoard_id: str = uuid4().hex + TimeUtil.get_time_as_('%Y%m%d%H%M%S%f') + SecurityUtil.get_random_alphabet()  # table 내 unique id
        id = Column(Integer, primary_key=True, autoincrement=True)  # index 로 이용
        writer = Column(VARCHAR(length=30))
        title = Column(VARCHAR(length=30))
        contents = Column(VARCHAR(length=13))
        date_reg = Column(DateTime, nullable=False, default=datetime.now)
        del_yn = Column(VARCHAR(length=50))

    class FaqBoardBase(BaseModel):  # pydantic validator 설정에는 FaqBoard_id 없음
        id: str
        writer: str
        title: str
        contents: Optional[str]
        date_reg: str
        del_yn: str

        @staticmethod
        @field_validator('id')
        def validate_id(value):
            FaqBoardUtil.validate_id(value)

        @staticmethod
        @field_validator('writer')
        def validate_writer(value):
            FaqBoardUtil.validate_writer(value)

        @staticmethod
        @field_validator('title')
        def validate_title(value):
            FaqBoardUtil.validate_title(value)

        @staticmethod
        @field_validator('contents')
        def validate_contents(value):
            FaqBoardUtil.validate_contents(value)

        @staticmethod
        @field_validator('date_reg')
        def validate_date_reg(value):
            # datetime.strptime(date_join, '%Y-%m-%d %H:%M %S%f')
            if len(value) != 18:
                raise HTTPException(status_code=400, detail="유효한 날짜가 아닙니다.")
            FaqBoardUtil.validate_date_reg(value)

        @staticmethod
        @field_validator('del_yn')
        def validate_del_yn(value):
            FaqBoardUtil.validate_del_yn(value)

    @staticmethod
    def get_faq_board_validated(faq_board):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")

        # FaqBoard 클래스의 필드 개수 확인
        field_count = len(FaqBoardUtil.FaqBoard.__table__.c)
        print(rf'''field_count : {field_count}''')

        targets_validated = [
            {"field_en": 'id', "field_ko": "아이디", "field_validation_func": FaqBoardUtil.validate_id, "field_length_limit": FaqBoardUtil.FaqBoard.__table__.c.id.type.length},
            {"field_en": 'writer', "field_ko": "이름", "field_validation_func": FaqBoardUtil.validate_writer, "field_length_limit": FaqBoardUtil.FaqBoard.__table__.c.writer.type.length},
            {"field_en": 'title', "field_ko": "이메일", "field_validation_func": FaqBoardUtil.validate_title, "field_length_limit": FaqBoardUtil.FaqBoard.__table__.c.title.type.length},
            {"field_en": 'contents', "field_ko": "전화번호", "field_validation_func": FaqBoardUtil.validate_contents, "field_length_limit": FaqBoardUtil.FaqBoard.__table__.c.contents.type.length},
            {"field_en": 'date_reg', "field_ko": "주소", "field_validation_func": FaqBoardUtil.validate_date_reg, "field_length_limit": FaqBoardUtil.FaqBoard.__table__.c.date_reg.type.length},
            {"field_en": 'del_yn', "field_ko": "가입일", "field_validation_func": FaqBoardUtil.validate_del_yn, "field_length_limit": FaqBoardUtil.FaqBoard.__table__.c.del_yn.type.length},
        ]
        for target in targets_validated:
            field_en = target["field_en"]
            field_ko = target["field_ko"]
            field_validation_func = target["field_validation_func"]
            field_length_limit = target["field_length_limit"]
            if len(faq_board[field_en]) > target['field_length_limit']:
                raise HTTPException(status_code=400, detail=f"{field_ko}({field_en})의 길이제한은 {field_length_limit}자 이하여야 합니다.")
            else:
                field_validation_func(faq_board[field_en])  # success, 호출할 수 없는 함수의 내부에 구현된 부분이 필요한것 이므로 내부에 구현된 것을 다른 클래스에 구현해서 참조하도록 로직 분리,
        return faq_board

    @staticmethod
    def validate_faq_board(faq_board):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        FaqBoardUtil.get_faq_board_validated(faq_board)

    class FaqBoardCreate(FaqBoardBase):
        pass

    class FaqBoardExtendedFaqBoardBase(FaqBoardBase):
        id: str
        writer: str
        title: str
        contents: str
        date_reg: str
        del_yn: str

        class Config:
            from_attributes = True

    @staticmethod
    def get_faq_boards(db: Session):
        return db.query(FaqBoardUtil.FaqBoard).all()

    @staticmethod
    def get_faq_board(db: Session, id: int):
        # return db.query(FaqBoardUtil.FaqBoard).filter(FaqBoardUtil.FaqBoard.id == id).first() # success , 그러나 타입힌팅 에러가...
        MySqlUtil.execute(f'''SELECT * FROM faq_boards where id= {id} ORDER BY del_yn LIMIT 2;''')  # LIMIT 2 로 쿼리 성능 향상 기대, 2인 이유는 id가 2개면
        # 네이티브 쿼리를 한번 더 작성한 이유는 쿼리 디버깅
        return select(FaqBoardUtil.FaqBoard).where(FaqBoardUtil.FaqBoard.id.in_([id]))  # try

    @staticmethod
    def insert_faq_board(db: Session, faq_board):
        faq_board_ = FaqBoardUtil.FaqBoard(**faq_board)
        db.add(faq_board_)
        db.flush()  # flush() 메서드 없이 바로 commit() 메서드를 호출하면, 롤백할 수 있는 포인트가 만들어지지 않습니다. (# 나중에 롤백을 수행할 수 있는 포인트가 만들어짐)
        db.commit()
        db.refresh(faq_board_)  # 데이터베이스에 업데이트된 최신내용을 세션에 가져오는 것.
        return faq_board_

    @staticmethod
    def update_faq_board(db: Session, faq_board, updated_faq_board):
        for key, value in updated_faq_board.model_dump().faq_boards():
            setattr(faq_board, key, value)
        db.commit()
        db.refresh(faq_board)
        return faq_board

    @staticmethod
    def delete_faq_board(db: Session, faq_board):
        db.delete(faq_board)
        db.commit()

    @staticmethod
    def is_faq_board_joined_by_id(id, request):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")

        result = MySqlUtil.get_session_local().query(FaqBoardUtil.FaqBoard).filter(FaqBoardUtil.FaqBoard.id == id).limit(2)
        for faq_board in result:
            print(f"faq_board.name: {faq_board.name}, faq_board.id: {faq_board.id}")
        faq_board_count = result.count()
        print(rf'''faq_board_count : {faq_board_count}''')
        if faq_board_count == 1:
            for faq_board_joined in result:
                print(f'faq_board_joined.name: {faq_board_joined.name}  faq_board_joined.id: {faq_board_joined.id}')
                request.session['name'] = faq_board_joined.name
            return True
        else:
            return False

    @staticmethod
    def validate_id(value):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        BusinessLogicUtil.raise_exception_after_special_charcater_check(value, inspect.currentframe().f_code.co_name)
        return True

    @staticmethod
    def validate_writer(value):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        BusinessLogicUtil.raise_exception_after_special_charcater_check(value, inspect.currentframe().f_code.co_name)
        return True

    @staticmethod
    def validate_title(value):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        BusinessLogicUtil.raise_exception_after_special_charcater_check(value, inspect.currentframe().f_code.co_name, ignore_list=["@"])
        FaqBoardUtil.validate_title(value)
        pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
        if not re.match(pattern, value):
            # if not date_reg_title.endswith('@kakao.com'):
            #     raise HTTPException(status_code=400, detail="유효한 카카오 이메일이 아닙니다.")
            # return date_reg_title
            raise HTTPException(status_code=400, detail=f"유효한 이메일 주소가 아닙니다. {value}")
        return value

    @staticmethod
    def validate_contents(value):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        # r'^\d{3}-\d{3,4}-\d{4}$'
        # r'^\d{2}-\d{3,4}-\d{4}$' 둘다
        if not re.match(r'^\d{2,3}-\d{3,4}-\d{4}$', value):
            raise HTTPException(status_code=400, detail=f"유효한 전화번호가 아닙니다. {value}")
        return value

    @staticmethod
    def validate_date_reg(value):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        return True

    @staticmethod
    def validate_del_yn(value):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        return True


class JwtUtil:
    """
    mysql / sqlalchemy / fastapi 의존하는 유틸리티 객체
    """

    class JwtModel(MySqlUtil.Base):  # orm 설정에는 Jwt_id 있음
        __tablename__ = "jwts"
        __table_args__ = {'extend_existing': True}
        # __table_args__ = {'extend_existing': True, 'mysql_collate': 'utf8_general_ci'} # encoding 안되면 비슷하게 방법을 알아보자  mysql 에 적용이 가능한 코드로 보인다.
        jwt_id: UUID = uuid4().hex + TimeUtil.get_time_as_('%Y%m%d%H%M%S%f') + SecurityUtil.get_random_alphabet()  # table 내 unique id
        index = Column(Integer, primary_key=True, autoincrement=True)  # index 로 이용
        # token = Column(VARCHAR(length=149))
        # token = Column(VARCHAR(length=152))
        token = Column(VARCHAR(length=255))

    class JwtBase(BaseModel):  # pydantic validator 설정에는 Jwt_id 없음
        index: str
        token: str

        @staticmethod
        @field_validator('index')
        def validate_index(value):
            JwtUtil.validate_index(value)

        @staticmethod
        @field_validator('token')
        def validate_token(value):
            JwtUtil.validate_token(value)

    @staticmethod
    def get_jwt_validated(jwt):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")

        # Jwt 클래스의 필드 개수 확인
        field_count = len(JwtUtil.JwtModel.__table__.c)
        print(rf'''field_count : {field_count}''')

        targets_validated = [
            {"field_en": 'index', "field_ko": "아이디", "field_validation_func": JwtUtil.validate_index, "field_length_limit": JwtUtil.JwtModel.__table__.c.index.type.length},
            {"field_en": 'token', "field_ko": "토큰", "field_validation_func": JwtUtil.validate_token, "field_length_limit": JwtUtil.JwtModel.__table__.c.token.type.length},
        ]
        for target in targets_validated:
            field_en = target["field_en"]
            field_ko = target["field_ko"]
            field_validation_func = target["field_validation_func"]
            field_length_limit = target["field_length_limit"]
            if len(jwt[field_en]) > target['field_length_limit']:
                raise HTTPException(status_code=400, detail=f"{field_ko}({field_en})의 길이제한은 {field_length_limit}자 이하여야 합니다.")
            else:
                field_validation_func(jwt[field_en])  # success, 호출할 수 없는 함수의 내부에 구현된 부분이 필요한것 이므로 내부에 구현된 것을 다른 클래스에 구현해서 참조하도록 로직 분리,
        return jwt

    @staticmethod
    def validate_jwt(jwt):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        JwtUtil.get_jwt_validated(jwt)

    class JwtCreate(JwtBase):
        pass

    class JwtExtendedJwtBase(JwtBase):
        index: str
        token: str

        class Config:
            from_attributes = True

    @staticmethod
    def get_jwts(db: Session):
        return db.query(JwtUtil.JwtModel).all()

    @staticmethod
    def get_jwts_by_token(db: Session, token: str):
        # 쿼리검증 테스트
        # result =  MySqlUtil.execute(f'''SELECT * FROM jwts where token="{token}" LIMIT 2;''')
        # rs = result.fetchall()

        # rs = db.query(JwtUtil.JwtModel).filter_by(JwtUtil.JwtModel.token == token).all()

        # result = db.excute(select(JwtUtil.JwtModel).where(JwtUtil.JwtModel.token.in_([token])))
        # rs = result.fetchall()

        # rs = db.query(JwtUtil.JwtModel).filter(JwtUtil.JwtModel.token == token).all()# success
        rs = db.query(JwtUtil.JwtModel).filter(JwtUtil.JwtModel.token == token).limit(2).all()  # try
        return rs

    @staticmethod
    def insert_jwt_encoded(db: Session, jwt_data):
        jwt_ = JwtUtil.JwtModel(**jwt_data)
        db.add(jwt_)
        db.flush()  # flush() 메서드 없이 바로 commit() 메서드를 호출하면, 롤백할 수 있는 포인트가 만들어지지 않습니다. (# 나중에 롤백을 수행할 수 있는 포인트가 만들어짐)
        db.commit()
        db.refresh(jwt_)  # 데이터베이스에 업데이트된 최신내용을 세션에 가져오는 것.
        return jwt_

    @staticmethod
    def update_jwt(db: Session, jwt, updated_jwt):
        for key, value in updated_jwt.model_dump().jwts():
            setattr(jwt, key, value)
        db.commit()
        db.refresh(jwt)
        return jwt

    @staticmethod
    def delete_jwt(db: Session, jwt):
        db.delete(jwt)
        db.commit()

    @staticmethod
    def is_jwt_joined_by_index(index, request):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")

        result = MySqlUtil.get_session_local().query(JwtUtil.JwtModel).filter(JwtUtil.JwtModel.index == index).limit(2)
        for jwt in result:
            print(f"jwt.name: {jwt.name}, jwt.index: {jwt.index}")
        jwt_count = result.count()
        print(rf'''jwt_count : {jwt_count}''')
        if jwt_count == 1:
            for jwt_joined in result:
                print(f'jwt_joined.name: {jwt_joined.name}  jwt_joined.index: {jwt_joined.index}')
                request.session['name'] = jwt_joined.name
            return True
        else:
            return False

    @staticmethod
    def validate_index(value):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        BusinessLogicUtil.raise_exception_after_special_charcater_check(value, inspect.currentframe().f_code.co_name)
        return True

    @staticmethod
    def validate_token(value):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        BusinessLogicUtil.raise_exception_after_special_charcater_check(value, inspect.currentframe().f_code.co_name, ignore_list=["@"])

        # value 를 처리해서
        # raise HTTPException(status_code=400, detail=f"유효한 토큰 이 아닙니다. {value}")
        return value


class CustomerServiceBoardUtil:
    """
    mysql / sqlalchemy / fastapi 의존하는 유틸리티 객체
    """

    class CustomerServiceBoard(MySqlUtil.Base):  # orm 설정에는 id 있음
        __tablename__ = "customer_service_board"
        __table_args__ = {'extend_existing': True}
        # __table_args__ = {'extend_existing': True, 'mysql_collate': 'utf8_general_ci'} # encoding 안되면 비슷하게 방법을 알아보자  mysql 에 적용이 가능한 코드로 보인다.
        CustomerServiceBoard_id: str = uuid4().hex + TimeUtil.get_time_as_('%Y%m%d%H%M%S%f') + SecurityUtil.get_random_alphabet()  # table 내 unique id
        id = Column(Integer, primary_key=True, autoincrement=True)  # index 로 이용
        writer = Column(VARCHAR(length=30))
        title = Column(VARCHAR(length=30))
        contents = Column(VARCHAR(length=13))
        date_reg = Column(DateTime, nullable=False, default=datetime.now)
        del_yn = Column(VARCHAR(length=50))

    class CustomerServiceBoardBase(BaseModel):  # pydantic validator 설정에는 CustomerServiceBoard_id 없음
        id: str
        writer: str
        title: str
        contents: Optional[str]
        date_reg: str
        del_yn: str

        @staticmethod
        @field_validator('id')
        def validate_id(value):
            CustomerServiceBoardUtil.validate_id(value)

        @staticmethod
        @field_validator('writer')
        def validate_writer(value):
            CustomerServiceBoardUtil.validate_writer(value)

        @staticmethod
        @field_validator('title')
        def validate_title(value):
            CustomerServiceBoardUtil.validate_title(value)

        @staticmethod
        @field_validator('contents')
        def validate_contents(value):
            CustomerServiceBoardUtil.validate_contents(value)

        @staticmethod
        @field_validator('date_reg')
        def validate_date_reg(value):
            # datetime.strptime(date_join, '%Y-%m-%d %H:%M %S%f')
            if len(value) != 18:
                raise HTTPException(status_code=400, detail="유효한 날짜가 아닙니다.")
            CustomerServiceBoardUtil.validate_date_reg(value)

        @staticmethod
        @field_validator('del_yn')
        def validate_del_yn(value):
            CustomerServiceBoardUtil.validate_del_yn(value)

    @staticmethod
    def get_customer_service_board_validated(customer_service_board):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")

        # CustomerServiceBoard 클래스의 필드 개수 확인
        field_count = len(CustomerServiceBoardUtil.CustomerServiceBoard.__table__.c)
        print(rf'''field_count : {field_count}''')

        targets_validated = [
            {"field_en": 'id', "field_ko": "아이디", "field_validation_func": CustomerServiceBoardUtil.validate_id, "field_length_limit": CustomerServiceBoardUtil.CustomerServiceBoard.__table__.c.id.type.length},
            {"field_en": 'writer', "field_ko": "이름", "field_validation_func": CustomerServiceBoardUtil.validate_writer, "field_length_limit": CustomerServiceBoardUtil.CustomerServiceBoard.__table__.c.writer.type.length},
            {"field_en": 'title', "field_ko": "이메일", "field_validation_func": CustomerServiceBoardUtil.validate_title, "field_length_limit": CustomerServiceBoardUtil.CustomerServiceBoard.__table__.c.title.type.length},
            {"field_en": 'contents', "field_ko": "전화번호", "field_validation_func": CustomerServiceBoardUtil.validate_contents, "field_length_limit": CustomerServiceBoardUtil.CustomerServiceBoard.__table__.c.contents.type.length},
            {"field_en": 'date_reg', "field_ko": "주소", "field_validation_func": CustomerServiceBoardUtil.validate_date_reg, "field_length_limit": CustomerServiceBoardUtil.CustomerServiceBoard.__table__.c.date_reg.type.length},
            {"field_en": 'del_yn', "field_ko": "가입일", "field_validation_func": CustomerServiceBoardUtil.validate_del_yn, "field_length_limit": CustomerServiceBoardUtil.CustomerServiceBoard.__table__.c.del_yn.type.length},
        ]
        for target in targets_validated:
            field_en = target["field_en"]
            field_ko = target["field_ko"]
            field_validation_func = target["field_validation_func"]
            field_length_limit = target["field_length_limit"]
            if len(customer_service_board[field_en]) > target['field_length_limit']:
                raise HTTPException(status_code=400, detail=f"{field_ko}({field_en})의 길이제한은 {field_length_limit}자 이하여야 합니다.")
            else:
                field_validation_func(customer_service_board[field_en])  # success, 호출할 수 없는 함수의 내부에 구현된 부분이 필요한것 이므로 내부에 구현된 것을 다른 클래스에 구현해서 참조하도록 로직 분리,
        return customer_service_board

    @staticmethod
    def validate_customer_service_board(customer_service_board):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        CustomerServiceBoardUtil.get_customer_service_board_validated(customer_service_board)

    class CustomerServiceBoardCreate(CustomerServiceBoardBase):
        pass

    class CustomerServiceBoardExtendedCustomerServiceBoardBase(CustomerServiceBoardBase):
        id: str
        writer: str
        title: str
        contents: str
        date_reg: str
        del_yn: str

        class Config:
            from_attributes = True

    @staticmethod
    def get_customer_service_boards(db: Session):
        return db.query(CustomerServiceBoardUtil.CustomerServiceBoard).all()

    @staticmethod
    def get_customer_service_board(id: int):
        # return db.query(CustomerServiceBoardUtil.CustomerServiceBoard).filter(CustomerServiceBoardUtil.CustomerServiceBoard.id == id).first() # success , 그러나 타입힌팅 에러가...
        MySqlUtil.execute(f'''SELECT * FROM customer_service_board where id= {id} ORDER BY del_yn LIMIT 2;''')  # LIMIT 2 로 쿼리 성능 향상 기대, 2인 이유는 id가 2개면
        # 네이티브 쿼리를 한번 더 작성한 이유는 쿼리 디버깅
        db_data = MySqlUtil.get_session_local().query(CustomerServiceBoardUtil.CustomerServiceBoard).filter_by(id=id).limit(4).all()
        return db_data  # try

    @staticmethod
    def insert_customer_service_board(db: Session, customer_service_board):
        customer_service_board_ = CustomerServiceBoardUtil.CustomerServiceBoard(**customer_service_board)
        db.add(customer_service_board_)
        db.flush()  # flush() 메서드 없이 바로 commit() 메서드를 호출하면, 롤백할 수 있는 포인트가 만들어지지 않습니다. (# 나중에 롤백을 수행할 수 있는 포인트가 만들어짐)
        db.commit()
        db.refresh(customer_service_board_)  # 데이터베이스에 업데이트된 최신내용을 세션에 가져오는 것.
        return customer_service_board_

    @staticmethod
    def update_customer_service_board(db: Session, customer_service_board, updated_customer_service_board):
        for key, value in updated_customer_service_board.model_dump().customer_service_boards():
            setattr(customer_service_board, key, value)
        db.commit()
        db.refresh(customer_service_board)
        return customer_service_board

    @staticmethod
    def delete_customer_service_board(db: Session, customer_service_board):
        db.delete(customer_service_board)
        db.commit()

    @staticmethod
    def is_customer_service_board_joined_by_id(id, request):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")

        result = MySqlUtil.get_session_local().query(CustomerServiceBoardUtil.CustomerServiceBoard).filter(CustomerServiceBoardUtil.CustomerServiceBoard.id == id).limit(2)
        for customer_service_board in result:
            print(f"customer_service_board.name: {customer_service_board.name}, customer_service_board.id: {customer_service_board.id}")
        customer_service_board_count = result.count()
        print(rf'''customer_service_board_count : {customer_service_board_count}''')
        if customer_service_board_count == 1:
            for customer_service_board_joined in result:
                print(f'customer_service_board_joined.name: {customer_service_board_joined.name}  customer_service_board_joined.id: {customer_service_board_joined.id}')
                request.session['name'] = customer_service_board_joined.name
            return True
        else:
            return False

    @staticmethod
    def validate_id(value):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        BusinessLogicUtil.raise_exception_after_special_charcater_check(value, inspect.currentframe().f_code.co_name)
        return True

    @staticmethod
    def validate_writer(value):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        BusinessLogicUtil.raise_exception_after_special_charcater_check(value, inspect.currentframe().f_code.co_name)
        return True

    @staticmethod
    def validate_title(value):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        BusinessLogicUtil.raise_exception_after_special_charcater_check(value, inspect.currentframe().f_code.co_name, ignore_list=["@"])
        CustomerServiceBoardUtil.validate_title(value)
        pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
        if not re.match(pattern, value):
            # if not date_reg_title.endswith('@kakao.com'):
            #     raise HTTPException(status_code=400, detail="유효한 카카오 이메일이 아닙니다.")
            # return date_reg_title
            raise HTTPException(status_code=400, detail=f"유효한 이메일 주소가 아닙니다. {value}")
        return value

    @staticmethod
    def validate_contents(value):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        # r'^\d{3}-\d{3,4}-\d{4}$'
        # r'^\d{2}-\d{3,4}-\d{4}$' 둘다
        if not re.match(r'^\d{2,3}-\d{3,4}-\d{4}$', value):
            raise HTTPException(status_code=400, detail=f"유효한 전화번호가 아닙니다. {value}")
        return value

    @staticmethod
    def validate_date_reg(value):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        return True

    @staticmethod
    def validate_del_yn(value):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        return True


class FinanceStockTickerUtil:
    """
    mysql / sqlalchemy / fastapi 의존하는 유틸리티 객체
    """

    class FinanceStockTicker(MySqlUtil.Base):  # orm 설정에는 id 있음
        __tablename__ = "finance_stock_ticker"
        __table_args__ = {'extend_existing': True}
        # __table_args__ = {'extend_existing': True, 'mysql_collate': 'utf8_general_ci'} # encoding 안되면 비슷하게 방법을 알아보자  mysql 에 적용이 가능한 코드로 보인다.
        FinanceStockTicker_id: str = uuid4().hex + TimeUtil.get_time_as_('%Y%m%d%H%M%S%f') + SecurityUtil.get_random_alphabet()  # table 내 unique id
        id = Column(Integer, primary_key=True, autoincrement=True)  # index 로 이용
        ticker = Column(VARCHAR(length=6))
        stock_name = Column(VARCHAR(length=100))
        market_name = Column(VARCHAR(length=13))
        date_reg = Column(DateTime, nullable=False, default=datetime.now)
        date_del = Column(DateTime, nullable=True)

    class FinanceStockTickerBase(BaseModel):  # pydantic validator 설정에는 FinanceStockTicker_id 없음
        id: str
        ticker: str
        stock_name: str
        market_name: Optional[str]
        date_reg: str
        date_del: str

        @staticmethod
        @field_validator('id')
        def validate_id(value):
            FinanceStockTickerUtil.validate_id(value)

        @staticmethod
        @field_validator('ticker')
        def validate_ticker(value):
            FinanceStockTickerUtil.ticker(value)

        @staticmethod
        @field_validator('stock_name')
        def validate_stock_name(value):
            FinanceStockTickerUtil.validate_stock_name(value)

        @staticmethod
        @field_validator('market_name')
        def validate_market_name(value):
            FinanceStockTickerUtil.validate_market_name(value)

        @staticmethod
        @field_validator('date_reg')
        def validate_date_reg(value):
            # datetime.strptime(date_join, '%Y-%m-%d %H:%M %S%f')
            if len(value) != 18:
                raise HTTPException(status_code=400, detail="유효한 날짜가 아닙니다.")
            FinanceStockTickerUtil.validate_date_reg(value)

        @staticmethod
        @field_validator('date_del')
        def validate_date_del(value):
            FinanceStockTickerUtil.validate_date_del(value)

    @staticmethod
    def get_finance_stock_ticker_validated(finance_stock_ticker):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")

        # FinanceStockTicker 클래스의 필드 개수 확인
        field_count = len(FinanceStockTickerUtil.FinanceStockTicker.__table__.c)
        print(rf'''field_count : {field_count}''')

        targets_validated = [
            {"field_en": 'id', "field_ko": "아이디", "field_validation_func": FinanceStockTickerUtil.validate_id, "field_length_limit": FinanceStockTickerUtil.FinanceStockTicker.__table__.c.id.type.length},
            {"field_en": 'ticker', "field_ko": "티커", "field_validation_func": FinanceStockTickerUtil.validate_ticker, "field_length_limit": FinanceStockTickerUtil.FinanceStockTicker.__table__.c.ticker.type.length},
            {"field_en": 'stock_name', "field_ko": "주식명", "field_validation_func": FinanceStockTickerUtil.validate_stock_name, "field_length_limit": FinanceStockTickerUtil.FinanceStockTicker.__table__.c.stock_name.type.length},
            {"field_en": 'market_name', "field_ko": "주식시장명", "field_validation_func": FinanceStockTickerUtil.validate_market_name, "field_length_limit": FinanceStockTickerUtil.FinanceStockTicker.__table__.c.market_name.type.length},
            {"field_en": 'date_reg', "field_ko": "등록일", "field_validation_func": FinanceStockTickerUtil.validate_date_reg, "field_length_limit": FinanceStockTickerUtil.FinanceStockTicker.__table__.c.date_reg.type.length},
            {"field_en": 'date_del', "field_ko": "삭제일", "field_validation_func": FinanceStockTickerUtil.validate_date_del, "field_length_limit": FinanceStockTickerUtil.FinanceStockTicker.__table__.c.date_del.type.length},
        ]
        for target in targets_validated:
            field_en = target["field_en"]
            field_ko = target["field_ko"]
            field_validation_func = target["field_validation_func"]
            field_length_limit = target["field_length_limit"]
            if len(finance_stock_ticker[field_en]) > target['field_length_limit']:
                raise HTTPException(status_code=400, detail=f"{field_ko}({field_en})의 길이제한은 {field_length_limit}자 이하여야 합니다.")
            else:
                field_validation_func(finance_stock_ticker[field_en])  # success, 호출할 수 없는 함수의 내부에 구현된 부분이 필요한것 이므로 내부에 구현된 것을 다른 클래스에 구현해서 참조하도록 로직 분리,
        return finance_stock_ticker

    @staticmethod
    def validate_finance_stock_ticker(finance_stock_ticker):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        FinanceStockTickerUtil.get_finance_stock_ticker_validated(finance_stock_ticker)

    class FinanceStockTickerCreate(FinanceStockTickerBase):
        pass

    class FinanceStockTickerExtendedFinanceStockTickerBase(FinanceStockTickerBase):
        id: str
        ticker: str
        stock_name: str
        market_name: str
        date_reg: str
        date_del: str

        class Config:
            from_attributes = True

    @staticmethod
    def get_finance_stock_tickers(db: Session):
        # db_data = MySqlUtil.execute(f'''SELECT * FROM finance_stock_ticker;''')
        db_data = db.query(FinanceStockTickerUtil.FinanceStockTicker).all()
        db.close()
        return db_data

    @staticmethod
    def get_finance_stock_ticker(id: int, db: Session):
        # db_data = MySqlUtil.execute(f'''SELECT * FROM finance_stock_ticker where id= {id} ORDER BY del_yn LIMIT 4;''')
        db_data = db.query(FinanceStockTickerUtil.FinanceStockTicker).filter_by(id=id).limit(4).all()
        db.close()
        return db_data

    @staticmethod
    def insert_finance_stock_ticker(finance_stock_ticker, db: Session):
        # db_data = MySqlUtil.execute(f'''insert into finance_stock_ticker () values () as new on duplicate key update ~~~~~~~~~ ;''')
        finances_tockt_icker_ = FinanceStockTickerUtil.FinanceStockTicker(**finance_stock_ticker)
        db.add(finances_tockt_icker_)
        db.flush()  # flush() 메서드 없이 바로 commit() 메서드를 호출하면, 롤백할 수 있는 포인트가 만들어지지 않습니다. (# 나중에 롤백을 수행할 수 있는 포인트가 만들어짐)
        db.commit()
        db.refresh(finances_tockt_icker_)  # 데이터베이스에 업데이트된 최신내용을 세션에 가져오는 것.
        db.close()

    @staticmethod
    def update_finance_stock_ticker(finance_stock_ticker, updated_finance_stock_ticker, db: Session):
        # db_data = MySqlUtil.execute(f'''update ~;''')
        for key, value in updated_finance_stock_ticker.model_dump().finances_tockt_ickers():
            setattr(finance_stock_ticker, key, value)
        db.commit()
        db.refresh(finance_stock_ticker)
        db.close()
        return finance_stock_ticker

    @staticmethod
    def delete_finance_stock_ticker(finance_stock_ticker, db: Session):
        # db_data = MySqlUtil.execute(f'''delete ~;''')
        db.delete(finance_stock_ticker)
        db.commit()
        db.close()

    @staticmethod
    def validate_id(value):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        BusinessLogicUtil.raise_exception_after_special_charcater_check(value, inspect.currentframe().f_code.co_name)
        return True

    @staticmethod
    def validate_ticker(value):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        BusinessLogicUtil.raise_exception_after_special_charcater_check(value, inspect.currentframe().f_code.co_name)
        return True

    @staticmethod
    def validate_stock_name(value):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        BusinessLogicUtil.raise_exception_after_special_charcater_check(value, inspect.currentframe().f_code.co_name, ignore_list=["@"])
        FinanceStockTickerUtil.validate_stock_name(value)
        pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
        if not re.match(pattern, value):
            # if not date_reg_stock_name.endswith('@kakao.com'):
            #     raise HTTPException(status_code=400, detail="유효한 카카오 이메일이 아닙니다.")
            # return date_reg_stock_name
            raise HTTPException(status_code=400, detail=f"유효한 이메일 주소가 아닙니다. {value}")
        return value

    @staticmethod
    def validate_market_name(value):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        # r'^\d{3}-\d{3,4}-\d{4}$'
        # r'^\d{2}-\d{3,4}-\d{4}$' 둘다
        if not re.match(r'^\d{2,3}-\d{3,4}-\d{4}$', value):
            raise HTTPException(status_code=400, detail=f"유효한 전화번호가 아닙니다. {value}")
        return value

    @staticmethod
    def validate_date_reg(value):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        return True

    @staticmethod
    def validate_date_del(value):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        return True


class StockInfoUtil:
    """
    mysql / sqlalchemy / fastapi 의존하는 유틸리티 객체
    """

    class StockInfo(MySqlUtil.Base):  # orm 설정에는 id 있음
        __tablename__ = "stock_info"
        __table_args__ = {'extend_existing': True}
        # __table_args__ = {'extend_existing': True, 'mysql_collate': 'utf8_general_ci'} # encoding 안되면 비슷하게 방법을 알아보자  mysql 에 적용이 가능한 코드로 보인다.
        StockInfo_id: str = uuid4().hex + TimeUtil.get_time_as_('%Y%m%d%H%M%S%f') + SecurityUtil.get_random_alphabet()  # table 내 unique id
        id = Column(Integer, primary_key=True, autoincrement=True)  # index 로 이용
        date = Column(DateTime, nullable=False)
        open = Column(Integer)
        high = Column(Integer)
        low = Column(Integer)
        close = Column(Integer)
        volume = Column(Integer)
        ticker = Column(VARCHAR(length=6))
        date_reg = Column(DateTime, nullable=False, default=datetime.now)
        date_del = Column(DateTime, nullable=True)

    class StockInfoBase(BaseModel):  # pydantic validator 설정에는 StockInfo_id 없음
        id: str
        date: str
        open: int
        high: int
        low: int
        close: int
        volume: int
        ticker: str
        date_reg: str
        date_del: str

        @staticmethod
        @field_validator('id')
        def validate_id(value):
            StockInfoUtil.validate_id(value)

        @staticmethod
        @field_validator('date_reg')
        def validate_date_reg(value):
            # datetime.strptime(date_join, '%Y-%m-%d %H:%M %S%f')
            if len(value) != 18:
                raise HTTPException(status_code=400, detail="유효한 날짜가 아닙니다.")
            StockInfoUtil.validate_date_reg(value)

    @staticmethod
    def get_stock_info_validated(stock_info):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")

        # StockInfo 클래스의 필드 개수 확인
        field_count = len(StockInfoUtil.StockInfo.__table__.c)
        print(rf'''field_count : {field_count}''')



        targets_validated = [
            {"field_en": 'id', "field_ko": "아이디", "field_validation_func": StockInfoUtil.validate_id, "field_length_limit": StockInfoUtil.StockInfo.__table__.c.id.type.length},
            {"field_en": 'date', "field_ko": "날짜", "field_validation_func": StockInfoUtil.validate_date_reg, "field_length_limit": StockInfoUtil.StockInfo.__table__.c.date.type.length},
            {"field_en": 'open', "field_ko": "시가", "field_validation_func": StockInfoUtil.validate_open, "field_length_limit": StockInfoUtil.StockInfo.__table__.c.open.type.length},
            {"field_en": 'high', "field_ko": "고가", "field_validation_func": StockInfoUtil.validate_high, "field_length_limit": StockInfoUtil.StockInfo.__table__.c.high.type.length},
            {"field_en": 'low', "field_ko": "저가", "field_validation_func": StockInfoUtil.validate_low, "field_length_limit": StockInfoUtil.StockInfo.__table__.c.low.type.length},
            {"field_en": 'close', "field_ko": "종가", "field_validation_func": StockInfoUtil.validate_low, "field_length_limit": StockInfoUtil.StockInfo.__table__.c.close.type.length},
            {"field_en": 'volume', "field_ko": "거래량", "field_validation_func": StockInfoUtil.validate_low, "field_length_limit": StockInfoUtil.StockInfo.__table__.c.volume.type.length},
            {"field_en": 'ticker', "field_ko": "종목코드", "field_validation_func": StockInfoUtil.validate_date_reg, "field_length_limit": StockInfoUtil.StockInfo.__table__.c.ticker.type.length},
            {"field_en": 'date_reg', "field_ko": "등록일", "field_validation_func": StockInfoUtil.validate_date_reg, "field_length_limit": StockInfoUtil.StockInfo.__table__.c.date_reg.type.length},
            {"field_en": 'date_del', "field_ko": "삭제일", "field_validation_func": StockInfoUtil.validate_date_del, "field_length_limit": StockInfoUtil.StockInfo.__table__.c.date_del.type.length},
        ]
        for target in targets_validated:
            field_en = target["field_en"]
            field_ko = target["field_ko"]
            field_validation_func = target["field_validation_func"]
            field_length_limit = target["field_length_limit"]
            if len(stock_info[field_en]) > target['field_length_limit']:
                raise HTTPException(status_code=400, detail=f"{field_ko}({field_en})의 길이제한은 {field_length_limit}자 이하여야 합니다.")
            else:
                field_validation_func(stock_info[field_en])  # success, 호출할 수 없는 함수의 내부에 구현된 부분이 필요한것 이므로 내부에 구현된 것을 다른 클래스에 구현해서 참조하도록 로직 분리,
        return stock_info

    @staticmethod
    def validate_stock_info(stock_info):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        StockInfoUtil.get_stock_info_validated(stock_info)

    class StockInfoCreate(StockInfoBase):
        pass

    class StockInfoExtendedStockInfoBase(StockInfoBase):
        id: str
        date: str
        open: int
        high: int
        low: int
        close: int
        volume: int
        ticker: str
        date_reg: str
        date_del: str

        class Config:
            from_attributes = True

    @staticmethod
    def get_stock_infos(db: Session):
        # db_data = MySqlUtil.execute(f'''SELECT * FROM stock_info;''')
        db_data = db.query(StockInfoUtil.StockInfo).all()
        db.close()
        return db_data

    @staticmethod
    def get_stock_info(id: int, db: Session):
        # db_data = MySqlUtil.execute(f'''SELECT * FROM stock_info where id= {id} ORDER BY del_yn LIMIT 4;''')
        db_data = db.query(StockInfoUtil.StockInfo).filter_by(id=id).limit(4).all()
        db.close()
        return db_data

    @staticmethod
    def insert_stock_info(stock_info, db: Session):
        # db_data = MySqlUtil.execute(f'''insert into stock_info () values () as new on duplicate key update ~~~~~~~~~ ;''')
        stock_info_ = StockInfoUtil.StockInfo(**stock_info)
        db.add(stock_info_)
        db.flush()  # flush() 메서드 없이 바로 commit() 메서드를 호출하면, 롤백할 수 있는 포인트가 만들어지지 않습니다. (# 나중에 롤백을 수행할 수 있는 포인트가 만들어짐)
        db.commit()
        db.refresh(stock_info_)  # 데이터베이스에 업데이트된 최신내용을 세션에 가져오는 것.
        db.close()

    @staticmethod
    def insert_stock_infos(stock_infos, db: Session):
        model = StockInfoUtil.StockInfo
        db.bulk_insert_mappings(model, stock_infos)
        db.flush()  # flush() 메서드 없이 바로 commit() 메서드를 호출하면, 롤백할 수 있는 포인트가 만들어지지 않습니다. (# 나중에 롤백을 수행할 수 있는 포인트가 만들어짐)
        db.commit()
        db.refresh(model)  # 데이터베이스에 업데이트된 최신내용을 세션에 가져오는 것.
        db.close()

    @staticmethod
    def update_stock_info(stock_info, updated_stock_info, db: Session):
        # db_data = MySqlUtil.execute(f'''update ~;''')
        for key, value in updated_stock_info.model_dump().stock_infos():
            setattr(stock_info, key, value)
        db.commit()
        db.refresh(stock_info)
        db.close()
        return stock_info

    @staticmethod
    def delete_stock_info(stock_info, db: Session):
        # db_data = MySqlUtil.execute(f'''delete ~;''')
        db.delete(stock_info)
        db.commit()
        db.close()



    @staticmethod
    def validate_id(value):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        BusinessLogicUtil.raise_exception_after_special_charcater_check(value, inspect.currentframe().f_code.co_name)
        return True


class BusinessLogicUtil:

    @staticmethod
    def get_none_count_of_list(list):
        count = sum(element is None for element in list)
        return count

    @staticmethod
    def get_validated(target: any):
        if target == None:
            target = "논값"
        if target == "":
            target = "공백"
        if type(target) == str:
            if BusinessLogicUtil.is_regex_in_contents(target=target, regex=r'[^a-zA-Z0-9가-힣\s]'):  # 특수문자 패턴 정의( 알파벳, 숫자, 한글, 공백을 제외한 모든 문자)
                target = BusinessLogicUtil.get_str_replaced_special_characters(target, "$특수문자$")
            return target
        else:
            return target

    @staticmethod
    def is_validated(target: any):
        if target == None:
            return False
        if target == "":
            return False
        if BusinessLogicUtil.is_regex_in_contents(target=target, regex=r'[^a-zA-Z0-9가-힣\s]'):  # 특수문자 패턴 정의( 알파벳, 숫자, 한글, 공백을 제외한 모든 문자)
            return False
        else:
            return True

    @staticmethod
    # def guide_user_input_intended_by_developer(user_input: str):
    def is_user_input_sanitized(user_input: str):
        # 수정 필요.
        user_input = user_input.strip()
        if TextToSpeechUtil.is_only_no(user_input):
            user_input = int(user_input)
        else:
            TextToSpeechUtil.speak_ments("you can input only number, please input only number again", sleep_after_play=0.65)
            return None
        return user_input

    @staticmethod
    def validate_and_return(value: str):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        try:
            print(rf'value : {value}')
            print(rf'type(value) : {type(value)}')
            print(rf'len(value) : {len(value)}')
        except:
            pass
        if value == None:
            print(rf'벨리데이션 테스트 결과 : None')
            return False
        if value == "":
            print(rf'벨리데이션 테스트 결과 : 공백')
            return False
        # if 전화번호만 같아 보이는지
        # if 특수문자만 같아 보이는지
        return value

    @staticmethod
    def get_biological_age(birth_day):
        # 2023-1994=29(생일후)
        # 2024-1994=30(생일후)
        # 만나이 == 생물학적나이
        # 생일 전 만나이
        # 생일 후 만나이
        pass

    # @staticmethod
    # def should_i_download_youtube_as_webm_alt():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     previous_text = ""
    #     while True:
    #         dialog = UiUtil.CustomQdialog(ment="다운로드하고 싶은 URL을 입력해주세요", btns=["입력", "입력하지 않기"], is_input_box=True, input_box_text_default=clipboard.paste())
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #         DebuggingUtil.print_magenta(btn_text_clicked)
    #         if btn_text_clicked == "입력":
    #             url = dialog.input_box.text()
    #             BusinessLogicUtil.download_from_youtube_to_webm_alt(url)
    #         else:
    #             break
    #
    # @staticmethod
    # def should_i_shutdown_this_computer():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #         dialog = UiUtil.CustomQdialog(ment="시스템을 종료할까요?", btns=["종료", "종료하지 않기"])
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #         DebuggingUtil.print_magenta(btn_text_clicked)
    #         if btn_text_clicked == "종료":
    #             FileSystemUtil.shutdown_this_computer()
    #         else:
    #             break
    #
    # @staticmethod
    # def should_i_back_up_target():  # please_type_abspath_to_back_up
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #
    #     while True:
    #         previous_input = clipboard.paste()
    #         previous_input = previous_input.strip()
    #         if previous_input == "":
    #             previous_input = StateManagementUtil.SERVICE_DIRECTORY
    #         dialog = UiUtil.CustomQdialog(ment="빽업할 타겟경로를 입력하세요", btns=["입력", "입력하지 않기"], is_input_box=True, input_box_text_default=previous_input)
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #         DebuggingUtil.print_magenta(btn_text_clicked)
    #         if btn_text_clicked == "입력":
    #             BusinessLogicUtil.back_up_target(dialog.input_box.text())
    #         else:
    #             break
    #
    # @staticmethod
    # def should_i_start_test_core():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #         dialog = UiUtil.CustomQdialog(ment="테스트를 시작할까요?", btns=["시작하기", "시작하지 않기"])
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #         if btn_text_clicked == "시작하기":
    #             DebuggingUtil.print_magenta("아무 테스트도 정의되지 않았습니다.")
    #             pass
    #         else:
    #             break
    #
    # @staticmethod
    # def should_i_enter_to_power_saving_mode():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #         dialog = UiUtil.CustomQdialog(ment="절전모드로 진입할까요?", btns=["진입", "10분 뒤 진입", "20분 뒤 진입", "30분 뒤 진입", "1시간 뒤 진입", "진입하지 않기"])
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #         DebuggingUtil.print_magenta(btn_text_clicked)
    #         if btn_text_clicked == "진입":
    #             FileSystemUtil.enter_power_saving_mode()
    #         elif btn_text_clicked == "10분 뒤 진입":
    #             UiUtil.pop_up_as_complete(title_="작업중간보고", ment="10분 뒤 절전모드 진입을 시도합니다", input_text_default="", auto_click_positive_btn_after_seconds=3)
    #             BusinessLogicUtil.sleep(min=10)
    #             FileSystemUtil.enter_power_saving_mode()
    #         elif btn_text_clicked == "20분 뒤 진입":
    #             UiUtil.pop_up_as_complete(title_="작업중간보고", ment="20분 뒤 절전모드 진입을 시도합니다", input_text_default="", auto_click_positive_btn_after_seconds=3)
    #             BusinessLogicUtil.sleep(min=20)
    #             FileSystemUtil.enter_power_saving_mode()
    #         elif btn_text_clicked == "30분 뒤 진입":
    #             UiUtil.pop_up_as_complete(title_="작업중간보고", ment="30분 뒤 절전모드 진입을 시도합니다", input_text_default="", auto_click_positive_btn_after_seconds=3)
    #             BusinessLogicUtil.sleep(min=30)
    #             FileSystemUtil.enter_power_saving_mode()
    #         elif btn_text_clicked == "1시간 뒤 진입":
    #             UiUtil.pop_up_as_complete(title_="작업중간보고", ment="1시간 뒤 절전모드 진입을 시도합니다", input_text_default="", auto_click_positive_btn_after_seconds=3)
    #             BusinessLogicUtil.sleep(min=60)
    #             FileSystemUtil.enter_power_saving_mode()
    #         else:
    #             break
    #
    # @staticmethod
    # def should_i_translate_eng_to_kor():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #         dialog = UiUtil.CustomQdialog(ment="WRITE SOMETHING YOU WANT TO TRANSLATE", btns=["Translate this to Korean", "Don't"], is_input_box=True)
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #         DebuggingUtil.print_magenta(btn_text_clicked)
    #         if btn_text_clicked == "Translate this to Korean":
    #             BusinessLogicUtil.translate_eng_to_kor(dialog.input_box.text())
    #         else:
    #             break
    #
    # @staticmethod
    # def should_i_translate_kor_to_eng():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #         dialog = UiUtil.CustomQdialog(ment='번역하고 싶은 내용을 입력하세요', btns=["영어로 번역", "번역하지 않기"], is_input_box=True)
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #         DebuggingUtil.print_magenta(btn_text_clicked)
    #         if btn_text_clicked == "영어로 번역":
    #             BusinessLogicUtil.translate_kor_to_eng(dialog.input_box.text())
    #         else:
    #             break
    #
    # @staticmethod
    # def should_i_exit_this_program():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #         # dialog = UiUtil.CustomQdialog(context="앱을 종료할까요?", buttons=["종료", "종료하지 않기"])
    #         dialog = UiUtil.CustomQdialog(ment="프로그램을 종료할까요?", btns=["종료", "종료하지 않기"])
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #         DebuggingUtil.print_magenta(btn_text_clicked)
    #         if btn_text_clicked == "종료":
    #             # app.quit()
    #             # Park4139.taskkill("python.exe")
    #             # self.close()
    #             sys.exit()
    #             # break
    #         else:
    #             break
    #
    # @staticmethod
    # def should_i_show_animation_information_from_web():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #         dialog = UiUtil.CustomQdialog(ment="nyaa.si 에서 검색할 내용을 입력하세요", btns=["검색", "검색하지 않기"], is_input_box=True)
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #         DebuggingUtil.print_magenta(btn_text_clicked)
    #         if btn_text_clicked == "검색":
    #             BusinessLogicUtil.search_animation_data_from_web(dialog.input_box.text())
    #         else:
    #             break
    #
    # @staticmethod
    # def should_i_find_direction_via_naver_map():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #         dialog = UiUtil.CustomQdialog(ment="어디로 길을 찾아드릴까요?", btns=["찾아줘", "찾아주지 않아도 되"], is_input_box=True)
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #         DebuggingUtil.print_magenta(btn_text_clicked)
    #         if btn_text_clicked == "찾아줘":
    #             BusinessLogicUtil.find_direction_via_naver_map(dialog.input_box.text())
    #         else:
    #             TextToSpeechUtil.speak_ments("네, 알겠습니다", sleep_after_play=0.65)
    #             break

    # @staticmethod
    # def should_i_reboot_this_computer():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #         dialog = UiUtil.CustomQdialog(ment='시스템을 재시작할까요?', btns=[MentsUtil.YES, MentsUtil.NO])
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #         DebuggingUtil.print_magenta(btn_text_clicked)
    #         if btn_text_clicked == MentsUtil.YES:
    #             FileSystemUtil.reboot_this_computer()
    #         else:
    #             break
    #
    # @staticmethod
    # def ask_something_to_ai():
    #     previous_question = None
    #     if previous_question == None:
    #         previous_question = clipboard.paste()
    #     while True:
    #         dialog = UiUtil.CustomQdialog(ment='AI 에게 할 질문을 입력하세요', btns=["질문하기", "질문하지 않기"], is_input_box=True, input_box_text_default=previous_question)
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #         DebuggingUtil.print_magenta(btn_text_clicked)
    #         if btn_text_clicked == "질문하기":
    #             BusinessLogicUtil.ask_to_web(dialog.input_box.text())
    #         else:
    #             break
    #         previous_question = dialog.input_box.text()
    #
    # @staticmethod
    # def should_i_connect_to_rdp1():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #         dialog = UiUtil.CustomQdialog(ment='rdp1에 원격접속할까요?', btns=[MentsUtil.YES, MentsUtil.NO])
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #         DebuggingUtil.print_magenta(btn_text_clicked)
    #         if btn_text_clicked == MentsUtil.YES:
    #             BusinessLogicUtil.connect_remote_rdp1()
    #         else:
    #             break
    #
    # @staticmethod
    # def should_i_record_macro():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #         dialog = UiUtil.CustomQdialog(ment="매크로 레코드를 어떻게 시작할까요?", btns=["새로 시작하기", "이어서 시작하기", "시작하지 않기"])
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #         if btn_text_clicked == "새로 시작하기":
    #             if os.path.exists(StateManagementUtil.MACRO_LOG):
    #                 FileSystemUtil.move_target_to_trash_bin(StateManagementUtil.MACRO_LOG)
    #             FileSystemUtil.make_leaf_file(StateManagementUtil.MACRO_LOG)
    #             FileSystemUtil.explorer(StateManagementUtil.MACRO_LOG)
    #         elif btn_text_clicked == "이어서 시작하기":
    #             if os.path.exists(StateManagementUtil.MACRO_LOG):
    #                 FileSystemUtil.make_leaf_file(StateManagementUtil.MACRO_LOG)
    #                 FileSystemUtil.explorer(StateManagementUtil.MACRO_LOG)
    #         elif btn_text_clicked == "시작하지 않기":
    #             break
    #         macro_window = UiUtil.MacroWindow()
    #         macro_window.show()
    #         macro_window.activateWindow()
    #         break

    # @staticmethod
    # def download_video_from_web1():
    #     while True:
    #         # Park4139.press("ctrl", "0")
    #         file_png = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_png\download_video_via_chrome_extensions.png"
    #         BusinessLogicUtil.click_center_of_img_recognized_by_mouse_left(img_abspath=file_png, recognize_loop_limit_cnt=10)
    #
    #         BusinessLogicUtil.sleep(1000)
    #
    #         BusinessLogicUtil.press("tab")
    #         BusinessLogicUtil.sleep(30)
    #
    #         BusinessLogicUtil.press("enter")
    #         BusinessLogicUtil.sleep(30)
    #
    #         BusinessLogicUtil.press("ctrl", "shift", "tab")
    #
    #         BusinessLogicUtil.press("ctrl", "0")
    #         BusinessLogicUtil.press("ctrl", "-")
    #         BusinessLogicUtil.press("ctrl", "-")
    #         break

    # @staticmethod
    # # def should_i_do_that(ment: str, function, auto_click_negative_btn_after_seconds: int = 30, auto_click_positive_btn_after_seconds: int = None):
    # def should_i_do(ment: str, function: Callable = None, is_void_mode: bool = True, auto_click_negative_btn_after_seconds: int = None, auto_click_positive_btn_after_seconds: int = None):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     # function() 이 void 함수 일때 시도, is_void_mode= True 설정인 경우, is_void_mode 파라미터 제거 대기
    #     if is_void_mode != True:
    #         if auto_click_negative_btn_after_seconds == None and auto_click_positive_btn_after_seconds == None:
    #             auto_click_negative_btn_after_seconds = 30
    #         while True:
    #             dialog = UiUtil.CustomQdialog(ment=ment, btns=[MentsUtil.YES, MentsUtil.NO], auto_click_negative_btn_after_seconds=auto_click_negative_btn_after_seconds, auto_click_positive_btn_after_seconds=auto_click_positive_btn_after_seconds)
    #             dialog.exec()
    #             if dialog.btn_text_clicked == MentsUtil.YES:
    #                 if function != None:
    #                     function()
    #             else:
    #                 break
    #             break
    #     # function() 이 return 함수 일때 시도, is_void_mode= false 설정인 경우,
    #     # else:
    #     #     if auto_click_negative_btn_after_seconds == None and auto_click_positive_btn_after_seconds == None:
    #     #         auto_click_negative_btn_after_seconds = 30
    #     #     while True:
    #     #         dialog = UiUtil.CustomQdialog(ment=ment, btns=[MentsUtil.YES, MentsUtil.NO], auto_click_negative_btn_after_seconds=auto_click_negative_btn_after_seconds, auto_click_positive_btn_after_seconds=auto_click_positive_btn_after_seconds, is_input_box=True)
    #     #         dialog.exec()
    #     #         btn_text_clicked = dialog.btn_text_clicked
    #     #         if btn_text_clicked == MentsUtil.YES:
    #     #             if function != None:
    #     #                 print(rf'function(url=dialog.input_box.text()) : {function(url=dialog.input_box.text())}')
    #     #                 # function(url=dialog.input_box.text()) # fail
    #     #                 partial(function, url=dialog.input_box.text()) # fail
    #     #         else:
    #     #             break
    #     #         break

    # @staticmethod
    # def should_i_check_your_promised_routines_before_coding(routines: [str]):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     """
    #     예정된 routines 수행 다하면 보상을 주는 게임을 실행할지 묻는 함수 mkmk
    #     """
    #     # 여기서 deepcopy() 를 쓴 이유
    #     # 원본의 len(routines) 을 알아야 하는데
    #     # deepcopy 를 하지않으면 step 마다  routine: str 이 줄어든 routines: [routine] 의 len 을 참조하게 되는데
    #     # 이는, 의도한 초기의 routines 의 len 을 참조하는 것과 다르므로, routines_deep_copied 를 수행하였다
    #
    #     # 여기서는 routines 를 수행된 routine을 제거하고 routine이 제거된 routines 를 관리하는데
    #     # 그동안 평소 구현했던 일반적으로 리스트를 순환할때와 enumerate를 통하여 cursor 를 움직이며 동작하는 것과 달리,
    #     # routines: [str] 에서 routine 을 하나씩 없애도록 만들었다 .
    #     # step= 1, routines = [ "routine1", "routine2", "routine3" ]
    #     # step= 2, routines = [ "routine2", "routine3" ]
    #     # step= 3, routines = [ "routine3" ]
    #     # 코드 실험을 해보면, 첫번째 원소가 계속 빠지도록 만들었다는 것을 알수 있다. 큐 자료구조와 일부 비슷한 부분이 있는 구조이다.
    #     # 의도하고 만든건 아니지만. 자료구조적으로는 FIFO 가 활용된 것이다.
    #     # cursor 는 routines[0] 만 계속 가리키게 한다. routines[0]을 수행했다면 routines 에서 routines[0](리스트의 첫 원소)를 계속 빼어 버린다.
    #     routines_deep_copied = copy.deepcopy(routines)
    #     max = len(routines_deep_copied)
    #     cursor_position = 0
    #     is_first_entry = True
    #     while True:
    #         if is_first_entry == True:
    #             ment = '혹시 코딩을 할 계획이 있으신가요?, 코딩 전 루틴을 가이드할까요?'
    #             # TextToSpeechUtil.speak_ments(ment=ment, sleep_after_play=0.55)
    #             dialog = UiUtil.CustomQdialog(ment=ment, btns=["하나씩 가이드해줘", "한번에 가이드해줘", "모두 미리 수행했어"], auto_click_positive_btn_after_seconds=30)
    #             is_first_entry = False
    #         else:
    #             TextToSpeechUtil.speak_ments(ment='루틴을 하나씩 가이드할까요?', sleep_after_play=0.65)
    #             dialog = UiUtil.CustomQdialog(ment='루틴을 하나씩 가이드할까요?', btns=["하나씩 가이드해줘", "한번에 가이드해줘", "모두 미리 수행했어"], auto_click_positive_btn_after_seconds=30)
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #         if btn_text_clicked == "모두 미리 수행했어":
    #             routines = []
    #             if len(routines) == 0:
    #                 ment = "모든 루틴을 계획대로 잘 수행하셨군요, 계획대로 잘 하셨습니다, 랜덤보상을 하나드릴게요"
    #                 TextToSpeechUtil.speak_ments(ment, sleep_after_play=0.65)
    #                 UiUtil.pop_up_as_complete(title_="작업성공보고", ment=ment, auto_click_positive_btn_after_seconds=5)
    #                 # 랜덤보상들 여기에 작성하기
    #                 FileSystemUtil.explorer(StateManagementUtil.MUSIC_FOR_WORK)
    #                 break
    #         elif btn_text_clicked == "하나씩 가이드해줘":
    #             # TtsUtil.speak("네 가이드 하겠습니다", sleep_after_play=0.65)
    #             for i in range(0, max):
    #                 print(rf'routines_deep_copied : {routines_deep_copied}')
    #                 print(rf'type(routines_deep_copied) : {type(routines_deep_copied)}')
    #                 print(rf'len(routines_deep_copied) : {len(routines_deep_copied)}')
    #                 print(rf'routines : {routines}')
    #                 print(rf'type(routines) : {type(routines)}')
    #                 print(rf'len(routines) : {len(routines)}')
    #
    #                 dialog = UiUtil.CustomQdialog(ment=rf"{routines[cursor_position]} 하셨어요?", btns=[MentsUtil.DONE, MentsUtil.I_WANT_TO_TO_DO_NEXT_TIME], auto_click_negative_btn_after_seconds=3600)
    #                 dialog.exec()
    #                 if dialog.btn_text_clicked == MentsUtil.DONE:
    #                     del routines[cursor_position]  # 리스트의 특정 인덱스 삭제
    #                     if len(routines) == 0:
    #                         ment = "루틴을 하나하나 계획대로 잘 수행하셨군요,\n 계획대로 잘 하셨습니다,\n 미루지 않았으니 내일도 좋은일 가득할 거에요,\n 랜덤보상을 하나드릴게요"
    #                         TextToSpeechUtil.speak_ments(ment, sleep_after_play=0.65)
    #                         UiUtil.pop_up_as_complete(title_="작업성공보고", ment=ment, auto_click_positive_btn_after_seconds=5)
    #                         # 랜덤보상들 여기에 작성하기
    #                         FileSystemUtil.explorer(StateManagementUtil.MUSIC_FOR_WORK)
    #                         break
    #                 elif dialog.btn_text_clicked == MentsUtil.I_WANT_TO_TO_DO_NEXT_TIME:
    #                     # ment = f'정말로 "{routines[cursor_position]}" 하는 것을 다음으로 미루시겠어요?, 미루지 않는 습관은 인생에 중요한데도요?'
    #                     ment = f'정말로 "{routines[cursor_position]}" 하는 것을 다음으로 미루시겠어요?,\n 미루지 않는 습관은 인생에 중요한데도요?'
    #                     TextToSpeechUtil.speak_ments(ment, sleep_after_play=0.65)
    #                     dialog = UiUtil.CustomQdialog(ment=ment, btns=[MentsUtil.OK_I_WILL_DO_IT_NOW, MentsUtil.I_WANT_TO_TO_DO_NEXT_TIME], auto_click_negative_btn_after_seconds=3600)
    #                     dialog.exec()
    #                     if dialog.btn_text_clicked == MentsUtil.OK_I_WILL_DO_IT_NOW:
    #                         dialog = UiUtil.CustomQdialog(ment=f"{routines[cursor_position]} 하셨어요?", btns=[MentsUtil.DONE, MentsUtil.I_WANT_TO_TO_DO_NEXT_TIME], auto_click_negative_btn_after_seconds=3600)
    #                         dialog.exec()
    #                         if dialog.btn_text_clicked == MentsUtil.DONE:
    #                             del routines[cursor_position]  # 리스트의 특정 인덱스 삭제 # routines의 cursor_position에 대한 인덱스 삭제
    #                         elif dialog.btn_text_clicked == MentsUtil.I_WANT_TO_TO_DO_NEXT_TIME:
    #                             dialog = UiUtil.CustomQdialog(ment="이것만 미룰까요?", btns=["이것만 미뤄", "전부 미뤄"], auto_click_negative_btn_after_seconds=3600)
    #                             dialog.exec()
    #                             if dialog.btn_text_clicked == "이것만 미뤄":
    #                                 cursor_position = cursor_position + 1
    #                                 pass
    #                             elif dialog.btn_text_clicked == "전부 미뤄":
    #                                 TextToSpeechUtil.speak_ments("네 알겠어요. 전부 미룰게요. 기억해둘게요. 나중에 다시하실 수 있게요", sleep_after_play=0.65)
    #                                 routines_left: str = "\n".join(routines)
    #                                 dialog = UiUtil.CustomQdialog(ment=f"<남은 루틴목록>\n\n{routines_left}", btns=[MentsUtil.CHECKED], auto_click_positive_btn_after_seconds=10)
    #                                 dialog.exec()
    #                                 break
    #                     elif dialog.btn_text_clicked == MentsUtil.I_WANT_TO_TO_DO_NEXT_TIME:
    #                         dialog = UiUtil.CustomQdialog(ment="이것만 미룰까요?", btns=["남은 루틴 이것만 미뤄", "남은 루틴 전부 미뤄"], auto_click_negative_btn_after_seconds=3600)
    #                         dialog.exec()
    #                         if dialog.btn_text_clicked == "남은 루틴 이것만 미뤄":
    #                             cursor_position = cursor_position + 1
    #                             pass
    #                         elif dialog.btn_text_clicked == "남은 루틴 전부 미뤄":
    #                             TextToSpeechUtil.speak_ments("네 알겠어요 남은 루틴을 전부 미룰게요. 기억해둘게요 나중에 다시하실 수 있게요", sleep_after_play=0.65)
    #                             routines_left: str = "\n".join(routines)
    #                             dialog = UiUtil.CustomQdialog(ment=f"<남은 루틴목록>\n\n{routines_left}", btns=[MentsUtil.CHECKED], auto_click_positive_btn_after_seconds=10)
    #                             dialog.exec()
    #                             break
    #             break
    #             # '응, 아니 지금할게'
    #         elif btn_text_clicked == "한번에 가이드해줘":
    #             TextToSpeechUtil.speak_ments("네 한번에 가이드를 할게요 예정된 루틴목록입니다", sleep_after_play=0.65)
    #             routines_promised: str = "\n".join(routines)
    #             dialog = UiUtil.CustomQdialog(ment=f"<예정된 루틴목록>\n\n{routines_promised}", btns=[MentsUtil.DONE, MentsUtil.I_WANT_TO_TO_DO_NEXT_TIME], auto_click_negative_btn_after_seconds=10)
    #             dialog.exec()
    #             if dialog.btn_text_clicked == MentsUtil.DONE:
    #                 ment = "계획대로 잘 하셨습니다, 미루지 않았으니 내일도 좋은일 가득할 거에요, 랜덤보상을 하나드릴게요"
    #                 TextToSpeechUtil.speak_ments(ment, sleep_after_play=0.65)
    #                 UiUtil.pop_up_as_complete(title_="작업성공보고", ment=ment, auto_click_positive_btn_after_seconds=5)
    #                 # 랜덤보상들 여기에 작성하기
    #                 FileSystemUtil.explorer(StateManagementUtil.MUSIC_FOR_WORK)
    #                 break
    #             elif dialog.btn_text_clicked == MentsUtil.I_WANT_TO_TO_DO_NEXT_TIME:
    #                 TextToSpeechUtil.speak_ments("네 알겠어요 남은 루틴을 전부 미룰게요. 기억해둘게요 나중에 다시하실 수 있게요", sleep_after_play=0.65)
    #                 routines_left: str = "\n".join(routines)
    #                 dialog = UiUtil.CustomQdialog(ment=f"<남은 루틴목록>\n\n{routines_left}", btns=[MentsUtil.CHECKED], auto_click_positive_btn_after_seconds=10)
    #                 dialog.exec()
    #                 break
    #         break
    #
    # @staticmethod
    # def should_i_download_youtube_as_webm():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     text_previous_from_clipboard = None
    #     while True:
    #         if text_previous_from_clipboard != clipboard.paste():
    #             text_previous_from_clipboard = clipboard.paste()
    #         else:
    #             text_previous_from_clipboard = ""
    #         dialog = UiUtil.CustomQdialog(ment="다운로드하고 싶은 URL을 입력해주세요", btns=["입력", "입력하지 않기"], is_input_box=True, input_box_text_default=text_previous_from_clipboard)
    #         dialog.exec()
    #         if dialog.btn_text_clicked == "입력":
    #             BusinessLogicUtil.download_from_youtube_to_webm(urls=dialog.input_box.text())
    #         else:
    #             break
    #
    # @staticmethod
    # def should_i_merge_directories():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     previous_text = ""
    #     while True:
    #         previous_text = clipboard.paste()
    #         dialog = UiUtil.CustomQdialog(ment="머지할 타겟경로들을 입력하세요", btns=["입력", "입력하지 않기"], is_input_box=True, input_box_text_default=previous_text, auto_click_negative_btn_after_seconds=10)
    #         dialog.exec()
    #         directoryies = dialog.input_box.text()
    #         if dialog.btn_text_clicked == "입력":
    #             FileSystemUtil.merge_directories(directoryies=directoryies)
    #             break
    #         else:
    #             break
    #
    # @staticmethod
    # def should_i_convert_mkv_to_wav():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #         dialog = UiUtil.CustomQdialog(ment="convert 할 MKV파일경로를 입력하세요", btns=["입력", "입력하지 않기"], is_input_box=True, auto_click_negative_btn_after_seconds=10)
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #         file_mkv = dialog.input_box.text()
    #         if btn_text_clicked == "입력":
    #             FileSystemUtil.convert_mkv_to_wav(file_mkv=file_mkv)
    #             break
    #         else:
    #             break
    #
    # @staticmethod
    # def should_i_gather_empty_directory():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #         dialog = UiUtil.CustomQdialog(ment="빈폴더를 순회하며 수집할 타겟경로를 입력하세요", btns=["입력", "입력하지 않기"], is_input_box=True)
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #         if btn_text_clicked == "입력":
    #             target_abspath = dialog.input_box.text()
    #             target_abspath = target_abspath.strip()
    #             target_abspath = target_abspath.replace("\"", "")
    #             target_abspath = target_abspath.replace("\'", "")
    #             connected_drives = []
    #             for drive_letter in string.ascii_uppercase:
    #                 drive_path = drive_letter + ":\\"
    #                 if os.path.exists(drive_path):
    #                     connected_drives.append(drive_path)
    #                     if target_abspath == drive_path:
    #                         TextToSpeechUtil.speak_ments("입력된 타겟경로는 너무 광범위하여 진행할 수 없도록 설정되어 있습니다", sleep_after_play=0.65)
    #                         break
    #             if not os.path.exists(target_abspath):
    #                 TextToSpeechUtil.speak_ments("입력된 타겟경로가 존재하지 않습니다", sleep_after_play=0.65)
    #                 continue
    #             if target_abspath == "":
    #                 continue
    #
    #             BusinessLogicUtil.gather_empty_directory(target_abspath)
    #             DebuggingUtil.print_ment_success(ment=rf"타겟경로를 순회하며 빈폴더를 모았습니다")
    #             # TextToSpeechUtil.speak_ments("타겟경로를 순회하며 빈폴더를 약속된 폴더로 모았습니다", sleep_after_play=0.65) # 시끄러웠다,
    #             UiUtil.pop_up_as_complete(title_="작업성공보고", ment=f"타겟경로를 순회하며 빈폴더를 모았습니다", auto_click_positive_btn_after_seconds=2)
    #         else:
    #             break
    #
    # @staticmethod
    # def should_i_gather_useless_files():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     try:
    #         while True:
    #             dialog = UiUtil.CustomQdialog(ment="타겟경로를 순회하며 약속된 불필요한 파일을 약속된 폴더로 모을까요?", btns=["모으기", "모으지 않기"], is_input_box=True)
    #             dialog.exec()
    #             btn_text_clicked = dialog.btn_text_clicked
    #             if btn_text_clicked == "모으기":
    #                 directory_abspath = dialog.input_box.text()
    #                 directory_abspath = directory_abspath.strip()
    #                 directory_abspath = directory_abspath.replace("\"", "")
    #                 directory_abspath = directory_abspath.replace("\'", "")
    #                 connected_drives = []
    #                 for drive_letter in string.ascii_uppercase:
    #                     drive_path = drive_letter + ":\\"
    #                     if os.path.exists(drive_path):
    #                         connected_drives.append(drive_path)
    #                         if directory_abspath == drive_path:
    #                             TextToSpeechUtil.speak_ments("입력된 타겟경로는 너무 광범위하여 진행할 수 없도록 설정되어 있습니다", sleep_after_play=0.65)
    #                             break
    #                 if not os.path.exists(directory_abspath):
    #                     TextToSpeechUtil.speak_ments("입력된 타겟경로가 존재하지 않습니다", sleep_after_play=0.65)
    #                     continue
    #                 if directory_abspath == "":
    #                     continue
    #                 try:
    #                     # useless_files 수집 [ using clipboard ] # os 명령어를 사용하기에 빠를 것으로 기대
    #                     useless_directories = StateManagementUtil.USELESS_DIRECTORIES
    #                     FileSystemUtil.make_leaf_directory(useless_directories)
    #                     useless_files = []
    #                     is_target_moved_done = False
    #                     lines = FileSystemUtil.get_lines_of_file(StateManagementUtil.USELESS_FILES)
    #                     tmp = clipboard.paste()
    #                     os.chdir(directory_abspath)  # cwd 를 target_abspath 로 이동
    #                     for file_name in lines:
    #                         file_name = file_name.strip()
    #                         file_name = file_name.strip("\n")
    #                         # Park4139.get_cmd_output(cmd = f'dir /b /s "{file_name}"')
    #                         FileSystemUtil.get_cmd_output(cmd=f'dir /b /s "{file_name}" | clip.exe')
    #                         useless_files.append(clipboard.paste())
    #                     useless_files = "\n".join(useless_files)  # from [str] to str
    #                     useless_files = useless_files.replace("\n\n", "\n")  # from "\n\n" to "\n"
    #                     useless_files = useless_files.split("\n")  # from str to [str] (개행을 시킨)
    #                     useless_files = [x for x in useless_files if x.strip()]  # from [""] to []
    #                     # useless_files = [os.path.basename(x) for x in useless_files]  # from abspath to basename
    #                     useless_files = [x.replace("\r", "") for x in useless_files]  # from ["\r"] to [""]
    #                     useless_files = [x.replace("\t", "") for x in useless_files]  # from ["\t"] to [""]
    #                     clipboard.copy(tmp)
    #                     os.chdir(StateManagementUtil.PROJECT_DIRECTORY)
    #                     [print(sample) for sample in useless_files]
    #                     print(rf'len(useless_files) : {len(useless_files)}')
    #                     if len(useless_files) == 0:
    #                         ment = "약속된 불필요한 파일이 검출되지 않았습니다"
    #                         print(ment)
    #                         # mkmk 팝업 띄우기
    #                         TextToSpeechUtil.speak_ments(ment=ment, sleep_after_play=0.65)
    #                         break
    #                     for useless_file in useless_files:
    #                         FileSystemUtil.move_target_without_overwrite(target_abspath=useless_file, dst=useless_directories)  # mkmk
    #                         is_target_moved_done = True
    #                     if is_target_moved_done == True:
    #                         UiUtil.pop_up_as_complete(title_="작업성공보고", ment="타겟경로를 순회하며, 약속된 불필요한 파일을, 약속된 디렉토리로 모았습니다", auto_click_positive_btn_after_seconds=2)
    #                 except:
    #                     DebuggingUtil.trouble_shoot("%%%FOO%%%")
    #
    #                 # try:
    #                 #     useless_files = []
    #                 #     isSpoken = False
    #                 #     lines = Park4139.get_lines_of_file(Park4139.USELESS_FILES)
    #                 #     os.chdir(directory_abspath)
    #                 #     for file_name in lines:
    #                 #         file_name = file_name.strip()
    #                 #         file_name = file_name.strip("\n")
    #                 #         useless_files.append(file_name)
    #                 #     dst = rf"D:\[noe] [8TB] [ext]\$useless_directories"
    #                 #     Park4139.make_leaf_directory(dst)
    #                 #     # os.chmod(file_path, 0o777) # 0o777 소유자만 7의 권한부여, 나머지는 권한 없앰 # 777 누구나
    #                 #     os.chmod(directory_abspath, 0o777)
    #                 #     os.chmod(dst, 0o777)
    #                 #     Park4139.make_leaf_directory(dst)
    #                 #     if os.path.isdir(directory_abspath):
    #                 #         for root, dirs, files in os.walk(directory_abspath, topdown=False):  # os.walk()는 with walking 으로 동작한다
    #                 #             for file in files:
    #                 #                 file_path = os.path.join(root, file)
    #                 #                 for useless_file in useless_files:
    #                 #                     if useless_file in os.path.basename(useless_file):
    #                 #                         print(rf"useless file : {file_path}")
    #                 #                         os.chmod(file_path, 777)
    #                 #                         Park4139.move_without_overwrite(src=file_path, dst=dst)
    #                 #                         if isSpoken == False:
    #                 #                             TtsUtil.speak_ments("타겟경로를 순회하며 약속된 불필요한 파일을 약속된 폴더로 모았습니다", sleep_after_play=0.65)
    #                 #                             isSpoken = True
    #                 # except:
    #                 #     DebuggingUtil.trouble_shoot("%%%FOO%%%")
    #             elif btn_text_clicked == "모으지 않기":
    #                 break
    #     except:
    #         DebuggingUtil.trouble_shoot("%%%FOO%%%")
    #
    # @staticmethod
    # def should_i_gather_special_files():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     """분류되지 않은 하나의 폴더로 모음"""
    #     # while True:
    #     #     dialog = UiUtil.CustomQdialog(ment="타겟경로를 순회하며 특별한 파일을 약속된 폴더로 모을까요?", buttons=["모으기", "모으지 않기"], is_input_text_box=True)
    #     #     dialog.exec()
    #     #     btn_text_clicked = dialog.btn_text_clicked
    #     #     if btn_text_clicked == "모으기":
    #     #         target_abspath = dialog.input_box.text()
    #     #         target_abspath = target_abspath.strip()
    #     #         target_abspath = target_abspath.replace("\"", "")
    #     #         target_abspath = target_abspath.replace("\'", "")
    #     #         connected_drives = []
    #     #         for drive_letter in string.ascii_uppercase:
    #     #             drive_path = drive_letter + ":\\"
    #     #             if os.path.exists(drive_path):
    #     #                 connected_drives.append(drive_path)
    #     #                 if target_abspath == drive_path:
    #     #                     TextToSpeechUtil.speak_ments("입력된 타겟경로는 너무 광범위하여 진행할 수 없도록 설정되어 있습니다", sleep_after_play=0.65)
    #     #                     break
    #     #         if not os.path.exists(target_abspath):
    #     #             TextToSpeechUtil.speak_ments("입력된 타겟경로가 존재하지 않습니다", sleep_after_play=0.65)
    #     #             continue
    #     #         if target_abspath == "":
    #     #             continue
    #     #         special_files = [
    #     #             "[subplease]",
    #     #         ]
    #     #         dst = rf"D:\[noe] [8TB] [ext]\$special_files"
    #     #         FileSystemUtil.make_leaf_directory(dst)
    #     #         isSpoken = False
    #     #         if os.path.isdir(target_abspath):
    #     #             for root, dirs, files in os.walk(target_abspath, topdown=False):  # os.walk()는 with walking 으로 동작한다
    #     #                 for file in files:
    #     #                     file_path = os.path.join(root, file)
    #     #                     for special_file in special_files:
    #     #                         if special_file in os.path.basename(file_path):
    #     #                             # print(rf'file_path : {file_path}')
    #     #                             print(rf"special file : {file_path}")
    #     #                             FileSystemUtil.move_target_without_overwrite(target_abspath=file_path, dst=dst)
    #     #                             if isSpoken == False:
    #     #                                 TextToSpeechUtil.speak_ments("타겟경로를 순회하며 특별한 파일을 약속된 폴더로 모았습니다", sleep_after_play=0.65)
    #     #                                 isSpoken = True
    #     #     else:
    #     #         break
    #     pass

    # @staticmethod
    # def should_i_classify_special_files():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     """분류된 여러 폴더로 모음"""
    #     is_nested_loop_broken = False
    #     previous_abspath = ""
    #     while True:
    #         while True:
    #             dialog1 = UiUtil.CustomQdialog(ment="파일분류를 위해 순회할 디렉토리의 타겟경로를 입력해주세요", btns=["제출", "제출하지 않기"], is_input_box=True, input_box_text_default=previous_abspath)
    #             dialog1.exec()
    #             btn_text_clicked = dialog1.btn_text_clicked
    #             if btn_text_clicked == "제출하지 않기":
    #                 is_nested_loop_broken = True
    #                 break
    #             if btn_text_clicked == "제출":
    #                 target_abspath = dialog1.input_box.text()
    #                 target_abspath = target_abspath.strip()
    #                 target_abspath = target_abspath.replace("\"", "")
    #                 target_abspath = target_abspath.replace("\'", "")
    #                 previous_abspath = target_abspath
    #                 connected_drives = []
    #                 for drive_letter in string.ascii_uppercase:
    #                     drive_path = drive_letter + ":\\"
    #                     if os.path.exists(drive_path):
    #                         connected_drives.append(drive_path)
    #                         if target_abspath == drive_path:
    #                             TextToSpeechUtil.speak_ments("입력된 타겟경로는 너무 광범위하여 진행할 수 없도록 설정되어 있습니다", sleep_after_play=0.65)
    #                             is_nested_loop_broken = True
    #                             break
    #                 if not os.path.exists(target_abspath):
    #                     TextToSpeechUtil.speak_ments("입력된 타겟경로가 존재하지 않습니다", sleep_after_play=0.65)
    #                     continue
    #                 if target_abspath == "":
    #                     continue
    #                 special_dirs_promised = [
    #                     # "blahblahblah_boom_boom_boom",
    #                 ]
    #                 previous_keyword = clipboard.paste()
    #                 if previous_keyword == target_abspath:
    #                     previous_keyword = ""
    #
    #                 dialog2 = UiUtil.CustomQdialog(ment="분류할 파일에 포함되어 있을 키워드를 입력해주세요", btns=["제출", "제출하지 않기"], is_input_box=True, input_box_text_default=previous_keyword)
    #                 dialog2.exec()
    #                 btn_text_clicked2 = dialog2.btn_text_clicked
    #                 if btn_text_clicked2 == "제출":
    #                     file_path = dialog2.input_box.text()
    #                     file_path = file_path.strip()
    #                     if file_path == "":
    #                         TextToSpeechUtil.speak_ments("입력된 타겟경로가 존재하지 않습니다", sleep_after_play=0.65)
    #                         break
    #                     if "\n" in file_path:
    #                         file_paths = file_path.split("\n")
    #                         TextToSpeechUtil.speak_ments(f"{len(file_paths)}개의 키워드들이 입력되었습니다, 파일분류를 시작합니다", sleep_after_play=0.65)
    #                     else:
    #                         file_paths = [file_path]
    #                     for file_path in file_paths:
    #                         file_path = file_path.strip()
    #                         if file_path != "":
    #                             special_dirs_promised.append(file_path)
    #                         # destination_directory_of_files_classified = StateManagementUtil.SPECIAL_DIRECTORY # e: drive 에 저장
    #                         dst_dir_of_files_classified = rf"{target_abspath}\`"  # 입력된 타겟경로에 ` 폴더를 만들어 저장
    #                         for special_file in special_dirs_promised:
    #                             FileSystemUtil.make_leaf_directory(rf"{dst_dir_of_files_classified}/{special_file}")
    #                         # FileSystemUtil.make_leaf_directory(destination_directory_of_files_classified)
    #                         file_abspaths_searched = []
    #                         if os.path.isdir(target_abspath):
    #                             for root, dirs, files in os.walk(target_abspath, topdown=False):  # os.walk()는 with walking 으로 동작한다
    #                                 for file in files:
    #                                     file_abspath = os.path.join(root, file)
    #                                     for file_path in special_dirs_promised:
    #                                         if file_path in os.path.basename(file_abspath):
    #                                             file_abspaths_searched.append(file_abspath)
    #                         file_abspaths_searched_for_print = "\n".join(file_abspaths_searched)  # [str] to str with 개행
    #                         # dialog3 = UiUtil.CustomQdialog(ment=f"<검색된 파일 목록>\n\n검색된 파일 개수 : {len(file_abspaths_searched)}\n{file_abspaths_searched_for_print}\n\n 검색된 내용대로 계속진행을 할까요?", btns=[MentsUtil.YES, MentsUtil.NO] )
    #                         dialog3 = UiUtil.CustomQdialog(ment=f"<검색된 파일 목록>\n\n검색된 파일 개수 : {len(file_abspaths_searched)}\n{file_abspaths_searched_for_print}\n\n 검색된 내용대로 계속진행을 할까요?", btns=[MentsUtil.YES, MentsUtil.NO], auto_click_positive_btn_after_seconds=0)  # 자동화를 위해서 대체
    #                         dialog3.exec()
    #                         if dialog3.btn_text_clicked == MentsUtil.NO:
    #                             break
    #                         if dialog3.btn_text_clicked == MentsUtil.YES:
    #                             for index, special_dir in enumerate(special_dirs_promised):
    #                                 for file_abspath_searched in file_abspaths_searched:
    #                                     if special_dir in os.path.basename(file_abspath_searched):
    #                                         # UiUtil.pop_up_as_complete(title="디버깅", ment=f"index : {index} \n special_dir: {special_dir}", auto_click_positive_btn_after_seconds=600)
    #                                         FileSystemUtil.move_target_without_overwrite(target_abspath=file_abspath_searched, dst=rf"{dst_dir_of_files_classified}/{special_dirs_promised[index]}")
    #                         special_dirs_promised = []
    #         if is_nested_loop_broken == True:
    #             break

    @staticmethod
    def do_random_schedules():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        int_random = random.randint(0, 7)
        # Park4139Park4139.Tts.speak(f'랜덤숫자 {int_random} 나왔습니다')
        # mkmk
        if int_random == 0:
            pass
        elif int_random == 1:
            pass
        elif int_random == 2:
            pass
        elif int_random == 3:
            BusinessLogicUtil.change_console_color()
        elif int_random == 4:
            pass
        elif int_random == 5:
            pass
        elif int_random == 6:
            pass
        elif int_random == 7:
            pass

    # @staticmethod
    # def make_me_go_to_sleep():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     StateManagementUtil.count_of_make_me_go_to_sleep.append("!!")
    #     TextToSpeechUtil.speak_ments(ment=f'이제 그만 주무세요, 새벽 {TimeUtil.get_time_as_('%H')}입니다', sleep_after_play=0.95, thread_join_mode=True)
    #     TextToSpeechUtil.speak_ments(ment='건강을 위해서 자는 것이 좋습니다, 더 나은 삶을 위해 주무셔야 합니다', sleep_after_play=0.95, thread_join_mode=True)
    #     TextToSpeechUtil.speak_ments(ment='개발을 하고 있는 것이라면 지금 자고 나서, 일찍일어나 코드를 작성하는 것이 좋습니다', sleep_after_play=0.95, thread_join_mode=True)
    #     TextToSpeechUtil.speak_ments(ment='늦게까지 개발하고 일찍일어 나지 못할 것이라면요', sleep_after_play=0.95, thread_join_mode=True)
    #     # TextToSpeechUtil.speak_ments(ment='늦게까지 개발하고 못 일어날 거잖아요', sleep_after_play=0.95, thread_join_mode=True)
    #     TextToSpeechUtil.speak_ments(ment='내일 너무 피곤하지 않을까요, 하루를 망쳐버릴 수도 있어요', sleep_after_play=0.95, thread_join_mode=True)
    #     TextToSpeechUtil.speak_ments(ment='', sleep_after_play=0.95, thread_join_mode=True)
    #     if len(StateManagementUtil.count_of_make_me_go_to_sleep) == 10:
    #         TextToSpeechUtil.speak_ments(ment='아무래도 안되겠군요, 이제 하나씩 종료를 할겁니다', sleep_after_play=0.75, thread_join_mode=True)
    #
    #         TextToSpeechUtil.speak_ments(ment='팟플레이어를 종료합니다', sleep_after_play=0.95, thread_join_mode=True)
    #         # BusinessLogicUtil.should_i_do(ment="알송을 종료합니다", function=partial(BusinessLogicUtil.taskkill, 'ALSong.exe'), auto_click_positive_btn_after_seconds=5)
    #
    #         TextToSpeechUtil.speak_ments(ment='알송을 종료합니다', sleep_after_play=0.95, thread_join_mode=True)
    #         BusinessLogicUtil.should_i_do(ment="알송을 종료합니다", function=partial(BusinessLogicUtil.taskkill, 'ALSong.exe'), auto_click_positive_btn_after_seconds=5)
    #
    #         TextToSpeechUtil.speak_ments(ment='개발 도구를 종료합니다', sleep_after_play=0.95, thread_join_mode=True)
    #         BusinessLogicUtil.should_i_do(ment="개발 도구를 종료합니다", function=partial(BusinessLogicUtil.taskkill, 'ALSong.exe'), auto_click_negative_btn_after_seconds=20)
    #
    #         ment = "프로젝트 백업을 하고 최대절전모드로 전환해 드릴까요?"
    #         TextToSpeechUtil.speak_ments(ment=ment, sleep_after_play=0.95, thread_join_mode=True)
    #         BusinessLogicUtil.should_i_do(ment=ment, function=BusinessLogicUtil.back_up_project_and_change_to_power_saving_mode(), auto_click_negative_btn_after_seconds=20)

    @staticmethod
    def sleep(milliseconds=None, sec=None, min=None, hour=None, print_mode=True):
        # DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        params_list = [milliseconds, sec, min, hour]
        # 리스트 내 요소가 3개만 None 인지 확인
        if BusinessLogicUtil.get_none_count_of_list(params_list) == 3:
            # milliseconds = None, sec = None, min = None, hour = None, 이 중 하나만 None 이 아니면 동작
            time_to_sleep = None
            if milliseconds is not None:
                seconds = milliseconds / 1000
                if print_mode:
                    DebuggingUtil.print_ment_via_colorama(f"{inspect.currentframe().f_code.co_name}(ms={milliseconds})", colorama_color=ColoramaUtil.WHITE)
                time_to_sleep = seconds
            elif sec is not None:
                if print_mode:
                    DebuggingUtil.print_ment_via_colorama(f"{inspect.currentframe().f_code.co_name}(seconds={sec})", colorama_color=ColoramaUtil.WHITE)
                time_to_sleep = sec
            elif min is not None:
                if print_mode:
                    DebuggingUtil.print_ment_via_colorama(f"{inspect.currentframe().f_code.co_name}(minutes={min})", colorama_color=ColoramaUtil.WHITE)
                time_to_sleep = 60 * min
            elif hour is not None:
                if print_mode:
                    DebuggingUtil.print_ment_via_colorama(f"{inspect.currentframe().f_code.co_name}(hours={hour})", colorama_color=ColoramaUtil.WHITE)
                time_to_sleep = 60 * 60 * hour
            if time_to_sleep is not None:
                time.sleep(time_to_sleep)
        else:
            DebuggingUtil.print_ment_via_colorama(f'sleep() 메소드는 milliseconds, sec, min, hour 중 하나만 단위만 설정할 수 있습니다', colorama_color=ColoramaUtil.RED)

            # raise Error() 위의 코드말고 이런식으로 처리할까 싶은데

    @staticmethod
    def get_os_sys_environment_variable(environment_variable_name: str):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        DebuggingUtil.commentize("모든 시스템 환경변수 출력")
        for i in os.environ:
            DebuggingUtil.print_magenta(i)
        return os.environ.get(environment_variable_name)

    @staticmethod
    def update_os_sys_environment_variable(environment_variable_name: str, new_path: str):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        """시스템 환경변수 path 업데이트"""
        DebuggingUtil.commentize("테스트가 필요한 함수를 적용하였습니다")
        DebuggingUtil.commentize("기대한 결과가 나오지 않을 수 있습니다")
        DebuggingUtil.commentize("업데이트 전 시스템 환경변수")
        for i in sys.path:
            DebuggingUtil.print_magenta(i)
        sys.path.insert(0, new_path)
        sys.path.append(new_path)
        DebuggingUtil.commentize("업데이트 전 시스템 환경변수")
        for i in sys.path:
            DebuggingUtil.print_magenta(i)

    @staticmethod
    def get_name_space():  # name space # namespace # 파이썬 네임스페이스
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        dir()
        return dir()

    # @staticmethod
    # def back_up_target(target_abspath):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     try:
    #         while True:
    #             # 전처리
    #             target_abspath = target_abspath.replace("\n", "")
    #             target_abspath = target_abspath.replace("\"", "")
    #
    #             if target_abspath.strip() == "":
    #                 TextToSpeechUtil.speak_ments(ment="빽업할 대상이 입력되지 않았습니다", sleep_after_play=0.65)
    #                 break
    #
    #             target_dirname = os.path.dirname(target_abspath)
    #             target_dirname_dirname = os.path.dirname(target_dirname)
    #             target_basename = os.path.basename(target_abspath).split(".")[0]
    #             target_zip = rf'{target_dirname}\$zip_{target_basename}.zip'
    #             target_yyyy_mm_dd_hh_mm_ss_zip_basename = rf'{target_basename} - {TimeUtil.get_time_as_("%Y %m %d %H %M %S")}.zip'
    #             DebuggingUtil.print_magenta(f"target_abspath : {target_abspath}")
    #             DebuggingUtil.print_magenta(f"target_dirname : {target_dirname}")
    #             DebuggingUtil.print_magenta(f"target_dirname_dirname : {target_dirname_dirname}")
    #             DebuggingUtil.print_magenta(f"target_basename : {target_basename}")
    #             DebuggingUtil.print_magenta(f"target_zip : {target_zip}")
    #             DebuggingUtil.print_magenta(f"target_yyyy_mm_dd_HH_MM_SS_zip_basename : {target_yyyy_mm_dd_hh_mm_ss_zip_basename}")
    #
    #             DebuggingUtil.commentize(f"{target_zip} 로 빽업")
    #             cmd = f'bz.exe c "{target_zip}" "{target_abspath}"'
    #             FileSystemUtil.get_cmd_output(cmd)
    #
    #             DebuggingUtil.commentize(rf'{target_yyyy_mm_dd_hh_mm_ss_zip_basename} 로 이름변경')
    #             cmd = rf'ren "{target_zip}" "{target_yyyy_mm_dd_hh_mm_ss_zip_basename}"'
    #             FileSystemUtil.get_cmd_output(cmd)
    #
    #             # DebuggingUtil.commentize(f'현재 프로그램 pid 출력')
    #             # # os.system(rf'powershell (Get-WmiObject Win32_Process -Filter ProcessId=$PID).ParentProcessId')
    #
    #             DebuggingUtil.commentize(rf'현재 디렉토리에서 zip 확장자 파일만 문자열 리스트로 출력')
    #             # 파일이 위치한 드라이브로 이동
    #             drives = [
    #                 "C",
    #                 "D",
    #                 "E",
    #                 "F",
    #                 "G",
    #             ]
    #             drive_where_target_is_located = target_abspath.split(":")[0].upper()
    #             for drive in drives:
    #                 if (drive_where_target_is_located == drive):
    #                     os.system(rf"cd {drive}:")
    #             DebuggingUtil.commentize("target_dirname 로 이동")
    #             try:
    #                 os.chdir(target_dirname)
    #             except:
    #                 TextToSpeechUtil.speak_ments(ment="경로를 이해할 수 없습니다", sleep_after_play=0.65)
    #                 os.chdir(StateManagementUtil.PROJECT_DIRECTORY)
    #                 break
    #             lines = FileSystemUtil.get_cmd_output('dir /b /a-d *.zip')
    #             for line in lines:
    #                 if line != "":
    #                     if os.getcwd() != line:  # 여기 os.getcwd() 이게 들어가네... 나중에 수정하자
    #                         # 2023-12-04 월 12:14 SyntaxWarning: invalid escape sequence '\d'
    #                         # r 을 사용 Raw String(원시 문자열),  \를 모두 제거
    #                         # 정규식은 r 쓰면 안된다. \ 써야한다?.
    #                         # 2023-12-12 화 14:23 SyntaxWarning: invalid escape sequence '\d'
    #                         # 가상환경 재설치 후 또 문제가 나타남,
    #                         # regex = 'd{4} d{2} d{2} d{2} d{2} d{2}'
    #                         # regex = r'\d{4} \d{2} \d{2} \d{2} \d{2} \d{2}'
    #                         regex = r'd{4} d{2} d{2} d{2} d{2} d{2}'
    #                         # Park4139.debug_as_cli(line)
    #                         if BusinessLogicUtil.is_regex_in_contents(line, regex):
    #                             DebuggingUtil.commentize(f"zip 파일 목록에 대하여 {regex} 타임스탬프 정규식 테스트를 통과했습니다")
    #                             # Park4139.debug_as_cli(line)
    #                             # 2023-12-03 일 20:03 trouble shooting 성공
    #                             # 빽업 시 타임스탬프에 언더바 넣도록 변경했는데 regex 는 변경 하지 않아서 난 실수 있었음.
    #                             time_to_backed_up = re.findall(regex, line)
    #                             time_to_backed_up_ = time_to_backed_up[0][0:10].replace(" ", "-") + " " + time_to_backed_up[0][11:16].replace(" ", ":") + ".00"
    #                             time_to_backed_up__ = datetime.strptime(str(time_to_backed_up_), '%Y-%m-%d %H:%M.%S')
    #                             time_current = datetime.now()
    #                             try:
    #                                 target_dirname_old = rf'{target_dirname}\$cache_zip'
    #                                 if not os.path.exists(target_dirname_old):
    #                                     os.makedirs(target_dirname_old)
    #                             except Exception:
    #                                 DebuggingUtil.trouble_shoot("%%%FOO%%%")
    #                                 os.chdir(StateManagementUtil.PROJECT_DIRECTORY)
    #                                 break
    #                             # 지금부터 7일 이전의 파일만
    #                             # diff = time_to_backed_up__ - time_current
    #                             # if diff.days <-7:
    #                             # Park4139.debug_as_cli(f"line : {line}")
    #
    #                             # DebuggingUtil.commentize(f"1분(60 seconds) 이전의 파일자동정리 시도...")
    #                             DebuggingUtil.commentize(f"파일자동정리 시도...")
    #                             change_min = time_current - timedelta(seconds=60)
    #                             diff = time_to_backed_up__ - change_min
    #                             if 60 < diff.seconds:
    #                                 try:
    #                                     file_with_time_stamp_zip = os.path.abspath(line.strip())
    #                                     file_dirname_old_abspath = os.path.abspath(target_dirname_old)
    #                                     DebuggingUtil.print_magenta(rf'move "{file_with_time_stamp_zip}" "{file_dirname_old_abspath}"')
    #                                     shutil.move(file_with_time_stamp_zip, file_dirname_old_abspath)
    #                                 except Exception:
    #                                     DebuggingUtil.trouble_shoot("%%%FOO%%%")
    #                                     os.chdir(StateManagementUtil.PROJECT_DIRECTORY)
    #                                     break
    #             os.chdir(StateManagementUtil.PROJECT_DIRECTORY)
    #             UiUtil.pop_up_as_complete(title_="작업성공보고", ment=f"약속된 백업이 성공되었습니다\n{target_abspath}", auto_click_positive_btn_after_seconds=3)
    #             break
    #     except:
    #         DebuggingUtil.trouble_shoot("%%%FOO%%%")
    #         os.chdir(StateManagementUtil.PROJECT_DIRECTORY)

    @staticmethod
    def upzip_target(target_abspath):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        try:
            while True:
                # 전처리
                target_abspath = target_abspath.replace("\n", "")
                target_abspath = target_abspath.replace("\"", "")

                if target_abspath.strip() == "":
                    TextToSpeechUtil.speak_ments(ment="빽업할 대상이 입력되지 않았습니다", sleep_after_play=0.65)
                    break

                target_dirname = os.path.dirname(target_abspath)
                target_basename = os.path.basename(target_abspath).split(".")[0]
                target_zip = rf'{target_dirname}/{target_basename}.zip'

                os.chdir(target_dirname)

                if os.path.exists(target_zip):
                    # cmd = f'bandizip.exe bx "{target_zip}"'
                    cmd = f'bz.exe x -aoa "{target_zip}"'  # x 는 경로 보존, -aoa :Overwrite All existing files without prompt
                    FileSystemUtil.get_cmd_output(cmd)
                    if os.path.exists(target_abspath):
                        cmd = rf'echo y | del /f "{target_zip}"'
                        FileSystemUtil.get_cmd_output(cmd)
                    else:
                        DebuggingUtil.print_ment_fail("압축해제 후 압축파일을 삭제에 실패")
                else:
                    DebuggingUtil.print_ment_light_yellow("압축해제할 파일이 없었습니다")
                os.chdir(StateManagementUtil.PROJECT_DIRECTORY)
                DebuggingUtil.print_ment_success("압축해제 성공")
                break
        except:
            DebuggingUtil.trouble_shoot("%%%FOO%%%")
            os.chdir(StateManagementUtil.PROJECT_DIRECTORY)

    @staticmethod
    def back_up_target_without_timestamp(target_abspath):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        try:
            while True:
                # 전처리
                target_abspath = target_abspath.replace("\n", "")
                target_abspath = target_abspath.replace("\"", "")

                if target_abspath.strip() == "":
                    TextToSpeechUtil.speak_ments(ment="빽업할 대상이 입력되지 않았습니다", sleep_after_play=0.65)
                    break

                target_dirname = os.path.dirname(target_abspath)
                target_dirname_dirname = os.path.dirname(target_dirname)
                target_basename = os.path.basename(target_abspath).split(".")[0]
                special_target_zip = rf'{target_dirname}\$zip_{target_basename}.zip'
                target_abspath_zip = rf'{target_dirname}/{target_basename}.zip'
                target_basename_zip = rf'{target_basename}.zip'
                DebuggingUtil.print_magenta(f"target_abspath : {target_abspath}")
                DebuggingUtil.print_magenta(f"target_dirname : {target_dirname}")
                DebuggingUtil.print_magenta(f"target_dirname_dirname : {target_dirname_dirname}")
                DebuggingUtil.print_magenta(f"target_basename : {target_basename}")
                DebuggingUtil.print_magenta(f"special_target_zip : {special_target_zip}")

                cmd = f'bz.exe c "{special_target_zip}" "{target_abspath}"'
                FileSystemUtil.get_cmd_output(cmd)

                cmd = rf'ren "{special_target_zip}" "{target_basename_zip}"'
                FileSystemUtil.get_cmd_output(cmd)

                if os.path.exists(target_abspath_zip):
                    cmd = rf'echo y | rmdir /s "{target_abspath}"'
                    FileSystemUtil.get_cmd_output(cmd)
                else:
                    DebuggingUtil.print_ment_fail("압축에 실패")
                    break

                os.chdir(StateManagementUtil.PROJECT_DIRECTORY)
                DebuggingUtil.print_ment_success("압축에 성공")
                break
        except:
            DebuggingUtil.trouble_shoot("%%%FOO%%%")
            os.chdir(StateManagementUtil.PROJECT_DIRECTORY)

    @staticmethod
    def monitor_target_edited_and_back_up(target_abspath: str):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        try:
            if os.path.isfile(target_abspath):
                if FileSystemUtil.is_file_changed(target_abspath):
                    BusinessLogicUtil.back_up_target(target_abspath)
            elif os.path.isdir(target_abspath):
                if FileSystemUtil.is_directory_changed(target_abspath):
                    BusinessLogicUtil.back_up_target(target_abspath)
        except:
            DebuggingUtil.trouble_shoot("%%%FOO%%%")

    @staticmethod
    def monitor_target_edited_and_sync(target_abspath: str):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        try:
            if os.path.isfile(target_abspath):
                if FileSystemUtil.is_file_changed(target_abspath):
                    FileSystemUtil.sync_directory_local(target_abspath)
            elif os.path.isdir(target_abspath):
                FileSystemUtil.sync_directory_local(target_abspath)
        except:
            DebuggingUtil.trouble_shoot("%%%FOO%%%")

    @staticmethod
    def print_and_open_py_pkg_global_path():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        for path in sys.path:
            DebuggingUtil.print_magenta(path)
            if BusinessLogicUtil.is_regex_in_contents(target=path, regex='site-packages') == True:
                DebuggingUtil.print_magenta(rf'echo "{path}"')
                FileSystemUtil.explorer(path)

    @staticmethod
    def taskkill(program_img_name):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        program_img_name = program_img_name.replace("\'", "")
        program_img_name = program_img_name.replace("\"", "")
        FileSystemUtil.get_cmd_output(f'taskkill /f /im "{program_img_name}"')
        FileSystemUtil.get_cmd_output(f'wmic process where name="{program_img_name}" delete ')

    # 2023-12-07 목요일 16:51 최신화 함수
    @staticmethod
    def afterpause(function):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")

        def wrapper(*args, **kwargs):
            DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
            function(*args, **kwargs)
            TestUtil.pause()
            pass

        return wrapper

    @staticmethod
    def get_time_as_(pattern: str):
        #  %Y-%m-%d %H:%M:%S 의 형태도 모두 쓸 수 있고, 특정해둔 키워드도 쓸 수 있다
        # 어느날 보니 이 함수가 의도한대로 동작이 안됬다, 테스트를 해서 고쳤는데, for 문이 다른 값을 반환하는 것이 문제였다.
        # 테스트를 하지 않고 넘어 갔던 것이 문제였다. 꼭 함수를 만들었을 때는 함수에 대한 기능 테스트 하자
        # time 은 ms 지원않한다. 즉 %f 사용불가...
        # datetime 은 ms 지원!.  즉 %f 가능
        import time
        now = time
        localtime = now.localtime()
        from datetime import datetime
        time_styles = {
            'now': str(now.time()),
            'yyyy': str(localtime.tm_year),
            'MM': str(localtime.tm_mon),
            'dd': str(localtime.tm_mday),
            'HH': str(localtime.tm_hour),
            'mm': str(localtime.tm_min),
            'ss': str(localtime.tm_sec),
            'weekday': str(localtime.tm_wday),
            'elapsed_days_from_jan_01': str(localtime.tm_yday),  # 이거 테스트해보자 네이밍으로는 금년 1월 1일 부터 오늘까지 몇일이 지났는가를 출력하는 것 같아 보인다
        }
        for key, value in time_styles.items():
            if key == pattern:
                return time_styles[pattern]
        return str(datetime.now().strftime(pattern))

    @staticmethod
    def parse_youtube_video_id(url):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        """
        code written by stack overflow (none regex way)
        파이썬에 내장 url 구문 분석 기능
        this method return String or None
        Strng 을 리턴했다면 유튜브 비디오 아이디로 기대할 수 있다
        """
        query = urllib_parser.urlparse(url=url)
        if query.hostname == 'youtu.be':
            # DebuggingUtil.print_magenta()(f"query.path[1:] : {query.path[1:]}")
            return query.path[1:]
        if query.hostname in ('www.youtube.com', 'youtube.com'):
            # Park4139.debug_as_cli(query.scheme)
            # Park4139.debug_as_cli(query.netloc)
            # Park4139.debug_as_cli(query.hostname)
            # Park4139.debug_as_cli(query.port)
            # Park4139.debug_as_cli(query._replace(fragment="").geturl())
            # Park4139.debug_as_cli(query)
            # Park4139.debug_as_cli(query["v"][0])
            if query.path == '/watch':
                p = urllib_parser.parse_qs(query.query)
                # DebuggingUtil.print_magenta()(f"p['v'][0] : {p['v'][0]}")
                return p['v'][0]
            if query.path[:7] == '/embed/':
                # DebuggingUtil.print_magenta()(f"query.path.split('/')[2] : {query.path.split('/')[2]}")
                return query.path.split('/')[2]
            if query.path[:3] == '/v/':
                # DebuggingUtil.print_magenta()(f"query.path.split('/')[2] : {query.path.split('/')[2]}")
                return query.path.split('/')[2]

    # @staticmethod
    # def download_clip(url: str):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #         if url.strip() == "":
    #             DebuggingUtil.print_magenta(rf"다운로드할 url이 아닙니다 {url}")
    #             break
    #         # 다운로드가 안되면 주석 풀어 시도
    #         # os.system(rf'{StateManagementUtil.YT_DLP_CMD} -U')
    #
    #         # Park4139.raise_error('의도적으로 에러를 발생 중...')
    #         # if
    #         #     Park4139.debug_as_cli(rf'다운로드가 된 url 입니다 {url}')
    #         #
    #
    #         print('다운로드 옵션 파싱 중...')
    #         video_id = ''
    #         # lines = subprocess.check_output(rf'{StateManagementUtil.YT_DLP_CMD} -F {url}', shell=True).decode('utf-8').split("\n")
    #
    #         cmd = rf'{StateManagementUtil.YT_DLP_CMD} -F {url}'
    #         lines = FileSystemUtil.get_cmd_output(cmd=cmd)
    #         # 순서는 우선순위에 입각해 설정되었다. 순서를 바꾸어서는 안된다.
    #         video_ids_allowed = [
    #             '616',
    #             '315',
    #             '313',
    #             '303',
    #             '302',
    #             '308',
    #             '616',
    #             '248',
    #             '247',
    #             '244',
    #             '137',
    #             '136',
    #         ]
    #         audio_ids_allowed = [
    #             '250',
    #             '251',
    #         ]
    #         audio_id = ""
    #         for line in lines:  # mkmk
    #             if 'video only' in line or 'audio only' in line:
    #                 DebuggingUtil.print_magenta(line)
    #                 # video_id 설정
    #                 for id in video_ids_allowed:
    #                     if id in line:
    #                         video_id = id
    #                         if video_id.strip() == "":
    #                             DebuggingUtil.print_magenta(rf"다운로드 할 수 있는 video_id가 아닙니다 {video_id.strip()}")
    #                             break
    #                 # audio_id 설정
    #                 for id in audio_ids_allowed:
    #                     if id in line:
    #                         audio_id = id
    #                         if audio_id.strip() == "":
    #                             DebuggingUtil.print_magenta(rf"다운로드 할 수 있는 audio_id가 아닙니다 {audio_id.strip()}")
    #                             break
    #
    #         # 다운로드 가능 옵션 ID 설정
    #         # if video_id not in video_ids and audio_id not in audio_ids:
    #         #     video_id = str(input('video option:'))
    #         #     audio_id = str(input('audio option:'))
    #         #     speak(rf'다운로드 옵션이 선택되었습니다')
    #         #     Park4139.debug_as_cli(rf'video option: {video_id}  audio option: {audio_id}')
    #         #     speak(rf'video option: {video_id}  audio option: {audio_id}')
    #         # else:
    #         #     pass
    #
    #         # directories = ["storage"]
    #         # for directory in directories:
    #         #     if not os.path.isdir(rf'{os.getcwd()}/{directory}'):
    #         #         print(rf'storage 디렉토리 생성 중...')
    #         #         os.makedirs(rf'{directory}')
    #
    #         # 2023년 12월 12일 (화) 16:02:06
    #         # 다운로드의 최고 품질이 아닐 수 있다. 그래도
    #         # 차선책으로 두는 것이 낫겠다
    #         # cmd = rf'{StateManagementUtil.YT_DLP_CMD} -f best "{url}"'
    #
    #         # cmd = rf'{StateManagementUtil.YT_DLP_CMD} -f {video_id}+{audio_id} {url}' # 초기에 만든 선택적인 방식
    #         # cmd = rf'{StateManagementUtil.YT_DLP_CMD} -f "best[ext=webm]" {url}' # 마음에 안드는 결과
    #         cmd = rf'{StateManagementUtil.YT_DLP_CMD} -f "bestvideo[ext=webm]+bestaudio[ext=webm]" {url}'  # 지금 가장 마음에 드는 방법
    #         # cmd = rf'{StateManagementUtil.YT_DLP_CMD} -f "bestvideo[ext=webm]+bestaudio[ext=webm]/best[ext=webm]/best" {url}' # 아직 시도하지 않은 방법
    #         if video_id == "" or audio_id == "" == 1:
    #             # text = "다운로드를 진행할 수 없습니다\n다운로드용 video_id 와 audio_id를 설정 후\nurl을 다시 붙여넣어 다운로드를 다시 시도하세요\n{url}"
    #             print("불완전한 다운로드 명령어가 감지되었습니다....")
    #             TextToSpeechUtil.speak_ments(ment="불완전한 다운로드 명령어가 감지되었습니다", sleep_after_play=0.65)
    #             dialog = UiUtil.CustomQdialog(ment=f"에러코드[E004]\n아래의 비디오 아이디를 저장하고 에러코드를 관리자에게 문의해주세요\nvideo id: {url}", btns=["확인"], is_input_box=True, input_box_text_default=url)
    #             dialog.exec()
    #             print(cmd)
    #             break
    #
    #         try:
    #             lines = FileSystemUtil.get_cmd_output(cmd=cmd)
    #         except:
    #             print(cmd)
    #
    #         # storage = rf'{os.path.dirname(StateManagementUtil.PROJECT_DIRECTORY)}\storage'
    #         storage = rf"D:\[noe] [8TB] [ext]\`"
    #
    #         if not os.path.exists(storage):
    #             os.makedirs(storage)
    #
    #         print("다운로드 파일 이동 시도 중...")
    #         file = ""
    #         try:
    #             clip_id = BusinessLogicUtil.parse_youtube_video_id(url)
    #             if clip_id == None:
    #                 clip_id = url
    #
    #             lines = os.listdir()
    #             for line in lines:
    #                 if BusinessLogicUtil.is_regex_in_contents(str(line), str(clip_id)):
    #                     file = line
    #
    #             src = os.path.abspath(file)
    #             src_renamed = rf"{storage}/{os.path.basename(file)}"
    #
    #             DebuggingUtil.print_magenta(f'src_renamed : {src_renamed}')
    #             if src == os.getcwd():  # 여기 또 os.getcwd() 있는 부분 수정하자..
    #                 # E001 : 실행불가능한 명령어입력 감지
    #                 # S001 : 다운로드 가능한 video_id 와 audio_id 를 가용목록에 추가해주세요.
    #                 dialog = UiUtil.CustomQdialog(ment=f"에러코드[E001]\n아래의 비디오 아이디를 저장하고 에러코드를 관리자에게 문의해주세요\nvideo id: {url}", btns=["확인"], is_input_box=True, input_box_text_default=url)
    #                 dialog.exec()
    #                 DebuggingUtil.print_magenta("cmd")
    #                 DebuggingUtil.print_magenta(cmd)
    #                 break
    #             # shutil.move(src, storage)
    #             if src != os.getcwd():  # 여기 또 os.getcwd() 있는 부분 수정하자..
    #                 FileSystemUtil.move_target_without_overwrite(src, src_renamed)
    #
    #         except:
    #             DebuggingUtil.trouble_shoot("202312030009x")
    #         print(rf'다운로드 결과 확인 중...')
    #         try:
    #             src_moved = rf'{storage}/{file}'
    #
    #             # 4 줄 주석처리
    #             # dialog = UiUtil.CustomQdialog(ment="다운로드된 영상을 재생할까요?", btns=["재생하기", "재생하지 않기"], auto_click_negative_btn_after_seconds=3)
    #             # dialog.exec()
    #             # if dialog.btn_text_clicked == "재생하기":
    #             #     FileSystemUtil.explorer(target_abspath=src_moved)
    #
    #             # UiUtil.pop_up_as_complete(title="작업성공보고", ment=f"다운로드가 성공되었습니다\n{src_moved}", auto_click_positive_btn_after_seconds=2) # 성공 뜨는게 귀찮아서 주석처리함,
    #
    #         except Exception:
    #             DebuggingUtil.trouble_shoot("202312030013")
    #
    #         # 다운로드 로깅 처리
    #         # cmd = f'echo "{url}" >> success_yt_dlp.log'
    #         # FileSystemUtil.get_cmd_output(cmd=cmd)
    #         break
    #
    # @staticmethod
    # def download_clip_alt(url: str):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #         if url.strip() == "":
    #             DebuggingUtil.print_magenta(rf"다운로드할 url이 아닙니다 {url}")
    #             break
    #
    #         DebuggingUtil.print_magenta('다운로드 옵션 파싱 중...')
    #         video_id = ''
    #         cmd = rf'{StateManagementUtil.YT_DLP_CMD} -F {url}'
    #         lines = FileSystemUtil.get_cmd_output(cmd=cmd)
    #         video_ids_allowed = [
    #             '315',
    #             '313',
    #             '303',
    #             '302',
    #             '308',
    #             '616',
    #             '248',
    #             '247',
    #             '244',
    #             '137',
    #             '136',
    #         ]
    #         audio_ids_allowed = [
    #             '250',
    #             '251',
    #         ]
    #         audio_id = ""
    #         for line in lines:
    #             if 'video only' in line or 'audio only' in line:
    #                 DebuggingUtil.print_magenta(line)
    #                 # video_id 설정
    #                 for id in video_ids_allowed:
    #                     if id in line:
    #                         video_id = id
    #                         if video_id.strip() == "":
    #                             DebuggingUtil.print_magenta(rf"다운로드 할 수 있는 video_id가 아닙니다 {video_id.strip()}")
    #                             break
    #                 # audio_id 설정
    #                 for id in audio_ids_allowed:
    #                     if id in line:
    #                         audio_id = id
    #                         if audio_id.strip() == "":
    #                             DebuggingUtil.print_magenta(rf"다운로드 할 수 있는 audio_id가 아닙니다 {audio_id.strip()}")
    #                             break
    #         cmd = rf'{StateManagementUtil.YT_DLP_CMD} -f {video_id}+{audio_id} {url}'  # 초기에 만든 선택적인 방식
    #         if video_id == "" or audio_id == "" == 1:
    #             print("불완전한 다운로드 명령어가 감지되었습니다....")
    #             TextToSpeechUtil.speak_ments(ment="불완전한 다운로드 명령어가 감지되었습니다", sleep_after_play=0.65)
    #             dialog = UiUtil.CustomQdialog(ment=f"에러코드[E000]\n불완전한 다운로드 명령어가 감지되었습니다. 에러코드를 관리자에게 문의해주세요", btns=["확인"], is_input_box=True, input_box_text_default=url)
    #             dialog.exec()
    #             DebuggingUtil.print_magenta(cmd)
    #             break
    #
    #         print(rf'명령어 실행 중...')
    #         FileSystemUtil.get_cmd_output(cmd=cmd)
    #
    #         print(rf'storage 생성 중...')
    #         storage = rf'{os.path.dirname(StateManagementUtil.PROJECT_DIRECTORY)}\storage'
    #         if not os.path.exists(storage):
    #             os.makedirs(storage)
    #
    #         print("다운로드 파일 이동 시도 중...")
    #         file = ""
    #         try:
    #             clip_id = BusinessLogicUtil.parse_youtube_video_id(url)
    #             if clip_id == None:
    #                 clip_id = url
    #
    #             lines = os.listdir()
    #             for line in lines:
    #                 if BusinessLogicUtil.is_regex_in_contents(str(line), str(clip_id)):
    #                     file = line
    #
    #             src = os.path.abspath(file)
    #             dst_ = rf"{storage}/{os.path.basename(file)}"
    #             DebuggingUtil.print_magenta(f'src : {src}')
    #
    #             DebuggingUtil.print_magenta(f'dst_ : {dst_}')
    #             if src == os.getcwd():  # 여기 또 os.getcwd() 있는 부분 수정하자..
    #                 print("실행불가능한 명령어가 실행되어 다운로드 할 수 없었습니다")
    #                 TextToSpeechUtil.speak_ments(ment="실행불가능한 명령어가 실행되어 다운로드 할 수 없었습니다", sleep_after_play=0.65)
    #                 dialog = UiUtil.CustomQdialog(ment=f"에러코드[E003]\n아래의 비디오 아이디를 저장하고 에러코드를 관리자에게 문의해주세요\nvideo id: {url}", btns=["확인"], is_input_box=True, input_box_text_default=url)
    #                 dialog.exec()
    #                 DebuggingUtil.print_magenta(f"{StateManagementUtil.UNDERLINE_PROMISED}다운로드 가능한 video_id 와 audio_id 를 가용목록에 추가해주세요")
    #                 DebuggingUtil.print_magenta(cmd)
    #                 break
    #
    #             if src != os.getcwd():  # 여기 또 os.getcwd() 있는 부분 수정하자..
    #                 FileSystemUtil.move_target_without_overwrite(src, dst_)
    #         except:
    #             DebuggingUtil.trouble_shoot("202312030009")
    #             # break
    #
    #         print(rf'다운로드 결과 확인 중...')
    #         try:
    #             src_moved = rf'{storage}/{file}'
    #             # UiUtil.pop_up_as_complete(title_="작업성공보고", ment=f"{src_moved} 가 다운로드 되었습니다", auto_click_positive_btn_after_seconds=5)
    #             # 다운로드된 영상을 재생할까요? 모드
    #             dialog = UiUtil.CustomQdialog(ment=f"다운로드된 영상을 재생할까요?{src_moved}", btns=[MentsUtil.YES], is_input_box=True)
    #             dialog.exec()
    #             if dialog.input_box.text() == MentsUtil.YES:
    #                 TextToSpeechUtil.speak_ment(ment="다운로드된 영상을 재생합니다", sleep_after_play=0.65)
    #                 FileSystemUtil.get_cmd_output(cmd=rf'explorer "{src_moved}"')
    #             break  # download clip alt() 에서는 break 해야할듯.
    #
    #         except Exception:
    #             DebuggingUtil.trouble_shoot("202312030013")
    #
    #         print(rf'다운로드 결과 로깅 중...')
    #         FileSystemUtil.get_cmd_output(cmd=f'echo "{url}" >> success_yt_dlp.log')
    #         break

    # @staticmethod
    # def download_from_youtube_to_webm(urls):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #         urls = str(urls).strip()
    #         if urls == None:
    #             TextToSpeechUtil.speak_ments(ment="다운로드할 대상 목록에 아무것도 입력되지 않았습니다", sleep_after_play=0.65)
    #             break
    #         if urls == "None":
    #             TextToSpeechUtil.speak_ments(ment="다운로드할 대상 목록에 이상한 것이 입력되었습니다", sleep_after_play=0.65)
    #             break
    #
    #         if "\n" in urls:
    #             urls = urls.split("\n")
    #         else:
    #             urls = [urls]
    #
    #         urls = [x for x in urls if x.strip("\n")]  # 리스트 요소 "" 제거,  from ["", A] to [A]       [""] to []
    #         UiUtil.pop_up_as_complete(title_="작업중간보고", ment=f"{len(urls)} 개의 url이 입력되었습니다", auto_click_positive_btn_after_seconds=1)
    #
    #         try:
    #             urls.append(sys.argv[1])
    #         except IndexError:
    #             pass
    #         except Exception:
    #             DebuggingUtil.trouble_shoot("202312071455")
    #             pass
    #
    #         # urls = list(set(urls)) # urls 중복제거(orderless way)    # remove duplicated elements of list ( orderless way ) # 파이썬 3.7 부터는 ordered 라는 것 같은데 명시적인 방법이 아닌것 같다..
    #
    #         # urls 중복제거(ordered way) # remove duplicated elements of list ( ordered way )
    #         urls_removed_duplicated_element: [str] = []
    #         for url in urls:
    #             if url not in urls_removed_duplicated_element:
    #                 if url is not None:
    #                     # if url is not "None":
    #                     urls_removed_duplicated_element.append(url)
    #         urls = urls_removed_duplicated_element
    #         DebuggingUtil.print_magenta(f"len(urls) : {len(urls)}")
    #         for i in urls:
    #             DebuggingUtil.print_magenta(i)
    #
    #         # DebuggingUtil.commentize(f'다운로드 할게 없으면 LOOP break')
    #         if len(urls) == 0:
    #             UiUtil.pop_up_as_complete(title_="작업성공보고", ment=f"다운로드할 대상이 없습니다", auto_click_positive_btn_after_seconds=5)
    #             # TtsUtil.speak_ments(ment="다운로드할 대상이 없습니다", sleep_after_play=0.65)
    #             break
    #         if len(urls) != 1:
    #             TextToSpeechUtil.speak_ments(f"{str(len(urls))}개의 다운로드 대상이 확인되었습니다", sleep_after_play=0.65)
    #         for url in urls:
    #             url = url.strip()  # url에 공백이 있어도 다운로드 가능하도록 설정
    #             if '&list=' in url:
    #                 DebuggingUtil.commentize(f' clips mode')
    #                 clips = Playlist(url)  # 이걸로도 parsing 기능 수행 생각 중
    #                 DebuggingUtil.print_magenta(f"predicted clips cnt : {len(clips.video_urls)}")
    #                 TextToSpeechUtil.speak_ments(ment=f"{len(clips.video_urls)}개의 다운로드 목록이 확인되었습니다", sleep_after_play=0.65)
    #                 # os.system(f'echo "여기서부터 비디오 리스트 시작 {url}" >> success_yt_dlp.log')
    #                 for clip in clips.video_urls:
    #                     try:
    #                         BusinessLogicUtil.download_clip(clip)
    #                     except Exception:
    #                         DebuggingUtil.trouble_shoot("%%%FOO%%%")
    #                         continue
    #                 # os.system(f'echo "여기서부터 비디오 리스트 종료 {url}" >> success_yt_dlp.log')
    #             else:
    #                 if BusinessLogicUtil.parse_youtube_video_id(url) != None:
    #                     DebuggingUtil.commentize(f'{StateManagementUtil.UNDERLINE_PROMISED} youtube video id parsing mode')
    #                     try:
    #                         BusinessLogicUtil.download_clip(f'https://www.youtube.com/watch?v={BusinessLogicUtil.parse_youtube_video_id(url)}')
    #                     except Exception:
    #                         DebuggingUtil.trouble_shoot("%%%FOO%%%")
    #                         continue
    #                 else:
    #                     DebuggingUtil.commentize(f'{StateManagementUtil.UNDERLINE_PROMISED} experimental mode')
    #                     print("???????????????????")
    #                     try:
    #                         url_parts_useless = [
    #                             "https://youtu.be/",
    #                             "https://www.youtube.com/shorts/",
    #                         ]
    #                         try:
    #                             for index, useless_str in enumerate(url_parts_useless):
    #                                 if useless_str in url:
    #                                     print(rf'url.split(useless_str)[1] : {url.split(useless_str)[1]}')
    #                                     BusinessLogicUtil.download_clip(url=url.split(useless_str)[1])
    #                         except Exception:
    #                             BusinessLogicUtil.download_clip(url)
    #                             DebuggingUtil.trouble_shoot("%%%FOO%%%")
    #                     except Exception:
    #                         DebuggingUtil.trouble_shoot("%%%FOO%%%")
    #                     continue
    #         break

    # @staticmethod
    # def get_display_setting():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     height = ''
    #     width = ''
    #     for monitor_info in get_monitors():
    #         for info in str(monitor_info).split(','):
    #             if 'width=' in info.strip():
    #                 width = info.split('=')[1]
    #             elif 'height=' in info.strip():
    #                 height = info.split('=')[1]
    #     display_setting = {
    #         'height': int(height),
    #         'width': int(width)
    #     }
    #     return display_setting

    # deprecated method by Park4139
    # def print_police_line(police_line_ment):

    #     police_line = ''
    #     for i in range(0, 255 // len(police_line_ment)):
    #         police_line = police_line + f'{police_line_ment} '
    #     Park4139.debug_as_cli(f'{police_line.upper()}')
    @staticmethod
    def is_regex_in_contents(target, regex):
        # DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        pattern = re.compile(regex)
        m = pattern.search(target)
        if m:
            # DebuggingUtil.commentize("function name   here here here")
            # Park4139.debug_as_cli(rf"contents: {contents}")
            # Park4139.debug_as_cli(rf"regex: {regex}")
            # Park4139.debug_as_cli(rf"True")
            return True
        else:
            # Park4139.debug_as_cli(rf"contents: {contents}")
            # Park4139.debug_as_cli(rf"regex: {regex}")
            # Park4139.debug_as_cli(rf"False")
            return False

    @staticmethod
    def is_regex_in_contents_with_case_ignored(contents, regex):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        pattern = re.compile(regex, re.IGNORECASE)
        m = pattern.search(contents)
        if m:
            return True
        else:
            return False

    @staticmethod
    # 출력의 결과는 tasklist /svc 와 유사하다.
    def print_python_process_for_killing_zombie_process():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        import psutil  # 실행중인 프로세스 및 시스템 활용 라이브러리
        for process in psutil.process_iter():
            print(rf'str(process.pid) : {str(process.pid)}')
            print(rf'process.status() : {process.status()}')
            print(rf'process.name() : {process.name()}')

    @staticmethod
    def toogle_console_color(color_bg, colors_texts):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        to_right_nbsp = ''
        to_right_nbsp_cnt = 150
        for i in range(0, to_right_nbsp_cnt):
            to_right_nbsp = to_right_nbsp + ' '
        for color_text in colors_texts:
            if color_bg != color_text:
                os.system(f"color {color_bg}{color_text}")

    @staticmethod
    def recommand_console_color():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        colors = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', 'a', 'b', 'c', 'd', 'e', 'f']
        while True:
            try:
                for color_bg in colors:
                    for color_text in colors:
                        if color_bg != color_text:
                            os.system('cls')
                            for setting_key, setting_value in FileSystemUtil.get_display_info().items():
                                pass
                                # Park4139.debug_as_cli(f'setting_key: {setting_key}  ,setting_value: {setting_value}  ')
                            # Park4139.debug_as_cli(f"color {color_bg}{color_text}")
                            for i in range(0, 32):
                                DebuggingUtil.print_magenta('')
                            to_right_nbsp = ''
                            for i in range(0, 150):
                                to_right_nbsp = to_right_nbsp + ' '
                            DebuggingUtil.print_magenta(f"{to_right_nbsp}color {color_bg}{color_text}")
                            for i in range(0, 32):
                                DebuggingUtil.print_magenta('')
                            os.system(f"color {color_bg}{color_text}")
                            import clipboard
                            clipboard.copy(f'color {color_bg}{color_text}')
                            TestUtil.pause()

            except Exception as e:
                DebuggingUtil.trouble_shoot("202312071431")
                # ctrl c 가 입력이 제대로 되지 않는 현상이 있어 ctrl c 로 콘솔을 종료하는데 불편...이는 어떻게 해결하지? 일단 코드 반응속도는 마음에 드는데...

    @staticmethod
    def make_matrix_console():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        os.system('color 0A')
        os.system('color 02')
        while True:
            lines = subprocess.check_output('dir /b /s /o /a-d', shell=True).decode('utf-8').split("\n")
            for line in lines:
                if "" != line:
                    if os.getcwd() != line:
                        DebuggingUtil.print_magenta(lines)
            time.sleep(60)

    @staticmethod
    def make_party_console():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        commands = [
            'color 03',
            'color 09',
            'color 4F',
            'color cF',
            'color bf',
            'color 10',
            'color 1f',
            'color 08',
            'color f0',
            'color f8',
            'color 4c',
            'color c4',
            'color 09',
            'color 0a',
            'color d4',
            'color a4',
            'color 4a',
            'color 51',
            'color 48',
            'color 4c',
            'color 5d',
            'color 3b',
            'color e6',
            'color 07',
            'color 3F',
            'color 13',
            'color a2',
            'color 2a',
            'color 01',
            'color 02',
            'color 03',
            'color 04',
            'color 05',
            'color 06',
            'color 8f',
            'color 9b',
            'color 4a',
            'color da',
            'color ad',
            'color a4',
            'color b2',
            'color 0c',
            'color 0d',
            'color e3',
            'color eb',
            'color e7',
            'color 0f',
        ]
        while (True):
            for command in commands:
                os.system(f'{command}')
                # os.system('echo "202312031429" && pause')

    @staticmethod
    # 이 메소드를 만들면서 권한을 얻는 여러가지 방법을 stack over flow 를 따라 시도해보았으나 적다한 해결책을 찾지 못함. pyautogui 로 시도 방법은 남아있으나
    # 일단은 regacy 한 방법으로 임시로 해결해두었다.
    def update_global_pkg_park4139_for_linux():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        local_pkg = rf"{StateManagementUtil.PROJECT_DIRECTORY}\pkg_park4139_for_linux"
        global_pkg = rf"C:\Python312\Lib\site-packages\pkg_park4139_for_linux"
        updateignore_txt = rf"{StateManagementUtil.PROJECT_DIRECTORY}\pkg_park4139_for_linux\updateignore.txt"
        try:
            if os.path.exists(global_pkg):
                # 삭제시도
                # shutil.rmtree(global_pkg)

                # 삭제시도
                # for file in os.scandir(global_pkg):
                # os.remove(file.path)

                # 덮어쓰기
                # src= local_pkg
                # dst =os.path.dirname(global_pkg)
                # os.system(f"echo y | copy {src} {dst}")
                # shutil.copytree(local_pkg, os.path.dirname(global_pkg))
                cmd = f'echo y | xcopy "{local_pkg}" "{global_pkg}" /k /e /h /exclude:{updateignore_txt} >nul'
                os.system(cmd)

                # 디버깅
                # Park4139TestUtil.pause()
                DebuggingUtil.print_magenta(f'{cmd}')
                DebuggingUtil.print_magenta(f"{StateManagementUtil.UNDERLINE_PROMISED}")
                return "REPLACED global pkg_park4139_for_linux AS local_pkg"
            else:
                return "pkg_park4139_for_linux NOT FOUND AT GLOBAL LOCATION"

        except Exception as e:
            DebuggingUtil.trouble_shoot("202312030016")

    @staticmethod
    def merge_video_and_sound(file_v_abspath, file_a_abspath):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        DebuggingUtil.commentize(f'다운로드 디렉토리 생성')
        directories = ["storage"]
        for directory in directories:
            if not os.path.isdir(rf'./{directory}'):
                os.makedirs(rf'{directory}')

        DebuggingUtil.commentize(rf'yotube 에서 고해상도 음성 없는 영상과 음성을 받아 하나의 영상으로 merge')
        DebuggingUtil.commentize(f'비디오 파일, 음성 파일 절대주소 get')
        dst = rf'{StateManagementUtil.PROJECT_DIRECTORY}\storage'
        paths = [os.path.abspath(dst), os.path.basename(file_v_abspath)]
        file_va = os.path.join(*paths)
        DebuggingUtil.print_magenta(rf'file_v_abspath : {file_v_abspath}')
        DebuggingUtil.print_magenta(rf'file_a_abspath : {file_a_abspath}')
        DebuggingUtil.print_magenta(rf'file_va : {file_va}')

        DebuggingUtil.commentize(f'ffmpeg.exe 위치 설정')
        location_ffmpeg = rf"{StateManagementUtil.USERPROFILE}\Desktop\`workspace\tool\LosslessCut-win-x64\resources\ffmpeg.exe"
        trouble_characters = ['Ä']
        trouble_characters_alternatives = {'Ä': 'A'}
        for trouble_character in trouble_characters:
            file_v_abspath = file_v_abspath.replace(trouble_character, trouble_characters_alternatives[trouble_character])
            file_a_abspath = file_a_abspath.replace(trouble_character, trouble_characters_alternatives[trouble_character])
            file_va = file_va.replace(trouble_character, trouble_characters_alternatives[trouble_character])
            DebuggingUtil.commentize(f'파일명 변경 시도')
            try:
                if trouble_character in file_va:
                    os.rename(file_v_abspath,
                              file_v_abspath.replace(trouble_character, trouble_characters_alternatives[trouble_character]))
                    os.rename(file_a_abspath,
                              file_a_abspath.replace(trouble_character, trouble_characters_alternatives[trouble_character]))
            except Exception as e:
                DebuggingUtil.trouble_shoot("202312030017")

        DebuggingUtil.commentize(f' 파일머지 시도')
        try:
            DebuggingUtil.print_magenta(rf'echo y | "{location_ffmpeg}" -i "{file_v_abspath}" -i "{file_a_abspath}" -c copy "{file_va}"')
            lines = subprocess.check_output(
                rf'echo y | "{location_ffmpeg}" -i "{file_v_abspath}" -i "{file_a_abspath}" -c copy "{file_va}"', shell=True).decode(
                'utf-8').split("\n")
            for line in lines:
                DebuggingUtil.print_magenta(line)
        except Exception as e:
            DebuggingUtil.trouble_shoot("202312030018")

        DebuggingUtil.print_magenta(rf'다운로드 및 merge 결과 확인 시도')
        try:
            DebuggingUtil.print_magenta(rf'explorer "{file_va}"')
            subprocess.check_output(rf'explorer "{file_va}"', shell=True).decode('utf-8').split("\n")
        except Exception as e:
            DebuggingUtil.trouble_shoot("202312030019a")

        DebuggingUtil.commentize(f' 불필요 리소스 삭제 시도')
        try:
            if os.path.exists(file_va):
                subprocess.check_output(rf'echo y | del /f "{file_v_abspath}"', shell=True).decode('utf-8').split("\n")
                lines = subprocess.check_output(rf'echo y | del /f "{file_a_abspath}"', shell=True).decode('utf-8').split("\n")
                for line in lines:
                    DebuggingUtil.print_magenta(line)
        except Exception as e:
            DebuggingUtil.trouble_shoot("202312030020")

    # @staticmethod
    # def elapsed(function):

    #     def wrapper(*args, **kwargs):

    #         import time
    #         time_s = time.time()
    #         function(*args, **kwargs)
    #         time_e = time.time()
    #         mesured_time = time_e - time_s
    #         Park4139.debug_as_cli(f'측정시간은 {mesured_time} 입니다')
    #         DebuggingUtil.commentize(f'측정시간은 {mesured_time} 입니다')
    #     return wrapper

    @staticmethod
    def cls():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        import os
        os.system('cls')

    @staticmethod
    def replace_with_auto_no(contents: str, unique_word: str, auto_cnt_starting_no=0):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        """

        input   = "-----1-----1----1------"
        Output  = "-----1-----2----3------"

        input
        --------
        -----1--
        ---1----
        ---1----
        --------
         ouput
        --------
        -----1--
        ---2----
        ---3----
        --------

        """
        tmp = []
        for index, element in enumerate(contents.split(unique_word)):
            if index != len(contents.split(unique_word)) - 1:
                tmp.append(element + str(auto_cnt_starting_no))
                auto_cnt_starting_no = auto_cnt_starting_no + 1
            else:
                tmp.append(element)
        lines_new_as_str = "".join(tmp)
        return lines_new_as_str

    @staticmethod
    def replace_with_auto_no_orderless(contents: str, unique_word: str, auto_cnt_starting_no=0):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        # DebuggingUtil.commentize("항상 필요했던 부분인데 만들었다. 편하게 개발하자. //웹 서비스 형태로 아무때서나 접근이되면 더 좋을 것 같다.  웹 개발툴 을 만들어 보자")
        before = unique_word
        after = 0 + auto_cnt_starting_no
        contents_new = []
        # lines = contents.split("\n")
        lines = contents.strip().split("\n")  # 문제 없긴 했는데,  어떻게 되나 실험해보자 안되면 위의 코드로 주석 스와핑할것.
        for line in lines:
            # Park4139.debug_as_cli(line)
            # Park4139.debug_as_cli(before)
            # Park4139.debug_as_cli(str(after))
            after = after + 1

            line_new = re.sub(str(before), str(after), str(line))
            # Park4139.debug_as_cli(line_new)
            contents_new.append(line_new)

        # DebuggingUtil.commentize("str list to str")
        delimiter = "\n"
        contents_new_as_str = delimiter.join(contents_new)
        return contents_new_as_str

    # @staticmethod
    # def move_with_overwrite(src: str, dst: str):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     try:
    #         # 목적지에 있는 중복타겟 삭제
    #         os.remove(dst)
    #     except FileNotFoundError as e:
    #         pass
    #     except Exception:
    #         DebuggingUtil.trouble_shoot("%%%FOO%%%")
    #     try:
    #         # 목적지로 타겟 이동
    #         os.rename(src, dst)
    #     except Exception:
    #         DebuggingUtil.trouble_shoot("%%%FOO%%%")

    @staticmethod
    def raise_error(str: str):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        raise shutil.Error(str)

    # @staticmethod
    # def git_push_by_auto():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     GIT_HUB_ADDRESS = StateManagementUtil.GIT_HUB_ADDRESS
    #     GIT_HUB_REPOSITORY_URL = rf"{GIT_HUB_ADDRESS}/archive_py"
    #     while True:
    #         try:
    #             lines = FileSystemUtil.get_cmd_output(cmd='git add * ')
    #             if lines == subprocess.CalledProcessError:
    #                 UiUtil.pop_up_as_complete(title_="테스트보고", ment="혹시 여기 발생됩니까? 안되길 바랍니다. - 2023 01 28 -", auto_click_positive_btn_after_seconds=5)
    #                 break
    #             if "" in lines:
    #                 DebuggingUtil.print_ment_success(f"add success")
    #                 commit_ment = f"refer to README.md ( pushed at {TimeUtil.get_time_as_('%Y-%m-%d %H:%M:%S')}"
    #                 lines = FileSystemUtil.get_cmd_output(cmd=rf'git commit -m "{commit_ment}"')
    #                 lines = FileSystemUtil.get_cmd_output(cmd='git status | findstr "nothing to commit, working tree clean"')
    #                 if "nothing to commit, working tree clean" in lines:
    #                     DebuggingUtil.print_ment_success(f"commit success")
    #                     lines = FileSystemUtil.get_cmd_output(cmd='git push -u origin main')
    #                     if "Everything up-to-date" or "branch 'main' set up to track 'origin/main'." in lines:
    #                         # os.system('color df')  # OPERATION
    #                         DebuggingUtil.print_ment_success(f"push success")
    #                         if int('08') <= int(TimeUtil.get_time_as_('%H')) <= int('23'):  # 자는데 하도 시끄러워서 추가한 코드
    #                             # TextToSpeechUtil.speak_ments("깃허브에 프로젝트 푸쉬를 성공했습니다", sleep_after_play=0.65)
    #                             UiUtil.pop_up_as_complete(title_="작업성공보고", ment="깃허브에 프로젝트 푸쉬를 성공했습니다", auto_click_positive_btn_after_seconds=5)
    #                             BusinessLogicUtil.should_i_do(ment="깃허브에서 직접확인하시겠요요?", function=partial(FileSystemUtil.explorer, GIT_HUB_REPOSITORY_URL), auto_click_negative_btn_after_seconds=5)
    #                             break
    #                         break
    #                     else:
    #                         DebuggingUtil.print_ment_fail(f"push fail")
    #                 else:
    #                     DebuggingUtil.print_ment_fail(f"commit fail")
    #             else:
    #                 DebuggingUtil.print_ment_fail(f"add fail")
    #         except:
    #             UiUtil.pop_up_as_complete(title_="작업성공보고", ment="깃허브에 프로젝트 푸쉬를 시도 중 에러가 발생하였습니다'", auto_click_positive_btn_after_seconds=5)
    #             break

    @staticmethod
    def save_all_list():  # 수정 필요, 모든 파일 디렉토리 말고, 특정디렉토리로 수정한는 것도 할것
        """모든 파일 디렉토리에 대한 정보를 텍스트 파일로 저장하는 함수"""
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")

        # 윈도우냐 아니야에 따라
        # os.system('chcp 65001 >nul')
        # os.system('export LANG=en_US.UTF-8 >nul')

        opening_directory = os.getcwd()
        proper_tree_txt = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_all_tree\proper_tree.txt"
        all_tree_txt = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_all_tree\all_tree.txt"
        if not os.path.exists(os.path.dirname(all_tree_txt)):
            os.makedirs(os.path.dirname(all_tree_txt))
            # os.system(f'echo. >> "{all_tree_txt}"')
            os.system(f'echo. >> "{all_tree_txt}" >nul')
            os.system(f'echo. >> "{proper_tree_txt}" >nul')
        with open(all_tree_txt, 'w', encoding="utf-8") as f:
            f.write(" ")
        with open(proper_tree_txt, 'w', encoding="utf-8") as f:
            f.write(" ")

        drives = "foo"

        file_cnt = 0
        f = open(StateManagementUtil.PROJECT_DIRECTORY + '\\all_list.txt', 'a', encoding="utf-8")  # >>  a    > w   각각 대응됨.
        for drive in drives:
            os.chdir(drive)
            for dirpath, subdirs, files in os.walk(os.getcwd()):  # 여기 또 os.getcwd() 있는 부분 수정하자..
                for file in files:
                    file_cnt = file_cnt + 1
                    f.write(str(file_cnt) + " " + os.path.join(dirpath, file) + "\n")
        f.close()  # close() 를 사용하지 않으려면 with 문을 사용하는 방법도 있다.
        DebuggingUtil.print_magenta(f"{StateManagementUtil.UNDERLINE_PROMISED}{StateManagementUtil.UNDERLINE_PROMISED} all_list.txt writing e")
        DebuggingUtil.print_magenta(f"{StateManagementUtil.UNDERLINE_PROMISED}{StateManagementUtil.UNDERLINE_PROMISED} all_list_proper.txt rewriting s")
        texts_black = [
            "C:\\$WinREAgent",
            "C:\\mingw64",
            "C:\\PerfLogs",
            "C:\\Program Files (x86)",
            "C:\\Program Files",
            "C:\\ProgramData",
            "C:\\Temp",
            "C:\\Users\\All Users",
            "C:\\Windows\\servicing",
            "C:\\Windows\\SystemResources",
            "C:\\Windows\\WinSxS",
            "C:\\Users\\Default",
            "C:\\Users\\Public",
            "C:\\Windows.old",
            "C:\\Windows",
            "C:\\$Recycle.Bin",
            "D:\\$RECYCLE.BIN",
            "E:\\$RECYCLE.BIN",
            "E:\\$Recycle.Bin",
            "F:\\$RECYCLE.BIN",
            rf"{StateManagementUtil.USERPROFILE}\\AppData",
        ]
        texts_white = [
            ".mkv",
        ]
        f = open(rf'{StateManagementUtil.PROJECT_DIRECTORY}\all_list.txt', 'r+', encoding="utf-8")
        f2 = open(rf'{StateManagementUtil.PROJECT_DIRECTORY}\all_list_proper.txt', 'a', encoding="utf-8")
        lines_cnt = 0
        while True:
            line = f.readline()
            if not line:
                break
            lines_cnt = lines_cnt + 1
            if any(text_black not in line for text_black in texts_black):
                # Park4139.debug_as_cli(line)
                if any(text_white in line for text_white in texts_white):
                    # Park4139.debug_as_cli(line.split("\n")[0] + " o")
                    f2.write(line.split("\n")[0] + " o " + "\n")
                    # Park4139.debug_as_cli('o')
                    pass
                else:
                    # Park4139.debug_as_cli(line.split("\n")[0] + " x")
                    # f2.write(line.split("\n")[0] + " x "+"\n")
                    # Park4139.debug_as_cli('x')
                    pass
        f.close()
        f2.close()
        DebuggingUtil.print_magenta(f"{StateManagementUtil.UNDERLINE_PROMISED}{StateManagementUtil.UNDERLINE_PROMISED} all_list_proper.txt rewriting e")

        DebuggingUtil.print_magenta(f"{StateManagementUtil.UNDERLINE_PROMISED}{StateManagementUtil.UNDERLINE_PROMISED} files opening s")
        os.chdir(os.getcwd())

        # 윈도우냐 아니냐
        os.system("chcp 65001 >nul")
        os.system('export LANG=en_US.UTF-8 >nul')

        # os.system("type all_list.txt")
        # os.system("explorer all_list.txt")
        os.system("explorer all_list_proper.txt")
        DebuggingUtil.print_magenta(f"{StateManagementUtil.UNDERLINE_PROMISED}{StateManagementUtil.UNDERLINE_PROMISED} files opening e")

        # os.system('del "'+os.getcwd()+'\\all_list.txt"')
        # mk("all_list.txt")
        DebuggingUtil.print_magenta(f"{StateManagementUtil.UNDERLINE_PROMISED}{StateManagementUtil.UNDERLINE_PROMISED} e")

    @staticmethod
    def get_line_cnt_of_file(target_abspath: str):
        try:
            DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
            line_cnt = 0
            # 파일 변경 감지 이슈: linecache 모듈은 파일의 변경을 감지하지 못합니다.
            # 파일이 변경되었을 때에도 이전에 캐시된 내용을 반환하여 오래된 정보를 사용할 수 있습니다.
            # 실시간으로 파일의 변경을 감지해야 하는 경우에는 정확한 결과를 얻기 어려울 수 있습니다.
            # line_cnt = len(linecache.getlines(target_abspath))
            # Park4139.debug_as_cli(f'line_cnt:{line_cnt}')  캐시된 내용을 반환하기 때문에. 실시간 정보가 아니다

            # 이 코드는 실시간으로 파일의 변경을 감지 처리 되도록 수정, 단, 파일이 크면 성능저하 이슈 있을 수 있다.
            with open(target_abspath, 'r', encoding="UTF-8") as file:
                # whole_contents = file.readlines()
                # Park4139.debug_as_cli(whole_contents)
                # line_cnt = len(whole_contents)
                # line_cnt = list(en umerate(file))[-1][0] + 1
                line_cnt = file.read().count("\n") + 1
            return line_cnt
        except FileNotFoundError:
            DebuggingUtil.print_magenta("파일을 찾을 수 없었습니다")
            pass
        except Exception:
            DebuggingUtil.trouble_shoot("%%%FOO%%%")

    @staticmethod
    def back_up_by_manual(target_abspath):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        starting_directory = os.getcwd()
        try:
            target_dirname = os.path.dirname(target_abspath)
            target_dirname_dirname = os.path.dirname(target_dirname)
            target_basename = os.path.basename(target_abspath).split(".")[0]
            target_zip = rf'$zip_{target_basename}.zip'
            target_yyyy_mm_dd_HH_MM_SS_zip = rf'{target_basename} - {TimeUtil.get_time_as_("%Y %m %d %H %M %S")}.zip'
            # DebuggingUtil.commentize(rf'# target_dirname_dirname 로 이동')
            os.chdir(target_dirname_dirname)
            # DebuggingUtil.commentize(rf'부모디렉토리로 빽업')
            cmd = f'bandizip.exe c "{target_zip}" "{target_abspath}"'
            FileSystemUtil.get_cmd_output(cmd)
            # DebuggingUtil.commentize(rf'이름변경')
            cmd = f'ren "{target_zip}" "{target_yyyy_mm_dd_HH_MM_SS_zip}"'
            FileSystemUtil.get_cmd_output(cmd)
            # DebuggingUtil.commentize(rf'부모디렉토리에서 빽업될 디렉토리로 이동')
            cmd = f'move "{target_yyyy_mm_dd_HH_MM_SS_zip}" "{target_dirname}"'
            FileSystemUtil.get_cmd_output(cmd)
            # DebuggingUtil.commentize(rf'빽업될 디렉토리로 이동')
            # os.chdir(target_dirname)
            os.chdir(starting_directory)
            # DebuggingUtil.commentize("os.getcwd()")
            # Park4139.debug_as_cli(os.getcwd())

        except:
            DebuggingUtil.trouble_shoot("202312030000d")
        finally:
            DebuggingUtil.commentize(rf'프로젝트 디렉토리로 이동')
            os.chdir(starting_directory)

    # @staticmethod
    # def move_mouse(abs_x: float, abs_y: float):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     pyautogui.moveTo(abs_x, abs_y)
    #
    # @staticmethod
    # def move_mouse_rel_x(rel_x: float, rel_y: float):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     pyautogui.move(rel_x, rel_y)
    #
    # @staticmethod
    # def get_current_mouse_abs_info():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     return pyautogui.position()  # x, y = get_current_mouse_abs_info() 이런식으로 받을수 있나 테스트

    @staticmethod
    def open_mouse_info():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        # pyautogui.mouseInfo()

    # @staticmethod
    # def press(*presses: str, interval=0.0):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #         key = None
    #         if len(presses) == 1:
    #             if presses == "pgup":
    #                 presses = "pageup"
    #             elif presses == "pgdn":
    #                 presses = "pagedown"
    #             for i in pyautogui.KEYBOARD_KEYS:
    #                 if str(i) == str(presses[0]):
    #                     pyautogui.press(str(presses[0]), interval=interval)
    #                     DebuggingUtil.print_magenta(rf"{str(i)} 눌렸습니다3")
    #                     break
    #                 else:
    #                     pass
    #         else:
    #             pyautogui.hotkey(*presses, interval=interval)
    #             tmp = ' + '.join(i for i in presses)
    #             DebuggingUtil.print_magenta(rf"{tmp} 눌렸습니다")
    #         # Park4139.sleep(milliseconds=100)
    #         break

    # @staticmethod
    # def get_400px_screenshot(miliseconds=0):

    #     """pyautogui, 마우스의 위치 주변 가로 세로 400 px  400 px 로 스크린샷 찍어서 저장하는 코드"""
    #     # 재우기
    #     Park4139.sleep(milliseconds=miliseconds)
    #
    #     # 현재 마우스 위치 가져오기
    #     x, y = pyautogui.position()
    #     width = 400
    #     height = 400
    #     left = x - width / 2  # height/2 일수도 있음
    #     top = y - height / 2  # 여기도 마찬가지 일수 있음
    #
    #     # 스크린샷 찍기
    #     pygui = pyautogui.screenshot(region=(left, top, width, height))
    #
    #     # 스크린샷 저장
    #     server_time = Park4139.get_time_as_('%Y_%m_%d_%H_%M_%S')
    #     screenshot_png = rf'{Park4139.PROJECT_DIRECTORY}\$cache_png\screenshot_{server_time}.png'
    #     try:
    #         os.makedirs(os.path.dirname(screenshot_png))
    #     except FileExistsError:
    #         pass
    #     pygui.save(screenshot_png)
    #     pygui.show(screenshot_png)

    @staticmethod
    def shoot_full_screenshot():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        try:
            BusinessLogicUtil.taskkill("SnippingTool.exe")

            # Snipping Tool.ink 실행
            cmd = rf' explorer "C:\ProgramData\Microsoft\Windows\Start Menu\Programs\Accessories\Snipping Tool.lnk" >nul'
            FileSystemUtil.get_cmd_output(cmd)
            BusinessLogicUtil.sleep(milliseconds=150)
            BusinessLogicUtil.press("alt", "m", interval=0.15)
            BusinessLogicUtil.press("s", interval=0.15)
            BusinessLogicUtil.press("ctrl", "s", interval=0.15)
            server_time = TimeUtil.get_time_as_('%Y_%m_%d_%H_%M_%S')
            file_ = rf'{StateManagementUtil.PROJECT_DIRECTORY}\storage\screenshot_full\by_screenshot_full_rpa_{server_time}.png'
            try:
                os.makedirs(os.path.dirname(file_))
            except FileExistsError:
                pass
            BusinessLogicUtil.write_fast(file_)
            BusinessLogicUtil.press("enter")
            try:
                cmd = rf'explorer "{file_}"'
                FileSystemUtil.get_cmd_output(cmd=cmd)
            except FileNotFoundError:
                pass

            BusinessLogicUtil.press("esc")

            BusinessLogicUtil.taskkill(program_img_name="SnippingTool.exe")
            BusinessLogicUtil.sleep(milliseconds=50)

            # Snipping Tool.ink 종료 try3
            # 이미지를 인식해서 닫는방법

            TextToSpeechUtil.speak_ments(ment="스크린샷 저장을 성공했습니다", sleep_after_play=0.65)
        except:
            traceback.print_exc(file=sys.stdout)

    @staticmethod
    def shoot_img_for_rpa():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        while True:
            try:
                # Snipping Tool.ink 종료
                BusinessLogicUtil.taskkill(program_img_name="SnippingTool.exe")

                # Snipping Tool.ink 실행
                cmd = rf' explorer "C:\ProgramData\Microsoft\Windows\Start Menu\Programs\Accessories\Snipping Tool.lnk" >nul'
                FileSystemUtil.get_cmd_output(cmd)
                BusinessLogicUtil.sleep(milliseconds=150)
                BusinessLogicUtil.press("alt", "m", interval=0.15)
                BusinessLogicUtil.press("s", interval=0.15)

                if BusinessLogicUtil.is_shortcut_pressed_within_10_secs("ctrl+s"):
                    server_time = TimeUtil.get_time_as_('%Y_%m_%d_%H_%M_%S')
                    file_png = rf'{StateManagementUtil.PROJECT_DIRECTORY}\$cache_png\{server_time}_for_rpa.png'  # 맨 언더바를 붙였다는 것은 rpa에 사용중이지 않은 이미지란 의미
                    try:
                        os.makedirs(os.path.dirname(file_png))
                    except FileExistsError:
                        pass
                    BusinessLogicUtil.sleep(250)
                    BusinessLogicUtil.write_fast(file_png)

                    # enter 눌러 저장
                    BusinessLogicUtil.sleep(250)
                    BusinessLogicUtil.press("enter")

                    try:
                        cmd = rf'explorer "{file_png}"'
                        FileSystemUtil.get_cmd_output(cmd=cmd)
                    except FileNotFoundError:
                        pass
                    # Snipping Tool.ink 종료
                    BusinessLogicUtil.taskkill(program_img_name="SnippingTool.exe")

                    BusinessLogicUtil.sleep(milliseconds=1000)
                    BusinessLogicUtil.press("esc")
                    break
                else:
                    break
            except:
                traceback.print_exc(file=sys.stdout)

    # @staticmethod
    # def get_abs_x_and_y_from_img(img_abspath):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     return pyautogui.locateOnScreen(img_abspath)  # x, y = get_current_mouse_abs_info() 이런식으로 받을수 있나 테스트

    @staticmethod
    def write_fast(presses: str):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        BusinessLogicUtil.sleep(milliseconds=500)
        BusinessLogicUtil.copy_and_paste_with_keeping_clipboard_current_contents(presses)
        DebuggingUtil.print_magenta(rf"{str(presses)} 눌렸어요")

    # @staticmethod
    # def write_slow(presses: str):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     pyautogui.write(presses, interval=0.09)
    #     DebuggingUtil.print_magenta(rf"{str(presses)} 눌렸어요")

    @staticmethod
    def ask_to_wrtn(question: str):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        while True:
            # 페이지 열기
            url = "https://wrtn.ai/"
            cmd = f'explorer  "{url}"  >nul'
            FileSystemUtil.get_cmd_output(cmd)

            # 크롬 창 활성화
            target_pid: int = BusinessLogicUtil.get_target_pid_by_process_name(target_process_name="chrome.exe")  # chrome.exe pid 가져오기
            BusinessLogicUtil.activate_window_by_pid(pid=target_pid)

            # 크롬 기본 배율로 변경
            BusinessLogicUtil.press('ctrl', '0')

            # 광고닫기 버튼 클릭
            file_png = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_png\ask_to_wrtn_ad_close.png"
            BusinessLogicUtil.click_center_of_img_recognized_by_mouse_left(img_abspath=file_png, recognize_loop_limit_cnt=10, is_zoom_toogle_mode=True)

            # 프롬프트 콘솔 클릭(광고 없어도 진행)
            file_png = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_png\ask_to_wrtn.png"
            if BusinessLogicUtil.click_center_of_img_recognized_by_mouse_left(img_abspath=file_png, recognize_loop_limit_cnt=50, is_zoom_toogle_mode=True):
                # 질문 작성 및 확인
                BusinessLogicUtil.write_fast(question)
                BusinessLogicUtil.press('enter')

            # 뤼튼 프롬프트 콘솔 최하단 이동 버튼 클릭
            break

    @staticmethod
    def find_direction_via_naver_map(destination: str):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        while True:
            # 배경화면으로 나가기(옵션로직)
            # Park4139.press("win", "m")
            # Park4139.press("win", "m")
            # Park4139.sleep(10)

            # 페이지 열기
            # url = "https://map.naver.com/"
            url = "https://map.naver.com/p/directions"
            cmd = f'explorer  "{url}"  >nul'
            FileSystemUtil.get_cmd_output(cmd)
            BusinessLogicUtil.sleep(300)

            # 크롬 창 활성화
            target_pid: int = BusinessLogicUtil.get_target_pid_by_process_name(target_process_name="chrome.exe")  # chrome.exe pid 가져오기
            BusinessLogicUtil.activate_window_by_pid(pid=target_pid)
            BusinessLogicUtil.sleep(30)

            # 반쪽화면 만들기(옵션로직)
            # Park4139.press("alt", "up")
            # Park4139.press("alt", "left")

            # 출발지 입력 클릭
            file_png = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_png\find_direction_via_naver_direction.png"
            BusinessLogicUtil.click_center_of_img_recognized_by_mouse_left(img_abspath=file_png, recognize_loop_limit_cnt=100, is_zoom_toogle_mode=True)
            BusinessLogicUtil.sleep(30)

            # 한가람한양아파트상가 입력
            BusinessLogicUtil.write_fast("한가람한양아파트상가")
            BusinessLogicUtil.sleep(30)
            BusinessLogicUtil.press('enter')
            BusinessLogicUtil.sleep(300)
            BusinessLogicUtil.press('tab')
            BusinessLogicUtil.sleep(30)

            # 목적지 입력
            BusinessLogicUtil.write_fast(destination)
            BusinessLogicUtil.sleep(30)
            BusinessLogicUtil.press('down')
            BusinessLogicUtil.press('enter')

            # 길찾기 클릭
            BusinessLogicUtil.press('tab')
            BusinessLogicUtil.press('tab')
            BusinessLogicUtil.press('tab')
            BusinessLogicUtil.press('enter')

            # 작업마침 알림
            TextToSpeechUtil.speak_ments(ment='길찾기가 시도되었습니다', sleep_after_play=0.65)
            break

    # @staticmethod
    # def search_animation_data_from_web(btn_text_clicked):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     try:
    #         while 1:
    #             # answer = "tate no"
    #             DebuggingUtil.commentize(f"{btn_text_clicked} 입력되었습니다")
    #             DebuggingUtil.print_magenta(btn_text_clicked)
    #             special_prefixes = " ".join(["subsplease", "1080"]) + " "
    #             query = urllib.parse.quote(f"{special_prefixes}{btn_text_clicked}")
    #             if query == "":
    #                 TextToSpeechUtil.speak_ments(ment="아무것도 입력되지 않았습니다", sleep_after_play=0.65)
    #                 break
    #             # url = f'https://nyaa.si/?f=0&c=0_0&q={query}&p=1'
    #             url = f'https://nyaa.si/?f=0&c=0_0&q={query}'
    #             DebuggingUtil.print_magenta(url)
    #
    #             driver = SeleniumUtil.get_driver_for_selenium()
    #             driver.get(url)
    #             DebuggingUtil.commentize("모든 웹소스 우회 설정 확인")
    #             # setting_for_proxy = ""
    #             # setting_for_proxy = setting_for_proxy  + f"user_agent     : {driver.find_element(By.CSS_SELECTOR,'#user-agent').text}\n"
    #             # setting_for_proxy = setting_for_proxy  + f"plugins_length : {driver.find_element(By.CSS_SELECTOR,  '#plugins-length').text}\n"
    #             # setting_for_proxy = setting_for_proxy  + f"languages : {driver.find_element(By.CSS_SELECTOR, '#languages').text}\n"
    #             # setting_for_proxy = setting_for_proxy  + f"webgl_vendor : {driver.find_element(By.CSS_SELECTOR, '#webgl-vendor').text}\n"
    #             # setting_for_proxy = setting_for_proxy  + f"webgl_renderer : {driver.find_element(By.CSS_SELECTOR, '#webgl-renderer').text}\n"
    #             # print(context=setting_for_proxy, is_app_instance_mode=False)
    #
    #             # 필요할까?
    #             # driver.implicitly_wait(2)
    #             # driver.implicitly_wait(3)
    #             # driver.get_screenshot_as_file('hlah.png')
    #
    #             DebuggingUtil.commentize("페이지 소스 RAW")
    #             page_src = driver.page_source
    #             DebuggingUtil.print_magenta(page_src)
    #             # print(context=f"\n{page_src}")
    #             # print(context=f"페이지 소스 RAW:\n\n{page_src}")
    #
    #             DebuggingUtil.commentize("web parser 설정")
    #             # soup = BeautifulSoup(page_src, "html5lib")
    #             # soup = BeautifulSoup(page_src, "html.parser")  # 파이썬 기본 파서
    #             soup = BeautifulSoup(page_src, "lxml")
    #
    #             DebuggingUtil.commentize("web src 분석시작")
    #             # title 가져오기
    #             # num = 0  # num 초기값
    #             # Park4139.debug_as_cli(soup.get_text())
    #             for i in soup.find_all(name="a"):
    #                 DebuggingUtil.print_magenta(i)
    #
    #             results = []
    #             DebuggingUtil.commentize("web src 분석시작2")
    #             lines = soup.find_all(name="a")
    #             for line in lines:
    #                 results.append(line.get('title'))
    #                 results.append(line.get('href'))
    #                 DebuggingUtil.print_magenta(line)
    #                 # pass
    #             DebuggingUtil.commentize("titles")
    #             titles = []
    #             for i in results:
    #                 if type(i) is str:
    #                     if BusinessLogicUtil.is_regex_in_contents_with_case_ignored(contents=str(i).strip(), regex=btn_text_clicked):
    #                         if not BusinessLogicUtil.is_regex_in_contents_with_case_ignored(contents=str(i).strip(), regex="&q="):
    #                             if not BusinessLogicUtil.is_regex_in_contents_with_case_ignored(contents=str(i).strip(), regex="xt="):
    #                                 DebuggingUtil.print_magenta(f'{type(i)} : {i}')
    #                                 titles.append(i)
    #             DebuggingUtil.commentize("magnets")
    #             magnets = []
    #             for i in results:
    #                 if type(i) is str:
    #                     if BusinessLogicUtil.is_regex_in_contents(target=i, regex="magnet*"):
    #                         DebuggingUtil.print_magenta(f'{type(i)} : {i}')
    #                         magnets.append(i)
    #             DebuggingUtil.commentize("finally crawled result")
    #             for i in results:
    #                 if type(i) is str:
    #                     if BusinessLogicUtil.is_regex_in_contents(target=i, regex="magnet*"):
    #                         DebuggingUtil.print_magenta(f'{type(i)} : {i}')
    #                     if BusinessLogicUtil.is_regex_in_contents_with_case_ignored(contents=str(i).strip(), regex=btn_text_clicked):
    #                         if not BusinessLogicUtil.is_regex_in_contents_with_case_ignored(contents=str(i).strip(), regex="&q="):
    #                             if not BusinessLogicUtil.is_regex_in_contents_with_case_ignored(contents=str(i).strip(), regex="xt="):
    #                                 DebuggingUtil.print_magenta(f'{type(i)} : {i}')
    #             results_about_titles_and_magnets = ""
    #             if len(magnets) == len(titles):
    #                 for i in range(0, len(magnets)):
    #                     DebuggingUtil.print_magenta(f"{i}   :    {titles[i]}    :     {magnets[i]}")
    #                     results_about_titles_and_magnets = results_about_titles_and_magnets + f"{i}   :    {titles[i]}    :     {magnets[i]}\n"
    #             # dialog = CustomDialog(contents="웹 크롤링 결과를 화면에 시현할까요?", buttons=["시현하기", "시현하지 않기"])
    #             # dialog.exec()
    #             # btn_text_clicked = dialog.btn_text_clicked
    #             # Park4139.debug_as_cli(btn_text_clicked)
    #             # if btn_text_clicked == "시현하기":
    #             dialog = UiUtil.CustomQdialog(ment=results_about_titles_and_magnets, btns=["확인"])
    #             dialog.exec()
    #             btn_text_clicked = dialog.btn_text_clicked
    #             DebuggingUtil.print_magenta(btn_text_clicked)
    #             if btn_text_clicked == "확인":
    #                 pass
    #             if len(magnets) == len(titles):
    #                 for i in range(0, len(magnets) - 1):
    #                     DebuggingUtil.print_magenta(f"{i}   :    {titles[i]}    :     {magnets[i]}")
    #             else:
    #                 ment = "크롤링 결과가 이상합니다\n크롤링한 마그넷주소의 개수와 타이틀의 개수가 일치하지 않습니다\n크롤링을 계속할까요"
    #                 TextToSpeechUtil.speak_ments(ment.replace("\n", " "), sleep_after_play=0.65)
    #                 dialog = UiUtil.CustomQdialog(ment=ment, btns=["계속하기", "계속하지 않기"])
    #                 dialog.exec()
    #                 btn_text_clicked = dialog.btn_text_clicked
    #                 DebuggingUtil.print_magenta(btn_text_clicked)
    #                 if btn_text_clicked == "계속하지 않기":
    #                     TextToSpeechUtil.speak_ments(ment="네 계속하지 않겠습니다", sleep_after_play=0.65)
    #                     break
    #
    #             while True:
    #                 ment = '토렌트에 추가하고 싶은 항목의 숫자인덱스를 입력하세요.\n전부를 원하시면 * 을 입력해주세요'
    #                 TextToSpeechUtil.speak_ments(ment, sleep_after_play=0.65)
    #                 dialog = UiUtil.CustomQdialog(ment=ment, btns=["계속하기", "그만하기", "크롤링 결과 다시보기"], is_input_box=True)
    #                 dialog.exec()
    #                 btn_text_clicked = dialog.btn_text_clicked
    #                 text_of_from_input_box = str(dialog.input_box.text()).strip()
    #                 DebuggingUtil.print_magenta(f"text_of_input_box : {text_of_from_input_box}")
    #                 DebuggingUtil.print_magenta(f"btn_text_clicked : {btn_text_clicked}")
    #                 if btn_text_clicked == "그만하기":
    #                     TextToSpeechUtil.speak_ments(ment="네 그만하겠습니다", sleep_after_play=0.65)
    #                     break
    #                 elif btn_text_clicked == "크롤링 결과 다시보기":
    #                     dialog = UiUtil.CustomQdialog(ment=results_about_titles_and_magnets, btns=["확인"])
    #                     dialog.exec()
    #                 elif btn_text_clicked == "계속하기":
    #                     if text_of_from_input_box.isdigit():
    #                         mg_num = int(text_of_from_input_box)
    #                         try:
    #                             if magnets[int(mg_num)]:
    #                                 DebuggingUtil.print_magenta(f"{StateManagementUtil.UNDERLINE_PROMISED}다운로드 가능한 범위의 숫자인덱스입니다")
    #                         except IndexError:
    #                             TextToSpeechUtil.speak_ments(ment="다운로드 가능한 범위의 숫자인덱스가 아닙니다", sleep_after_play=0.65)
    #                         ment = f"1건에 대한 마그넷 추가를 진행할까요?\n"  # 단건 마그넷 추가 진행
    #                         ment = ment + f"\n"
    #                         ment = ment + f"index  : {int(mg_num)}\n"
    #                         ment = ment + f"title  : {titles[int(mg_num)]}\n"
    #                         ment = ment + f"magnet : {magnets[int(mg_num)][0:30]} ...\n"
    #                         for i in range(0, 4):
    #                             ment = ment + f"\n"
    #                         dialog = UiUtil.CustomQdialog(ment=ment, btns=["계속하기", "그만하기"])
    #                         dialog.exec()
    #                         btn_text_clicked = dialog.btn_text_clicked
    #                         if btn_text_clicked == "그만하기":
    #                             TextToSpeechUtil.speak_ments(ment="네 그만하겠습니다", sleep_after_play=0.65)
    #                             break
    #                         webbrowser.open(magnets[int(mg_num)])  # 토렌트 추가
    #                         break
    #                     elif text_of_from_input_box.strip() == "*":
    #                         ment = f"{len(magnets)}건에 대한 마그넷 추가를 진행할까요?\n"  # 복수 마그넷 추가 진행
    #                         for i in range(0, 4):
    #                             ment = ment + f"\n"
    #                         dialog = UiUtil.CustomQdialog(ment=ment, btns=["계속하기", "그만하기"])
    #                         dialog.exec()
    #                         btn_text_clicked = dialog.btn_text_clicked
    #                         if btn_text_clicked == "그만하기":
    #                             TextToSpeechUtil.speak_ments(ment="네 그만하겠습니다", sleep_after_play=0.65)
    #                             break
    #                         for i in range(0, len(magnets)):
    #                             # Park4139.sleep(milliseconds=random.randint(20, 40))
    #                             BusinessLogicUtil.sleep(milliseconds=random.randint(10, 20))
    #                             # Park4139.sleep(milliseconds=random.randint(1,3))
    #                             DebuggingUtil.print_magenta(f"{i}  : title  {titles[i]}")
    #                             DebuggingUtil.print_magenta(f"{i}  : magnet {magnets[i]}")
    #                             webbrowser.open(magnets[i])  # 토렌트 추가
    #                         break
    #                     else:
    #                         TextToSpeechUtil.speak_ments(ment="입력하신 내용이 숫자나 별이 아닌 것 같아요", sleep_after_play=0.65)
    #             break
    #     except:
    #         traceback.print_exc(file=sys.stdout)

    @staticmethod  # move_target_recycle_bin()
    def empty_recycle_bin():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        DebuggingUtil.commentize("휴지통 을 비우는 중...")
        FileSystemUtil.get_cmd_output('PowerShell.exe -NoProfile -Command Clear-RecycleBin -Confirm:$false')
        # 휴지통 삭제 (외장하드까지)
        # for %%a in (cdefghijk L mnopqrstuvwxyz) do (
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        # 존재하는 경우 %%a:\$RECYCLE.BIN for /f "tokens=* usebackq" %%b in (`"dir /a:d/b %%a:\$RECYCLE.BIN\"`) do rd / q/s "%%a:\$RECYCLE.BIN\%%~b"
        # 존재하는 경우 %%a:\RECYCLER for /f "tokens=* usebackq" %%b in (`"dir /a:d/b %%a:\RECYCLER\"`) do rd /q/s "%% a:\RECYCLER\%%~b"
        # )
        TextToSpeechUtil.speak_ments(f'휴지통을 비웠습니다', sleep_after_play=0.65, thread_join_mode=True)

        # 가끔 휴지통을 열어볼까요?
        # DebuggingUtil.commentize("숨김 휴지통 열기")
        # cmd = 'explorer c:\$RECYCLE.BIN'
        # Park4139.get_cmd_output(cmd=cmd)
        # 외장하드 숨김 휴지통 을 보여드릴까요
        # explorer c:\$RECYCLE.BIN
        # explorer d:\$RECYCLE.BIN
        # explorer e:\$RECYCLE.BIN
        # explorer f:\$RECYCLE.BIN

    @staticmethod
    def debug(context: str, is_app_instance_mode=False, input_text_default="", auto_click_positive_btn_after_seconds=3):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        """
        이 함수는 특별한 사용요구사항이 있습니다
        pyside6 앱 내에서 해당 함수를 호출할때는 is_app_instance_mode 를 파라미터에 넣지 않고 쓰는 것을 default 로 디자인했습니다.
        pyside6 앱 밖에서 해당 함수를 호출할때는 is_app_instance_mode 를 True 로 설정하고 쓰십시오.

        사용 요구사항이 생기게 된 이유는 다음과 같습니다
        pyside6 는 app을 singletone으로 instance를 구현합니다. 즉, instance는 반드시 pyside6 app 내에서 하나여야 합니다.
        pyside6의 QApplication()이 앱 내/외에서도 호출이 될 수 있도록 디자인했습니다.
        앱 내에서 호출 시에는 is_app_instance_mode 파라미터를 따로 설정하지 않아도 되도록 디자인되어 있습니다.
        앱 외에서 호출 시에는 is_app_instance_mode 파라미터를 True 로 설정해야 동작하도록 디자인되어 있습니다.
        """
        # app_foo: QApplication = None
        # if is_app_instance_mode == True:
        #     app_foo: QApplication = QApplication()
        # if input_text_default == "":
        #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        #     is_input_text_box = False
        # else:
        #     is_input_text_box = True
        # dialog = UiUtil.CustomQdialog(ment=f"{context}", buttons=["확인"], is_input_box=is_input_text_box, input_box_text_default=input_text_default)
        # DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        # dialog.exec()
        # btn_text_clicked = dialog.btn_text_clicked
        # if btn_text_clicked == "":
        #     DebuggingUtil.print_magenta()(f'누르신 버튼은 {btn_text_clicked} 입니다')
        # if is_app_instance_mode == True:
        #     if isinstance(app_foo, QApplication):
        #         app_foo.exec()
        # if is_app_instance_mode == True:
        #     # app_foo.quit()# QApplication 인스턴스 제거시도 : fail
        #     # app_foo.deleteLater()# QApplication 인스턴스 파괴시도 : fail
        #     # del app_foo # QApplication 인스턴스 파괴시도 : fail
        #     # app_foo = None # QApplication 인스턴스 파괴시도 : fail
        #     app_foo.shutdown()  # QApplication 인스턴스 파괴시도 : success  # 성공요인은 app.shutdown()이 호출이 되면서 메모리를 해제까지 수행해주기 때문
        #     # sys.exit()
        # # return app_foo
        DebuggingUtil.print_ment_via_colorama(traceback.print_exc(file=sys.stdout), colorama_color=ColoramaUtil.RED)
        # traceback.print_exc(file=sys.stdout)
        # if FileSystemUtil.is_os_windows():
        #     UiUtil.pop_up_as_complete(title_="디버깅결과보고", ment=context, input_text_default=input_text_default, auto_click_positive_btn_after_seconds=auto_click_positive_btn_after_seconds)
        # else:
        DebuggingUtil.print_magenta(ment=context)

    # @staticmethod
    # def click_mouse_left_btn(abs_x=None, abs_y=None):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     if abs_x and abs_y:
    #         pyautogui.click(button='left', clicks=1, interval=0)
    #     else:
    #         pyautogui.click(button='left', clicks=1, interval=0, x=abs_x, y=abs_y)
    #
    # @staticmethod
    # def click_mouse_right_btn(abs_x=None, abs_y=None):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     if abs_x and abs_y:
    #         pyautogui.click(button='right', clicks=1, interval=0)
    #     else:
    #         pyautogui.click(button='right', clicks=1, interval=0, x=abs_x, y=abs_y, )

    @staticmethod
    def ask_to_bard(question: str):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        while True:
            cmd = f"explorer https://bard.google.com/chat  >nul"
            FileSystemUtil.get_cmd_output(cmd)

            # 이시간은 web rendering time 대기 해주는 시간정도로 생각하고 있는데
            # 시간 성능에 대한 마지노선을 설정하는게 여간 편차가 크다.
            # web 크롤링 을 한뒤 연산하는게 훨씬 빠르겠다.
            # Park4139.sleep(milliseconds=3000) # 전부 성공
            # Park4139.sleep(milliseconds=1500) #
            BusinessLogicUtil.sleep(milliseconds=1060)
            # Park4139.sleep(milliseconds=1030)  # 실패있음
            # Park4139.sleep(milliseconds=1000)  #실패있음. 질문입력 안됨
            # Park4139.sleep(milliseconds=500)

            # 질문 작성 및 확인
            BusinessLogicUtil.write_fast(question)
            BusinessLogicUtil.sleep(milliseconds=30)
            BusinessLogicUtil.press('enter')
            break

    @staticmethod
    def ask_to_google(question: str):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        while True:
            question_ = question.replace(" ", "+")
            cmd = f'explorer "https://www.google.com/search?q={question_}"  >nul'
            FileSystemUtil.get_cmd_output(cmd)
            break

    # @staticmethod
    # def get_img_when_img_recognized_succeed(img_abspath, recognize_loop_limit_cnt=0, is_zoom_toogle_mode=False):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     # TextToSpeechUtil.speak_ment("화면 이미지 분석을 시도합니다")
    #     UiUtil.pop_up_as_complete(title_="화면이미지분석 시도 전 보고", ment="화면 이미지 분석을 시도합니다", auto_click_positive_btn_after_seconds=1)
    #
    #     BusinessLogicUtil.press('ctrl', '0', interval=0.15)  # 이미지 분석 시 크롬 zoom 초기화(ctrl+0)
    #
    #     # 고해상도/다크모드 에서 시도
    #     for i in range(0, 9):
    #         BusinessLogicUtil.press('ctrl', '+')
    #
    #     chrome_zoom_step = 0
    #
    #     # 루프카운트 제한 n번 default 0번 으로 설정
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     loop_cnt = 0
    #     # recognize_loop_limit_cnt = recognize_loop_limit_cnt
    #     if recognize_loop_limit_cnt == 0:
    #         while True:
    #             # 속도개선 시도
    #             # pip install opencv-python # 이것은 고급 기능이 포함되지 않은 Python용 OpenCV의 미니 버전입니다. 우리의 목적에는 충분합니다.
    #             # confidence=0.7(70%)유사도를 낮춰 인식률개선시도, region 낮춰 속도개선시도, grayscale 흑백으로 판단해서 속도개선시도,
    #             # open cv 설치했는데 적용안되고 있음. 재부팅도 하였는 데도 안됨.
    #             # pycharm 에서 import 하는 부분에서 cv2 설치 시도 중에 옵션으로 opencv-python 이 있길래 설치했더니 결국 됨. 혹쉬 경로 설정 필요했나?
    #             # xy_infos_of_imgs = pyautogui.locateOnScreen(img_abspath, confidence=0.7, grayscale=True)
    #             # print(xy_infos_of_imgs == None)
    #             DebuggingUtil.commentize("화면 이미지 인식 시도 중...")
    #             loop_cnt = loop_cnt + 1
    #             try:
    #                 img = pyautogui.locateOnScreen(img_abspath, confidence=0.7, grayscale=True)
    #                 DebuggingUtil.print_magenta(type(img))
    #                 DebuggingUtil.print_magenta(img)
    #                 DebuggingUtil.print_magenta(img != None)
    #                 if img != None:
    #                     return img
    #                 else:
    #                     DebuggingUtil.commentize(f"화면 이미지 분석 중...")
    #                     DebuggingUtil.print_magenta(img_abspath)
    #                     BusinessLogicUtil.sleep(milliseconds=15)
    #                     if is_zoom_toogle_mode == True:
    #                         if chrome_zoom_step == 14:
    #                             chrome_zoom_step = 0
    #                         elif chrome_zoom_step < 7:
    #                             BusinessLogicUtil.press('ctrl', '-')
    #                             chrome_zoom_step = chrome_zoom_step + 1
    #                         elif 7 <= chrome_zoom_step:
    #                             BusinessLogicUtil.press('ctrl', '+')
    #                             chrome_zoom_step = chrome_zoom_step + 1
    #             except pyautogui.ImageNotFoundException:
    #                 DebuggingUtil.commentize(f"{loop_cnt}번의 화면인식시도를 했지만 인식하지 못하였습니다")
    #                 pass
    #     else:
    #         while True:
    #             loop_cnt = loop_cnt + 1
    #             DebuggingUtil.commentize("화면 이미지 인식 시도 중...")
    #             if recognize_loop_limit_cnt == loop_cnt:
    #                 DebuggingUtil.commentize(f"{loop_cnt}번의 화면인식시도를 했지만 인식하지 못하였습니다")
    #                 return None
    #             try:
    #                 img = pyautogui.locateOnScreen(img_abspath, confidence=0.7, grayscale=True)
    #                 DebuggingUtil.print_magenta(type(img))
    #                 DebuggingUtil.print_magenta(img)
    #                 DebuggingUtil.print_magenta(img != None)
    #                 if img != None:
    #                     return img
    #             except:
    #                 DebuggingUtil.print_magenta(img_abspath)
    #                 BusinessLogicUtil.sleep(milliseconds=10)
    #                 if is_zoom_toogle_mode == True:
    #                     if chrome_zoom_step == 14:
    #                         chrome_zoom_step = 0
    #                     elif chrome_zoom_step < 7:
    #                         BusinessLogicUtil.press('ctrl', '-')
    #                         chrome_zoom_step = chrome_zoom_step + 1
    #                     elif 7 <= chrome_zoom_step:
    #                         BusinessLogicUtil.press('ctrl', '+')
    #                         chrome_zoom_step = chrome_zoom_step + 1

    @staticmethod
    def click_center_of_img_recognized_by_mouse_left(img_abspath: str, recognize_loop_limit_cnt=0, is_zoom_toogle_mode=False):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        img = BusinessLogicUtil.get_img_when_img_recognized_succeed(img_abspath, recognize_loop_limit_cnt, is_zoom_toogle_mode)
        if not img == None:
            center_x = img.left + (img.width / 2)
            center_y = img.top + (img.height / 2)
            BusinessLogicUtil.move_mouse(abs_x=center_x, abs_y=center_y)
            BusinessLogicUtil.click_mouse_left_btn(abs_x=center_x, abs_y=center_y)
            return True
        else:
            return False

    @staticmethod
    def connect_remote_rdp1():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        FileSystemUtil.get_cmd_output(cmd=rf"{StateManagementUtil.RDP_82106_BAT}")

    @staticmethod
    def shoot_custom_screenshot():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        while True:
            try:
                # Snipping Tool.ink 종료
                BusinessLogicUtil.taskkill(program_img_name="SnippingTool.exe")

                # Snipping Tool.ink 실행
                cmd = rf' explorer "C:\ProgramData\Microsoft\Windows\Start Menu\Programs\Accessories\Snipping Tool.lnk" >nul'
                FileSystemUtil.get_cmd_output(cmd)
                BusinessLogicUtil.sleep(milliseconds=150)
                BusinessLogicUtil.press("alt", "m", interval=0.15)
                BusinessLogicUtil.press("s", interval=0.15)

                if BusinessLogicUtil.is_shortcut_pressed_within_10_secs("ctrl+s"):
                    server_time = TimeUtil.get_time_as_('%Y_%m_%d_%H_%M_%S')
                    file_png = rf'{StateManagementUtil.PROJECT_DIRECTORY}\storage\screenshot_custom\by_rpa_{server_time}.png'
                    try:
                        os.makedirs(os.path.dirname(file_png))
                    except FileExistsError:
                        pass
                    BusinessLogicUtil.sleep(250)
                    BusinessLogicUtil.write_fast(file_png)

                    # enter 눌러 저장
                    BusinessLogicUtil.sleep(250)
                    BusinessLogicUtil.press("enter")

                    try:
                        cmd = rf'explorer "{file_png}"'
                        FileSystemUtil.get_cmd_output(cmd=cmd)
                    except FileNotFoundError:
                        pass
                    # Snipping Tool.ink 종료
                    BusinessLogicUtil.taskkill(program_img_name="SnippingTool.exe")

                    BusinessLogicUtil.sleep(milliseconds=1000)
                    BusinessLogicUtil.press("esc")
                    break
                else:
                    break
            except:
                traceback.print_exc(file=sys.stdout)

    @staticmethod
    def download_from_youtube_to_webm_alt(url: str):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        while True:
            url_ = str(url).strip()
            if url_ == None:
                TextToSpeechUtil.speak_ments(ment="아무것도 입력되지 않았습니다.", sleep_after_play=0.65)
                break
            if url_ == "":
                TextToSpeechUtil.speak_ments(ment="아무것도 입력되지 않았습니다", sleep_after_play=0.65)
                break
            if url_ == "None":
                TextToSpeechUtil.speak_ments(ment="이상한 것이 입력되었습니다", sleep_after_play=0.65)
                break
            urls = [url.strip() for url in url_.split("\n")]

            DebuggingUtil.print_magenta(rf'urls : {urls}')
            DebuggingUtil.print_magenta(rf'type(urls) : {type(urls)}')
            DebuggingUtil.print_magenta(rf'len(urls) : {len(urls)}')

            try:
                useless_parts_of_url = [
                    "https://youtu.be/",
                    "https://www.youtube.com/shorts/",
                ]
                for index, useless_str in enumerate(useless_parts_of_url):
                    if useless_str in url_:
                        # DebuggingUtil.print_ment_light_white("모드 1")
                        # BusinessLogicUtil.download_clip(url=url_.split(useless_str)[1])
                        # DebuggingUtil.print_ment_light_white("모드 2")
                        # BusinessLogicUtil.download_clip_alt(url=url_.split(useless_str)[1])
                        # DebuggingUtil.print_ment_light_white("모드 3")
                        # BusinessLogicUtil.download_clip_alt(url_)
                        DebuggingUtil.print_ment_light_white("모드 4")
                        BusinessLogicUtil.download_clip_alt(f'https://www.youtube.com/watch?v={BusinessLogicUtil.parse_youtube_video_id(url_)}')
                        return
            except Exception:
                DebuggingUtil.trouble_shoot("%%%FOO%%%")

    @staticmethod
    def speak_that_service_is_in_preparing():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        TextToSpeechUtil.speak_ments(ment="아직 준비되지 않은 서비스 입니다", sleep_after_play=0.65)

    # @staticmethod
    # def ask_to_web(question):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #         # 윈도우즈 모든창 최소화
    #         # Park4139.press('win', 'm')
    #
    #         # wrtn에게 묻기 (chatGPT 3 기반)
    #         BusinessLogicUtil.ask_to_wrtn(question=question)
    #
    #         # 구글에게 묻기
    #         # Park4139.ask_to_google(question=question)
    #
    #         # bard에게 묻기
    #         # Park4139.ask_to_bard(question=question)
    #
    #         # 크롬 최대화
    #         # Park4139.press('win', 'up')
    #
    #         # 크롬 화면 적당한 배율로 변경
    #         BusinessLogicUtil.press('ctrl', '0')
    #         BusinessLogicUtil.press('ctrl', '-')
    #         BusinessLogicUtil.press('ctrl', '-')
    #         BusinessLogicUtil.press('ctrl', '-')
    #
    #         # 크롬 이전 시트로 이동
    #         # Park4139.press("ctrl", "shift", "tab")
    #
    #         BusinessLogicUtil.press("ctrl", "0")
    #         for i in range(1, 5):
    #             BusinessLogicUtil.press("ctrl", "-")
    #
    #         # 사용자에게 질문 답변
    #         TextToSpeechUtil.speak_ments(ment="AI 에게 질문을 성공했습니다", sleep_after_play=0.65)
    #         break

    @staticmethod
    def is_shortcut_pressed_within_10_secs(key_plus_key: str):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        """
        is_shortcut_pressed("ctrl+s")
        """
        time_s = time.time()

        keys = key_plus_key.split("+")
        DebuggingUtil.print_magenta(f"{keys[0]}+{keys[1]}가 눌릴때까지 기다리고 있습니다")
        while True:
            if all(keyboard.is_pressed(key) for key in keys):
                # 단축키 조합이 모두 눌렸을 때 실행할 코드를 여기에 작성해주세요
                DebuggingUtil.print_magenta(f"{keys[0]}+{keys[1]}가 눌렸습니다")
                return True
            else:
                time_e = time.time()
                time_diff = time_e - time_s
                if time_diff == 10:  # 10초 간만 시도하고 종료
                    return False
                BusinessLogicUtil.sleep(30)

    # @staticmethod
    # def translate_eng_to_kor(question: str):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     try:
    #         while True:
    #             try:
    #                 question = question.strip('""')
    #             except AttributeError:
    #                 break
    #             BusinessLogicUtil.press("win", "m")
    #
    #             # 구글 번역 페이지로 이동
    #             url = "https://www.google.com/search?q=eng+to+kor"
    #             cmd = f'explorer "{url}" >nul'
    #             FileSystemUtil.get_cmd_output(cmd)
    #
    #             # 크롬 창 활성화
    #             target_pid: int = BusinessLogicUtil.get_target_pid_by_process_name(target_process_name="chrome.exe")  # chrome.exe pid 가져오기
    #             BusinessLogicUtil.activate_window_by_pid(pid=target_pid)
    #
    #             # 텍스트를 입력하세 클릭
    #             file_png = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_png\eng to kor.png"
    #             BusinessLogicUtil.click_center_of_img_recognized_by_mouse_left(img_abspath=file_png, is_zoom_toogle_mode=True, recognize_loop_limit_cnt=100)
    #
    #             # 번역할 내용 작성
    #             BusinessLogicUtil.write_fast(question)
    #
    #             # 글자수가 많으면 text to voice icon 이 잘려서 보이지 않음. 이는 이미지의 객체 인식이 불가능해지는데
    #             # 스크롤를 내려서 이미지 인식을 가능토록
    #             if len(question) > 45:
    #                 pyautogui.vscroll(-15)
    #             BusinessLogicUtil.sleep(30)
    #
    #             # text to voice icon
    #             file_png = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_png\text to voice icon.png"
    #             BusinessLogicUtil.click_center_of_img_recognized_by_mouse_left(img_abspath=file_png, is_zoom_toogle_mode=True, recognize_loop_limit_cnt=100)
    #
    #             # 종료
    #             break
    #     except:
    #         traceback.print_exc(file=sys.stdout)

    # @staticmethod
    # def translate_kor_to_eng(question: str):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     try:
    #         while True:
    #             try:
    #                 question = question.strip('""')
    #             except AttributeError:
    #                 break
    #             BusinessLogicUtil.press("win", "m")
    #
    #             # 페이지 열기
    #             url = "https://www.google.com/search?q=kor+to+eng"
    #             cmd = f'explorer "{url}" >nul'
    #             FileSystemUtil.get_cmd_output(cmd)
    #
    #             # 크롬 창 활성화
    #             target_pid: int = BusinessLogicUtil.get_target_pid_by_process_name(target_process_name="chrome.exe")  # chrome.exe pid 가져오기
    #             BusinessLogicUtil.activate_window_by_pid(pid=target_pid)
    #
    #             # Enter Text 클릭
    #             file_png = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_png\kor to eng.png"
    #             BusinessLogicUtil.click_center_of_img_recognized_by_mouse_left(img_abspath=file_png, is_zoom_toogle_mode=True, recognize_loop_limit_cnt=100)
    #
    #             # 번역할 내용 작성
    #             BusinessLogicUtil.write_fast(question)
    #
    #             # 글자수가 많으면 text to voice icon 이 잘려서 보이지 않음. 이는 이미지의 객체 인식이 불가능해지는데
    #             # 스크롤를 내려서 이미지 인식을 가능토록
    #             if len(question) > 45:
    #                 pyautogui.vscroll(-15)
    #             BusinessLogicUtil.sleep(30)
    #
    #             # text to voice icon
    #             file_png = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_png\text to voice icon.png"
    #             BusinessLogicUtil.click_center_of_img_recognized_by_mouse_left(img_abspath=file_png, is_zoom_toogle_mode=True, recognize_loop_limit_cnt=100)
    #
    #             break
    #     except:
    #         traceback.print_exc(file=sys.stdout)

    @staticmethod
    def run_cmd_exe_as_admin():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        try:
            while True:
                # run.exe 관리자모드로 실행
                FileSystemUtil.get_cmd_output('PowerShell -Command "Start-Process cmd -Verb RunAs"')

                # 네 클릭
                file_png = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_png\run cmd exe.png"
                BusinessLogicUtil.click_center_of_img_recognized_by_mouse_left(img_abspath=file_png)

                # 루프 종료
                break
        except:
            traceback.print_exc(file=sys.stdout)

    # @staticmethod
    # def copy_and_paste_with_keeping_clipboard_current_contents(contents_new):
    #     # 제 프로그램의 클립보드 문제 발견했습니다, 클립보드 copy 를 수행전 내용을 변수에 저장해두었다가 copy 기능이 끝나면 되돌릴 수 있도록 만들었습니다
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     # 변수에 클립보드에 있던 이전내용 저장
    #     clipboard_current_contents = clipboard.paste()
    #
    #     # 클립보드에 새로운 내용 저장하고
    #     clipboard.copy(contents_new)
    #
    #     # ctrl v 로 새로운 내용 붙여넣기
    #     BusinessLogicUtil.press("ctrl", "v")
    #
    #     # 클립보드에 클립보드에 있던 이전내용 저장
    #     clipboard.copy(clipboard_current_contents)

    @staticmethod
    def is_void_function(func):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        """
            함수가 void 함수인지 아닌지 판단하는 함수입니다.

            Args:
              function: 함수

            Returns:
              함수가 void 함수이면 True, 아니면 False
        """
        function_code = func.__code__
        return function_code.co_argcount == 0 and function_code.co_flags & 0x20 == 0

    @staticmethod
    def get_count_args(func):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        return func.__code__.co_argcount

    @staticmethod
    def get_list_replaced_from_list_that_have_special_characters(target: [str], replacement: str):  # from [str] to [str]
        return [re.sub(pattern=r'[^a-zA-Z0-9가-힣\s]', repl='', string=string) for string in target]

    @staticmethod
    def get_str_replaced_special_characters(target: str, replacement: str):  # str to str
        target_replaced = re.sub(pattern=r'[^a-zA-Z0-9가-힣\s]', repl=replacement, string=target)  # 정규표현식을 사용하여 특수문자(알파벳, 숫자, 한글, 공백을 제외한 모든 문자) 제거
        return target_replaced

    # @staticmethod
    # def speak_alt_for_emergency(contents: str):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     BusinessLogicUtil.translate_eng_to_kor(contents)

    # @staticmethod
    # def get_all_pid_and_process_name():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     """모든 프로세스명 돌려주는 함수"""
    #     process_info = ""
    #
    #     def enum_windows_callback(hwnd, _):
    #         DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #         import psutil  # 실행중인 프로세스 및 시스템 활용 라이브러리
    #         nonlocal process_info
    #         if win32gui.IsWindowVisible(hwnd):
    #             _, pid = win32process.GetWindowThreadProcessId(hwnd)
    #             try:
    #                 process = psutil.Process(pid)
    #                 process_name = process.name()
    #                 process_info += f"창 핸들: {hwnd}, pid: {pid}, process_name: {process_name}\n"
    #             except psutil.NoSuchProcess:
    #                 pass
    #
    #     win32gui.EnumWindows(enum_windows_callback, None)
    #     return process_info

    @staticmethod
    def get_process_name_by_pid(pid):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        try:
            data = FileSystemUtil.get_cmd_output(cmd=f'tasklist | findstr "{pid}"')
            return data[0].split(" ")[0]
        except:
            return 0

    @staticmethod
    def move_window_to_front_by_pid(pid):  # deprecated
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        pass

    # @staticmethod
    # def activate_window_by_pid(pid: int):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     import psutil  # 실행중인 프로세스 및 시스템 활용 라이브러리
    #     if not str(pid).isdigit():
    #         DebuggingUtil.print_magenta(f"pid 분석결과 숫자가 아닌 것으로 판단됨 \n\n{pid}")
    #     pid = int(pid)
    #     try:
    #         process = psutil.Process(pid)
    #         if process.is_running() and process.status() != psutil.STATUS_ZOMBIE:
    #             hwnd = win32gui.FindWindow(None, None)  # PID를 기반으로 창 핸들을 가져옵니다.
    #             while hwnd:
    #                 _, found_pid = win32process.GetWindowThreadProcessId(hwnd)  # 창이 속한 프f로세스의 PID를 가져옵니다.
    #                 if found_pid == pid:
    #                     win32gui.SetForegroundWindow(hwnd)  # 창을 활성화합니다.
    #                     break
    #                 hwnd = win32gui.FindWindowEx(None, hwnd, None, None)
    #         else:
    #             DebuggingUtil.print_magenta(f"{StateManagementUtil.UNDERLINE_PROMISED}프로세스가 실행 중이지 않습니다.")
    #     except psutil.NoSuchProcess:
    #         DebuggingUtil.print_magenta(f"{StateManagementUtil.UNDERLINE_PROMISED}유효하지 않은 PID입니다.")

    # @staticmethod
    # def get_target_pid_by_process_name_legacy(target_process_name: str):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     # Q.how to activate certain program window at python?
    #     pids: str = BusinessLogicUtil.get_all_pid_and_process_name()
    #     # print(f"pids:\n\n{pids}")
    #     pids: list[str] = [i for i in pids.split("\n") if BusinessLogicUtil.is_regex_in_contents_with_case_ignored(contents=i, regex=target_process_name)]  # 프로세스명이 target_process_name 인 경우만 추출
    #     pids: str = pids[0].split(",")[1].replace("pid:", "").strip()  # strip() 은 특정 문자를 제거를 위해서 만들어짐. 단어를 제거하기 위해서는 replace() 가 더 적절하다고 chatGPT 는 말한다.
    #     target_pid = int(pids)  # 추출된 target_process_name 의 pid
    #     DebuggingUtil.print_magenta(f"target_process_name 프로세스 정보\n\n{target_pid}")
    #     return target_pid

    @staticmethod
    def get_target_pid_by_process_name(target_process_name: str):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        while True:
            pids = FileSystemUtil.get_cmd_output(f"tasklist | findstr {target_process_name}")
            # print(f"pids:\n\n{pids}")
            cnts: [int] = []
            for i in pids:
                cnts.append(i.count(" "))  # str 내의 특정문자의 개수를 cnts에 저장
            try:
                max_cnt = max(cnts)
            except ValueError:
                # Park4139.debug_as_cli(traceback.format_exc())
                DebuggingUtil.print_magenta(f"{target_process_name}에 대한 현재 실행중인 pid 가 없는 것 같습니다")
                break
            for i in range(0, max_cnt):  # max(cnts):  pids 요소별 가장 " " 가 많이 든 요소의 개수
                pids: [str] = [i.replace("  ", " ") for i in pids]  # 공백 제거

            pids_ = []
            for i in pids:
                if i.strip() != "\"":
                    if i.strip() != "":
                        # pids_.append(i.split(" "))  # 리스트 요소 내 "" 인 경우 제거, \" 인 경우 제거
                        pids_.append(list(map(str, i.split(" "))))  # 리스트 요소 내 "" 인 경우 제거, \" 인 경우 제거
            pids = pids_

            pid_and_size = {}
            for i in pids:
                pid_and_size[i[1]] = i[4]  # pid 와 size 를 튜플로 나눠 담았는데 이부분 생략하고 처음부터 남아 담아도 되었을 것 같은데

            pids = []
            sizes = []
            for pid, size in pid_and_size.items():
                pids.append(pid)
                sizes.append(size)

            for pid, size in pid_and_size.items():  # 프로세스명이 중복일 때 size 가 큰 프로세스의 pid를 리턴
                if size == max(sizes):
                    DebuggingUtil.print_magenta(f"pid : {pid}")
                    return int(pid)

    @staticmethod
    def print_today_time_info():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        yyyy = TimeUtil.get_time_as_('%Y')
        MM = TimeUtil.get_time_as_('%m')
        dd = TimeUtil.get_time_as_('%d')
        HH = TimeUtil.get_time_as_('%H')
        mm = TimeUtil.get_time_as_('%M')
        week_name = BusinessLogicUtil.return_korean_week_name()
        DebuggingUtil.print_magenta(f'현재 시각 {int(yyyy)}년 {int(MM)}월 {int(dd)}일 {week_name}요일 {int(HH)}시 {int(mm)}분')

    @staticmethod
    def connect_to_remote_computer_via_chrome_desktop():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        url = 'https://remotedesktop.google.com/access'
        FileSystemUtil.get_cmd_output(cmd=f'explorer "{url}"')

    @staticmethod
    def translate_eng_to_kor_via_googletrans(text: str):  # 수정할것 update 되었다는데 googletrans 업데이트 해보자
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        # from googletrans import Translator
        # translater = Translator()
        # text_translated = translater.translate([text_eng.split(".")], dest='ko')
        # for translation in text_translated:
        #     print(translation.origin, ' -> ', translation.text)
        # print(context=text_translated, is_app_instance_mode=True)

        # from papago_translate import Translator
        # naver papago api 서비스는 2024 년 종료가 된다...
        # translator = Translator(service_urls=['translate.google.com', 'translate.google.co.kr'],user_agent='Mozilla/5.0 (Windows NT 10.0;  x64; Win64)', timeout=random.randint(0, 99) / 10)
        # tmp = translator.translate(text, src="ko", dest="en")
        # print(rf'tmp : {tmp}')
        # print(rf'type(tmp) : {type(tmp)}')
        # print(rf'len(tmp) : {len(tmp)}')
        # Park4139TestUtil.pause()

        # 한수훈 씨가 만든 모듈 같다. 무료이다. 단, 안정성 보장은 안되며, google 로 부터 ip가 차단이 될 수 있다.
        # 단일 텍스트의 최대 글자수 제한은 15k입니다. 이거 로직으로 제한 해야 한다.
        # 약간 시간적인 트릭도 넣을 것
        # <Translated src=ko dest=en text=Good evening. pronunciation=Good evening.>
        # translator.translate('안녕하세요.', dest='ja')
        # <Translated src=ko dest=ja text=こんにちは。 pronunciation=Kon'nichiwa.>
        # translator.translate('veritas lux mea', src='la')
        # <Translated src=la dest=en text=The truth is my light pronunciation=The truth is my light>
        # 2023 12 31 01 11 현재시각기준 안된다

        # from googletrans import Translator
        # translator = Translator()
        # translator.detect('이 문장은 한글로 쓰여졌습니다.')
        # # <Detected lang=ko confidence=0.27041003>
        # translator.detect('この文章は日本語で書かれました。')
        # # <Detected lang=ja confidence=0.64889508>
        # translator.detect('This sentence is written in English.')
        # # <Detected lang=en confidence=0.22348526>
        # translator.detect('Tiu frazo estas skribita en Esperanto.')
        # # <Detected lang=eo confidence=0.10538048>

    @staticmethod
    def translate_kor_to_eng_foo(request: str):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        translator = Translator()
        text_translated = translator.translate(str(request), src='ko', dest='en')
        DebuggingUtil.print_magenta(context=text_translated, is_app_instance_mode=True)

    # @staticmethod
    # def keyDown(key: str):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     pyautogui.keyDown(key)
    #
    # @staticmethod
    # def keyUp(key: str):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     pyautogui.keyUp(key)

    @staticmethod
    def get_font_name_for_mataplot(font_abspath):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        import matplotlib
        import matplotlib.font_manager as mataplotlig_fontmanager

        # mataplot 폰트캐시 삭제
        cache_dir = matplotlib.get_cachedir()
        for file in os.listdir(cache_dir):
            if file.endswith('.json'):
                os.remove(os.path.join(cache_dir, file))

        # local font 를 mataplot 폰트 위치로 복사
        future_abspath = rf"{os.path.dirname(matplotlib.matplotlib_fname())}\fonts\ttf\{os.path.basename(font_abspath)}"
        if not os.path.exists(future_abspath):
            FileSystemUtil.xcopy_with_overwrite(target_abspath=font_abspath, future_target_abspath=future_abspath)
        font_name = mataplotlig_fontmanager.FontProperties(fname=future_abspath).get_name()
        return font_name

    @staticmethod
    def return_korean_week_name():
        weekday: str
        weekday = TimeUtil.get_time_as_('weekday')
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        if weekday == "0":
            return "월"
        elif weekday == "1":
            return "화"
        elif weekday == "2":
            return "수"
        elif weekday == "3":
            return "목"
        elif weekday == "4":
            return "금"
        elif weekday == "5":
            return "토"
        elif weekday == "6":
            return "일"
        else:
            return None

    # @staticmethod
    # def back_up_biggest_targets():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     DebuggingUtil.commentize(f"biggest_targets에 대한 빽업을 시도합니다")
    #     for biggest_target in StateManagementUtil.biggest_targets:
    #         BusinessLogicUtil.back_up_target(f'{biggest_target}')
    #
    # @staticmethod
    # def back_up_smallest_targets():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     DebuggingUtil.commentize(f"smallest_targets에 대한 빽업을 시도합니다")
    #     for target in StateManagementUtil.smallest_targets:
    #         BusinessLogicUtil.back_up_target(f'{target}')

    @staticmethod
    def classify_targets_between_smallest_targets_biggest_targets():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        DebuggingUtil.commentize(f'빽업할 파일들의 크기를 분류합니다.')
        targets = [
            rf"{StateManagementUtil.USERPROFILE}\Desktop\services\helper-from-youtube-url-to-webm",
            rf"{StateManagementUtil.USERPROFILE}\Desktop\services",
        ]
        DebuggingUtil.commentize(f'biggest_targets(300 메가 초과), smallest_targets(300 메가 이하) 분류 시도')
        for target in targets:
            target_size_megabite = FileSystemUtil.get_target_megabite(target.strip())
            print(target_size_megabite)
            if target_size_megabite <= 300:
                StateManagementUtil.smallest_targets.append(target.strip())

            elif 300 < target_size_megabite:
                StateManagementUtil.biggest_targets.append(target.strip())
            else:
                DebuggingUtil.print_magenta(f'{target.strip()}pass')

        DebuggingUtil.commentize(f'smallest_target 출력')
        # targets 에서 biggest_targets 과 일치하는 것을 소거 시도
        smallest_targets = [i for i in targets if i not in StateManagementUtil.biggest_targets]
        for target in StateManagementUtil.smallest_targets:
            print(target)

        DebuggingUtil.commentize(f'biggest_target 출력')
        for target in StateManagementUtil.biggest_targets:
            print(target)
        pass

    @staticmethod
    def gather_storages():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        starting_directory = os.getcwd()
        dst = StateManagementUtil.DIRECTORYIES_GATHERED
        if not os.path.exists(dst):
            return
        services = os.path.dirname(dst)
        os.chdir(services)
        storages = []
        cmd = rf'dir /b /s "{StateManagementUtil.USERPROFILE}\Downloads"'
        lines = FileSystemUtil.get_cmd_output(cmd)
        for line in lines:
            if line.strip() != "":
                storages.append(line.strip())

        DebuggingUtil.commentize(rf'archive_py 는 storage 목록 에서 제외')
        withouts = ['archive_py']
        for storage in storages:
            for without in withouts:
                if BusinessLogicUtil.is_regex_in_contents(target=storage, regex=without):
                    storages.remove(storage)
        for storage in storages:
            print(storage)

        DebuggingUtil.commentize(rf'이동할 storage 목록 중간점검 출력 시도')
        for storage in storages:
            print(os.path.abspath(storage))

        if not storages:
            DebuggingUtil.commentize(rf'이동할 storage 목록 이 없어 storage 이동을 할 수 없습니다')
        else:
            DebuggingUtil.commentize(rf'이동할 storage 목록 출력 시도')
            for storage in storages:
                print(os.path.abspath(storage))
            DebuggingUtil.commentize(rf'목적지 생성 시도')
            if not os.path.exists(dst):
                os.makedirs(dst)
            for storage in storages:
                # print(src)
                try:
                    DebuggingUtil.commentize(rf'storage 이동 시도')
                    FileSystemUtil.move_target_without_overwrite(storage, dst)
                except FileNotFoundError:
                    DebuggingUtil.trouble_shoot('202312071430')

                except Exception as e:
                    DebuggingUtil.trouble_shoot('20231205095308')

        os.chdir(starting_directory)

    @staticmethod
    def run_targets_promised():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        try:
            targets = [
                # Park4139.MUSIC_FOR_WORK,
                StateManagementUtil.PYCHARM64_EXE,
                StateManagementUtil.PROJECT_DIRECTORY,
                StateManagementUtil.SERVICE_DIRECTORY,
            ]
            for target in targets:
                FileSystemUtil.explorer(target)
        except:
            DebuggingUtil.trouble_shoot("%%%FOO%%%")

    @staticmethod
    def change_console_color():
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        TextToSpeechUtil.speak_ments(ment="테스트 콘솔의 색상을 바꿔볼게요", sleep_after_play=0.65)
        BusinessLogicUtil.toogle_console_color(color_bg='0', colors_texts=['7', 'f'])

    @staticmethod
    def taskkill_useless_programs():
        targets = [
            # "python.exe", # 프로그램 스스로를 종료시켜는 방법
            "alsong.exe",
            "cortana.exe",
            "mysqld.exe",
            "KakaoTalk.exe",
            "OfficeClickToRun.exe",
            "TEWebP.exe",
            "TEWeb64.exe",
            "TEWebP64.exe",
            "AnySign4PC.exe",
        ]
        for target in targets:
            BusinessLogicUtil.taskkill(target)

    # @staticmethod
    # def run_up_and_down_game():
    #     """업 앤 다운 게임 mkmk"""
    #     correct_answer: int = random.randint(1, 100)
    #     left_oportunity: int = 10
    #     ment = f"<UP AND DOWN GAME>\n\nFIND CORRECT NUMBER"
    #     TextToSpeechUtil.speak_ments(ment, sleep_after_play=1.0)
    #     dialog = UiUtil.CustomQdialog(ment=ment, btns=["START", "EXIT"], is_input_box=False)
    #     dialog.exec()
    #     user_input = None
    #     is_game_strated = False
    #     btn_text_clicked = dialog.btn_text_clicked
    #
    #     # def get_input_remove_not_intended(user_input : str):
    #     # def remove_weired_input(user_input : str):
    #
    #     if btn_text_clicked == "START":
    #         ment = f"START IS PRESSED, LETS START GAME"
    #         TextToSpeechUtil.speak_ments(ment, sleep_after_play=0.65, thread_join_mode=True)
    #         while left_oportunity >= 0:
    #             if left_oportunity == 0:
    #                 ment = f"LEFT CHANCE IS {left_oportunity} \nTAKE YOUR NEXT CHANCE."
    #                 TextToSpeechUtil.speak_ments(ment, sleep_after_play=0.65)
    #                 dialog = UiUtil.CustomQdialog(ment=ment, btns=["EXIT"])
    #                 dialog.exec()
    #                 break
    #             elif is_game_strated == False or user_input is None:
    #                 ment = f"TYPE NUMBER BETWEEN 1 TO 100"
    #                 if user_input is None:
    #                     ment = ment + ", AGAIN"
    #                 TextToSpeechUtil.speak_ments(ment, sleep_after_play=0.65)
    #                 dialog = UiUtil.CustomQdialog(ment=ment, btns=["SUBMIT"], is_input_box=True)
    #                 dialog.exec()
    #                 user_input = BusinessLogicUtil.is_user_input_sanitized(dialog.input_box.text())
    #                 if user_input is not None:
    #                     left_oportunity = left_oportunity - 1
    #                 is_game_strated = True
    #             elif user_input == correct_answer:
    #                 ment = f"CONGRATULATIONS, YOUR NUMBER IS {correct_answer}\nTHIS IS ANSWER, SEE YOU AGAIN"
    #                 TextToSpeechUtil.speak_ments(ment, sleep_after_play=0.65)
    #                 dialog = UiUtil.CustomQdialog(ment=ment, btns=["EXIT"])
    #                 dialog.exec()
    #                 break
    #             elif correct_answer < user_input:
    #                 ment = f"YOUR NUMBER IS {user_input}\n\nYOU NEED DOWN\nYOUR LEFT CHANCE IS {left_oportunity}"
    #                 TextToSpeechUtil.speak_ments(ment, sleep_after_play=0.65)
    #                 dialog = UiUtil.CustomQdialog(ment=ment, btns=["SUBMIT"], is_input_box=True)
    #                 dialog.exec()
    #                 user_input = BusinessLogicUtil.is_user_input_sanitized(dialog.input_box.text())
    #                 if user_input is not None:
    #                     left_oportunity = left_oportunity - 1
    #             elif correct_answer > user_input:
    #                 ment = f"YOUR NUMBER IS {user_input}\n\nYOU NEED UP\nYOUR LEFT CHANCE IS {left_oportunity}"
    #                 TextToSpeechUtil.speak_ments(ment, sleep_after_play=0.65)
    #                 dialog = UiUtil.CustomQdialog(ment=ment, btns=["submit"], is_input_box=True)
    #                 dialog.exec()
    #                 user_input = BusinessLogicUtil.is_user_input_sanitized(dialog.input_box.text())
    #                 if user_input is not None:
    #                     left_oportunity = left_oportunity - 1
    #     else:
    #         return
    #
    # @staticmethod
    # def run_console_blurred_as_scheduler_as_thread(q_application: QApplication):
    #     async def run_scheduler(q_application):
    #         # schedule.every(30).minutes.do(partial(Park4139.speak_after_x_min, 30))
    #         # schedule.every().tuesday.at("15:00").do(Park4139.speak_today_time_info)
    #         # schedule.every().wednesday.at("15:00").do(Park4139.speak_today_time_info)
    #         # schedule.every().thursday.at("15:00").do(Park4139.speak_today_time_info)
    #         # schedule.every().friday.at("15:00").do(Park4139.speak_today_time_info)
    #         # schedule.every().saturday.at("15:00").do(Park4139.speak_today_time_info)
    #         # schedule.every().sunday.at("15:00").do(Park4139.speak_today_time_info)
    #         # schedule.every(1).hour.do(Park4139.speak_server_hh_mm)
    #         # schedule.every(1).day.do(Park4139.speak_server_hh_mm)
    #         # schedule.every(1).day.at("12:30").do(Park4139.speak_server_hh_mm)
    #         # while True:
    #         # schedule.run_pending()
    #         # Park4139.debug_as_cli(f"async def {inspect.currentframe().f_code.co_name}() is running...")
    #         # await ....sleep(1000)
    #         BusinessLogicUtil.run_console_blurred_core_as_scheduler(q_application)
    #         pass
    #
    #     def run_async_loop(q_application):
    #         loop = asyncio.new_event_loop()
    #         asyncio.set_event_loop(loop)
    #         loop.run_until_complete(run_scheduler(q_application))
    #
    #     thread = threading.Thread(target=partial(run_async_loop, q_application))
    #     thread.start()
    #
    # @staticmethod
    # def run_console_blurred_as_gui_as_thread(q_application: QApplication):
    #     # 비동기 이벤트 함수 설정 (simple for void async processing)
    #     async def run_rpa_program(q_application):
    #         dialog = UiUtil.CustomDialog(q_application=q_application, q_wiget=UiUtil.RpaProgramMainWindow(q_application=q_application), is_app_instance_mode=True)
    #         # dialog = UiUtil.CustomDialog(q_application=q_application, q_wiget=UiUtil.RpaProgramMainWindow(), is_app_instance_mode=True)
    #         # dialog = UiUtil.CustomDialog(q_application=q_application, q_wiget=UiUtil.RpaProgramMainWindow, is_app_instance_mode=True)
    #
    #     # 비동기 이벤트 루프 설정 (simple for void async processing)
    #     def run_async_event_loop(q_application):
    #         loop = asyncio.new_event_loop()
    #         asyncio.set_event_loop(loop)
    #         loop.run_until_complete(run_rpa_program(q_application))
    #
    #     # 비동기 이벤트 루프 실행할 쓰레드 설정 (simple for void async processing)
    #     thread_for_run_scheduler = threading.Thread(target=partial(run_async_event_loop, q_application))
    #     thread_for_run_scheduler.start()
    #
    # @staticmethod
    # def run_console_blurred_core_as_gui(q_application: QApplication):
    #     window = UiUtil.RpaProgramMainWindow(q_application)
    #     q_application.exec()
    #
    # @staticmethod
    # def run_console_blurred_core_as_scheduler(q_application: QApplication):
    #     scheduler_loop_cnt = 0
    #     yyyy = TimeUtil.get_time_as_('%Y')
    #     while True:
    #         # 루프마다 == 가능한 짧은 시간 마다
    #         yyyy_previous_loop = yyyy
    #         yyyy = TimeUtil.get_time_as_('%Y')
    #         MM = TimeUtil.get_time_as_('%m')
    #         dd = TimeUtil.get_time_as_('%d')
    #         HH = TimeUtil.get_time_as_('%H')
    #         mm = TimeUtil.get_time_as_('%M')
    #         ss = TimeUtil.get_time_as_('%S')
    #         server_time = TimeUtil.get_time_as_(rf'%Y-%m-%d %H:%M:%S')
    #         # 첫번째 루프만 mkmk
    #         if scheduler_loop_cnt == 0:
    #             # QThread 설정
    #             # thread = CustomQthread(q_application=q_application)
    #             # thread.finished.connect(lambda: Park4139Park4139.Tts.speak("QThread 작업이 성공되었습니다"))
    #             # thread.start()
    #
    #             # 프로그램 외 전역감지 단축키 이벤트 설정
    #             # shortcut_keys_up_promised = {
    #             #     # "<ctrl>+h": partial(q_wiget.toogle_rpa_window, "<ctrl>+h"),  # fail
    #             #     "<ctrl>+A": Park4139.ask_something_to_ai, # fail
    #             # "<ctrl>+H": Park4139.단축키 목록 토글, # 이거 하나만이라도 되게 만들자. # 진짜 그래도 안되면 python rpa 프로그램을 따로 하나 더 띄우게 명령어를 설정하자.
    #             # Park4139.should_i_merge_directories,
    #             # Park4139.should_i_download_youtube_as_webm,
    #             # }
    #             # keyboard_main_listener = pynput.keyboard.GlobalHotKeys(shortcut_keys_up_promised)
    #             # keyboard_main_listener.start()
    #
    #             BusinessLogicUtil.should_i_check_your_promised_routines_before_coding(StateManagementUtil.routines_promised)
    #
    #             BusinessLogicUtil.should_i_do(ment="약속된 타겟들을 실행할까요?", function=BusinessLogicUtil.run_targets_promised, auto_click_negative_btn_after_seconds=5)
    #
    #             BusinessLogicUtil.should_i_do(ment="약속된 불필요한 프로그램을 종료할까요?", function=BusinessLogicUtil.taskkill_useless_programs, auto_click_negative_btn_after_seconds=5)
    #
    #             # TtsUtil.speak_with_queue("랜덤스케쥴을 실행했습니다") # 루프 돌때마다 1번만 speak_queue 에 쌓인 내용을 speak_as_sync 읇어주도록
    #             BusinessLogicUtil.should_i_do(ment="랜덤 스케쥴을 실행할까요?", function=BusinessLogicUtil.do_random_schedules, auto_click_negative_btn_after_seconds=5)
    #
    #             if not DbTomlUtil.is_accesable_local_database():
    #                 TextToSpeechUtil.speak_ments("로컬 데이터베이스에 접근할 수 없어 접근이 가능하도록 설정했습니다", sleep_after_play=0.65)
    #
    #             # BusinessLogicUtil.should_i_do(ment="백업할 타겟들을 크기에 따라 분류를 해둘까요?", function=BusinessLogicUtil.classify_targets_between_smallest_targets_biggest_targets, auto_click_positive_btn_after_seconds=10)
    #             BusinessLogicUtil.should_i_do(ment="백업할 타겟들을 크기에 따라 분류를 해둘까요?", function=BusinessLogicUtil.classify_targets_between_smallest_targets_biggest_targets, auto_click_negative_btn_after_seconds=5)
    #
    #         # 루프 카운트 갱신
    #         scheduler_loop_cnt = scheduler_loop_cnt + 1
    #         print(f"scheduler_loop_cnt:{scheduler_loop_cnt} scheduler_time:{server_time}")
    #
    #         # 0시에서 24시 사이, # 분단위 스케쥴
    #         if 0 <= int(HH) <= 24 and int(ss) == 0:
    #             BusinessLogicUtil.monitor_target_edited_and_back_up(target_abspath=StateManagementUtil.PARK4139_ARCHIVE_TOML)  # seconds_performance_test_results : ['11.95sec', '26.78sec', '11.94sec', '3.7sec', '11.72sec']
    #             if int(HH) == 6 and int(mm) == 30:
    #                 # TextToSpeechUtil.speak_ments(f'{int(HH)} 시 {int(mm)}분 루틴을 시작합니다', sleep_after_play=0.65, thread_join_mode=True)  # 쓰레드가 많아지니 speak() 하면서 부정확한 재생이 늘어났다. 쓰레드의 정확한 타이밍 제어가 필요한 것 같다. 급한대로 thread_join_mode 를 만들었다)
    #                 # TextToSpeechUtil.speak_ments(f'아침음악을 준비합니다, 아침음악을 재생할게요', sleep_after_play=0.65, thread_join_mode=True)
    #                 pass
    #             if int(HH) == 7 and int(mm) == 30:
    #                 # TextToSpeechUtil.speak_ments(f'{int(HH)} 시 {int(mm)}분 루틴을 시작합니다', sleep_after_play=0.65, thread_join_mode=True)
    #                 # TextToSpeechUtil.speak_ments('지금 나가지 않는다면 지각할 수 있습니다, 더이상 나가는 것을 지체하기 어렵습니다', sleep_after_play=0.65, thread_join_mode=True)
    #                 pass
    #
    #             if int(HH) == 8 and int(mm) == 50:
    #                 # TextToSpeechUtil.speak_ments(f'{int(HH)} 시 {int(mm)}분 루틴을 시작합니다', sleep_after_play=0.65, thread_join_mode=True)
    #                 # TextToSpeechUtil.speak_ments('업무시작 10분전입니다, 업무준비를 시작하세요, 업무 전에 세수와 양치는 하셨나요', sleep_after_play=0.65, thread_join_mode=True)
    #                 pass
    #
    #             if int(HH) == 9:
    #                 # TextToSpeechUtil.speak_ments(f'{int(mm)}시 정각, 루틴을 시작합니다', sleep_after_play=0.65, thread_join_mode=True)
    #                 # TextToSpeechUtil.speak_ments('근무시간이므로 음악을 종료합니다', sleep_after_play=0.65, thread_join_mode=True)
    #                 BusinessLogicUtil.taskkill('ALSong.exe')
    #             if int(HH) == 11 and int(mm) == 30:
    #                 # TextToSpeechUtil.speak_ments(f'{int(HH)} 시 {int(mm)}분 루틴을 시작합니다', sleep_after_play=0.65, thread_join_mode=True)
    #                 # TextToSpeechUtil.speak_ments('점심시간입니다, 음악을 재생합니다', sleep_after_play=0.65, thread_join_mode=True)
    #                 # TextToSpeechUtil.speak_ments('용량이 큰 약속된 타겟들을 빽업을 수행 시도합니다', sleep_after_play=0.65, thread_join_mode=True)
    #                 BusinessLogicUtil.back_up_biggest_targets()
    #                 # TextToSpeechUtil.speak_ments('용량이 작은 약속된 타겟들을 빽업을 수행 시도합니다', sleep_after_play=0.65, thread_join_mode=True)
    #                 BusinessLogicUtil.back_up_smallest_targets()
    #                 # TextToSpeechUtil.speak_ments('흩어져있는 storage 들을 한데 모으는 시도를 합니다', sleep_after_play=0.65, thread_join_mode=True)
    #                 BusinessLogicUtil.gather_storages()
    #             if int(HH) == 22 and int(mm) == 10:
    #                 # TextToSpeechUtil.speak_ments(f'{int(HH)} 시 {int(mm)}분 루틴을 시작합니다', sleep_after_play=0.65, thread_join_mode=True)
    #                 TextToSpeechUtil.speak_ments('씻으실 것을 추천드립니다, 샤워루틴을 수행하실 것을 추천드립니다', sleep_after_play=0.65, thread_join_mode=True)  # 샤워루틴 확인창 띄우기
    #             if int(HH) == 22 and int(mm) == 30:
    #                 # TextToSpeechUtil.speak_ments(f'{int(HH)} 시 {int(mm)}분 루틴을 시작합니다', sleep_after_play=0.65, thread_join_mode=True)
    #                 TextToSpeechUtil.speak_ments('건강을 위해서 지금 씻고 주무실 것을 추천드려요', sleep_after_play=0.65, thread_join_mode=True)
    #             if int(HH) == 24:
    #                 TextToSpeechUtil.speak_ments(f'자정이 되었습니다', sleep_after_play=0.65, thread_join_mode=True)
    #             if int(mm) % 15 == 0:
    #                 # TextToSpeechUtil.speak_ments(f'15분 간격 루틴을 시작합니다', sleep_after_play=0.65, thread_join_mode=True)
    #                 # TextToSpeechUtil.speak_ments(f'랜덤 스케줄을 시작합니다', sleep_after_play=0.65, thread_join_mode=True)
    #                 BusinessLogicUtil.do_random_schedules()
    #                 # TextToSpeechUtil.speak_ments(f'프로젝트 디렉토리 싱크를 시작합니다', sleep_after_play=0.65, thread_join_mode=True)
    #             if int(mm) % 30 == 0:
    #                 # TextToSpeechUtil.speak_ments(f'30분 간격 루틴을 시작합니다', sleep_after_play=0.65, thread_join_mode=True)
    #                 # TextToSpeechUtil.speak_ments(f'깃허브로 파이썬 아카이브 프로젝트 빽업을 시도합니다', sleep_after_play=0.65, thread_join_mode=True)
    #                 BusinessLogicUtil.git_push_by_auto()
    #                 BusinessLogicUtil.monitor_target_edited_and_sync(target_abspath=StateManagementUtil.PROJECT_DIRECTORY)  # seconds_performance_test_results : ['28.46sec', '27.53sec', '2.85sec', '2.9sec', '2.91sec']
    #             if int(mm) % 60 == 0:
    #                 # TextToSpeechUtil.speak_ments(f'1시간 간격 루틴을 시작합니다', sleep_after_play=0.65, thread_join_mode=True)
    #                 BusinessLogicUtil.should_i_do(ment="쓰레기통을 비울까요?", function=BusinessLogicUtil.empty_recycle_bin, auto_click_negative_btn_after_seconds=10)
    #                 BusinessLogicUtil.should_i_do(ment="오늘 시간정보를 말씀드릴까요?", function=TextToSpeechUtil.speak_today_time_info, auto_click_negative_btn_after_seconds=10)
    #                 BusinessLogicUtil.monitor_target_edited_and_sync(target_abspath=StateManagementUtil.SERVICE_DIRECTORY)  # seconds_performance_test_results : ['28.46sec', '27.53sec', '2.85sec', '2.9sec', '2.91sec']
    #
    #         # 0시에서 4시 사이, 30초 마다
    #         if 0 <= int(HH) <= 4 and int(ss) % 30 == 0:
    #             BusinessLogicUtil.make_me_go_to_sleep()
    #
    #         if BusinessLogicUtil.is_same_time(time1=server_time, time2=datetime(year=2024, month=1, day=2, hour=23, minute=50, second=0)):
    #             TextToSpeechUtil.speak_server_hh_mm()
    #
    #         # 하루인사
    #         unique_id = "already_said_about_today_greeting"
    #         already_said_about_today_greeting = DbTomlUtil.select_db_toml(key=DbTomlUtil.get_db_toml_key(unique_id=unique_id))
    #         if already_said_about_today_greeting == None:
    #             DbTomlUtil.insert_db_toml(key=DbTomlUtil.get_db_toml_key(unique_id=unique_id), value=False)
    #             already_said_about_today_greeting = DbTomlUtil.select_db_toml(key=DbTomlUtil.get_db_toml_key(unique_id=unique_id))
    #         if already_said_about_today_greeting == False:
    #             ment = f'오늘도 좋은 하루 되세요!'
    #             TextToSpeechUtil.speak_ments(ment=ment, sleep_after_play=0.65, thread_join_mode=True)
    #             BusinessLogicUtil.should_i_do(ment=ment, auto_click_positive_btn_after_seconds=100)
    #             DbTomlUtil.update_db_toml(key=DbTomlUtil.get_db_toml_key(unique_id=unique_id), value=True)  # DB_TOML 업데이트
    #         if BusinessLogicUtil.is_midnight():  # 자정이면 초기화
    #             DbTomlUtil.update_db_toml(key=DbTomlUtil.get_db_toml_key(unique_id=unique_id), value=False)  # 자정에 초기화 잘되면, 리펙토링 수행하자
    #
    #         # 신년 축하 인사하기
    #         # print(rf'yyyy_previous_loop : {yyyy_previous_loop}')
    #         # print(rf'yyyy : {yyyy}')
    #         if yyyy != yyyy_previous_loop:
    #             unique_id = "already_said_about_new_year_today"
    #             already_said_about_new_year_today = DbTomlUtil.select_db_toml(key=DbTomlUtil.get_db_toml_key(unique_id=unique_id))
    #             if already_said_about_new_year_today == None:
    #                 DbTomlUtil.insert_db_toml(key=DbTomlUtil.get_db_toml_key(unique_id=unique_id), value=yyyy_previous_loop)
    #                 already_said_about_new_year_today = DbTomlUtil.select_db_toml(key=DbTomlUtil.get_db_toml_key(unique_id=unique_id))
    #             if already_said_about_new_year_today == False:
    #                 ment = f'{yyyy} 신년입니다! 새해에는 반드시 좋은일이 가득하시길 바랄게요!!'
    #                 TextToSpeechUtil.speak_ments(ment=ment, sleep_after_play=0.65, thread_join_mode=True)
    #                 BusinessLogicUtil.should_i_do(ment=ment, auto_click_positive_btn_after_seconds=100)
    #                 DbTomlUtil.update_db_toml(key=DbTomlUtil.get_db_toml_key(unique_id=unique_id), value=yyyy)  # DB_TOML 업데이트
    #             if BusinessLogicUtil.is_midnight():  # 자정이면 초기화
    #                 DbTomlUtil.update_db_toml(key=DbTomlUtil.get_db_toml_key(unique_id=unique_id), value=False)
    #
    #         # 크리스마스 축하 인사하기
    #         if BusinessLogicUtil.is_christmas():
    #             unique_id = "already_said_about_christmas_today"
    #             already_said_about_christmas_today = DbTomlUtil.select_db_toml(key=DbTomlUtil.get_db_toml_key(unique_id=unique_id))
    #             if already_said_about_christmas_today == None:
    #                 DbTomlUtil.insert_db_toml(key=DbTomlUtil.get_db_toml_key(unique_id=unique_id), value=False)
    #                 already_said_about_christmas_today = DbTomlUtil.select_db_toml(key=DbTomlUtil.get_db_toml_key(unique_id=unique_id))
    #             if already_said_about_christmas_today == False:
    #                 ment = f'{yyyy}년 크리스마스입니다! 즐거운 크리스마스 되세요!'
    #                 TextToSpeechUtil.speak_ments(ment=ment, sleep_after_play=0.65, thread_join_mode=True)
    #                 BusinessLogicUtil.should_i_do(ment=ment, auto_click_positive_btn_after_seconds=100)
    #                 DbTomlUtil.update_db_toml(key=DbTomlUtil.get_db_toml_key(unique_id=unique_id), value=True)  # DB_TOML 업데이트
    #             if BusinessLogicUtil.is_midnight():  # 자정이면 초기화
    #                 DbTomlUtil.update_db_toml(key=DbTomlUtil.get_db_toml_key(unique_id=unique_id), value=False)
    #
    #         BusinessLogicUtil.sleep(1000)
    #         # await asyncio.sleep(1)  # 비동기 처리된 루프 sleep() 시 , 비동기 이벤트 함수 내에서는 time.sleep() 대신 await asyncio.sleep(5)를 사용해야한다
    #
    #         # 1년에 한번 수행 아이디어
    #         # random_schedule.json 에서 leaved_max_count를 읽어온다
    #         # leaved_max_count=1 이면 년에 1번 수행 하도록
    #         # leaved_max_count=10 이면 년에 10번 수행 하도록
    #         # leaved_max_count=0 이면 올해에는 더이상 수행하지 않음
    #         # leaved_max_count 를 random_schedule_tb.toml 에 저장
    #
    #         # - 1시간 뒤 시스템 종료 예약 기능
    #         # - 즉시 시스템 종료 시도 기능
    #         # - 시간 시현기능 기능(autugui 이용)
    #         #   ment ='pc 정밀검사를 한번 수행해주세요'
    #         #   commentize(ment)
    #         # - 하드코딩된 스케줄 작업 수행 기능
    #         # - 미세먼지 웹스크래핑 기능
    #         # - 초미세먼지 웹스크래핑 기능
    #         # - 종합날씨 웹스크래핑 기능
    #         # - 습도 웹스크래핑 기능
    #         # - 체감온도 웹스크래핑 기능
    #         # - 현재온도 웹스크래핑 기능
    #         # - 음악재생 기능
    #         # - 영상재생 기능

    # @staticmethod
    # def run_console_blurred_as_scheduler():
    #
    #     # pyautogui 페일세이프 모드 설정
    #     # pyautogui.FAILSAFE = True # 페일세이프 너무 불편한데. 매크로 작업 시에만 False 하면 어떨까 싶다
    #     pyautogui.FAILSAFE = False
    #
    #     # global app
    #     # app = QApplication(sys.argv)
    #     # global 로 app을 설정 하고 싶진 않았지만 app.primaryScreen() 동작에 필요했다.
    #     # app.primaryScreen()의 기능에 대한 대체 방법이 있다면 global app 없애고 싶다, 공유객체로 해소가 될 것 같은데 더 쉬운 방법을 못찾았다
    #     # 유사방법을 쓰긴 했는데 이상하게 동작한다. 대안모색필요
    #
    #     # print(context=f"TEST LOOP ERROR CNT REPORT:", is_app_instance_mode=True)
    #
    #     # 시작스케쥴러 설정
    #     # Park4139.run_scheduler_as_thread()# Qthread 가 아니라 Thread 형태 이면 동작은 하는데, 스케쥴러 내에서 pyside6 Qdialog 동작하지 않는다.
    #     # run_scheduler_as_qthread() # 이렇게 함수에 넣으면 Process finished with exit code -1073740791 (0xC0000409) 이 에러와 함께 바로 앱 종료.
    #
    #     # QThread 로 scheduler 동작 테스트 시도
    #     # thread = SchedulerAsQthread()
    #     # thread.finished.connect(lambda: print("쓰레드 작업이 성공되었습니다"))
    #     # thread.start()
    #     # 이 방법도 결국 되지않았다. Process finished with exit code -1073741819 (0xC0000005) 이 에러와 함께 바로 앱 종료.
    #
    #     # 3 가지 try 모두 실패하였다.
    #     # scheduler를 pyside6 안에서 돌리는 것 까지는 가능한데, scheduler 를 Qthread , Threading 시도 모두 실패하였다. 아예 엎기로 생각했다. 꼭 scheduler 에서 dialog 를 띄우고 말거다.
    #     # 이제는 pyside6 app 안에서 scheduler 를 띄우고 그 위에 QDialog 를 띄우는 시도는 하지 않을 것이다.
    #     # 다른 시도할 기획은 이렇다 scheduler 를 돌리고 Qapplication을 Qdialog를 wrapping 한뒤 이 wrap 된 Qdialog 를 띄우고 Qdialog 종료 시 Qapplicaiton 을 종료할 것이다.
    #     # pyside6를 이렇게 쓰라고 만든건 아닌것 같은데... # CustomDialog() 객체를 만들게 되었고, 지금까지는 의도한대로 동작 테스트 되었다.
    #     # 이제 구현 되었다.
    #     # CustomDialog() 인스턴스 생성 테스트를 해보니, 분명 q_application 과 q_wiget 는 named argument 인데도, positional argument 처럼 순서가 중요했다는 것을 알 수 있었다, q_application 다음에 q_wiget 을 초기화 하자! 이건 pyside6 의 공식 규칙이다.
    #
    #     q_application = QApplication(sys.argv)
    #     dialog = UiUtil.CustomDialog(q_application=q_application, q_wiget=UiUtil.CustomQdialog(ment=f"다음의 프로젝트 디렉토리에서 자동화 프로그램이 시작됩니다\n{StateManagementUtil.PROJECT_DIRECTORY}", btns=["실행", "실행하지 않기"], auto_click_positive_btn_after_seconds=0), is_app_instance_mode=True)
    #     # dialog = CustomDialog(q_application=q_application, q_wiget=UiUtil.CustomQdialog(context=f"다음의 프로젝트 디렉토리에서 자동화 프로그램이 시작됩니다\n{Park4139.PROJECT_DIRECTORY}", buttons=["실행", "실행하지 않기"], auto_starting_seconds=10), is_app_instance_mode=True)
    #     if dialog.btn_text_clicked == "실행":
    #         print(StateManagementUtil.PROJECT_DIRECTORY)
    #         os.chdir(StateManagementUtil.PROJECT_DIRECTORY)
    #
    #         # 프로그램 코어 진입지점
    #         BusinessLogicUtil.run_console_blurred_core_as_scheduler(q_application)
    #         # Park4139.run_console_blurred_core_as_gui(q_application)
    #
    #     if dialog.btn_text_clicked == "실행하지 않기":
    #         sys.exit()
    #
    #     # pyside6 app instance 를 삭제하고 새 instnace를 만드는 게 어려운 일인 것 같다.
    #     # q_application 를 scheduler 에 전달해서 써는 시도 방안이 떠올랐다.
    #     # pyside6 app 처럼 진정 singletone 설계를 하면 계속 이렇게 인스턴스를 붙들고 써야하나 보다.
    #
    #     # pyside6 창 간 통신 설정
    #     # shared_obj = SharedObject()
    #     # rpa_program_main_window = RpaProgramMainWindow(shared_obj=shared_obj) # 공유객체인 shared_obj 가 사용되는 곳이 없고 불필요할 것으로 판단하여 제거
    #
    # @staticmethod
    # def run_console_blurred_as_gui():
    #     # colorama 초기화, 이거 프로젝트의 root 근처의 상단에 설정 해두어야 함. 재호출 될때 많이엄청느려진다. root 근처에서 호출되는게 아니라, colorma 호출 직전 초기화 되도록 수정하였음.
    #     # init(autoreset=True)
    #     # init(autoreset=True)
    #
    #     pyautogui.FAILSAFE = False
    #     q_application = QApplication(sys.argv)
    #     dialog = UiUtil.CustomDialog(q_application=q_application, q_wiget=UiUtil.CustomQdialog(ment=f"다음의 프로젝트 디렉토리에서 자동화 프로그램이 시작됩니다\n{StateManagementUtil.PROJECT_DIRECTORY}", btns=["실행", "실행하지 않기"], auto_click_positive_btn_after_seconds=0), is_app_instance_mode=True)
    #     # dialog = CustomDialog(q_application=q_application, q_wiget=UiUtil.CustomQdialog(context=f"다음의 프로젝트 디렉토리에서 자동화 프로그램이 시작됩니다\n{Park4139.PROJECT_DIRECTORY}", buttons=["실행", "실행하지 않기"], auto_starting_seconds=10), is_app_instance_mode=True)
    #     if dialog.btn_text_clicked == "실행":
    #         print(StateManagementUtil.PROJECT_DIRECTORY)
    #         os.chdir(StateManagementUtil.PROJECT_DIRECTORY)
    #
    #         # 프로그램 코어 진입지점
    #         # Park4139.run_console_blurred_core_as_scheduler(q_application)
    #         BusinessLogicUtil.run_console_blurred_core_as_gui(q_application)
    #
    #     if dialog.btn_text_clicked == "실행하지 않기":
    #         sys.exit()

    @staticmethod
    def is_christmas():
        import datetime
        today = datetime.datetime.now()
        christmas = datetime.datetime(year=today.year, month=12, day=25)
        if today == christmas:
            return True
        else:
            return False

    @staticmethod
    def is_same_time(time1, time2):
        time2.strftime(rf'%Y-%m-%d %H:%M:%S')
        # print(rf'time1 : {time1} , time2 : {time2}')
        if time1 == time2:
            return True
        else:
            return False

    @staticmethod
    def is_midnight():
        import datetime
        now = datetime.datetime.now()
        if now.hour == 0 and now.minute == 0 and now.second == 0:
            return True
        else:
            return False

    # @staticmethod
    # def back_up_project_and_change_to_power_saving_mode():
    #     BusinessLogicUtil.back_up_target(target_abspath=StateManagementUtil.PROJECT_DIRECTORY)
    #     FileSystemUtil.enter_power_saving_mode()

    # @staticmethod
    # @TestUtil.measure_seconds_performance_once  #
    # def crawl_html_href(url: str):
    #     DebuggingUtil.print_ment_via_colorama(f"{StateManagementUtil.UNDERLINE_PROMISED} {inspect.currentframe().f_code.co_name}", colorama_color=ColoramaUtil.BLUE)
    #     # 최하단으로 스크롤 이동 처리를 추가로 해야함. 그렇지 않으면 기대하는 모든 영상을 크롤링 할 수 없음..귀찮..지만 처리했다.
    #
    #     # url 전처리
    #     url = url.strip()
    #
    #     # driver 설정
    #     DebuggingUtil.print_ment_via_colorama(f"SeleniumUtil.get_driver_for_selenium() 수행 중...", colorama_color=ColoramaUtil.BLUE)
    #     driver = SeleniumUtil.get_driver_for_selenium()
    #
    #     DebuggingUtil.print_ment_via_colorama(f"driver.get(target_url) 수행 중...", colorama_color=ColoramaUtil.BLUE)
    #     target_url = url
    #     driver.get(target_url)
    #
    #     # 자동제어 브라우저 화면 초기 로딩 random.randint(1,n) 초만큼 명시적 대기
    #     n = 2
    #     seconds = random.randint(1, n)
    #     DebuggingUtil.print_ment_via_colorama(f"자동제어 브라우저 화면 초기 로딩 중... {seconds} seconds", colorama_color=ColoramaUtil.BLUE)
    #     driver.implicitly_wait(seconds)  # 처음페이지 로딩이 끝날 때까지 약 random.randint(1,n)초 대기
    #
    #     # 최하단으로 자동 스크롤, 페이지 최하단에서 더이상 로딩될 dom 객체가 없을 때 까지
    #     DebuggingUtil.print_ment_via_colorama("스크롤 최하단으로 이동 중...", colorama_color=ColoramaUtil.BLUE)
    #     scroll_cnt = 0
    #     previous_scroll_h = None
    #     current_scroll_h = None
    #     scroll_maxs_monitored = []
    #     while True:
    #         if current_scroll_h is not None and previous_scroll_h is not None:
    #             if previous_scroll_h == current_scroll_h:
    #                 scroll_maxs_monitored.append(True)
    #                 # break
    #
    #         # 로딩타이밍 제어가 어려워 추가한 코드. n번 모니터링.
    #         n = 6  # success
    #         if len(scroll_maxs_monitored) == n:
    #             if all(scroll_maxs_monitored) == True:  # [bool] bool list 내 요소가 모두 true 인지 확인
    #                 DebuggingUtil.print_ment_via_colorama(ment="스크롤 최하단으로 이동되었습니다", colorama_color=ColoramaUtil.BLUE)
    #                 break
    #
    #         # previous_scroll_h 업데이트
    #         # previous_scroll_h = driver.execute_script("return document.body.scrollHeight")
    #         previous_scroll_h = driver.execute_script("return document.documentElement.scrollHeight")
    #
    #         # 가능한만큼 스크롤 최하단으로 이동
    #         # driver.find_element(By.CSS_SELECTOR, 'body').send_keys(Keys.PAGE_DOWN)  # page_down 을 누르는 방법, success
    #         # driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")# JavaScript 로 스크롤 최하단으로 이동, 네이버용 코드?
    #         driver.execute_script("window.scrollTo(0, document.documentElement.scrollHeight);")  # JavaScript 로 스크롤 최하단으로 이동, 유튜브용 코드?
    #         # time.sleep(2)  # 스크롤에 의한 추가적인 dom 객체 로딩 대기, 여러가지 예제를 보니, 일반적으로 2 초 정도 두는 것 같음. 2초 내에 로딩이 되지 않을 때도 있는데.
    #         BusinessLogicUtil.sleep(milliseconds=500, printing_mode=False)  # 스크롤에 의한 추가적인 dom 객체 로딩 대기, success, 지금껏 문제 없었음.
    #
    #         # previous_scroll_h = driver.execute_script("return document.body.scrollHeight")
    #         current_scroll_h = driver.execute_script("return document.documentElement.scrollHeight")
    #
    #         scroll_cnt = scroll_cnt + 1
    #
    #         DebuggingUtil.print_ment_via_colorama(f'{scroll_cnt}번 째 스크롤 성공 previous_scroll_h : {previous_scroll_h} current_scroll_h : {current_scroll_h}   previous_scroll_h==current_scroll_h : {previous_scroll_h == current_scroll_h}', colorama_color=ColoramaUtil.BLUE)
    #
    #     page_src = driver.page_source
    #     soup = BeautifulSoup(page_src, "lxml")
    #     driver.close()
    #
    #     # 모든 태그 가져오기
    #     # tags = soup.find_all()
    #     # for tag in tags:
    #     #     print(tag)
    #
    #     # body 리소스 확인 : success
    #     # bodys = soup.find_all("body")
    #     # for body in bodys:
    #     #     print(f"body:{body}")
    #
    #     # 이미지 태그 크롤링
    #     # images = soup.find_all("img")
    #     # for img in images:
    #     #     img_url = img.get("src")
    #     #     print("Image URL:", img_url)
    #     #
    #     # 스크립트 태그 크롤링
    #     # scripts = soup.find_all("script")
    #     # for script in scripts:
    #     #     script_url = script.get("src")
    #     #     print("Script URL:", script_url)
    #     #
    #     # # 스타일시트 크롤링
    #     # stylesheets = soup.find_all("link", rel="stylesheet")
    #     # for stylesheet in stylesheets:
    #     #     stylesheet_url = stylesheet.get("href")
    #     #     print("Stylesheet URL:", stylesheet_url)
    #
    #     # 특정 태그의 class 가 "어쩌구" 인
    #     # div_tags = soup.find_all("div", class_="어쩌구")
    #
    #     # a 태그 크롤링
    #     # a_tags = soup.find_all("a")
    #     # results = ""
    #     # for a_tag in a_tags:
    #     #     hrefs = a_tag.get("href")
    #     #     if hrefs != None and hrefs != "":
    #     #         # print("href", hrefs)
    #     #         results = f"{results}{hrefs}\n"
    #
    #     # 변수에 저장 via selector
    #     # name = soup.select('a#video-title')
    #     # video_url = soup.select('a#video-title')
    #     # view = soup.select('a#video-title')
    #
    #     # name, video_url 에 저장 via tag_name and id
    #     name = soup.find_all("a", id="video-title")
    #     video_url = soup.find_all("a", id="video-title")
    #
    #     # 유튜브 주소 크롤링 및 진행률 표시 via tqdm, 14 초나 걸리는데. 성능이 필요할때는 여러개의 thread 로 처리해보자.
    #     a_tags = soup.find_all("a")
    #
    #     # success
    #     # DebuggingUtil.print_magenta_as_gui(f"{len(a_tags)}")
    #
    #     # results를 str으로 처리
    #     # results = ""
    #     # a_tags_cnt  = 0
    #     # with tqdm(total=total_percent,ncols = 79 , desc= "웹 크롤링 진행률") as process_bar:
    #     #     for a_tag in a_tags:
    #     #         hrefs = a_tag.get("href")
    #     #         if hrefs != None and hrefs != "" and "/watch?v=" in hrefs :
    #     #             if hrefs not in results:
    #     #                 results = f"{results}{hrefs}\n"
    #     #                 a_tags_cnt = a_tags_cnt + 1
    #     #         time.sleep(0.06)
    #     #         process_bar.update(total_percent/len(a_tags))
    #
    #     # fail
    #     # if process_bar.total == 90:
    #     #     TextToSpeechUtil.speak_ments(ment='웹 크롤링이 90퍼센트 진행되었습니다. 잠시만 기다려주세요', sleep_after_play=0.65)
    #
    #     # results를 list 으로 처리, list 으로만 처리하고 str 으로 변형하는 처리를 추가했는데 3초나 빨라졌다. 항상 list 로 처리를 하자.
    #     results = []
    #     a_tags_cnt = 0
    #     for a_tag in a_tags:
    #         hrefs = a_tag.get("href")
    #         if hrefs != None and hrefs != "" and "/watch?v=" in hrefs:
    #             if hrefs not in results:
    #                 results.append(hrefs)
    #                 a_tags_cnt = a_tags_cnt + 1
    #     results = DataStructureUtil.add_prefix_to_string_list(results, 'https://www.youtube.com')  # string list 의 요소마다 suffix 추가
    #     results = "\n".join(results)  # list to str
    #
    #     # fail
    #     # dialog = UiUtil.CustomQdialog(title=f"크롤링결과보고", ment=f"{results}", btns=[MentsUtil.YES], auto_click_positive_btn_after_seconds="")
    #     # dialog.exec()
    #
    #     # fail
    #     # UiUtil.pop_up_as_complete(title="크롤링결과보고", ment=f"{results}")
    #
    #     # success
    #     # DebuggingUtil.print_magenta_as_gui(f"{results}") # 테스트용 팝업    UiUtil 로 옮기는 게 나을 지 고민 중이다.
    #
    #     # success
    #     # 비동기로 진행 가능
    #     global dialog
    #     dialog = UiUtil.CustomQdialog(title=f"크롤링결과보고", ment=f"({a_tags_cnt}개 추출됨)\n\n{results}")
    #     dialog.show()
    #
    # @staticmethod
    # def crawl_youtube_video_title_and_url(url: str):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #
    #     # url 전처리
    #     url = url.strip()
    #
    #     # driver 설정
    #     total_percent = 100
    #     driver = SeleniumUtil.get_driver_for_selenium()
    #     with tqdm(total=total_percent, ncols=79, desc="driver 설정 진행률") as process_bar:
    #         global title
    #         title = 'html  href 크롤링 결과'
    #         target_url = url
    #         driver.get(target_url)
    #         page_src = driver.page_source
    #         soup = BeautifulSoup(page_src, "lxml")
    #         time.sleep(0.0001)
    #         process_bar.update(total_percent)
    #     driver.close()
    #
    #     # 변수에 저장 via tag_name and id
    #     name = soup.find_all("a", id="video-title")
    #     video_url = soup.find_all("a", id="video-title")
    #
    #     # list 에 저장
    #     name_list = []
    #     url_list = []
    #     # view_list = []
    #     for i in range(len(name)):
    #         name_list.append(name[i].text.strip())
    #         # view_list.append(view[i].get('aria-label').split()[-1])
    #     for i in video_url:
    #         url_list.append('{}{}'.format('https://www.youtube.com', i.get('href')))
    #
    #     # dict 에 저장
    #     # youtubeDic = {
    #     #     '제목': name_list,
    #     #     '주소': url_list,
    #     #     # '조회수': view_list
    #     # }
    #
    #     # csv 에 저장
    #     # import pandas as pd
    #     # youtubeDf = pd.DataFrame(youtubeDic)
    #     # youtubeDf.to_csv(f'{keyword}.csv', encoding='', index=False)
    #
    #     # str 에 저장
    #     results_list = []
    #     for index, url in enumerate(url_list):
    #         results_list.append(f"{name_list[index]}   {url_list[index]}")
    #     results_str = "\n".join(results_list)  # list to str
    #
    #     # fail
    #     # dialog = UiUtil.CustomQdialog(title=f"크롤링결과보고", ment=f"{results}", btns=[MentsUtil.YES], auto_click_positive_btn_after_seconds="")
    #     # dialog.exec()
    #
    #     # fail
    #     # UiUtil.pop_up_as_complete(title="크롤링결과보고", ment=f"{results}")
    #
    #     # success
    #     # DebuggingUtil.print_magenta_as_gui(f"{results}") # 테스트용 팝업    UiUtil 로 옮기는 게 나을 지 고민 중이다.
    #
    #     # success
    #     # 비동기로 진행 가능
    #     global dialog
    #     dialog = UiUtil.CustomQdialog(title=f"크롤링결과보고", ment=f"({len(url_list)}개 url 추출됨)\n\n{results_str}")
    #     dialog.show()
    #
    # @staticmethod
    # def crawl_youtube_playlist(url: str):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #
    #     # url 전처리
    #     url = url.strip()
    #
    #     # driver 설정
    #     total_percent = 100
    #     driver = SeleniumUtil.get_driver_for_selenium()
    #     with tqdm(total=total_percent, ncols=79, desc="driver 설정 진행률") as process_bar:
    #         global title
    #         title = 'html  href 크롤링 결과'
    #         target_url = url
    #         driver.get(target_url)
    #         page_src = driver.page_source
    #         soup = BeautifulSoup(page_src, "lxml")
    #         time.sleep(0.0001)
    #         process_bar.update(total_percent)
    #     driver.close()
    #
    #     # 변수에 저장 via tag_name and href
    #     names = soup.find_all("a", id="video-title")
    #     hrefs = soup.find_all("a", id="video-title")
    #     # hrefs = copy.deepcopy(names)
    #
    #     # list 에 저장
    #     name_list = []
    #     hrefs_list = []
    #     # view_list = []
    #     for i in range(len(names)):
    #         name_list.append(names[i].text.strip())
    #         # view_list.append(view[i].get('aria-label').split()[-1])
    #     for i in hrefs:
    #         hrefs_list.append('{}{}'.format('https://www.youtube.com', i.get('href')))
    #
    #     # str 에 저장
    #     results_list = []
    #     for index, url in enumerate(hrefs_list):
    #         # results_list.append(f"{name_list[index]}   {hrefs_list[index]}")
    #         results_list.append(f"{hrefs_list[index]}")  # href 만 출력
    #     results_str = "\n".join(results_list)  # list to str
    #
    #     # fail
    #     # dialog = UiUtil.CustomQdialog(title=f"크롤링결과보고", ment=f"{results}", btns=[MentsUtil.YES], auto_click_positive_btn_after_seconds="")
    #     # dialog.exec()
    #
    #     # fail
    #     # UiUtil.pop_up_as_complete(title="크롤링결과보고", ment=f"{results}")
    #
    #     # success
    #     # DebuggingUtil.print_magenta_as_gui(f"{results}") # 테스트용 팝업    UiUtil 로 옮기는 게 나을 지 고민 중이다.
    #
    #     # success
    #     # 비동기로 진행 가능
    #     global dialog
    #     dialog = UiUtil.CustomQdialog(title=f"크롤링결과보고", ment=f"({len(hrefs_list)}개 playlist 추출됨)\n\n{results_str}")
    #     dialog.show()
    #
    # @staticmethod
    # def should_i_crawl_a_tag_href():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #
    #         # 테스트용
    #         # url = "https://www.youtube.com/@blahblah/videos"
    #         url = ""
    #
    #         dialog = UiUtil.CustomQdialog(ment="해당 페이지의 href 를 크롤링할까요?", btns=[MentsUtil.YES, MentsUtil.NO], is_input_box=True, input_box_text_default=url)
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #
    #         if btn_text_clicked == MentsUtil.YES:
    #             BusinessLogicUtil.crawl_html_href(url=dialog.input_box.text())
    #             break
    #         else:
    #             break
    #
    # @staticmethod
    # def should_i_crawl_youtube_video_title_and_url():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #
    #         # 테스트용
    #         keyword = 'blahblah'
    #         url = f'https://www.youtube.com/results?search_query={keyword}'
    #
    #         dialog = UiUtil.CustomQdialog(ment="해당 페이지의 video title, video url을 크롤링할까요?", btns=[MentsUtil.YES, MentsUtil.NO], is_input_box=True, input_box_text_default=url)
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #
    #         if btn_text_clicked == MentsUtil.YES:
    #             BusinessLogicUtil.crawl_youtube_video_title_and_url(url=dialog.input_box.text())
    #             break
    #         else:
    #             break
    #
    # @staticmethod
    # def should_i_crawl_youtube_playlist():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #         # 테스트용
    #         keyword = 'blahblah'
    #         url = f'https://www.youtube.com/@{keyword}/playlists'
    #
    #         dialog = UiUtil.CustomQdialog(ment="해당 페이지의 video title, video url을 크롤링할까요?", btns=[MentsUtil.YES, MentsUtil.NO], is_input_box=True, input_box_text_default=url)
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #
    #         if btn_text_clicked == MentsUtil.YES:
    #             BusinessLogicUtil.crawl_youtube_playlist(url=dialog.input_box.text())
    #             break
    #         else:
    #             break
    #
    # @staticmethod
    # def print_json_via_jq_pkg(json_str=None, json_file=None, json_list=None):
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     if FileSystemUtil.is_os_windows():
    #         if BusinessLogicUtil.get_none_count_of_list([json_str, json_file, json_list]) == 2:  # 2개가 NONE이면 1나는 BINDING 된것으로 판단하는 로직
    #             if json_str != None:
    #                 # lines = FileSystemUtil.get_cmd_output(cmd=rf'echo "{json_str}" | {StateManagementUtil.JQ_WIN64_EXE} "."') # 나오긴 하는데 한줄로 나온다
    #                 # lines = FileSystemUtil.get_cmd_output(cmd=rf'echo "{json_str}" | python -mjson.tool ')# 나오긴 하는데 한줄로 나온다
    #                 # [DebuggingUtil.print_as_success(line) for line in lines]
    #                 json_str = json.dumps(json_str, indent=4)  # json.dumps() 함수는 JSON 데이터를 문자열로 변환하는 함수이며, indent 매개변수를 사용하여 들여쓰기를 설정하여 json 형태의 dict 를 예쁘게 출력할 수 있습니다.
    #                 DebuggingUtil.print_ment_light_white(json_str)
    #             if json_file != None:
    #                 lines = FileSystemUtil.get_cmd_output(cmd=rf"type {json_file} | {StateManagementUtil.JQ_WIN64_EXE} ")
    #                 [DebuggingUtil.print_ment_light_white(line) for line in lines]
    #             if json_list != None:
    #                 lines = FileSystemUtil.get_cmd_output(cmd=rf'echo "{json_list}" | "{StateManagementUtil.JQ_WIN64_EXE}" ')
    #                 [DebuggingUtil.print_ment_light_white(line) for line in lines]
    #         else:
    #             DebuggingUtil.print_ment_fail(ment=rf"{inspect.currentframe().f_code.co_name}() 를 사용하려면 json_str/json_file/json_list 파라미터들 중 둘 중 하나만 데이터바인딩이 되어야합니다")
    #     else:
    #         print("리눅스 시스템에서 아직 지원되지 않는 함수입니다")
    #
    # @staticmethod
    # def should_i_explorer():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #         dialog = UiUtil.CustomQdialog(ment="해당위치의 타겟을 실행할까요?", btns=[MentsUtil.YES, MentsUtil.NO], is_input_box=True, input_box_text_default=clipboard.paste())
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #         input_box_text = dialog.input_box.text()
    #         if btn_text_clicked == MentsUtil.YES:
    #             FileSystemUtil.explorer(target_abspath=input_box_text)
    #             break
    #         else:
    #             break

    # @staticmethod
    # def should_i_sync():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     # text_promised = clipboard.paste()
    #     text_promised = StateManagementUtil.SERVICE_DIRECTORY
    #     while True:
    #         dialog = UiUtil.CustomQdialog(ment="해당위치의 타겟을 싱크할까요?", btns=[MentsUtil.YES, MentsUtil.NO], is_input_box=True, input_box_text_default=text_promised)
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #
    #         # dialog_input_box_text 에 데이터 저장
    #         dialog_input_box_text = dialog.input_box.text()
    #
    #         # dialog_input_box_text 데이터 전처리
    #         # "  C:\projects\services  " -> "C:\projects\services"
    #         dialog_input_box_text = dialog_input_box_text.strip()
    #         # "C:\projects\services" -> C:\projects\services
    #         if dialog_input_box_text.startswith("\""):
    #             if dialog_input_box_text.endswith("\""):
    #                 dialog_input_box_text = dialog_input_box_text.replace("\"", "", 1)
    #                 # dialog_input_box_text = dialog_input_box_text[:-(len("\""))] + "${add suffix test}" # 이코드는 add suffix 만들 때 활용하자
    #                 dialog_input_box_text = dialog_input_box_text[:-(len("\""))] + ""
    #
    #         target_abspath = dialog_input_box_text
    #         target_abspath_sync = rf"{target_abspath}_sync"
    #         target_abspath_sync_zip = rf"{target_abspath}_sync.zip"
    #
    #         if btn_text_clicked == MentsUtil.YES:
    #             FileSystemUtil.sync_directory_local(target_abspath=target_abspath)
    #             break
    #         else:
    #             break

    # @staticmethod
    # def should_i_sync_V_0_0_2():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     # text_promised = clipboard.paste()
    #     text_promised = StateManagementUtil.SERVICE_DIRECTORY
    #     while True:
    #         dialog = UiUtil.CustomQdialog(ment="해당위치의 타겟을 싱크할까요?", btns=[MentsUtil.YES, MentsUtil.NO], is_input_box=True, input_box_text_default=text_promised)
    #         dialog.exec()
    #         btn_text_clicked = dialog.btn_text_clicked
    #
    #         # dialog_input_box_text 에 데이터 저장
    #         dialog_input_box_text = dialog.input_box.text()
    #         dialog_input_box_text = dialog_input_box_text.strip()
    #         if dialog_input_box_text.startswith("\""):
    #             if dialog_input_box_text.endswith("\""):
    #                 dialog_input_box_text = dialog_input_box_text.replace("\"", "", 1)
    #                 dialog_input_box_text = dialog_input_box_text[:-(len("\""))] + ""
    #
    #         target_abspath = dialog_input_box_text
    #         target_abspath_sync = rf"{target_abspath}_sync"
    #         target_abspath_sync_zip = rf"{target_abspath}_sync.zip"
    #
    #         if btn_text_clicked == MentsUtil.YES:
    #             BusinessLogicUtil.upzip_target(target_abspath=target_abspath_sync_zip)
    #             FileSystemUtil.sync_directory_local(target_abspath=target_abspath)
    #             BusinessLogicUtil.back_up_target_without_timestamp(target_abspath=target_abspath_sync)
    #             break
    #         else:
    #             break

    # @staticmethod
    # def download_video_from_web2():
    #     DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
    #     while True:
    #         file_png = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_png\download_video_via_chrome_extensions1.png"
    #         is_image_finded = BusinessLogicUtil.click_center_of_img_recognized_by_mouse_left(img_abspath=file_png, recognize_loop_limit_cnt=100)
    #         if is_image_finded:
    #             BusinessLogicUtil.sleep(30)
    #             BusinessLogicUtil.press("ctrl", "f")
    #             BusinessLogicUtil.press("end")
    #             BusinessLogicUtil.press("ctrl", "a")
    #             BusinessLogicUtil.press("backspace")
    #             BusinessLogicUtil.write_fast("save")
    #             BusinessLogicUtil.press("enter")
    #             BusinessLogicUtil.press("enter")
    #             BusinessLogicUtil.press("esc")
    #             BusinessLogicUtil.press("enter")
    #             file_png = rf"{StateManagementUtil.PROJECT_DIRECTORY}\$cache_png\download_video_via_chrome_extensions2.png"
    #             is_image_finded = BusinessLogicUtil.click_center_of_img_recognized_by_mouse_left(img_abspath=file_png, recognize_loop_limit_cnt=100)
    #             if is_image_finded:
    #                 BusinessLogicUtil.press("shift", "w")
    #             else:
    #                 TextToSpeechUtil.speak_ments(ment="이미지를 찾을 수 없어 해당 자동화 기능을 마저 진행할 수 없습니다", sleep_after_play=0.65)
    #         else:
    #             TextToSpeechUtil.speak_ments(ment="이미지를 찾을 수 없어 해당 자동화 기능을 마저 진행할 수 없습니다", sleep_after_play=0.65)
    #         break
    #     pass

    @staticmethod
    def gather_empty_directory(target_abspath: str):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        # Park4139Park4139.Tts.speak("타겟경로를 순회하며 리프 디렉토리만 약속된 폴더로 모읍니다")
        # dst = rf"D:\[noe] [8TB] [ext]\$leaf_directories"
        # Park4139.make_leaf_directory(dst)
        # if os.path.isdir(target_abspath):
        #     for dirpath, dirnames, filenames in os.walk(target_abspath, topdown=False):
        #         if not dirnames and not filenames:
        #             # 하위디렉토리도 없고, 파일도 없는 경우만 leaf directory 로 간주
        #             print(f"leaf_directory : {dirpath}")
        #             Park4139.move_without_overwrite(src=dirpath, dst=dst)
        #     if Park4139.is_empty_directory(target_abspath) == True:
        #         Park4139.move_without_overwrite(src=target_abspath, dst=dst)

        # 빈 디렉토리 제거
        FileSystemUtil.make_leaf_directory(leaf_directory_abspath=StateManagementUtil.EMPTY_DIRECTORYIES)
        if os.path.isdir(target_abspath):
            for dirpath, dirnames, filenames in os.walk(target_abspath, topdown=False):
                if not dirnames and not filenames:
                    if os.path.isdir(dirpath):
                        FileSystemUtil.is_empty_directory(dirpath)
                        FileSystemUtil.move_target_without_overwrite(target_abspath=dirpath, dst=StateManagementUtil.EMPTY_DIRECTORYIES)
        DebuggingUtil.print_ment_via_colorama(ment=rf"빈디렉토리제거성공", colorama_color=ColoramaUtil.LIGHTCYAN_EX)

        # 빈 트리 리프디렉토리별로 해체한 뒤 제거
        for index, directory in enumerate(target_abspath):
            if FileSystemUtil.is_empty_tree(directory):
                for root, directories, files in os.walk(directory, topdown=True):
                    for directory in directories:
                        directory = os.path.abspath(os.path.join(root, directory))
                        if FileSystemUtil.is_leaf_directory(directory):
                            FileSystemUtil.move_target_without_overwrite(target_abspath=directory, dst=StateManagementUtil.EMPTY_DIRECTORYIES)
                            DebuggingUtil.print_ment_via_colorama(rf'directory : {directory}', colorama_color=ColoramaUtil.LIGHTWHITE_EX)
        DebuggingUtil.print_ment_via_colorama(ment=rf"빈트리제거성공", colorama_color=ColoramaUtil.LIGHTCYAN_EX)

        # 빈 디렉토리 제거
        FileSystemUtil.make_leaf_directory(leaf_directory_abspath=StateManagementUtil.EMPTY_DIRECTORYIES)
        if os.path.isdir(target_abspath):
            for dirpath, dirnames, filenames in os.walk(target_abspath, topdown=False):
                if not dirnames and not filenames:
                    if os.path.isdir(dirpath):
                        FileSystemUtil.is_empty_directory(dirpath)
                        FileSystemUtil.move_target_without_overwrite(target_abspath=dirpath, dst=StateManagementUtil.EMPTY_DIRECTORYIES)

    @staticmethod
    def is_containing_special_characters(text: str, ignore_list: [str] = None):
        pattern = "[~!@#$%^&*()_+|<>?:{}]"  # , 는 제외인가?
        if ignore_list is not None:
            for exception in ignore_list:
                pattern = pattern.replace(exception, "")
        if re.search(pattern, text):
            return True

    @staticmethod
    def get_kor_from_eng(english_word: str):
        translating_dictionary = {
            'name': "이름",
            "id": "아이디",
            "pw": "패스워드",
            "e mail": "이메일",
            'phone_no': "휴대폰번호",
            'e_mail': "이메일",
            'pw2': "비밀번호 확인",
            'address': "주소",
            'birthday': "생년월일",
            'writer': "작성자",
            'title': "제목",
            'contents': "내용",
        }
        result = ""
        try:
            result = translating_dictionary[english_word]
        except:
            result = english_word
        return result

    @staticmethod
    def raise_exception_after_special_charcater_check(value, inspect_currentframe_f_code_co_name, ignore_list: [str] = None):
        DebuggingUtil.commentize(f"{inspect.currentframe().f_code.co_name}()")
        if BusinessLogicUtil.is_containing_special_characters(value, ignore_list):
            word_english = inspect_currentframe_f_code_co_name
            word_english = word_english.replace('validate_', "")
            word_english = word_english.replace("_", " ")
            word_english = word_english.strip()
            word_korean = BusinessLogicUtil.get_kor_from_eng(english_word=word_english)
            ment = f"유효한 {word_korean}이(가) 아닙니다. 특수문자가 없어야 합니다 {value}"
            raise HTTPException(status_code=400, detail=ment)

    @staticmethod
    def merge_excel_files(dir_path):
        function_name = inspect.currentframe().f_code.co_name
        DebuggingUtil.commentize(f"{function_name}()")
        try:
            import openpyxl  # pip install openpyxl
            import os
            import pandas as pd
            from openpyxl.styles import Font, Alignment, PatternFill, Color, Border, Side

            print(rf'''dir_path : {dir_path}''')

            # xls 에서 xlsx로 변환 # openpyxl 은 xls 지원안함.  xlsx 가 더 최신기술.
            file_to_merge_ext = ".xls"
            files = [rf"{dir_path}/{file}" for file in os.listdir(dir_path) if file_to_merge_ext in FileSystemUtil.get_target_as_x(file)]
            for file in files:
                FileSystemUtil.convert_xls_to_xlsx(file)

            # 합병할 파일의 목록을 files_to_merge 에 저장
            file_to_merge_ext = ".xlsx"
            files_to_merge = [f"{dir_path}/{file}" for file in os.listdir(dir_path) if file_to_merge_ext in FileSystemUtil.get_target_as_x(file)]
            [print(sample) for sample in files_to_merge]
            print(rf'type(file_list) : {type(files_to_merge)}')
            print(rf'len(file_list) : {len(files_to_merge)}')

            # 파일합병할 작업공간 제어
            # wb_new = openpyxl.Workbook()
            # ws1 = wb_new.active
            # ws2 = wb_new.create_sheet("result")
            # files_cnt = len(files_to_merge)

            # 엑셀 병합 및 저장
            merged_file = StateManagementUtil.FILE_MERGED_EXCEL_XLSX
            merged_cnt = 0
            merged_df = pd.DataFrame()
            for file_path in files_to_merge:
                # df = pd.read_excel(file_path, engine = "openpyxl")  # 엑셀 파일 읽기
                # df = pd.read_excel(file_path, engine = "openpyxl", header=0, usecols = [0, 1, 2,3])  # fail, sheet_name="Sheet1"  여러 시트가 있을 경우 시트명을 직접 입력하여 dataframe화 # usecols = [0, 2]  컬럼선택
                # df = pd.read_excel(file_path, engine = "openpyxl", header=0, usecols = [1, 2,3])  # fail,   sheet_name="Sheet1"  여러 시트가 있을 경우 시트명을 직접 입력하여 dataframe화 # usecols = [0, 2]  컬럼선택
                df = pd.read_excel(file_path, engine="openpyxl")  # success, 근데 sheet1 만 되고 sheet2 는 무시 된다.

                # df = df.iloc[1:]  # 첫줄 제거, 첫줄 을 제외한 나머지 데이터 선택

                # df = df.iloc[:-1]  # 마지막줄 제거, 마지막줄 제외한 나머지 데이터 선택

                merged_df = pd.concat([merged_df, df], ignore_index=True)  # 두 데이터프레임 병합 # df 병합

                merged_cnt = merged_cnt + 1

            print(rf'''merge_files_cnt : {merged_cnt + 1}''')
            print(rf'''merged_df : ''')
            print(rf'''{merged_df}''')
            print(rf'''merged_cnt : {merged_cnt}''')
            print(rf'''merged_file : {merged_file}''')
            try:
                merged_df.to_excel(merged_file)  # success
                # pd.ExcelWriter(merged_file, engine= "openpyxl") # fail, 확장자 잘못 저장했나?
                # FileSystemUtil.explorer(merged_file)
            except PermissionError:
                DebuggingUtil.print_ment_fail(f"{function_name}(), fail, 엑셀파일이 열려있을 수 있습니다. 닫고 머지를 다시 시도해 주세요")
            except Exception as e:
                DebuggingUtil.print_ment_fail(f"{function_name}(), fail, \n {traceback.format_exc()}")
            DebuggingUtil.print_ment_success(f"{function_name}(), success")
        except:
            DebuggingUtil.trouble_shoot(trouble_id="%%%FOO%%%")

    @staticmethod
    def get_random_math_expression_english_input_for_code_test():
        random_cnt = random.randint(0, 10)
        param = random.choice(["ONE", "TWO", "THREE", "FOUR", "FIVE", "SIX", "SEVEN", "EIGHT", "NINE"])
        for i in range(0, random_cnt):
            random_num = random.choice(["ONE", "TWO", "THREE", "FOUR", "FIVE", "SIX", "SEVEN", "EIGHT", "NINE"])
            random_operand = random.choice(["PLUS", "MINUS", "TIMES"])
            nbsp = " "
            param += f"{nbsp}{random_operand}{nbsp}{random_num}"
        DebuggingUtil.print_magenta(rf'''param : {param}''')
        return param

    @staticmethod
    def convert_math_expression_english_to_math_expression_math(english_input):
        result = english_input.upper()
        result = result.replace("ONE", "1")
        result = result.replace("TWO", "2")
        result = result.replace("THREE", "3")
        result = result.replace("FOUR", "4")
        result = result.replace("FIVE", "5")
        result = result.replace("SIX", "6")
        result = result.replace("SEVEN", "7")
        result = result.replace("EIGHT", "8")
        result = result.replace("NINE", "9")
        result = result.replace("TIMES", "*")
        result = result.replace("PLUS", "+")
        result = result.replace("MINUS", "-")
        return result

    @staticmethod
    async def get_operation_result_of_parser_for_code_test(tokens: str):
        # dummy
        # tokens = '0 - 9 * 0 + 8 - 1 + 8 - 8 + 2 + 3 * 2'
        # tokens = '2 - 9 + 1 - 0 * 6 * 8'
        # tokens = '3 + 6 + 3 - 6 * 8 + 1'
        # tokens = '4 + 6 - 1 + 3 + 5 * 4 - 8 - 9 + 4 * 7'
        # tokens = '0 + 7 * 4 - 5 + 3 + 1 + 1 + 4 * 9'

        print(rf'''tokens : {tokens}''')

        tokens = tokens.split(" ")
        print(rf'''tokens : {tokens}''')

        # 연산
        for operand in ["*", "-", "+"]:  # 처리순서 고수해야 한다. tokens = '3 - 3 * 2 - 1 + 0 * 6 - 9 - 3 - 6 + 2' 샘플에서 문제확인됨
            while True:
                for index, token in enumerate(tokens):
                    if token == operand:
                        if tokens[index - 1] == None:
                            break
                        if token == "*":
                            result = int(tokens[index - 1]) * int(tokens[index + 1])
                        if token == "-":
                            result = int(tokens[index - 1]) - int(tokens[index + 1])
                        if token == "+":
                            result = int(tokens[index - 1]) + int(tokens[index + 1])
                        tokens[index - 1] = None
                        tokens[index] = result
                        tokens[index + 1] = None
                tokens_len = len(tokens)
                tokens = [token for token in tokens if token is not None]
                tokens_len2 = len(tokens)
                if tokens_len == tokens_len2:
                    break
                print(rf'''tokens : {tokens}''')
        print(rf'''tokens : {tokens}''')
        tokens = tokens[0]
        print(rf'''tokens : {tokens}''')
        return tokens

    @staticmethod
    def get_stock_name(ticker):
        df = MySqlUtil.execute(f"""select * from finance_stock_ticker where ticker="{ticker}" """)
        return df
    @staticmethod
    def get_ticker_by_search(stock_name: str):
        df1 = MySqlUtil.execute(f"""select * from finance_stock_ticker where stock_name like "%{stock_name}%" """)
        df2 = MySqlUtil.execute(f"""select * from finance_stock_ticker where stock_name like "%{stock_name.upper()}%" """)
        df3 = MySqlUtil.execute(f"""select * from finance_stock_ticker where stock_name like "%{stock_name.lower()}%" """)
        df4 = MySqlUtil.execute(f"""select * from finance_stock_ticker where stock_name like "%{stock_name.capitalize()}%" """)
        df_merged = pd.concat([df1, df2, df3, df4], ignore_index=True )  # df 세로병합
        df_dropped_duplication = df_merged.drop_duplicates()
        df_searched = df_dropped_duplication
        return df_searched
